import os
import csv
import requests
from dotenv import load_dotenv
import sqlite3
import datetime
import uuid
import random
import base64
import hashlib
import json
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, make_response
from werkzeug.utils import secure_filename
import openpyxl
from weasyprint import HTML, CSS

app = Flask(__name__)
app.secret_key = 'rrb-cbt-v104-secret-key-2024'
load_dotenv('apikey.env')
# Configuration
UPLOAD_FOLDER = 'uploads'
EXPORT_FOLDER = 'exports'
BRANDING_FOLDER = os.path.join('static', 'uploads', 'branding')
STUDENT_PIC_FOLDER = os.path.join('static', 'student_profile_pic')
TEACHER_PIC_FOLDER = os.path.join('static', 'Teacher_profile_picture')
ALLOWED_EXTENSIONS = {'csv'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)
os.makedirs(os.path.join('static', 'uploads'), exist_ok=True)
os.makedirs(BRANDING_FOLDER, exist_ok=True)
os.makedirs(STUDENT_PIC_FOLDER, exist_ok=True)
os.makedirs(TEACHER_PIC_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXPORT_FOLDER'] = EXPORT_FOLDER
app.config['BRANDING_FOLDER'] = BRANDING_FOLDER
app.config['STUDENT_PIC_FOLDER'] = STUDENT_PIC_FOLDER
app.config['TEACHER_PIC_FOLDER'] = TEACHER_PIC_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# ── helpers ──────────────────────────────────────────────
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    return hash_password(password) == hashed

def save_student_picture(file, student_id):
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"student_{student_id}.{ext}"
    filepath = os.path.join(STUDENT_PIC_FOLDER, filename)
    # Remove old picture if different extension
    for old_ext in ALLOWED_IMAGE_EXTENSIONS:
        old_path = os.path.join(STUDENT_PIC_FOLDER, f"student_{student_id}.{old_ext}")
        if os.path.exists(old_path) and old_path != filepath:
            os.remove(old_path)
    file.save(filepath)
    return os.path.join('student_profile_pic', filename).replace('\\', '/')

def save_teacher_picture(file, teacher_id):
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"teacher_{teacher_id}.{ext}"
    filepath = os.path.join(TEACHER_PIC_FOLDER, filename)
    for old_ext in ALLOWED_IMAGE_EXTENSIONS:
        old_path = os.path.join(TEACHER_PIC_FOLDER, f"teacher_{teacher_id}.{old_ext}")
        if os.path.exists(old_path) and old_path != filepath:
            os.remove(old_path)
    file.save(filepath)
    return os.path.join('Teacher_profile_picture', filename).replace('\\', '/')

def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def save_question_image(file, class_, subject, question_id):
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None

    folder_name = f"{class_.strip()}_{subject.strip()}".replace(' ', '_')
    folder_path = os.path.join('static', 'uploads', folder_name)
    os.makedirs(folder_path, exist_ok=True)

    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{folder_name}_{question_id}.{ext}"
    filepath = os.path.join(folder_path, filename)

    if os.path.exists(filepath):
        os.remove(filepath)

    file.save(filepath)
    return os.path.join('uploads', folder_name, filename).replace('\\', '/')

# Database initialization
def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA busy_timeout = 10000")

    c.execute('''CREATE TABLE IF NOT EXISTS students
                 (student_id TEXT PRIMARY KEY, name TEXT, class TEXT, subject TEXT, ip TEXT,
                  status TEXT DEFAULT 'Not Started', exam_started_at TIMESTAMP,
                  admission_no TEXT, section TEXT, dob TEXT, house TEXT, 
                  parents_name TEXT, address TEXT, picture TEXT)''')
    c.execute("PRAGMA table_info(students)")
    columns = [col[1] for col in c.fetchall()]
    if 'exam_started_at' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN exam_started_at TIMESTAMP")
    if 'admission_no' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN admission_no TEXT")
    if 'section' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN section TEXT")
    if 'dob' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN dob TEXT")
    if 'house' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN house TEXT")
    if 'parents_name' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN parents_name TEXT")
    if 'address' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN address TEXT")
    if 'picture' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN picture TEXT")

    c.execute('''CREATE TABLE IF NOT EXISTS questions
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, class TEXT, subject TEXT, question TEXT,
                  option_a TEXT, option_b TEXT, option_c TEXT, option_d TEXT, correct_answer TEXT,
                  image_path TEXT, chapter TEXT)''')
        # Add chapter column if not exists
    c.execute("PRAGMA table_info(questions)")
    columns = [col[1] for col in c.fetchall()]
    if 'chapter' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN chapter TEXT")
    if 'class' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN class TEXT")
    if 'subject' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN subject TEXT")
    if 'image_path' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN image_path TEXT")
    if 'chapter' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN chapter TEXT")

    c.execute('''CREATE TABLE IF NOT EXISTS responses
                 (student_id TEXT, question_id INTEGER, selected_option TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (student_id, question_id))''')
    c.execute("PRAGMA table_info(responses)")
    resp_columns = [col[1] for col in c.fetchall()]
    if 'created_at' not in resp_columns:
        c.execute("ALTER TABLE responses ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")

    c.execute('''CREATE TABLE IF NOT EXISTS results
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id TEXT, name TEXT, 
                  class TEXT, subject TEXT, score INTEGER, total_questions INTEGER,
                  percentage REAL, test_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  chapter TEXT)''')
    # Migrate old results table if it has the old schema (student_id was PRIMARY KEY)
    c.execute("PRAGMA table_info(results)")
    res_columns = [col[1] for col in c.fetchall()]
    if 'id' not in res_columns:
        # Old schema – rename and recreate
        c.execute("ALTER TABLE results RENAME TO results_old")
        c.execute('''CREATE TABLE results
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id TEXT, name TEXT,
                      class TEXT, subject TEXT, score INTEGER, total_questions INTEGER,
                      percentage REAL, test_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                      chapter TEXT)''')
        c.execute("""INSERT INTO results (student_id, name, class, subject, score, total_questions, percentage)
                     SELECT student_id, name, class, subject, score, score, 0 FROM results_old""")
        c.execute("DROP TABLE results_old")
    c.execute('''CREATE TABLE IF NOT EXISTS exam_control
                 (id INTEGER PRIMARY KEY CHECK (id=1), is_active INTEGER DEFAULT 0,
                  start_time TIMESTAMP, duration INTEGER DEFAULT 60,
                  negative_marking INTEGER DEFAULT 0,
                  negative_value REAL DEFAULT 0.33)''')
    c.execute("INSERT OR IGNORE INTO exam_control (id, is_active, start_time, duration, negative_marking, negative_value) VALUES (1, 0, NULL, 60, 0, 0.33)")
    # Add columns if upgrading from older schema
    c.execute("PRAGMA table_info(exam_control)")
    ec_cols = [r[1] for r in c.fetchall()]
    if 'negative_marking' not in ec_cols:
        c.execute("ALTER TABLE exam_control ADD COLUMN negative_marking INTEGER DEFAULT 0")
    if 'negative_value' not in ec_cols:
        c.execute("ALTER TABLE exam_control ADD COLUMN negative_value REAL DEFAULT 0.33")

    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('school_name', 'RRB Group of Schools')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('logo_path', '')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('school_address', '')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('academic_session', '')")

    c.execute('''CREATE TABLE IF NOT EXISTS reattempt_requests
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                class TEXT NOT NULL,
                subject TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                reviewed_at TIMESTAMP,
                admin_note TEXT)''')

    # Table for storing shuffled question order and option mapping per student
    c.execute('''CREATE TABLE IF NOT EXISTS shuffled_questions
                 (student_id TEXT,
                  question_id INTEGER,
                  shuffled_index INTEGER,
                  option_order TEXT,
                  PRIMARY KEY (student_id, question_id))''')

    # Teachers table
    c.execute('''CREATE TABLE IF NOT EXISTS teachers
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  mobile TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  email TEXT,
                  address TEXT,
                  picture TEXT,
                  status TEXT DEFAULT 'active',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Teacher assignments (class and subject assignments)
    c.execute('''CREATE TABLE IF NOT EXISTS teacher_assignments
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id INTEGER NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  is_class_teacher INTEGER DEFAULT 0,
                  FOREIGN KEY (teacher_id) REFERENCES teachers(id))''')
    c.execute("PRAGMA table_info(teacher_assignments)")
    ta_cols = [col[1] for col in c.fetchall()]
    if 'section' not in ta_cols:
        c.execute("ALTER TABLE teacher_assignments ADD COLUMN section TEXT DEFAULT ''")

    # Test papers table (CSV uploads with naming convention)
    c.execute('''CREATE TABLE IF NOT EXISTS test_papers
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  filename TEXT NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  test_no TEXT NOT NULL,
                  uploaded_by TEXT NOT NULL,
                  uploader_type TEXT DEFAULT 'admin',
                  is_active INTEGER DEFAULT 0,
                  question_count INTEGER DEFAULT 0,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Bulletins / Announcements (Admin + Teacher can post; shown on student login page)
    c.execute('''CREATE TABLE IF NOT EXISTS bulletins
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  title TEXT NOT NULL,
                  content TEXT NOT NULL,
                  posted_by TEXT NOT NULL,
                  poster_type TEXT DEFAULT 'admin',
                  target_class TEXT DEFAULT '',
                  is_active INTEGER DEFAULT 1,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Test Generation History (for AI-generated tests)
    c.execute('''CREATE TABLE IF NOT EXISTS test_generation_history
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id TEXT NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  chapter TEXT NOT NULL,
                  test_no TEXT NOT NULL,
                  output_mode TEXT DEFAULT 'cbt',
                  total_questions INTEGER DEFAULT 0,
                  mcq_count INTEGER DEFAULT 0,
                  assertion_count INTEGER DEFAULT 0,
                  very_short_count INTEGER DEFAULT 0,
                  short_count INTEGER DEFAULT 0,
                  long_count INTEGER DEFAULT 0,
                  case_study_count INTEGER DEFAULT 0,
                  remark TEXT DEFAULT '',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Add question_type column to questions if not exists
    c.execute("PRAGMA table_info(questions)")
    q_cols = [col[1] for col in c.fetchall()]
    if 'question_type' not in q_cols:
        c.execute("ALTER TABLE questions ADD COLUMN question_type TEXT DEFAULT 'MCQ'")
    if 'test_no' not in q_cols:
        c.execute("ALTER TABLE questions ADD COLUMN test_no TEXT DEFAULT ''")

    # Student class lock: once registered, class/section is locked
    c.execute('''CREATE TABLE IF NOT EXISTS student_class_lock
                 (student_id TEXT PRIMARY KEY,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  locked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Combined / Multi-Subject Test entities
    c.execute('''CREATE TABLE IF NOT EXISTS combined_tests
                 (id           INTEGER PRIMARY KEY AUTOINCREMENT,
                  test_no      TEXT NOT NULL,
                  class        TEXT NOT NULL,
                  section      TEXT DEFAULT '',
                  title        TEXT DEFAULT '',
                  created_by   TEXT NOT NULL,
                  creator_type TEXT DEFAULT 'teacher',
                  is_active    INTEGER DEFAULT 1,
                  created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    c.execute('''CREATE TABLE IF NOT EXISTS combined_test_subjects
                 (id               INTEGER PRIMARY KEY AUTOINCREMENT,
                  combined_test_id INTEGER NOT NULL,
                  subject          TEXT NOT NULL,
                  question_count   INTEGER DEFAULT 0,
                  FOREIGN KEY (combined_test_id) REFERENCES combined_tests(id) ON DELETE CASCADE)''')

    conn.commit()
    conn.close()

def reset_exam_on_startup():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute("UPDATE exam_control SET is_active=0, start_time=NULL WHERE id=1")
    conn.commit()
    conn.close()

init_db()
reset_exam_on_startup()

def get_db():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

def get_setting(key, default=''):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row['value'] if row else default

def set_setting(key, value):
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('student_id'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('teacher_logged_in'):
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated_function

# -------------------------------
# Student Routes
# -------------------------------
@app.route('/')
def index():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class != '' ORDER BY class, subject")
    rows = c.fetchall()

    class_subject_map = {}
    for row in rows:
        cls = row['class']
        sub = row['subject']
        if cls not in class_subject_map:
            class_subject_map[cls] = []
        class_subject_map[cls].append(sub)

    # Get active bulletins for the student login page bulletin board
    c.execute("""SELECT b.title, b.content, b.poster_type, b.target_class, b.created_at,
                        COALESCE(t.name, 'Admin') as poster_name
                 FROM bulletins b
                 LEFT JOIN teachers t ON b.posted_by = t.id AND b.poster_type = 'teacher'
                 WHERE b.is_active = 1
                 ORDER BY b.created_at DESC LIMIT 10""")
    bulletins = [dict(r) for r in c.fetchall()]

    # Get class teachers for display
    c.execute("""SELECT ta.class, ta.section, t.name, t.picture
                 FROM teacher_assignments ta
                 JOIN teachers t ON ta.teacher_id = t.id
                 WHERE ta.is_class_teacher = 1
                 ORDER BY ta.class, ta.section""")
    class_teachers = [dict(r) for r in c.fetchall()]

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    conn.close()

    return render_template('index.html',
                           class_subject_map=class_subject_map,
                           bulletins=bulletins,
                           class_teachers=class_teachers,
                           school_name=school_name,
                           logo_path=logo_path)

@app.route('/student_login', methods=['POST'])
def student_login():
    auto_submit_expired_exams()

    name        = request.form.get('name', '').strip()
    student_id  = request.form.get('student_id', '').strip()
    class_      = request.form.get('class', '').strip()
    section     = request.form.get('section', '').strip()
    subject     = request.form.get('subject', '').strip()
    test_no     = request.form.get('test_no', '').strip()   # Feature #4
    ip          = request.remote_addr

    if not all([name, student_id, class_, subject]):
        return "All fields required", 400

    conn = get_db()
    c = conn.cursor()

    # ── CLASS LOCK CHECK ──────────────────────────────────────────
    c.execute("SELECT class, section FROM student_class_lock WHERE student_id=?", (student_id,))
    lock = c.fetchone()
    if lock:
        if lock['class'] != class_ or lock['section'] != section:
            conn.close()
            return render_template('index.html',
                                   error=f"You are locked to Class {lock['class']} Section {lock['section']}. Contact Admin to change.",
                                   class_subject_map=_get_class_subject_map(),
                                   bulletins=[], class_teachers=[],
                                   school_name=get_setting('school_name','RRB Group of Schools'),
                                   logo_path=get_setting('logo_path',''))
    else:
        c.execute("INSERT OR IGNORE INTO student_class_lock (student_id, class, section) VALUES (?,?,?)",
                  (student_id, class_, section))

    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()

    # BUG-001 FIX: Only block if student submitted THIS specific test (same subject+test_no)
    # Different test_no = new exam, must never be blocked by a previous test's submission
    if student and student['status'] == 'Submitted':
        prev_subject = student['subject'] or ''
        # Check if previous submission was for the SAME test
        c.execute("""SELECT id FROM results
                     WHERE student_id=? AND subject=? AND (
                         (? != '' AND chapter=?) OR
                         (? = '' AND subject=?)
                     ) LIMIT 1""",
                  (student_id, subject,
                   test_no, test_no,
                   test_no, subject))
        same_test_result = c.fetchone()

        if same_test_result and prev_subject == subject:
            # Same test attempted again — show reattempt prompt
            conn.commit()
            conn.close()
            session.update({'student_id': student_id, 'student_name': name,
                            'class': class_, 'subject': subject, 'section': section, 'test_no': test_no})
            return redirect(url_for('submitted'))
        else:
            # Different test (new test_no or different subject) — reset status and allow
            c.execute("""UPDATE students SET status='Not Started', exam_started_at=NULL,
                          class=?, subject=?, section=?, ip=? WHERE student_id=?""",
                      (class_, subject, section, ip, student_id))
            c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
            conn.commit()

    if student:
        c.execute("UPDATE students SET name=?, class=?, subject=?, section=?, ip=? WHERE student_id=?",
                  (name, class_, subject, section, ip, student_id))
    else:
        c.execute("""INSERT INTO students (student_id, name, class, subject, section, ip, status)
                     VALUES (?,?,?,?,?,?,'Not Started')""",
                  (student_id, name, class_, subject, section, ip))
        conn.commit()
        c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
        student = c.fetchone()

    conn.commit()
    c.execute("SELECT status, subject FROM students WHERE student_id=?", (student_id,))
    updated_status = c.fetchone()
    if updated_status and updated_status['status'] == 'Submitted' and updated_status['subject'] == subject:
        conn.close()
        session.update({'student_id': student_id, 'student_name': name,
                        'class': class_, 'subject': subject, 'section': section, 'test_no': test_no})
        return redirect(url_for('submitted'))

    # Shuffle questions — filter by test_no if provided
    # For combined tests, detect and load ALL subjects
    combined_test_id = None
    combined_subjects = []

    if test_no:
        conn2 = get_db()
        c2    = conn2.cursor()
        # Check if this test_no is a combined test for this class
        c2.execute("""SELECT ct.id FROM combined_tests ct
                      WHERE ct.test_no=? AND ct.class=? AND ct.is_active=1""",
                   (test_no, class_))
        ct_row = c2.fetchone()
        if ct_row:
            combined_test_id = ct_row['id']
            c2.execute("""SELECT subject FROM combined_test_subjects
                          WHERE combined_test_id=? ORDER BY id""", (combined_test_id,))
            combined_subjects = [r['subject'] for r in c2.fetchall()]

            # Load ALL questions across all subjects for this combined test
            placeholders = ','.join(['?']*len(combined_subjects))
            c2.execute(f"""SELECT id FROM questions
                           WHERE class=? AND test_no=? AND subject IN ({placeholders})
                           ORDER BY subject, id""",
                       [class_, test_no] + combined_subjects)
            qids = [r['id'] for r in c2.fetchall()]
        else:
            c2.execute("SELECT id FROM questions WHERE class=? AND subject=? AND (test_no=? OR chapter=?) ORDER BY id",
                       (class_, subject, test_no, test_no))
            qids = [r['id'] for r in c2.fetchall()]
        conn2.close()
    else:
        conn_tmp = get_db()
        c_tmp    = conn_tmp.cursor()
        c_tmp.execute("SELECT id FROM questions WHERE class=? AND subject=? ORDER BY id", (class_, subject))
        qids = [r['id'] for r in c_tmp.fetchall()]
        conn_tmp.close()

    conn3 = get_db()
    c3    = conn3.cursor()
    random.shuffle(qids)
    c3.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    option_letters = ['A', 'B', 'C', 'D']
    for idx, qid in enumerate(qids):
        shuffled_opts = option_letters[:]
        random.shuffle(shuffled_opts)
        c3.execute("INSERT INTO shuffled_questions (student_id, question_id, shuffled_index, option_order) VALUES (?,?,?,?)",
                   (student_id, qid, idx, ''.join(shuffled_opts)))
    conn3.commit()
    conn3.close()

    session.update({'student_id': student_id, 'student_name': name,
                    'class': class_, 'subject': subject, 'section': section, 'test_no': test_no,
                    'combined_test_id': combined_test_id,
                    'combined_subjects': combined_subjects})
    return redirect(url_for('waiting'))

def _get_class_subject_map():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
    rows = c.fetchall()
    conn.close()
    csmap = {}
    for r in rows:
        csmap.setdefault(r['class'], []).append(r['subject'])
    return csmap

@app.route('/api/test_numbers')
def api_test_numbers():
    """Return distinct test_no values for a class+subject combo — includes combined tests."""
    class_   = request.args.get('class','').strip()
    subject  = request.args.get('subject','').strip()
    if not class_:
        return jsonify({'test_numbers': [], 'combined_tests': []})

    conn = get_db()
    c = conn.cursor()

    if subject:
        # Regular single-subject tests
        c.execute("""SELECT DISTINCT COALESCE(NULLIF(test_no,''), chapter) as tno
                     FROM questions
                     WHERE class=? AND subject=?
                       AND (test_no IS NOT NULL AND test_no != ''
                            OR chapter IS NOT NULL AND chapter != '')
                     ORDER BY tno""", (class_, subject))
        test_numbers = [r[0] for r in c.fetchall() if r[0]]
    else:
        test_numbers = []

    # Combined (multi-subject) tests for this class — shown as a separate group
    c.execute("""SELECT ct.id, ct.test_no, ct.title,
                        GROUP_CONCAT(cts.subject || '(' || cts.question_count || 'Q)', ', ') as subjects_info
                 FROM combined_tests ct
                 JOIN combined_test_subjects cts ON cts.combined_test_id = ct.id
                 WHERE ct.class=? AND ct.is_active=1
                 GROUP BY ct.id ORDER BY ct.created_at DESC""", (class_,))
    combined = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'test_numbers': test_numbers, 'combined_tests': combined})

@app.route('/api/generate_combined_test', methods=['POST'])
def generate_combined_test():
    """Generate a truly unified multi-subject test — one entity, multiple subjects."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    import urllib.request, json as json_lib

    class_   = request.form.get('class','').strip()
    section  = request.form.get('section','').strip()
    test_no  = request.form.get('test_no','').strip()
    title    = request.form.get('title','').strip() or test_no
    remark   = request.form.get('remark','').strip()

    # subjects_json: [{"subject":"Math","count":10}, ...]
    try:
        subjects = json_lib.loads(request.form.get('subjects_json','[]'))
    except Exception:
        return jsonify({'status':'error','message':'Invalid subjects data'}), 400

    if not class_ or not test_no or not subjects:
        return jsonify({'status':'error','message':'Class, Test No and at least one subject required'}), 400

    gemini_key = _get_gemini_key()
    if not gemini_key:
        return jsonify({'status':'error','message':'GEMINI_API_KEY not configured'}), 500

    conn = get_db()
    c = conn.cursor()

    # Check no combined test with same test_no+class already exists
    c.execute("SELECT id FROM combined_tests WHERE test_no=? AND class=?", (test_no, class_))
    if c.fetchone():
        conn.close()
        return jsonify({'status':'error',
                        'message':f'Combined test "{test_no}" already exists for Class {class_}. Choose a different Test No.'}), 400

    creator_id   = str(session.get('teacher_id','admin'))
    creator_type = 'teacher' if session.get('teacher_logged_in') else 'admin'

    # Create the combined_tests record first
    c.execute("""INSERT INTO combined_tests (test_no, class, section, title, created_by, creator_type)
                 VALUES (?,?,?,?,?,?)""",
             (test_no, class_, section, title, creator_id, creator_type))
    combined_id = c.lastrowid
    conn.commit()

    errors       = []
    total_inserted = 0

    for subj_entry in subjects:
        subject   = subj_entry.get('subject','').strip()
        mcq_count = int(subj_entry.get('count', 0))
        if not subject or mcq_count <= 0:
            continue

        # Build AI prompt for this subject
        prompt = _build_ai_prompt(class_, section, subject,
                                  f'{title} — {subject}', test_no,
                                  remark or 'Standard CBSE difficulty',
                                  [f"{mcq_count} MCQ (4 options, mark correct_answer as option_a/b/c/d)"])
        try:
            url     = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
            payload = json_lib.dumps({
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.7, "maxOutputTokens": 4096}
            }).encode('utf-8')
            req = urllib.request.Request(url, data=payload,
                                         headers={'Content-Type':'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json_lib.loads(resp.read().decode('utf-8'))

            raw = result['candidates'][0]['content']['parts'][0]['text'].strip()
            if '```' in raw:
                for part in raw.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): raw = part; break
            questions = json_lib.loads(raw)

            inserted = 0
            for q in questions:
                if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                    c.execute("""INSERT INTO questions
                               (class, subject, chapter, test_no, question_type,
                                question, option_a, option_b, option_c, option_d, correct_answer)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                             (class_, subject, title, test_no,
                              q.get('question_type','MCQ'), q.get('question',''),
                              q.get('option_a',''), q.get('option_b',''),
                              q.get('option_c',''), q.get('option_d',''),
                              q.get('correct_answer','')))
                    inserted += 1

            # Record subject in combined_test_subjects
            c.execute("""INSERT INTO combined_test_subjects (combined_test_id, subject, question_count)
                         VALUES (?,?,?)""", (combined_id, subject, inserted))
            total_inserted += inserted
            conn.commit()

        except Exception as e:
            errors.append(f"{subject}: {str(e)}")

    if errors and total_inserted == 0:
        # Complete failure — clean up
        c.execute("DELETE FROM combined_tests WHERE id=?", (combined_id,))
        conn.commit()
        conn.close()
        return jsonify({'status':'error','message':'Generation failed:\n' + '\n'.join(errors)}), 500

    conn.close()
    msg = f'Combined test "{test_no}" created with {total_inserted} questions across {len(subjects)} subjects.'
    if errors:
        msg += f'\nWarnings: ' + '; '.join(errors)
    return jsonify({'status':'success','message':msg,'combined_test_id':combined_id})

@app.route('/api/check_test_no')
def api_check_test_no():
    """Feature #5: Check if a test_no already exists for a subject (across all sections)."""
    subject = request.args.get('subject','').strip()
    test_no = request.args.get('test_no','').strip()
    class_  = request.args.get('class','').strip()
    if not subject or not test_no:
        return jsonify({'exists': False, 'details': []})
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT DISTINCT tp.class, tp.section, tp.test_no, t.name as teacher_name, tp.created_at
                 FROM test_papers tp
                 LEFT JOIN teachers t ON tp.uploaded_by = t.id AND tp.uploader_type='teacher'
                 WHERE tp.subject=? AND tp.test_no=? AND tp.class=?""",
             (subject, test_no, class_))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'exists': len(rows) > 0, 'details': rows})

@app.route('/waiting')
@student_required
def waiting():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    if row and row['is_active']:
        return redirect(url_for('guidelines'))
    return render_template('waiting.html')

@app.route('/check_exam_status')
@student_required
def check_exam_status():
    auto_submit_expired_exams()
    student_id = session.get('student_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active, start_time, duration FROM exam_control WHERE id=1")
    row = c.fetchone()
    c.execute("SELECT status FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    # BUG-002 Fix: detect admin force-submit so exam.js can redirect immediately
    force_submitted = bool(student and student['status'] == 'Submitted')
    if not row or not row['is_active']:
        return jsonify({'active': False, 'force_submitted': force_submitted})
    if force_submitted:
        return jsonify({'active': False, 'force_submitted': True})
    return jsonify({'active': True, 'force_submitted': False})

@app.route('/exam')
@student_required
def exam():
    import random
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    if not row or not row['is_active']:
        conn.close()
        return redirect(url_for('waiting'))

    student_id = session['student_id']
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if student and student['status'] == 'Submitted':
        conn.close()
        return redirect(url_for('waiting'))

    if student:
        c.execute("SELECT COUNT(*) FROM questions WHERE class=? AND subject=?",
                  (student['class'], student['subject']))
        q_count = c.fetchone()[0]
        if q_count == 0:
            conn.close()
            return "No questions available for this subject/class. Please contact admin.", 404

        c.execute("SELECT COUNT(*) FROM shuffled_questions WHERE student_id=?", (student_id,))
        count = c.fetchone()[0]
        if count == 0:
            c.execute("SELECT id FROM questions WHERE class=? AND subject=? ORDER BY id",
                      (student['class'], student['subject']))
            qids = [row['id'] for row in c.fetchall()]
            random.shuffle(qids)
            option_letters = ['A', 'B', 'C', 'D']
            for idx, qid in enumerate(qids):
                shuffled_opts = option_letters[:]
                random.shuffle(shuffled_opts)
                opt_str = ''.join(shuffled_opts)
                c.execute("INSERT INTO shuffled_questions (student_id, question_id, shuffled_index, option_order) VALUES (?,?,?,?)",
                          (student_id, qid, idx, opt_str))
            conn.commit()

        if student['exam_started_at'] is None or student['status'] == 'Not Started':
            now_str = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            c.execute("UPDATE students SET status='In Progress', exam_started_at=? WHERE student_id=?",
                      (now_str, student_id))
            conn.commit()

    # Get chapter/test_no for header
    c.execute("SELECT DISTINCT chapter FROM questions WHERE class=? AND subject=? AND chapter IS NOT NULL AND chapter!='' LIMIT 1",
              (student['class'], student['subject']))
    chapter_row = c.fetchone()
    test_no = chapter_row['chapter'] if chapter_row else ''

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    conn.close()

    now = datetime.datetime.now()
    combined_subjects = session.get('combined_subjects', [])
    is_combined       = bool(session.get('combined_test_id'))
    return render_template('exam.html',
                           school_name=school_name,
                           logo_path=logo_path,
                           gemini_key=_get_gemini_key(),
                           student_name=session.get('student_name',''),
                           student_id=student_id,
                           student_class=student['class'],
                           student_section=student['section'] if student['section'] else session.get('section', ''),
                           student_subject=student['subject'],
                           test_no=test_no,
                           exam_date=now.strftime('%d/%m/%Y'),
                           exam_day=now.strftime('%A'),
                           is_combined=is_combined,
                           combined_subjects=combined_subjects)

@app.route('/get_questions')
@student_required
def get_questions():
    student_id = session.get('student_id')
    class_     = session.get('class')
    subject    = session.get('subject')
    conn = get_db()
    c    = conn.cursor()

    c.execute('''
        SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d,
               q.image_path, q.subject, q.question_type,
               COALESCE(sq.option_order, 'ABCD') as option_order,
               r.selected_option
        FROM questions q
        LEFT JOIN shuffled_questions sq ON q.id = sq.question_id AND sq.student_id = ?
        LEFT JOIN responses r           ON q.id = r.question_id  AND r.student_id = ?
        WHERE sq.student_id = ?
        ORDER BY COALESCE(sq.shuffled_index, q.id)
    ''', (student_id, student_id, student_id))
    rows = c.fetchall()
    conn.close()

    if not rows:
        return jsonify([])

    questions = []
    for row in rows:
        opt_order     = row['option_order']
        original_opts = {'A': row['option_a'], 'B': row['option_b'],
                         'C': row['option_c'], 'D': row['option_d']}
        shuffled_options = {}
        for display_idx, orig_letter in enumerate(opt_order):
            display_letter = chr(ord('A') + display_idx)
            shuffled_options[display_letter] = original_opts[orig_letter]

        raw_image = row['image_path'] or ''
        if raw_image.startswith('__smiles__'):
            smiles_val   = raw_image[len('__smiles__'):]
            actual_image = None
        else:
            smiles_val   = ''
            actual_image = raw_image if raw_image else None

        questions.append({
            'id':            row['id'],
            'question':      row['question'],
            'option_a':      shuffled_options['A'],
            'option_b':      shuffled_options['B'],
            'option_c':      shuffled_options['C'],
            'option_d':      shuffled_options['D'],
            'image_path':    actual_image,
            'smiles':        smiles_val,
            'subject':       row['subject'],       # for tab grouping
            'question_type': row['question_type'] or 'MCQ',
            'selected':      row['selected_option'] or '',
        })
    return jsonify(questions)

    if not rows:
        return jsonify([])

    questions = []
    for row in rows:
        opt_order = row['option_order']
        original_opts = {
            'A': row['option_a'],
            'B': row['option_b'],
            'C': row['option_c'],
            'D': row['option_d']
        }
        shuffled_options = {}
        for display_idx, orig_letter in enumerate(opt_order):
            display_letter = chr(ord('A') + display_idx)
            shuffled_options[display_letter] = original_opts[orig_letter]

        # Decode image_path: if prefixed __smiles__ extract SMILES, else treat as image path
        raw_image = row['image_path'] or ''
        if raw_image.startswith('__smiles__'):
            smiles_val      = raw_image[len('__smiles__'):]
            actual_image    = None
        else:
            smiles_val      = ''
            actual_image    = raw_image if raw_image else None

        questions.append({
            'id':           row['id'],
            'question':     row['question'],
            'option_a':     shuffled_options['A'],
            'option_b':     shuffled_options['B'],
            'option_c':     shuffled_options['C'],
            'option_d':     shuffled_options['D'],
            'image_path':   actual_image,
            'smiles':       smiles_val,
            'content_type': row.get('content_type', '') if hasattr(row, 'get') else '',
            'image_prompt': '',
            'selected':     row['selected_option'],
            'option_order': opt_order
        })
    return jsonify(questions)

@app.route('/save_answer', methods=['POST'])
@student_required
def save_answer():
    data = request.get_json()
    question_id = data.get('question_id')
    displayed_option = data.get('selected_option')
    student_id = session['student_id']

    conn = get_db()
    c = conn.cursor()
    # Get the option order for this question
    c.execute("SELECT option_order FROM shuffled_questions WHERE student_id=? AND question_id=?",
              (student_id, question_id))
    row = c.fetchone()
    original_letter = ''
    if row and displayed_option:
        opt_order = row['option_order']
        # Map displayed letter back to original letter
        display_index = ord(displayed_option) - ord('A')
        if display_index < len(opt_order):
            original_letter = opt_order[display_index]

    c.execute("INSERT OR REPLACE INTO responses (student_id, question_id, selected_option, created_at) VALUES (?,?,?, CURRENT_TIMESTAMP)",
              (student_id, question_id, original_letter))
    c.execute("UPDATE students SET status='In Progress' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})

@app.route('/submit_exam', methods=['POST'])
@student_required
def submit_exam():
    student_id        = session['student_id']
    class_            = session.get('class')
    subject           = session.get('subject')
    name              = session.get('student_name')
    combined_subjects = session.get('combined_subjects', [])
    test_no           = session.get('test_no', '')

    conn = get_db()
    c    = conn.cursor()

    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec          = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    def _score_subject(subj):
        c.execute("""SELECT q.correct_answer, r.selected_option, q.subject
                     FROM shuffled_questions sq
                     JOIN questions q   ON q.id = sq.question_id
                     LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
                     WHERE sq.student_id = ? AND q.subject = ?""",
                 (student_id, student_id, subj))
        rows  = c.fetchall()
        total = len(rows)
        raw   = 0.0
        for r in rows:
            sel = r['selected_option']
            cor = r['correct_answer']
            if sel == cor:
                raw += 1.0
            elif sel and neg_enabled:
                raw -= neg_value
        score = max(0.0, raw)
        pct   = round((score / total * 100), 2) if total > 0 else 0.0
        return score, total, pct

    if combined_subjects:
        # Save one result row per subject
        for subj in combined_subjects:
            score, total, pct = _score_subject(subj)
            c.execute("""INSERT INTO results
                         (student_id, name, class, subject, score, total_questions, percentage, test_date)
                         VALUES (?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                     (student_id, name, class_, subj, round(score,2), total, pct))
    else:
        # Single-subject test
        score, total, pct = _score_subject(subject)
        c.execute("""INSERT INTO results
                     (student_id, name, class, subject, score, total_questions, percentage, test_date)
                     VALUES (?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                 (student_id, name, class_, subject, round(score,2), total, pct))

    c.execute("UPDATE students SET status='Submitted' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'submitted', 'redirect': url_for('submitted')})

@app.route('/submitted')
@student_required
def submitted():
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT status FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='pending'",
              (student_id, class_, subject))
    pending = c.fetchone()
    conn.close()
    return render_template('submitted.html', pending=bool(pending))

@app.route('/request_reattempt', methods=['POST'])
@student_required
def request_reattempt():
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='pending'",
              (student_id, class_, subject))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'You already have a pending request.'}), 400

    c.execute("INSERT INTO reattempt_requests (student_id, class, subject, status) VALUES (?,?,?, 'pending')",
              (student_id, class_, subject))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Request sent to admin.'})

@app.route('/check_reattempt_status')
@student_required
def check_reattempt_status():
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='approved'",
              (student_id, class_, subject))
    approved = c.fetchone()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    exam_active = bool(exam_row['is_active']) if exam_row else False
    c.execute("SELECT status FROM students WHERE student_id=?", (student_id,))
    student_row = c.fetchone()
    student_status = student_row['status'] if student_row else 'Unknown'
    conn.close()

    can_start = (approved is not None) and exam_active and (student_status == 'Not Started')
    return jsonify({
        'approved': approved is not None,
        'exam_active': exam_active,
        'student_status': student_status,
        'can_start': can_start
    })

@app.route('/get_exam_time')
@student_required
def get_exam_time():
    """Feature #2: Per-student timer — counts from student's own exam_started_at."""
    student_id = session.get('student_id')
    conn = get_db()
    c = conn.cursor()

    # Get exam duration from global control
    c.execute("SELECT duration, is_active FROM exam_control WHERE id=1")
    ctrl = c.fetchone()
    if not ctrl or not ctrl['is_active']:
        conn.close()
        return jsonify({'remaining': 0, 'duration': 0})

    duration = ctrl['duration']

    # Use THIS student's personal exam_started_at
    c.execute("SELECT exam_started_at, status FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    if not student or not student['exam_started_at']:
        return jsonify({'remaining': duration * 60, 'duration': duration})

    if student['status'] == 'Submitted':
        return jsonify({'remaining': 0, 'duration': duration})

    try:
        # Support both ISO format and SQLite CURRENT_TIMESTAMP format
        started_str = student['exam_started_at'].replace('T', ' ')
        start = datetime.datetime.fromisoformat(started_str)
    except Exception:
        return jsonify({'remaining': duration * 60, 'duration': duration})

    end_time  = start + datetime.timedelta(minutes=duration)
    remaining = int((end_time - datetime.datetime.utcnow()).total_seconds())
    if remaining < 0:
        remaining = 0

    return jsonify({'remaining': remaining, 'duration': duration})

# -------------------------------
# Admin Routes
# -------------------------------
@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if password == 'admin123':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin_login.html', error='Invalid credentials')
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM students WHERE status='In Progress'")
    active_students = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM teachers WHERE status='active'")
    active_teachers = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM reattempt_requests WHERE status='pending'")
    pending_reattempts = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM students")
    total_students = c.fetchone()[0]
    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    conn.close()
    return render_template('admin_dashboard.html',
                           active_students=active_students,
                           active_teachers=active_teachers,
                           pending_reattempts=pending_reattempts,
                           total_students=total_students,
                           school_name=school_name,
                           logo_path=logo_path)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/exam_status')
@admin_required
def exam_status():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    return jsonify({'is_active': bool(row['is_active']) if row else False})

@app.route('/admin/start_exam', methods=['POST'])
@admin_required
def start_exam():
    duration         = request.form.get('duration', 60, type=int)
    negative_marking = 1 if request.form.get('negative_marking') == 'yes' else 0
    negative_value   = request.form.get('negative_value', 0.33, type=float)
    conn = get_db()
    c = conn.cursor()
    start_time = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    c.execute("""UPDATE exam_control
                 SET is_active=1, start_time=?, duration=?,
                     negative_marking=?, negative_value=?
                 WHERE id=1""",
              (start_time, duration, negative_marking, negative_value))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/exam_settings')
@admin_required
def admin_exam_settings():
    """Return current exam control settings as JSON — used by admin dashboard."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT duration, negative_marking, negative_value FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    if row:
        return jsonify({'duration': row['duration'],
                        'negative_marking': bool(row['negative_marking']),
                        'negative_value': row['negative_value']})
    return jsonify({'duration': 60, 'negative_marking': False, 'negative_value': 0.33})

@app.route('/admin/student/force_submit/<student_id>', methods=['POST'])
@admin_required
def admin_force_submit_student(student_id):
    """Feature #3: Force-submit an individual student from monitoring page."""
    conn = get_db()
    c = conn.cursor()

    # Only act if student is In Progress
    c.execute("SELECT name, class, subject, status FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
    if student['status'] not in ('In Progress', 'Not Started'):
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student already submitted'}), 400

    name    = student['name']
    class_  = student['class']
    subject = student['subject']

    # Calculate score with negative marking
    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    c.execute("""SELECT q.correct_answer, r.selected_option
                 FROM questions q
                 LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
                 WHERE q.class = ? AND q.subject = ?""",
             (student_id, class_, subject))
    rows  = c.fetchall()
    total = len(rows)
    raw   = sum(
        1.0 if r['selected_option'] == r['correct_answer'] else
        (-neg_value if r['selected_option'] and neg_enabled else 0.0)
        for r in rows
    )
    score      = max(0.0, raw)
    percentage = round((score / total * 100), 2) if total > 0 else 0.0

    # Save result
    c.execute("""INSERT INTO results (student_id, name, class, subject, score, total_questions, percentage, test_date)
                 VALUES (?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
             (student_id, name, class_, subject, round(score, 2), total, percentage))
    c.execute("UPDATE students SET status='Submitted' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': f'{name} submitted. Score: {round(score,2)}/{total}'})

@app.route('/admin/stop_exam')
@admin_required
def stop_exam():
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE exam_control SET is_active=0 WHERE id=1")
    conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    if request.method == 'POST':
        for field in ['school_name', 'school_address', 'academic_session']:
            val = request.form.get(field, '').strip()
            if val:
                set_setting(field, val)

        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename != '' and allowed_image_file(logo_file.filename):
            ext = logo_file.filename.rsplit('.', 1)[1].lower()
            filename = f"school_logo.{ext}"
            filepath = os.path.join(app.config['BRANDING_FOLDER'], filename)
            for old_ext in ALLOWED_IMAGE_EXTENSIONS:
                old_path = os.path.join(app.config['BRANDING_FOLDER'], f"school_logo.{old_ext}")
                if os.path.exists(old_path) and old_path != filepath:
                    os.remove(old_path)
            logo_file.save(filepath)
            set_setting('logo_path', os.path.join('uploads', 'branding', filename).replace('\\', '/'))

        return redirect(url_for('admin_settings'))

    school_name      = get_setting('school_name', 'RRB Group of Schools')
    school_address   = get_setting('school_address', '')
    academic_session = get_setting('academic_session', '')
    logo_path        = get_setting('logo_path', '')
    return render_template('settings.html',
                           school_name=school_name,
                           school_address=school_address,
                           academic_session=academic_session,
                           logo_path=logo_path)

@app.route('/admin/generate_question_paper')
@admin_required
def generate_question_paper():
    class_ = request.args.get('class', '')
    subject = request.args.get('subject', '')
    if not class_ or not subject:
        return "Class and subject required", 400

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer, image_path FROM questions WHERE class=? AND subject=? ORDER BY id",
              (class_, subject))
    questions = [dict(row) for row in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode('utf-8')}"

    total_marks = len(questions)
    rendered_html = render_template(
        'question_paper_template.html',
        school_name=school_name,
        logo_base64=logo_base64,
        class_name=class_,
        subject=subject,
        questions=questions,
        total_marks=total_marks,
        date=datetime.datetime.now().strftime('%d/%m/%Y')
    )

    html = HTML(string=rendered_html, base_url=request.base_url)
    css = CSS(string='''
        @page { size: A4; margin: 2cm; }
        body { font-family: 'Poppins', sans-serif; }
        .header { text-align: center; margin-bottom: 30px; }
        .header img { max-height: 80px; }
        .school-name { font-size: 24px; font-weight: bold; }
        .exam-info { margin: 20px 0; }
        .question { margin-bottom: 20px; page-break-inside: avoid; }
        .options { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-left: 20px; }
        .answer-key { margin-top: 30px; page-break-before: always; }
        .answer-table { width: 100%; border-collapse: collapse; }
        .answer-table th, .answer-table td { border: 1px solid #333; padding: 8px; text-align: left; }
    ''')
    pdf = html.write_pdf(stylesheets=[css])

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=Question_Paper_{class_}_{subject}.pdf'
    return response

@app.route('/admin/questions')
@admin_required
def questions():
    return render_template('questions.html')

@app.route('/admin/questions/data')
@admin_required
def questions_data():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, class, subject, question, option_a, option_b, option_c, option_d, correct_answer, image_path FROM questions ORDER BY class, subject, id")
    questions = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(questions)

@app.route('/admin/question/add', methods=['POST'])
@admin_required
def add_question():
    if request.is_json:
        data = request.get_json()
        class_ = data['class']
        subject = data['subject']
        question = data['question']
        opt_a = data['option_a']
        opt_b = data['option_b']
        opt_c = data['option_c']
        opt_d = data['option_d']
        correct = data['correct_answer']
        image_file = None
    else:
        class_ = request.form.get('class')
        subject = request.form.get('subject')
        question = request.form.get('question')
        opt_a = request.form.get('option_a')
        opt_b = request.form.get('option_b')
        opt_c = request.form.get('option_c')
        opt_d = request.form.get('option_d')
        correct = request.form.get('correct_answer')
        image_file = request.files.get('image')

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO questions (class, subject, question, option_a, option_b, option_c, option_d, correct_answer) VALUES (?,?,?,?,?,?,?,?)",
              (class_, subject, question, opt_a, opt_b, opt_c, opt_d, correct))
    qid = c.lastrowid

    image_path = None
    if image_file and image_file.filename != '':
        image_path = save_question_image(image_file, class_, subject, qid)
        if image_path:
            c.execute("UPDATE questions SET image_path=? WHERE id=?", (image_path, qid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'id': qid})

@app.route('/admin/question/update/<int:qid>', methods=['POST'])
@admin_required
def update_question(qid):
    class_ = request.form.get('class')
    subject = request.form.get('subject')
    question = request.form.get('question')
    opt_a = request.form.get('option_a')
    opt_b = request.form.get('option_b')
    opt_c = request.form.get('option_c')
    opt_d = request.form.get('option_d')
    correct = request.form.get('correct_answer')
    image_file = request.files.get('image')
    remove_image = request.form.get('remove_image') == 'true'

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT image_path FROM questions WHERE id=?", (qid,))
    old_image = c.fetchone()
    old_path = old_image['image_path'] if old_image else None

    c.execute('''UPDATE questions 
                 SET class=?, subject=?, question=?, option_a=?, option_b=?, option_c=?, option_d=?, correct_answer=?
                 WHERE id=?''',
              (class_, subject, question, opt_a, opt_b, opt_c, opt_d, correct, qid))

    image_path = old_path
    if remove_image:
        if old_path and os.path.exists(os.path.join('static', old_path)):
            try:
                os.remove(os.path.join('static', old_path))
            except:
                pass
        image_path = None
        c.execute("UPDATE questions SET image_path=NULL WHERE id=?", (qid,))
    elif image_file and image_file.filename != '':
        if old_path and os.path.exists(os.path.join('static', old_path)):
            try:
                os.remove(os.path.join('static', old_path))
            except:
                pass
        new_path = save_question_image(image_file, class_, subject, qid)
        if new_path:
            image_path = new_path
            c.execute("UPDATE questions SET image_path=? WHERE id=?", (image_path, qid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'updated'})

@app.route('/admin/question/delete/<int:qid>', methods=['DELETE'])
@admin_required
def delete_question(qid):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT image_path FROM questions WHERE id=?", (qid,))
    row = c.fetchone()
    if row and row['image_path']:
        img_path = os.path.join('static', row['image_path'])
        if os.path.exists(img_path):
            try:
                os.remove(img_path)
            except:
                pass
    c.execute("DELETE FROM questions WHERE id=?", (qid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'deleted'})

@app.route('/admin/questions/delete_by_class_subject', methods=['POST'])
@admin_required
def delete_questions_by_class_subject():
    data = request.get_json()
    class_ = data.get('class')
    subject = data.get('subject')

    if not class_ or not subject:
        return jsonify({'status': 'error', 'message': 'Class and subject required'}), 400

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT image_path FROM questions WHERE class=? AND subject=?", (class_, subject))
    images = c.fetchall()

    for img in images:
        if img['image_path']:
            img_path = os.path.join('static', img['image_path'])
            if os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except:
                    pass

    c.execute("DELETE FROM questions WHERE class=? AND subject=?", (class_, subject))
    deleted_count = c.rowcount
    conn.commit()
    conn.close()

    return jsonify({'status': 'success', 'deleted': deleted_count})

@app.route('/admin/upload_csv', methods=['POST'])
@admin_required
def upload_csv():
    if 'csv_file' not in request.files:
        return redirect(request.url)
    file = request.files['csv_file']
    if file.filename == '':
        return redirect(request.url)
    if file and file.filename.endswith('.csv'):
        filename = secure_filename(file.filename)
        base = os.path.splitext(filename)[0]
        parts = base.split('_')
        # Parse class, subject, and optional test_no
        if len(parts) >= 3:
            class_ = parts[0].strip()
            subject = parts[1].strip()
            test_no = parts[2].strip()
        elif len(parts) == 2:
            class_ = parts[0].strip()
            subject = parts[1].strip()
            test_no = ''
        else:
            class_ = "Unknown"
            subject = base.strip()
            test_no = ''
        if not class_:
            class_ = "Unknown"

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        conn = get_db()
        c = conn.cursor()
        try:
            with open(filepath, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader)
                normalized = [h.strip().lower().replace(' ', '_') for h in header]

                try:
                    idx_q = normalized.index('question')
                    idx_a = normalized.index('option_a')
                    idx_b = normalized.index('option_b')
                    idx_c = normalized.index('option_c')
                    idx_d = normalized.index('option_d')
                    idx_correct = normalized.index('correct_answer')
                except ValueError as e:
                    return f"Missing required column: {e}. Headers found: {header}", 400

                for row in reader:
                    if not row or all(cell.strip() == '' for cell in row):
                        continue
                    if len(row) <= max(idx_q, idx_a, idx_b, idx_c, idx_d, idx_correct):
                        continue

                    question = row[idx_q].strip()
                    opt_a = row[idx_a].strip()
                    opt_b = row[idx_b].strip()
                    opt_c = row[idx_c].strip()
                    opt_d = row[idx_d].strip()
                    correct_value = row[idx_correct].strip()

                    correct_letter = None
                    if correct_value == opt_a:
                        correct_letter = 'A'
                    elif correct_value == opt_b:
                        correct_letter = 'B'
                    elif correct_value == opt_c:
                        correct_letter = 'C'
                    elif correct_value == opt_d:
                        correct_letter = 'D'
                    else:
                        if correct_value.upper() in ('A', 'B', 'C', 'D'):
                            correct_letter = correct_value.upper()
                        elif correct_value.lower() == 'option_a':
                            correct_letter = 'A'
                        elif correct_value.lower() == 'option_b':
                            correct_letter = 'B'
                        elif correct_value.lower() == 'option_c':
                            correct_letter = 'C'
                        elif correct_value.lower() == 'option_d':
                            correct_letter = 'D'
                        else:
                            print(f"Warning: Could not map correct_answer '{correct_value}' to any option. Skipping question.")
                            continue

                    if not question:
                        continue

                    # Insert with chapter (test_no)
                    c.execute("INSERT INTO questions (class, subject, question, option_a, option_b, option_c, option_d, correct_answer, chapter) VALUES (?,?,?,?,?,?,?,?,?)",
                              (class_, subject, question, opt_a, opt_b, opt_c, opt_d, correct_letter, test_no))
            conn.commit()
        except Exception as e:
            conn.rollback()
            return f"Error processing CSV: {e}", 500
        finally:
            conn.close()
            os.remove(filepath)
        return redirect(url_for('questions'))
    return "Invalid file", 400

@app.route('/admin/students')
@admin_required
def manage_students():
    return render_template('manage_students.html')

@app.route('/admin/students/data')
@admin_required
def students_data():
    filter_class = request.args.get('class', '')
    filter_subject = request.args.get('subject', '')
    conn = get_db()
    c = conn.cursor()
    query = """SELECT student_id, name, class, subject, ip, status,
                      admission_no, section, dob, house, parents_name, address, picture
               FROM students WHERE 1=1"""
    params = []
    if filter_class:
        query += " AND class=?"
        params.append(filter_class)
    if filter_subject:
        query += " AND subject=?"
        params.append(filter_subject)
    c.execute(query, params)
    students = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(students)

@app.route('/admin/student/delete/<student_id>', methods=['DELETE'])
@admin_required
def delete_student(student_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM students WHERE student_id=?",           (student_id,))
    c.execute("DELETE FROM responses WHERE student_id=?",          (student_id,))
    c.execute("DELETE FROM results WHERE student_id=?",            (student_id,))
    c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))  # Bug #3 fix
    c.execute("DELETE FROM reattempt_requests WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'deleted'})

@app.route('/admin/monitoring')
@admin_required
def monitoring():
    return render_template('monitoring.html')

@app.route('/admin/monitoring/data')
@admin_required
def monitoring_data():
    auto_submit_expired_exams()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, class, subject, student_id, ip, status, exam_started_at FROM students ORDER BY status, name")
    students = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(students)

@app.route('/admin/evaluate')
@admin_required
def evaluate():
    conn = get_db()
    c = conn.cursor()

    # BUG-003 Fix: fetch negative marking settings FIRST
    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    c.execute("DELETE FROM results")
    c.execute("SELECT student_id, name, class, subject FROM students WHERE status='Submitted'")
    students = c.fetchall()
    for student in students:
        student_id = student['student_id']
        name  = student['name']
        class_  = student['class']
        subject = student['subject']
        c.execute("SELECT id, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=?", (class_, subject))
        q_rows = c.fetchall()
        correct_map = {}
        for q in q_rows:
            correct = q['correct_answer']
            if correct is None:
                continue
            correct = correct.strip()
            if len(correct) == 1 and correct.upper() in ('A','B','C','D'):
                correct_map[q['id']] = correct.upper()
            else:
                if correct == q['option_a'] or correct.lower() == 'option_a':
                    correct_map[q['id']] = 'A'
                elif correct == q['option_b'] or correct.lower() == 'option_b':
                    correct_map[q['id']] = 'B'
                elif correct == q['option_c'] or correct.lower() == 'option_c':
                    correct_map[q['id']] = 'C'
                elif correct == q['option_d'] or correct.lower() == 'option_d':
                    correct_map[q['id']] = 'D'
        c.execute("SELECT question_id, selected_option FROM responses WHERE student_id=?", (student_id,))
        responses = c.fetchall()

        # BUG-003 Fix: apply negative marking to each wrong answer
        raw_score = 0.0
        total     = len(correct_map)
        for resp in responses:
            qid      = resp['question_id']
            selected = (resp['selected_option'] or '').strip().upper()
            if qid not in correct_map:
                continue
            if selected == correct_map[qid]:
                raw_score += 1.0
            elif selected and neg_enabled:
                raw_score -= neg_value

        score      = max(0.0, raw_score)
        percentage = round((score / total * 100), 2) if total > 0 else 0.0
        c.execute("""INSERT OR REPLACE INTO results (student_id, name, class, subject, score, total_questions, percentage)
                     VALUES (?,?,?,?,?,?,?)""",
                  (student_id, name, class_, subject, round(score, 2), total, percentage))
    conn.commit()
    conn.close()
    return redirect(url_for('results_page'))

@app.route('/admin/results')
@admin_required
def results_page():
    return render_template('results.html')

@app.route('/admin/results/data')
@admin_required
def results_data():
    conn = get_db()
    c = conn.cursor()
    c.execute('''
        SELECT r.id, r.student_id, r.name, r.class, r.subject, r.score,
               r.total_questions, r.percentage, r.test_date,
               s.exam_started_at
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        ORDER BY r.test_date DESC, r.class, r.subject, r.name
    ''')
    results = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(results)

@app.route('/admin/export_results_page')
@admin_required
def export_results_page():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class FROM students ORDER BY class")
    classes = [row['class'] for row in c.fetchall()]
    c.execute("SELECT DISTINCT subject FROM students ORDER BY subject")
    subjects = [row['subject'] for row in c.fetchall()]
    conn.close()
    return render_template('export_results_filter.html', classes=classes, subjects=subjects)

@app.route('/admin/export/results')
@admin_required
def export_results():
    class_filter = request.args.get('class', '')
    subject_filter = request.args.get('subject', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    single_date = request.args.get('date', '')

    conn = get_db()
    c = conn.cursor()

    query = '''
        SELECT r.student_id, r.name, r.class, r.subject, r.score,
               r.total_questions, r.percentage, r.test_date,
               s.exam_started_at
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        WHERE 1=1
    '''
    params = []
    if class_filter and class_filter != 'All':
        query += " AND r.class = ?"
        params.append(class_filter)
    if subject_filter and subject_filter != 'All':
        query += " AND r.subject = ?"
        params.append(subject_filter)

    if single_date:
        query += " AND DATE(r.test_date) = ?"
        params.append(single_date)
    else:
        if start_date:
            query += " AND DATE(r.test_date) >= ?"
            params.append(start_date)
        if end_date:
            query += " AND DATE(r.test_date) <= ?"
            params.append(end_date)

    query += " ORDER BY r.name ASC"

    c.execute(query, params)
    rows = c.fetchall()
    total_students = len(rows)

    school_name = get_setting('school_name', 'RRB Group of Schools')
    class_display = class_filter if class_filter and class_filter != 'All' else "All Classes"
    subject_display = subject_filter if subject_filter and subject_filter != 'All' else "All Subjects"

    if single_date:
        date_display = single_date
    elif start_date and end_date:
        date_display = f"{start_date} to {end_date}"
    elif start_date:
        date_display = f"From {start_date}"
    elif end_date:
        date_display = f"Up to {end_date}"
    else:
        date_display = "All Dates"

    export_time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')

    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    ws.merge_cells('A1:G1')
    ws['A1'].value = school_name
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'].value = f"Class: {class_display}  |  Subject: {subject_display}  |  Date: {date_display}"
    ws['A2'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.merge_cells('A3:G3')
    ws['A3'].value = f"Total Students Appeared: {total_students}  |  Generated: {export_time}"
    ws['A3'].font = openpyxl.styles.Font(italic=True)
    ws['A3'].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws.append([])

    headers = ['Student ID', 'Name', 'Class', 'Subject', 'Score', 'Percentage', 'Test Date']
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, start=header_row+1):
        ws.cell(row=row_idx, column=1, value=row['student_id'])
        ws.cell(row=row_idx, column=2, value=row['name'])
        ws.cell(row=row_idx, column=3, value=row['class'])
        ws.cell(row=row_idx, column=4, value=row['subject'])
        ws.cell(row=row_idx, column=5, value=f"{row['score']}/{row['total_questions']}" if row['total_questions'] else row['score'])
        ws.cell(row=row_idx, column=6, value=f"{row['percentage']:.1f}%" if row['percentage'] else 'N/A')
        test_date = row['test_date'] or row['exam_started_at'] or ''
        ws.cell(row=row_idx, column=7, value=test_date[:10] if test_date else 'N/A')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    file_parts = []
    file_parts.append(class_filter if class_filter and class_filter != 'All' else 'all')
    file_parts.append(subject_filter if subject_filter and subject_filter != 'All' else 'all')
    if start_date and end_date:
        file_parts.append(f"{start_date}_to_{end_date}")
    elif start_date:
        file_parts.append(f"from_{start_date}")
    elif end_date:
        file_parts.append(f"to_{end_date}")
    else:
        file_parts.append("all_dates")
    file_parts.append(timestamp)
    filename = f"results_{'_'.join(file_parts)}.xlsx"

    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)

@app.route('/admin/student/responses/<student_id>')
@admin_required
def view_student_responses(student_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        return "Student not found", 404

    # Fetch questions with all option texts
    c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=? ORDER BY id",
              (student['class'], student['subject']))
    questions = c.fetchall()

    c.execute("SELECT question_id, selected_option FROM responses WHERE student_id=?", (student_id,))
    responses = {row['question_id']: row['selected_option'] for row in c.fetchall()}

    conn.close()

    question_data = []
    for q in questions:
        selected = responses.get(q['id'], '')
        # Map the stored correct_answer to a letter (A/B/C/D)
        correct_raw = q['correct_answer']
        correct_letter = None
        if correct_raw:
            correct_raw = correct_raw.strip()
            if len(correct_raw) == 1 and correct_raw.upper() in ('A','B','C','D'):
                correct_letter = correct_raw.upper()
            else:
                # Try to match against option texts or "option_x"
                if correct_raw == q['option_a'] or correct_raw.lower() == 'option_a':
                    correct_letter = 'A'
                elif correct_raw == q['option_b'] or correct_raw.lower() == 'option_b':
                    correct_letter = 'B'
                elif correct_raw == q['option_c'] or correct_raw.lower() == 'option_c':
                    correct_letter = 'C'
                elif correct_raw == q['option_d'] or correct_raw.lower() == 'option_d':
                    correct_letter = 'D'
        # Determine if correct
        is_correct = False
        if selected and correct_letter:
            is_correct = (selected.strip().upper() == correct_letter)

        question_data.append({
            'id': q['id'],
            'question': q['question'],
            'options': {
                'A': q['option_a'],
                'B': q['option_b'],
                'C': q['option_c'],
                'D': q['option_d']
            },
            'correct': correct_letter or '?',  # display the letter, not raw text
            'selected': selected,
            'is_correct': is_correct
        })

    return render_template('student_responses.html', student=student, questions=question_data)

@app.route('/admin/export/answers')
@admin_required
def export_answers():
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT r.student_id, q.question, r.selected_option, q.correct_answer
                 FROM responses r JOIN questions q ON r.question_id = q.id
                 ORDER BY r.student_id, q.id''')
    rows = c.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.append(['Student ID', 'Question', 'Selected', 'Correct'])
    for row in rows:
        ws.append([row['student_id'], row['question'], row['selected_option'], row['correct_answer']])
    filename = f"answers_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)

@app.route('/admin/reattempt_requests')
@admin_required
def admin_reattempt_requests():
    return render_template('reattempt_requests.html')

@app.route('/admin/reattempt_requests/data')
@admin_required
def reattempt_requests_data():
    filter_status = request.args.get('status', '')
    conn = get_db()
    c = conn.cursor()
    query = """
        SELECT r.id, r.student_id, s.name, r.class, r.subject, r.status, r.requested_at, r.reviewed_at, r.admin_note
        FROM reattempt_requests r
        JOIN students s ON r.student_id = s.student_id
        WHERE 1=1
    """
    params = []
    if filter_status:
        query += " AND r.status = ?"
        params.append(filter_status)
    query += " ORDER BY r.requested_at DESC"
    c.execute(query, params)
    requests = [dict(row) for row in c.fetchall()]
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    exam_is_active = bool(exam_row['is_active']) if exam_row else False
    conn.close()
    return jsonify({'requests': requests, 'exam_is_active': exam_is_active})

def auto_submit_expired_exams():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, duration, is_active FROM exam_control WHERE id=1")
    exam = c.fetchone()

    if not exam or not exam['duration'] or not exam['is_active']:
        conn.close()
        return

    duration = exam['duration']

    # ✅ Use SQLite's own time for comparison — avoids timezone mismatch
    c.execute("""
        UPDATE students 
        SET status='Submitted'
        WHERE status='In Progress'
        AND exam_started_at IS NOT NULL
        AND datetime(exam_started_at, ? || ' minutes') < datetime('now')
    """, (str(duration),))

    conn.commit()
    conn.close()

@app.route('/guidelines')
@student_required
def guidelines():
    # Mark that the student has not yet accepted guidelines for this session
    session['guidelines_accepted'] = False
    return render_template('guidelines.html')

@app.route('/accept_guidelines', methods=['POST'])
@student_required
def accept_guidelines():
    session['guidelines_accepted'] = True
    return redirect(url_for('exam'))

@app.route('/admin/reattempt_request/<int:req_id>/<action>', methods=['POST'])
@admin_required
def handle_reattempt_request(req_id, action):
    if action not in ('approve', 'reject'):
        return jsonify({'status': 'error', 'message': 'Invalid action'}), 400

    admin_note = request.json.get('note', '') if request.is_json else ''

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT student_id, class, subject, status FROM reattempt_requests WHERE id=?", (req_id,))
    req = c.fetchone()
    if not req:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Request not found'}), 404
    if req['status'] != 'pending':
        conn.close()
        return jsonify({'status': 'error', 'message': 'Request already processed'}), 400

    student_id = req['student_id']
    class_ = req['class']
    subject = req['subject']

    if action == 'approve':
        c.execute("DELETE FROM responses WHERE student_id=?", (student_id,))
        c.execute("DELETE FROM results WHERE student_id=?", (student_id,))
        c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
        c.execute("UPDATE students SET status='Not Started', exam_started_at=NULL WHERE student_id=?", (student_id,))
        c.execute("UPDATE reattempt_requests SET status='approved', reviewed_at=CURRENT_TIMESTAMP, admin_note=? WHERE id=?",
                  (admin_note, req_id))
        
        conn.commit()
    else:
        c.execute("UPDATE reattempt_requests SET status='rejected', reviewed_at=CURRENT_TIMESTAMP, admin_note=? WHERE id=?",
                  (admin_note, req_id))
        conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

# ========================
# TEACHER ROUTES
# ========================

@app.route('/teacher', methods=['GET', 'POST'])
def teacher_login():
    if request.method == 'POST':
        mobile = request.form.get('mobile')
        password = request.form.get('password')
        
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM teachers WHERE mobile=? AND status='active'", (mobile,))
        teacher = c.fetchone()
        conn.close()
        
        if teacher and verify_password(password, teacher['password']):
            session['teacher_logged_in'] = True
            session['teacher_id'] = teacher['id']
            session['teacher_name'] = teacher['name']
            session['teacher_mobile'] = teacher['mobile']
            return redirect(url_for('teacher_dashboard'))
        else:
            return render_template('teacher_login.html', error="Invalid credentials")
    
    return render_template('teacher_login.html')

@app.route('/teacher/logout')
def teacher_logout():
    session.pop('teacher_logged_in', None)
    session.pop('teacher_id', None)
    session.pop('teacher_name', None)
    session.pop('teacher_mobile', None)
    return redirect(url_for('teacher_login'))

@app.route('/teacher/create_test', methods=['GET', 'POST'])
@teacher_required
def teacher_create_test():
    if request.method == 'POST':
        class_ = request.form.get('class')
        subject = request.form.get('subject')
        chapter = request.form.get('chapter')
        num_questions = int(request.form.get('num_questions', 10))
        prompt = request.form.get('prompt', '')
        method = request.form.get('method', 'ai')  # 'ai' or 'upload'
        
        if method == 'upload':
            # Handle CSV upload
            file = request.files.get('csv_file')
            if file and file.filename.endswith('.csv'):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Import questions from CSV
                with open(filepath, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    conn = get_db()
                    c = conn.cursor()
                    for row in reader:
                        c.execute("""INSERT INTO questions 
                                   (class, subject, chapter, question, option_a, option_b, option_c, option_d, correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?)""",
                                (class_, subject, chapter,
                                 row.get('question', ''),
                                 row.get('option a', row.get('option_a', '')),
                                 row.get('option b', row.get('option_b', '')),
                                 row.get('option c', row.get('option_c', '')),
                                 row.get('option d', row.get('option_d', '')),
                                 row.get('correct_answer', '')))
                    conn.commit()
                    conn.close()
                os.remove(filepath)
                return jsonify({'status': 'success', 'message': 'Questions uploaded successfully'})
        
        else:  # AI generation using Google Gemini
            try:
                import urllib.request
                import json as json_lib

                gemini_api_key = os.environ.get('GEMINI_API_KEY', '')
                if not gemini_api_key:
                    return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY not set. Use Option 7 in the batch menu to set it.'}), 500

                ai_prompt = f"""You are an expert CBSE/ICSE question paper generator.
Output ONLY a valid JSON array — no markdown, no text outside the array.

Generate exactly {num_questions} MCQ questions for:
Class: {class_} | Subject: {subject} | Chapter: {chapter}

Teacher instructions: {prompt if prompt else 'Standard difficulty, balanced coverage'}

SCHEMA — every object must have these exact keys:
{{
  "question":      "<question text, use LaTeX or mhchem where needed>",
  "option_a":      "<option text or LaTeX>",
  "option_b":      "<option text or LaTeX>",
  "option_c":      "<option text or LaTeX>",
  "option_d":      "<option text or LaTeX>",
  "correct_answer":"option_a"|"option_b"|"option_c"|"option_d",
  "smiles":        "<SMILES for structural diagram, else empty string>",
  "content_type":  "math"|"chemistry"|"physics"|"biology"|"text"
}}

MATHEMATICS — use MathJax LaTeX (escape backslashes in JSON):
• Inline: $\\frac{{d}}{{dx}}(x^2)$   Block: $$\\int_0^1 x^2\\,dx = \\frac{{1}}{{3}}$$
• Vectors: $\\vec{{F}} = m\\vec{{a}}$  Matrices: $\\begin{{pmatrix}} a & b \\\\\\\\ c & d \\end{{pmatrix}}$
• NEVER use raw Unicode math symbols — use LaTeX commands only.

CHEMISTRY — use mhchem inside $ $:
• Equations: "$\\ce{{H2 + O2 -> H2O}}$"   Equilibrium: "$\\ce{{N2 + 3H2 <=> 2NH3}}$"
• States: "$\\ce{{CaCO3(s) -> CaO(s) + CO2(g)}}$"   Ionic: "$\\ce{{Fe^{{2+}} + 2e- -> Fe}}$"
• Structural diagrams: set smiles field (e.g. benzene="c1ccccc1", ethanol="CCO")

Output the JSON array now:"""

                # Call Gemini API
                url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_api_key}"
                payload = json_lib.dumps({
                    "contents": [{"parts": [{"text": ai_prompt}]}],
                    "generationConfig": {"temperature": 0.7, "maxOutputTokens": 8192}
                }).encode('utf-8')

                req = urllib.request.Request(url, data=payload,
                                             headers={'Content-Type': 'application/json'}, method='POST')
                with urllib.request.urlopen(req, timeout=60) as resp:
                    result = json_lib.loads(resp.read().decode('utf-8'))

                response_text = result['candidates'][0]['content']['parts'][0]['text'].strip()

                # Clean markdown fences if present
                if '```' in response_text:
                    parts = response_text.split('```')
                    for part in parts:
                        part = part.strip()
                        if part.startswith('json'):
                            part = part[4:].strip()
                        if part.startswith('['):
                            response_text = part
                            break

                questions = json_lib.loads(response_text)

                # Save to database (including smiles and content_type for multimodal rendering)
                conn = get_db()
                c = conn.cursor()
                for q in questions:
                    # Store smiles in image_path column prefixed so renderer knows it's SMILES
                    smiles_val = q.get('smiles', '') or ''
                    # Store content_type in question_type if it's a pure text field question
                    c.execute("""INSERT INTO questions
                               (class, subject, chapter, question_type, question,
                                option_a, option_b, option_c, option_d, correct_answer, image_path)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                            (class_, subject, chapter,
                             'MCQ',
                             q.get('question', ''),
                             q.get('option_a', ''),
                             q.get('option_b', ''),
                             q.get('option_c', ''),
                             q.get('option_d', ''),
                             q.get('correct_answer', ''),
                             f'__smiles__{smiles_val}' if smiles_val else None))
                conn.commit()
                conn.close()

                return jsonify({'status': 'success', 'message': f'{len(questions)} questions generated and saved using Gemini!'})

            except urllib.error.HTTPError as e:
                err_body = e.read().decode('utf-8')
                return jsonify({'status': 'error', 'message': f'Gemini API error: {err_body}'}), 500
            except Exception as e:
                return jsonify({'status': 'error', 'message': str(e)}), 500
    
    # GET request - show form
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return render_template('teacher_create_test.html', assignments=assignments)
@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    
    # Get teacher assignments
    c.execute("""SELECT class, subject, is_class_teacher 
                 FROM teacher_assignments 
                 WHERE teacher_id=?""", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Get teacher info
    c.execute("SELECT * FROM teachers WHERE id=?", (teacher_id,))
    teacher = dict(c.fetchone())
    
    conn.close()
    return render_template('teacher_dashboard.html', 
                         assignments=assignments,
                         teacher=teacher)


@app.route('/teacher/students')
@teacher_required
def teacher_students():
    teacher_id = session.get('teacher_id')
    class_filter = request.args.get('class', '')
    subject_filter = request.args.get('subject', '')
    
    conn = get_db()
    c = conn.cursor()
    
    # Get teacher assignments to verify access
    c.execute("""SELECT class, subject, is_class_teacher 
                 FROM teacher_assignments 
                 WHERE teacher_id=?""", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Build student query based on assignments
    query = """SELECT DISTINCT s.student_id, s.name, s.admission_no, s.class, s.section, 
                      s.dob, s.house, s.parents_name, s.picture
               FROM students s
               WHERE 1=1"""
    params = []
    
    # If teacher has any class teacher assignment, they can see all students in those classes
    class_teacher_classes = [a['class'] for a in assignments if a['is_class_teacher']]
    
    if class_teacher_classes and class_filter:
        if class_filter in class_teacher_classes:
            query += " AND s.class = ?"
            params.append(class_filter)
    elif not class_teacher_classes:
        # Subject teacher - can only see students in their subject classes
        assigned_classes = list(set([(a['class'], a['subject']) for a in assignments]))
        if assigned_classes:
            class_conditions = " OR ".join(["(s.class=? AND s.subject=?)"] * len(assigned_classes))
            query += f" AND ({class_conditions})"
            for cls, subj in assigned_classes:
                params.extend([cls, subj])
    
    c.execute(query, params)
    students = [dict(row) for row in c.fetchall()]
    conn.close()
    
    return render_template('teacher_students.html', 
                         students=students,
                         assignments=assignments,
                         class_filter=class_filter,
                         subject_filter=subject_filter)

@app.route('/teacher/student/<student_id>')
@teacher_required
def teacher_student_profile(student_id):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    
    # Get student info
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return "Student not found", 404
    student = dict(student)
    
    # Get teacher assignments
    c.execute("SELECT class, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Check if teacher is class teacher for this student
    is_class_teacher = any(a['class'] == student['class'] and a['is_class_teacher'] for a in assignments)
    
    # Get test history with safe numeric values
    if is_class_teacher:
        # Class teacher sees all subjects
        c.execute("""
            SELECT 
                r.id, r.subject, r.chapter, r.score, r.test_date,
                COALESCE(r.total_questions, 
                    (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject)) AS total_questions,
                COALESCE(r.percentage,
                    CASE 
                        WHEN (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject) > 0 
                        THEN ROUND((r.score * 100.0 / (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject)), 2)
                        ELSE 0.0 
                    END) AS percentage
            FROM results r
            WHERE r.student_id = ?
            ORDER BY r.test_date DESC
        """, (student_id,))
    else:
        # Subject teacher sees only their subjects
        teacher_subjects = [a['subject'] for a in assignments if a['class'] == student['class']]
        if teacher_subjects:
            placeholders = ','.join(['?'] * len(teacher_subjects))
            c.execute(f"""
                SELECT 
                    r.id, r.subject, r.chapter, r.score, r.test_date,
                    COALESCE(r.total_questions, 
                        (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject)) AS total_questions,
                    COALESCE(r.percentage,
                        CASE 
                            WHEN (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject) > 0 
                            THEN ROUND((r.score * 100.0 / (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject)), 2)
                            ELSE 0.0 
                        END) AS percentage
                FROM results r
                WHERE r.student_id = ? AND r.subject IN ({placeholders})
                ORDER BY r.test_date DESC
            """, [student_id] + teacher_subjects)
        else:
            c.execute("""
                SELECT 
                    r.id, r.subject, r.chapter, r.score, r.test_date,
                    COALESCE(r.total_questions, 0) AS total_questions,
                    COALESCE(r.percentage, 0.0) AS percentage
                FROM results r
                WHERE 1=0
            """)
    
    test_history = [dict(row) for row in c.fetchall()]
    
    # Generate chart data for progress
    chart_data = {}
    for test in test_history:
        subj = test['subject']
        if subj not in chart_data:
            chart_data[subj] = {'dates': [], 'percentages': [], 'marks': []}
        chart_data[subj]['dates'].append(test['test_date'][:10] if test['test_date'] else '')
        chart_data[subj]['percentages'].append(test['percentage'] or 0)
        chart_data[subj]['marks'].append(f"{test['score']}/{test['total_questions']}")
    
    conn.close()
    
    return render_template('teacher_student_profile.html',
                         student=student,
                         test_history=test_history,
                         chart_data=chart_data,
                         is_class_teacher=is_class_teacher)

@app.route('/teacher/student/<student_id>/edit', methods=['POST'])
@teacher_required
def teacher_edit_student(student_id):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()

    # Get student class
    c.execute("SELECT class FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404

    # Both class teachers AND subject teachers in that class can edit
    c.execute("""SELECT id FROM teacher_assignments
                 WHERE teacher_id=? AND class=?""", (teacher_id, student['class']))
    assignment = c.fetchone()
    if not assignment:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    name = request.form.get('name')
    admission_no = request.form.get('admission_no')
    class_ = request.form.get('class')
    section = request.form.get('section')
    dob = request.form.get('dob')
    house = request.form.get('house')
    parents_name = request.form.get('parents_name')
    address = request.form.get('address')

    picture_path = save_student_picture(request.files.get('picture'), student_id)

    if picture_path:
        c.execute("""UPDATE students
                     SET name=?, admission_no=?, class=?, section=?, dob=?, house=?,
                         parents_name=?, address=?, picture=?
                     WHERE student_id=?""",
                 (name, admission_no, class_, section, dob, house, parents_name, address,
                  picture_path, student_id))
    else:
        c.execute("""UPDATE students
                     SET name=?, admission_no=?, class=?, section=?, dob=?, house=?,
                         parents_name=?, address=?
                     WHERE student_id=?""",
                 (name, admission_no, class_, section, dob, house, parents_name, address, student_id))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Student details updated'})

@app.route('/teacher/student/add', methods=['POST'])
@teacher_required
def teacher_add_student():
    teacher_id = session.get('teacher_id')
    name = request.form.get('name', '').strip()
    student_id = request.form.get('student_id', '').strip()
    class_ = request.form.get('class', '').strip()
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject', '').strip()
    admission_no = request.form.get('admission_no', '').strip()
    dob = request.form.get('dob', '').strip()
    house = request.form.get('house', '').strip()
    parents_name = request.form.get('parents_name', '').strip()
    address = request.form.get('address', '').strip()

    if not name or not student_id or not class_:
        return jsonify({'status': 'error', 'message': 'Name, Student ID and Class are required'}), 400

    # Verify teacher has access to this class
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM teacher_assignments WHERE teacher_id=? AND class=?", (teacher_id, class_))
    if not c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'You are not assigned to this class'}), 403

    # Check duplicate
    c.execute("SELECT student_id FROM students WHERE student_id=?", (student_id,))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student ID already exists'}), 400

    picture_path = save_student_picture(request.files.get('picture'), student_id)

    c.execute("""INSERT INTO students
                 (student_id, name, class, subject, section, admission_no, dob, house, parents_name, address, picture, status)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,'Not Started')""",
             (student_id, name, class_, subject, section, admission_no, dob, house, parents_name, address, picture_path))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': f'Student {name} added successfully'})

@app.route('/teacher/student/<student_id>/delete', methods=['DELETE'])
@teacher_required
def teacher_delete_student(student_id):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT class FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
    c.execute("SELECT id FROM teacher_assignments WHERE teacher_id=? AND class=?", (teacher_id, student['class']))
    if not c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    c.execute("DELETE FROM students WHERE student_id=?",           (student_id,))
    c.execute("DELETE FROM responses WHERE student_id=?",          (student_id,))
    c.execute("DELETE FROM results WHERE student_id=?",            (student_id,))
    c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))  # Bug #3 fix
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Student deleted'})

@app.route('/teacher/profile', methods=['GET', 'POST'])
@teacher_required
def teacher_profile():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        address = request.form.get('address', '').strip()
        new_password = request.form.get('new_password', '').strip()
        pic_path = save_teacher_picture(request.files.get('picture'), teacher_id)
        if new_password:
            hashed = hash_password(new_password)
            if pic_path:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, password=?, picture=? WHERE id=?",
                         (name, email, address, hashed, pic_path, teacher_id))
            else:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, password=? WHERE id=?",
                         (name, email, address, hashed, teacher_id))
        else:
            if pic_path:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, picture=? WHERE id=?",
                         (name, email, address, pic_path, teacher_id))
            else:
                c.execute("UPDATE teachers SET name=?, email=?, address=? WHERE id=?",
                         (name, email, address, teacher_id))
        conn.commit()
        session['teacher_name'] = name
        conn.close()
        return jsonify({'status': 'success', 'message': 'Profile updated'})

    c.execute("SELECT * FROM teachers WHERE id=?", (teacher_id,))
    teacher = dict(c.fetchone())
    c.execute("SELECT class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_profile.html', teacher=teacher, assignments=assignments)

@app.route('/teacher/test/<int:result_id>/responses')
@teacher_required
def teacher_view_test_responses(result_id):
    conn = get_db()
    c = conn.cursor()
    
    # Get test result details
    c.execute("SELECT * FROM results WHERE id=?", (result_id,))
    result = c.fetchone()
    if not result:
        conn.close()
        return "Test result not found", 404
    result = dict(result)
    
    # Get student responses with full question data
    c.execute("""
        SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d, 
               q.correct_answer, q.image_path, r.selected_option
        FROM questions q
        LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
        WHERE q.class = ? AND q.subject = ?
        ORDER BY q.id
    """, (result['student_id'], result['class'], result['subject']))
    rows = c.fetchall()
    conn.close()
    
    # Process questions to map correct_answer to a letter
    questions = []
    for row in rows:
        q = dict(row)
        selected = q['selected_option'] or ''
        correct_raw = q['correct_answer']
        correct_letter = None
        
        if correct_raw:
            correct_raw = correct_raw.strip()
            # If already a single letter, use it
            if len(correct_raw) == 1 and correct_raw.upper() in ('A','B','C','D'):
                correct_letter = correct_raw.upper()
            else:
                # Try to match against option texts or "option_x"
                if correct_raw == q['option_a'] or correct_raw.lower() == 'option_a':
                    correct_letter = 'A'
                elif correct_raw == q['option_b'] or correct_raw.lower() == 'option_b':
                    correct_letter = 'B'
                elif correct_raw == q['option_c'] or correct_raw.lower() == 'option_c':
                    correct_letter = 'C'
                elif correct_raw == q['option_d'] or correct_raw.lower() == 'option_d':
                    correct_letter = 'D'
        
        is_correct = False
        if selected and correct_letter:
            is_correct = (selected.strip().upper() == correct_letter)
        
        questions.append({
            'id': q['id'],
            'question': q['question'],
            'options': {
                'A': q['option_a'],
                'B': q['option_b'],
                'C': q['option_c'],
                'D': q['option_d']
            },
            'image_path': q['image_path'],
            'correct': correct_letter or '?',
            'selected': selected,
            'is_correct': is_correct
        })
    
    return render_template('teacher_test_responses.html',
                         result=result,
                         questions=questions)

@app.route('/teacher/generate_student_report/<student_id>')
@teacher_required
def teacher_generate_student_report(student_id):
    conn = get_db()
    c = conn.cursor()
    
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = dict(c.fetchone())
    
    c.execute("""SELECT subject, chapter, score, total_questions, percentage, test_date
                 FROM results WHERE student_id=?
                 ORDER BY test_date DESC""", (student_id,))
    test_history = [dict(row) for row in c.fetchall()]
    
    conn.close()
    
    # Generate PDF (simplified version)
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial; margin: 20px; }}
            h1 {{ color: #2e7d32; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #2e7d32; color: white; }}
        </style>
    </head>
    <body>
        <h1>Student Progress Report</h1>
        <p><strong>Name:</strong> {student.get('name', 'N/A')}</p>
        <p><strong>Admission No:</strong> {student.get('admission_no', 'N/A')}</p>
        <p><strong>Class:</strong> {student.get('class', 'N/A')} - {student.get('section', 'N/A')}</p>
        <p><strong>House:</strong> {student.get('house', 'N/A')}</p>
        
        <h2>Test History</h2>
        <table>
            <tr>
                <th>Date</th>
                <th>Subject</th>
                <th>Chapter</th>
                <th>Marks</th>
                <th>Percentage</th>
            </tr>
    """
    
    for test in test_history:
        html_content += f"""
            <tr>
                <td>{test.get('test_date', '')[:10]}</td>
                <td>{test.get('subject', 'N/A')}</td>
                <td>{test.get('chapter', 'N/A')}</td>
                <td>{test.get('score', 0)}/{test.get('total_questions', 0)}</td>
                <td>{test.get('percentage', 0):.1f}%</td>
            </tr>
        """
    
    html_content += """
        </table>
    </body>
    </html>
    """
    
    from weasyprint import HTML
    pdf = HTML(string=html_content).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=student_report_{student_id}.pdf'
    
    return response

# ========================
# ADMIN TEACHER MANAGEMENT
# ========================

@app.route('/admin/teachers')
@admin_required
def admin_teachers():
    return render_template('admin_teachers.html')

@app.route('/admin/teachers/data')
@admin_required
def admin_teachers_data():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT t.id, t.name, t.mobile, t.email, t.status, t.created_at,
                        GROUP_CONCAT(DISTINCT ta.class || ' - ' || ta.subject) as assignments
                 FROM teachers t
                 LEFT JOIN teacher_assignments ta ON t.id = ta.teacher_id
                 GROUP BY t.id
                 ORDER BY t.name""")
    teachers = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'teachers': teachers})

@app.route('/admin/teacher/add', methods=['POST'])
@admin_required
def admin_add_teacher():
    name = request.form.get('name')
    mobile = request.form.get('mobile')
    password = request.form.get('password')
    email = request.form.get('email')
    address = request.form.get('address')
    
    conn = get_db()
    c = conn.cursor()
    
    # Check if mobile exists
    c.execute("SELECT id FROM teachers WHERE mobile=?", (mobile,))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Mobile number already exists'}), 400
    
    hashed_pw = hash_password(password)
    c.execute("""INSERT INTO teachers (name, mobile, password, email, address)
                 VALUES (?,?,?,?,?)""",
             (name, mobile, hashed_pw, email, address))
    conn.commit()
    conn.close()
    
    return jsonify({'status': 'success', 'message': 'Teacher added successfully'})

@app.route('/admin/teacher/<int:teacher_id>/assign', methods=['POST'])
@admin_required
def admin_assign_teacher(teacher_id):
    class_ = request.form.get('class', '').strip()
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject', '').strip()
    is_class_teacher = request.form.get('is_class_teacher') == 'true'

    if not class_ or not subject:
        return jsonify({'status': 'error', 'message': 'Class and subject are required'}), 400

    conn = get_db()
    c = conn.cursor()

    # Enforce 1:1 class teacher constraint
    if is_class_teacher:
        c.execute("""SELECT ta.id, t.name FROM teacher_assignments ta
                     JOIN teachers t ON ta.teacher_id = t.id
                     WHERE ta.class=? AND ta.section=? AND ta.is_class_teacher=1
                     AND ta.teacher_id != ?""",
                 (class_, section, teacher_id))
        existing_ct = c.fetchone()
        if existing_ct:
            conn.close()
            return jsonify({'status': 'error',
                           'message': f"Class {class_} {section} already has a Class Teacher: {existing_ct['name']}"}), 400

    # Check if this exact assignment already exists
    c.execute("""SELECT id FROM teacher_assignments
                 WHERE teacher_id=? AND class=? AND section=? AND subject=?""",
             (teacher_id, class_, section, subject))
    existing = c.fetchone()

    if existing:
        c.execute("""UPDATE teacher_assignments
                     SET is_class_teacher=?
                     WHERE id=?""", (1 if is_class_teacher else 0, existing['id']))
    else:
        c.execute("""INSERT INTO teacher_assignments (teacher_id, class, section, subject, is_class_teacher)
                     VALUES (?,?,?,?,?)""",
                 (teacher_id, class_, section, subject, 1 if is_class_teacher else 0))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Assignment saved successfully'})

@app.route('/admin/teacher/<int:teacher_id>/remove_assignment/<int:assign_id>', methods=['DELETE'])
@admin_required
def admin_remove_assignment(teacher_id, assign_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM teacher_assignments WHERE id=? AND teacher_id=?", (assign_id, teacher_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/admin/teacher/<int:teacher_id>', methods=['GET'])
@admin_required
def admin_get_teacher(teacher_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name, mobile, email, address, picture, status FROM teachers WHERE id=?", (teacher_id,))
    teacher = c.fetchone()
    if not teacher:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
    c.execute("SELECT id, class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'teacher': dict(teacher), 'assignments': assignments})

@app.route('/admin/teacher/<int:teacher_id>/update', methods=['POST'])
@admin_required
def admin_update_teacher(teacher_id):
    conn = get_db()
    c = conn.cursor()
    name = request.form.get('name', '').strip()
    mobile = request.form.get('mobile', '').strip()
    email = request.form.get('email', '').strip()
    address = request.form.get('address', '').strip()
    new_password = request.form.get('password', '').strip()

    pic_file = request.files.get('picture')
    pic_path = save_teacher_picture(pic_file, teacher_id)

    if new_password:
        hashed = hash_password(new_password)
        if pic_path:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, password=?, picture=? WHERE id=?",
                     (name, mobile, email, address, hashed, pic_path, teacher_id))
        else:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, password=? WHERE id=?",
                     (name, mobile, email, address, hashed, teacher_id))
    else:
        if pic_path:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, picture=? WHERE id=?",
                     (name, mobile, email, address, pic_path, teacher_id))
        else:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=? WHERE id=?",
                     (name, mobile, email, address, teacher_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Teacher updated'})

@app.route('/admin/teacher/upload_csv', methods=['POST'])
@admin_required
def admin_upload_teachers_csv():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'status': 'error', 'message': 'Invalid file'}), 400
    
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    conn = get_db()
    c = conn.cursor()
    added = 0
    
    with open(filepath, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            name = row.get('name', '').strip()
            mobile = row.get('mobile', '').strip()
            password = row.get('password', 'teacher123')  # Default password
            email = row.get('email', '').strip()
            
            if not name or not mobile:
                continue
            
            # Check if exists
            c.execute("SELECT id FROM teachers WHERE mobile=?", (mobile,))
            if c.fetchone():
                continue
            
            hashed_pw = hash_password(password)
            c.execute("""INSERT INTO teachers (name, mobile, password, email)
                         VALUES (?,?,?,?)""",
                     (name, mobile, hashed_pw, email))
            added += 1
    
    conn.commit()
    conn.close()
    os.remove(filepath)
    
    return jsonify({'status': 'success', 'message': f'{added} teachers added successfully'})

# ===================================================
# TEST PAPER MODULE (CSV naming: Class_Subject_TestNo)
# ===================================================

@app.route('/admin/test_papers')
@admin_required
def admin_test_papers():
    return render_template('test_papers.html', role='admin')

@app.route('/teacher/test_papers')
@teacher_required
def teacher_test_papers():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('test_papers.html', role='teacher', assignments=assignments)

@app.route('/api/test_papers')
def api_test_papers():
    role = request.args.get('role', 'admin')
    teacher_id = session.get('teacher_id') if role == 'teacher' else None
    conn = get_db()
    c = conn.cursor()
    if role == 'teacher' and teacher_id:
        c.execute("""SELECT ta.class, ta.subject FROM teacher_assignments ta WHERE ta.teacher_id=?""", (teacher_id,))
        assigns = c.fetchall()
        papers = []
        for a in assigns:
            c.execute("""SELECT tp.*, t.name as uploader_name FROM test_papers tp
                         LEFT JOIN teachers t ON tp.uploaded_by = t.id AND tp.uploader_type='teacher'
                         WHERE tp.class=? AND tp.subject=?
                         ORDER BY tp.created_at DESC""", (a['class'], a['subject']))
            papers.extend([dict(r) for r in c.fetchall()])
    else:
        c.execute("""SELECT tp.* FROM test_papers tp ORDER BY tp.created_at DESC""")
        papers = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'papers': papers})

@app.route('/api/test_papers/upload', methods=['POST'])
def upload_test_paper():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    file = request.files.get('csv_file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'status': 'error', 'message': 'CSV file required'}), 400

    # Parse filename convention: Class_Subject_TestNo.csv
    raw_name = os.path.splitext(file.filename)[0]
    parts = raw_name.split('_')
    if len(parts) < 3:
        return jsonify({'status': 'error',
                        'message': 'Filename must follow format: Class_Subject_TestNo (e.g. 10A_Math_Test01)'}), 400

    class_ = parts[0].strip()
    test_no = parts[-1].strip()
    subject = '_'.join(parts[1:-1]).strip()
    section = ''
    # If class has letter suffix treat it as section: e.g. 10A -> class=10, section=A
    if class_ and class_[-1].isalpha():
        section = class_[-1].upper()
        class_ = class_[:-1]

    uploader_type = 'admin' if session.get('admin_logged_in') else 'teacher'
    uploaded_by = 'admin' if uploader_type == 'admin' else str(session.get('teacher_id'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # Import questions
    conn = get_db()
    c = conn.cursor()
    question_count = 0
    try:
        with open(filepath, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                q = row.get('question', '').strip()
                if not q:
                    continue
                c.execute("""INSERT INTO questions
                             (class, subject, chapter, question, option_a, option_b, option_c, option_d, correct_answer)
                             VALUES (?,?,?,?,?,?,?,?,?)""",
                         (class_, subject, test_no, q,
                          row.get('option a', row.get('option_a', '')),
                          row.get('option b', row.get('option_b', '')),
                          row.get('option c', row.get('option_c', '')),
                          row.get('option d', row.get('option_d', '')),
                          row.get('correct_answer', '')))
                question_count += 1

        c.execute("""INSERT INTO test_papers
                     (filename, class, section, subject, test_no, uploaded_by, uploader_type, question_count)
                     VALUES (?,?,?,?,?,?,?,?)""",
                 (raw_name, class_, section, subject, test_no, uploaded_by, uploader_type, question_count))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        os.remove(filepath)
        return jsonify({'status': 'error', 'message': str(e)}), 500
    conn.close()
    os.remove(filepath)
    return jsonify({'status': 'success',
                    'message': f'Uploaded {question_count} questions for {class_}{section} {subject} - {test_no}'})

@app.route('/api/test_papers/<int:paper_id>/toggle', methods=['POST'])
def toggle_test_paper(paper_id):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM test_papers WHERE id=?", (paper_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Not found'}), 404
    new_status = 0 if row['is_active'] else 1
    c.execute("UPDATE test_papers SET is_active=? WHERE id=?", (new_status, paper_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'is_active': new_status})

@app.route('/api/test_papers/<int:paper_id>/pdf')
def test_paper_pdf(paper_id):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return redirect(url_for('teacher_login'))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_papers WHERE id=?", (paper_id,))
    paper = c.fetchone()
    if not paper:
        conn.close()
        return "Not found", 404
    paper = dict(paper)

    # Bug #4 Fix: query by BOTH test_no and chapter columns to catch all upload scenarios
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                        correct_answer, image_path,
                        COALESCE(question_type, 'MCQ') as question_type
                 FROM questions
                 WHERE class=? AND subject=?
                   AND (test_no=? OR chapter=?)
                 ORDER BY question_type, id""",
             (paper['class'], paper['subject'], paper['test_no'], paper['test_no']))
    questions = [dict(r) for r in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext  = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"

    # Use v2 template (handles all question types properly)
    rendered = render_template('question_paper_template_v2.html',
        school_name=school_name,
        logo_base64=logo_base64,
        class_name=f"{paper['class']}{paper['section']}",
        subject=paper['subject'],
        chapter=paper['test_no'],
        test_no=paper['test_no'],
        questions=questions,
        total_marks=len(questions),
        date=datetime.datetime.now().strftime('%d/%m/%Y'))

    pdf = HTML(string=rendered, base_url=request.base_url).write_pdf()
    resp = make_response(pdf)
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename={paper["filename"]}.pdf'
    return resp

# ===================================================
# TEACHER MONITORING (own classes only)
# ===================================================

@app.route('/teacher/monitoring')
@teacher_required
def teacher_monitoring():
    return render_template('teacher_monitoring.html')

@app.route('/teacher/monitoring/data')
@teacher_required
def teacher_monitoring_data():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    # Get teacher's assigned classes
    c.execute("SELECT DISTINCT class FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    classes = [r['class'] for r in c.fetchall()]
    if not classes:
        conn.close()
        return jsonify({'students': [], 'exam_is_active': False})

    placeholders = ','.join(['?'] * len(classes))
    c.execute(f"""SELECT s.student_id, s.name, s.class, s.subject, s.section, s.picture,
                         s.status, s.exam_started_at,
                         (SELECT COUNT(*) FROM responses r WHERE r.student_id=s.student_id) as answered
                  FROM students s
                  WHERE s.class IN ({placeholders})
                  ORDER BY s.class, s.name""", classes)
    students = [dict(r) for r in c.fetchall()]
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    conn.close()
    return jsonify({'students': students, 'exam_is_active': bool(exam_row['is_active']) if exam_row else False})

# ===================================================
# CLASS REPORT (Excel + PDF)
# ===================================================

@app.route('/admin/class_report')
@admin_required
def admin_class_report():
    return render_template('class_report.html', role='admin')

@app.route('/teacher/class_report')
@teacher_required
def teacher_class_report():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('class_report.html', role='teacher', assignments=assignments)

@app.route('/api/class_report/generate', methods=['POST'])
def generate_class_report():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401

    data = request.get_json() or {}
    class_       = data.get('class', '').strip()
    section      = data.get('section', '').strip()
    subject      = data.get('subject', '').strip()
    date_from    = data.get('date_from', '').strip()
    date_to      = data.get('date_to', '').strip()
    export_format = data.get('format', 'excel')
    sort_by      = data.get('sort_by', 'name')

    # Bug #5 Fix: require class
    if not class_:
        return jsonify({'status': 'error', 'message': 'Please enter a class to generate the report.'}), 400

    conn = get_db()
    c = conn.cursor()

    query  = """SELECT DISTINCT s.student_id, s.name, s.admission_no, s.class, s.section,
                       s.house, s.parents_name, s.picture
                FROM students s WHERE s.class=?"""
    params = [class_]
    if section:
        query += " AND (s.section=? OR s.section IS NULL OR s.section='')"
        params.append(section)

    c.execute(query, params)
    students = [dict(r) for r in c.fetchall()]

    report_data = []
    for st in students:
        result_query  = """SELECT subject, chapter, score, total_questions, percentage, test_date
                           FROM results WHERE student_id=?"""
        result_params = [st['student_id']]
        if subject:
            result_query += " AND subject=?"
            result_params.append(subject)
        if date_from:
            result_query += " AND DATE(test_date)>=?"
            result_params.append(date_from)
        if date_to:
            result_query += " AND DATE(test_date)<=?"
            result_params.append(date_to)
        result_query += " ORDER BY test_date DESC"
        c.execute(result_query, result_params)
        results    = [dict(r) for r in c.fetchall()]
        avg_pct    = round(sum(r['percentage'] or 0 for r in results) / len(results), 1) if results else 0
        st['results']         = results
        st['avg_percentage']  = avg_pct
        st['total_tests']     = len(results)
        report_data.append(st)

    school_name = get_setting('school_name', 'RRB Group of Schools')
    conn.close()  # single close here

    if sort_by == 'percentage':
        report_data.sort(key=lambda x: x['avg_percentage'], reverse=True)
    else:
        report_data.sort(key=lambda x: x['name'])

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    label     = f"{class_}{section}_{subject or 'All'}_{timestamp}"

    # Bug #5 Fix: json preview no longer calls conn.close() again
    if export_format == 'json':
        return jsonify({'status': 'success', 'students': report_data})

    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class Report"

        # Header
        ws.merge_cells('A1:H1')
        ws['A1'] = school_name
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Class: {class_}{section}  |  Subject: {subject or 'All'}  |  Generated: {datetime.datetime.now().strftime('%d-%m-%Y')}"
        ws['A2'].font = openpyxl.styles.Font(bold=True, size=12)
        ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.append([])

        headers = ['#', 'Admission No', 'Name', 'Class', 'Section', 'House', 'Total Tests', 'Avg %']
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color='1b5e20', end_color='1b5e20', fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for i, st in enumerate(report_data, 1):
            ws.append([i, st.get('admission_no') or '', st['name'], st['class'],
                       st.get('section') or '', st.get('house') or '',
                       st['total_tests'], f"{st['avg_percentage']}%"])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12) + 2

        filepath = os.path.join(app.config['EXPORT_FOLDER'], f"class_report_{label}.xlsx")
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name=f"class_report_{label}.xlsx")

    else:  # PDF
        rows_html = ''
        for i, st in enumerate(report_data, 1):
            rows_html += f"""<tr>
                <td>{i}</td>
                <td>{st.get('admission_no') or '—'}</td>
                <td><strong>{st['name']}</strong></td>
                <td>{st['class']}{st.get('section') or ''}</td>
                <td>{st.get('house') or '—'}</td>
                <td>{st['total_tests']}</td>
                <td><strong>{st['avg_percentage']}%</strong></td>
            </tr>"""

        html = f"""<!DOCTYPE html><html><head>
        <style>
            body{{font-family:Arial,sans-serif;margin:20px;font-size:12px}}
            h1{{color:#1b5e20;text-align:center;margin-bottom:5px}}
            p{{text-align:center;color:#555;margin:3px}}
            table{{width:100%;border-collapse:collapse;margin-top:20px}}
            th{{background:#1b5e20;color:white;padding:8px;text-align:left}}
            td{{border:1px solid #ddd;padding:7px}}
            tr:nth-child(even){{background:#f9f9f9}}
        </style></head><body>
        <h1>{school_name}</h1>
        <p>Class: <strong>{class_}{section}</strong> &nbsp;|&nbsp; Subject: <strong>{subject or 'All'}</strong></p>
        <p>Generated: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}</p>
        <table>
            <tr><th>#</th><th>Adm No</th><th>Name</th><th>Class</th><th>House</th><th>Tests</th><th>Avg %</th></tr>
            {rows_html}
        </table></body></html>"""

        pdf = HTML(string=html).write_pdf()
        resp = make_response(pdf)
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename=class_report_{label}.pdf'
        return resp

# ═══════════════════════════════════════════════════════════════
# BULLETIN BOARD
# ═══════════════════════════════════════════════════════════════

@app.route('/api/bulletins')
def api_bulletins():
    conn = get_db()
    c = conn.cursor()
    class_filter = request.args.get('class', '')
    query = """SELECT b.id, b.title, b.content, b.poster_type, b.target_class, b.created_at,
                      COALESCE(t.name,'Admin') as poster_name
               FROM bulletins b
               LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
               WHERE b.is_active=1"""
    params = []
    if class_filter:
        query += " AND (b.target_class='' OR b.target_class=?)"
        params.append(class_filter)
    query += " ORDER BY b.created_at DESC LIMIT 15"
    c.execute(query, params)
    bulletins = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'bulletins': bulletins})

@app.route('/api/bulletins/post', methods=['POST'])
def post_bulletin():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401
    title   = request.form.get('title','').strip()
    content = request.form.get('content','').strip()
    target  = request.form.get('target_class','').strip()
    if not title or not content:
        return jsonify({'status': 'error', 'message': 'Title and content required'}), 400
    if session.get('admin_logged_in'):
        posted_by   = 'admin'
        poster_type = 'admin'
    else:
        posted_by   = str(session.get('teacher_id'))
        poster_type = 'teacher'
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO bulletins (title, content, posted_by, poster_type, target_class) VALUES (?,?,?,?,?)",
              (title, content, posted_by, poster_type, target))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Bulletin posted'})

@app.route('/api/bulletins/<int:bid>/delete', methods=['DELETE'])
def delete_bulletin(bid):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE bulletins SET is_active=0 WHERE id=?", (bid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

# ═══════════════════════════════════════════════════════════════
# ADMIN: UNLOCK STUDENT CLASS
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/student/<student_id>/unlock_class', methods=['POST'])
@admin_required
def admin_unlock_student_class(student_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Class lock removed. Student can re-register.'})

# ═══════════════════════════════════════════════════════════════
# ENHANCED AI TEST GENERATOR (Multi-type + Dual output mode)
# ═══════════════════════════════════════════════════════════════

@app.route('/teacher/create_test_v2', methods=['GET', 'POST'])
@teacher_required
def teacher_create_test_v2():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, section, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]

    if request.method == 'POST':
        import urllib.request, json as json_lib

        class_   = request.form.get('class','').strip()
        section  = request.form.get('section','').strip()
        subject  = request.form.get('subject','').strip()
        chapter  = request.form.get('chapter','').strip()
        test_no  = request.form.get('test_no','').strip()
        remark   = request.form.get('remark','').strip()
        output_mode = request.form.get('output_mode','cbt')   # cbt or print
        method   = request.form.get('method','ai')

        mcq_count         = int(request.form.get('mcq_count', 0) or 0)
        assertion_count   = int(request.form.get('assertion_count', 0) or 0)
        very_short_count  = int(request.form.get('very_short_count', 0) or 0)
        short_count       = int(request.form.get('short_count', 0) or 0)
        long_count        = int(request.form.get('long_count', 0) or 0)
        case_study_count  = int(request.form.get('case_study_count', 0) or 0)
        total_q = mcq_count + assertion_count + very_short_count + short_count + long_count + case_study_count

        if method == 'upload':
            file = request.files.get('csv_file')
            if file and file.filename.endswith('.csv'):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                with open(filepath, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    inserted = 0
                    for row in reader:
                        q = row.get('question','').strip()
                        if not q: continue
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,row.get('question_type','MCQ'),q,
                                  row.get('option a',row.get('option_a','')),
                                  row.get('option b',row.get('option_b','')),
                                  row.get('option c',row.get('option_c','')),
                                  row.get('option d',row.get('option_d','')),
                                  row.get('correct_answer','')))
                        inserted += 1
                conn.commit()
                os.remove(filepath)
                conn.close()
                return jsonify({'status':'success','message':f'{inserted} questions uploaded'})
            conn.close()
            return jsonify({'status':'error','message':'Invalid CSV file'}), 400

        # ── AI GENERATION ─────────────────────────────────────
        gemini_api_key = os.environ.get('GEMINI_API_KEY','')
        if not gemini_api_key:
            # Try reading from apikey.env file
            env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'apikey.env')
            if os.path.exists(env_path):
                gemini_api_key = open(env_path).read().strip()
        if not gemini_api_key:
            conn.close()
            return jsonify({'status':'error','message':'GEMINI_API_KEY not set'}), 500

        type_instructions = []
        if mcq_count:        type_instructions.append(f"{mcq_count} MCQ (4 options, mark correct_answer as option_a/b/c/d)")
        if assertion_count:  type_instructions.append(f"{assertion_count} Assertion-Reason (Assertion + Reason format)")
        if very_short_count: type_instructions.append(f"{very_short_count} Very Short Answer (1-2 lines)")
        if short_count:      type_instructions.append(f"{short_count} Short Answer (3-5 lines)")
        if long_count:       type_instructions.append(f"{long_count} Long Answer (detailed)")
        if case_study_count: type_instructions.append(f"{case_study_count} Case Study based")

        if not type_instructions:
            conn.close()
            return jsonify({'status':'error','message':'Enter at least 1 question quantity'}), 400

        ai_prompt = f"""You are an expert CBSE/ICSE academic question paper generator.
Output ONLY a valid JSON array — no markdown fences, no text outside the JSON array.

PAPER DETAILS:
Class: {class_}{section} | Subject: {subject} | Chapter: {chapter} | Test No: {test_no}
Teacher instructions: {remark if remark else 'Standard difficulty, balanced coverage of the chapter'}

QUESTION TYPES REQUIRED:
{chr(10).join(type_instructions)}

══════════════════════════════════════════════════════════
STRICT SCHEMA — every question must have ALL of these keys:
══════════════════════════════════════════════════════════
{{
  "question_type": "MCQ" | "Assertion-Reason" | "Very Short" | "Short" | "Long" | "Case Study",
  "content_type":  "math" | "chemistry" | "physics" | "biology" | "text",
  "question":      "<string with LaTeX or mhchem where needed>",
  "option_a":      "<string, empty for non-MCQ>",
  "option_b":      "<string, empty for non-MCQ>",
  "option_c":      "<string, empty for non-MCQ>",
  "option_d":      "<string, empty for non-MCQ>",
  "correct_answer":"option_a"|"option_b"|"option_c"|"option_d"|"N/A",
  "smiles":        "<SMILES string for 2-D molecular diagram, else empty>",
  "image_prompt":  "<Imagen 3 prompt for diagram, else empty>",
  "marks":         1|2|3|4|5
}}

══════════════════════════
MATHEMATICS — LATEX RULES
══════════════════════════
Use MathJax-compatible LaTeX. ALWAYS escape backslashes in JSON (\\frac not \frac).

• Inline expressions  →  $...$
  "Find $\\frac{{d}}{{dx}}(x^2 + 3x)$ at $x = 2$"
  "$\\vec{{F}} = 3\\hat{{i}} - 4\\hat{{j}} + 5\\hat{{k}}$"

• Display/block expressions  →  $$...$$
  "$$\\int_{{0}}^{{\\pi}} \\sin x\\, dx = 2$$"
  "$$\\begin{{pmatrix}} 1 & 2 \\\\\\\\ 3 & 4 \\end{{pmatrix}}$$"

• Limits:  "$\\lim_{{x \\to 0}} \\dfrac{{\\sin x}}{{x}} = 1$"
• Roots:   "$\\sqrt{{b^2 - 4ac}}$"
• Integrals: "$\\int_{{a}}^{{b}} f(x)\\, dx$"
• Summation: "$\\sum_{{n=1}}^{{\\infty}} \\frac{{1}}{{n^2}}$"
• Options containing math must also use $...$:  "option_a": "$x = \\frac{{1}}{{2}}$"
• NEVER use raw Unicode math symbols (×, ÷, √, ∫, ², ³) — use LaTeX commands.

══════════════════════════════
CHEMISTRY — mhchem RULES
══════════════════════════════
Use \\ce{{}} inside $ $ for ALL chemical formulas and equations.

• Basic equation:  "$\\ce{{H2SO4 + 2NaOH -> Na2SO4 + 2H2O}}$"
• Combustion:      "$\\ce{{CH4 + 2O2 -> CO2 + 2H2O}}$"
• Equilibrium:     "$\\ce{{N2 + 3H2 <=> 2NH3}}$"
• With states:     "$\\ce{{CaCO3(s) -> CaO(s) + CO2(g)}}$"
• Ionic:           "$\\ce{{Fe^{{2+}} + 2e- -> Fe}}$"
• Precipitation:   "$\\ce{{Ag+ + Cl- -> AgCl v}}$"
• Gas evolved:     "$\\ce{{Zn + H2SO4 -> ZnSO4 + H2 ^}}$"
• Named compound inline: "The reaction of $\\ce{{H2O}}$ with $\\ce{{CO2}}$..."

For 2-D structural diagrams set "smiles" field (SMILES notation):
  Benzene="c1ccccc1"  Ethanol="CCO"  Acetic acid="CC(=O)O"
  Toluene="Cc1ccccc1"  Aspirin="CC(=O)Oc1ccccc1C(=O)O"
  Naphthalene="c1ccc2ccccc2c1"  Cyclohexane="C1CCCCC1"

══════════════════════════════════════
BIOLOGY / PHYSICS — IMAGE PROMPT RULES
══════════════════════════════════════
Set image_prompt ONLY when a non-symbolic diagram is pedagogically required.
Write a specific, detailed Imagen 3 prompt:
  GOOD: "Labeled scientific diagram of human heart: left ventricle, right ventricle,
         aorta, pulmonary artery, vena cava, mitral valve, tricuspid valve.
         Medical textbook illustration, white background, clear black labels."
  BAD:  "a heart"

══════════════════════════
ASSERTION-REASON OPTIONS
══════════════════════════
Always use exactly these four options:
  option_a: "Both A and R are true and R is the correct explanation of A"
  option_b: "Both A and R are true but R is NOT the correct explanation of A"
  option_c: "A is true but R is false"
  option_d: "A is false but R is true"

Non-MCQ questions: set option_a/b/c/d="" and correct_answer="N/A"

Output the JSON array now:"""

        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_api_key}"
            payload = json_lib.dumps({
                "contents": [{"parts": [{"text": ai_prompt}]}],
                "generationConfig": {"temperature": 0.7, "maxOutputTokens": 8192}
            }).encode('utf-8')
            req = urllib.request.Request(url, data=payload,
                                          headers={'Content-Type':'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json_lib.loads(resp.read().decode('utf-8'))

            response_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
            if '```' in response_text:
                for part in response_text.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): response_text = part; break

            questions = json_lib.loads(response_text)

            if output_mode == 'cbt':
                # BUG-004 FIX: check if test_no already exists for this class+subject
                c.execute("""SELECT COUNT(*) as cnt FROM questions
                             WHERE class=? AND subject=? AND test_no=?""",
                         (class_, subject, test_no))
                existing_count = c.fetchone()['cnt']
                if existing_count > 0:
                    conn.close()
                    return jsonify({
                        'status': 'error',
                        'message': f'Test No "{test_no}" already has {existing_count} questions for {subject} Class {class_}. '
                                   f'Use a different Test No (e.g., Test02) to create a separate test.'
                    }), 400

                # Push to questions table (MCQ only for CBT)
                inserted = 0
                for q in questions:
                    if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,
                                  q.get('question_type','MCQ'),q.get('question',''),
                                  q.get('option_a',''),q.get('option_b',''),
                                  q.get('option_c',''),q.get('option_d',''),
                                  q.get('correct_answer','')))
                        inserted += 1
                # Register test paper
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                         (f"{class_}{section}_{subject}_{test_no}",class_,section,subject,test_no,
                          str(teacher_id),'teacher',inserted))
                conn.commit()

            # Save to history
            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (str(teacher_id),class_,section,subject,chapter,test_no,output_mode,
                      len(questions),mcq_count,assertion_count,very_short_count,
                      short_count,long_count,case_study_count,remark))
            conn.commit()
            history_id = c.lastrowid
            conn.close()

            if output_mode == 'print':
                # Feature #6: redirect to print preview page (browser print dialog)
                return jsonify({'status':'success','mode':'print',
                                'preview_url': f'/api/test_generation_history/{history_id}/print_preview',
                                'history_id': history_id,
                                'message':f'{len(questions)} questions generated — opening print preview'})
            else:
                return jsonify({'status':'success','mode':'cbt',
                                'message':f'{len(questions)} questions generated and pushed to CBT. Test No: {test_no}'})

        except urllib.error.HTTPError as e:
            conn.close()
            return jsonify({'status':'error','message':f'Gemini error: {e.read().decode()}'}), 500
        except Exception as e:
            conn.close()
            return jsonify({'status':'error','message':str(e)}), 500

    conn.close()
    # Build class→subjects map for dynamic subject filtering (Features #2, #3)
    class_subject_map = {}
    for a in assignments:
        cls = a['class']
        if cls not in class_subject_map:
            class_subject_map[cls] = []
        if a['subject'] not in class_subject_map[cls]:
            class_subject_map[cls].append(a['subject'])
    return render_template('teacher_create_test_v2.html',
                           assignments=assignments,
                           class_subject_map=class_subject_map)

@app.route('/teacher/print_test/<int:result_id>')
@teacher_required
def teacher_print_test(result_id):
    conn = get_db()
    c = conn.cursor()
    # Get the test result
    c.execute("SELECT * FROM results WHERE id=?", (result_id,))
    result = c.fetchone()
    if not result:
        conn.close()
        return "Test result not found", 404
    result = dict(result)

    # Check teacher access (optional, if you want to restrict)
    teacher_id = session['teacher_id']
    if not teacher_has_access(teacher_id, result['class'], result['subject']):
        conn.close()
        return "Access denied", 403

    # Fetch questions for this test's class/subject
    c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer, image_path FROM questions WHERE class=? AND subject=? ORDER BY id",
              (result['class'], result['subject']))
    questions = [dict(row) for row in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode('utf-8')}"

    total_marks = len(questions)
    return render_template('question_paper_print.html',
                           school_name=school_name,
                           logo_base64=logo_base64,
                           class_name=result['class'],
                           subject=result['subject'],
                           questions=questions,
                           total_marks=total_marks,
                           date=datetime.datetime.now().strftime('%d/%m/%Y'))

@app.route('/teacher/test_history')
@teacher_required
def teacher_test_history():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT * FROM test_generation_history
                 WHERE teacher_id=? ORDER BY created_at DESC""", (str(teacher_id),))
    history = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_test_history.html', history=history)

@app.route('/api/test_generation_history/<int:hid>/pdf')
@teacher_required
def test_gen_history_pdf(hid):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_generation_history WHERE id=? AND teacher_id=?", (hid, str(teacher_id)))
    hist = c.fetchone()
    if not hist:
        conn.close()
        return "Not found", 404
    hist = dict(hist)
    c.execute("""SELECT * FROM questions
                 WHERE class=? AND subject=? AND test_no=?
                 ORDER BY question_type, id""",
             (hist['class'], hist['subject'], hist['test_no']))
    questions = [dict(r) for r in c.fetchall()]
    conn.close()
    school_name = get_setting('school_name','RRB Group of Schools')
    logo_path   = get_setting('logo_path','')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path,'rb') as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = 'image/png' if ext=='.png' else 'image/jpeg'
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"
    rendered = render_template('question_paper_template_v2.html',
        school_name=school_name, logo_base64=logo_base64,
        class_name=f"{hist['class']}{hist['section']}", subject=hist['subject'],
        chapter=hist['chapter'], test_no=hist['test_no'],
        questions=questions, total_marks=len(questions),
        date=datetime.datetime.now().strftime('%d/%m/%Y'))
    pdf = HTML(string=rendered, base_url=request.base_url).write_pdf()
    resp = make_response(pdf)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename=test_{hist["test_no"]}.pdf'
    return resp

# ═══════════════════════════════════════════════════════════════
# ADMIN: Student class lock management
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/student_locks')
@admin_required
def admin_student_locks():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT l.student_id, l.class, l.section, l.locked_at, s.name
                 FROM student_class_lock l
                 LEFT JOIN students s ON l.student_id=s.student_id
                 ORDER BY l.class, l.section, s.name""")
    locks = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'locks': locks})

# ═══════════════════════════════════════════════════════════════
# MULTIMODAL: IMAGEN 3 IMAGE GENERATION PROXY
# ═══════════════════════════════════════════════════════════════

@app.route('/api/generate_diagram', methods=['POST'])
def generate_diagram():
    """Proxy endpoint: receives an image prompt, calls Gemini Imagen, returns base64 image."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    import urllib.request, json as json_lib

    data = request.get_json() or {}
    image_prompt = data.get('prompt', '').strip()
    if not image_prompt:
        return jsonify({'status': 'error', 'message': 'No prompt provided'}), 400

    gemini_api_key = _get_gemini_key()
    if not gemini_api_key:
        return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY not set'}), 500

    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/imagen-3.0-generate-002:predict?key={gemini_api_key}"
        payload = json_lib.dumps({
            "instances": [{"prompt": image_prompt}],
            "parameters": {"sampleCount": 1, "aspectRatio": "1:1", "safetyFilterLevel": "block_some"}
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload,
                                     headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=45) as resp:
            result = json_lib.loads(resp.read().decode('utf-8'))
        img_b64 = result['predictions'][0]['bytesBase64Encoded']
        mime = result['predictions'][0].get('mimeType', 'image/png')
        return jsonify({'status': 'success', 'image': f'data:{mime};base64,{img_b64}',
                        'prompt': image_prompt})
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8')
        return jsonify({'status': 'error', 'message': f'Imagen API error: {err}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

def _get_gemini_key():
    key = os.environ.get('GEMINI_API_KEY', '')
    if not key:
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'apikey.env')
        if os.path.exists(env_path):
            key = open(env_path).read().strip()
    return key

# ═══════════════════════════════════════════════════════════════
# BULLETIN POSTING API for Admin & Teacher dashboards
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/bulletins')
@admin_required
def admin_bulletins_page():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT b.*, COALESCE(t.name,'Admin') as poster_name
                 FROM bulletins b LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
                 WHERE b.is_active=1 ORDER BY b.created_at DESC""")
    bulletins = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('admin_bulletins.html', bulletins=bulletins)


@app.route('/api/test_generation_history/<int:hid>/delete', methods=['DELETE'])
def delete_test_history(hid):
    """Feature #1: Delete individual test history entry."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401
    conn = get_db()
    c = conn.cursor()
    if session.get('teacher_logged_in'):
        c.execute("DELETE FROM test_generation_history WHERE id=? AND teacher_id=?",
                  (hid, str(session.get('teacher_id'))))
    else:
        c.execute("DELETE FROM test_generation_history WHERE id=?", (hid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/api/test_generation_history/<int:hid>/print_preview')
def test_gen_print_preview(hid):
    """Feature #6: Return HTML page for browser print dialog (client-side PDF)."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return redirect(url_for('teacher_login'))
    conn = get_db()
    c = conn.cursor()

    if session.get('teacher_logged_in'):
        c.execute("SELECT * FROM test_generation_history WHERE id=? AND teacher_id=?",
                  (hid, str(session.get('teacher_id'))))
    else:
        c.execute("SELECT * FROM test_generation_history WHERE id=?", (hid,))
    hist = c.fetchone()
    if not hist:
        conn.close()
        return "Not found", 404
    hist = dict(hist)

    # Get questions by test_no OR chapter
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                        correct_answer, image_path,
                        COALESCE(question_type,'MCQ') as question_type
                 FROM questions
                 WHERE class=? AND subject=? AND (test_no=? OR chapter=?)
                 ORDER BY question_type, id""",
             (hist['class'], hist['subject'], hist['test_no'], hist['test_no']))
    questions = [dict(r) for r in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, 'rb') as f:
                lb = f.read()
                ext  = os.path.splitext(abs_path)[1].lower()
                mime = 'image/png' if ext == '.png' else 'image/jpeg'
                logo_base64 = f"data:{mime};base64,{base64.b64encode(lb).decode()}"

    return render_template('print_preview.html',
                           school_name=school_name,
                           logo_base64=logo_base64,
                           hist=hist,
                           questions=questions,
                           date=datetime.datetime.now().strftime('%d/%m/%Y'))

# ── ADMIN TEST GENERATION (Feature #8) ───────────────────────────────────────

@app.route('/admin/create_test', methods=['GET', 'POST'])
@admin_required
def admin_create_test():
    """Feature #8: Admin test generation - unrestricted class/subject access."""
    conn = get_db()
    c = conn.cursor()
    # Admin sees ALL classes and subjects
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
    all_pairs = [dict(r) for r in c.fetchall()]

    if request.method == 'POST':
        import urllib.request, json as json_lib

        class_    = request.form.get('class','').strip()
        section   = request.form.get('section','').strip()
        subject   = request.form.get('subject','').strip()
        chapter   = request.form.get('chapter','').strip()
        test_no   = request.form.get('test_no','').strip()
        remark    = request.form.get('remark','').strip()
        output_mode = request.form.get('output_mode','cbt')
        method    = request.form.get('method','ai')

        mcq_count        = int(request.form.get('mcq_count', 0) or 0)
        assertion_count  = int(request.form.get('assertion_count', 0) or 0)
        very_short_count = int(request.form.get('very_short_count', 0) or 0)
        short_count      = int(request.form.get('short_count', 0) or 0)
        long_count       = int(request.form.get('long_count', 0) or 0)
        case_study_count = int(request.form.get('case_study_count', 0) or 0)

        type_instructions = []
        if mcq_count:        type_instructions.append(f"{mcq_count} MCQ")
        if assertion_count:  type_instructions.append(f"{assertion_count} Assertion-Reason")
        if very_short_count: type_instructions.append(f"{very_short_count} Very Short Answer")
        if short_count:      type_instructions.append(f"{short_count} Short Answer")
        if long_count:       type_instructions.append(f"{long_count} Long Answer")
        if case_study_count: type_instructions.append(f"{case_study_count} Case Study")

        if not type_instructions:
            conn.close()
            return jsonify({'status':'error','message':'Enter at least 1 question quantity'}), 400

        if method == 'upload':
            file = request.files.get('csv_file')
            if file and file.filename.endswith('.csv'):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                inserted = 0
                with open(filepath, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        q = row.get('question','').strip()
                        if not q: continue
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,row.get('question_type','MCQ'),q,
                                  row.get('option a',row.get('option_a','')),
                                  row.get('option b',row.get('option_b','')),
                                  row.get('option c',row.get('option_c','')),
                                  row.get('option d',row.get('option_d','')),
                                  row.get('correct_answer','')))
                        inserted += 1
                conn.commit()
                os.remove(filepath)
                conn.close()
                return jsonify({'status':'success','message':f'{inserted} questions uploaded'})
            conn.close()
            return jsonify({'status':'error','message':'Invalid CSV file'}), 400

        # AI generation
        gemini_api_key = _get_gemini_key()
        if not gemini_api_key:
            conn.close()
            return jsonify({'status':'error','message':'GEMINI_API_KEY not set'}), 500

        ai_prompt = _build_ai_prompt(class_, section, subject, chapter, test_no, remark, type_instructions)

        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_api_key}"
            payload = json_lib.dumps({
                "contents": [{"parts": [{"text": ai_prompt}]}],
                "generationConfig": {"temperature": 0.7, "maxOutputTokens": 8192}
            }).encode('utf-8')
            req = urllib.request.Request(url, data=payload, headers={'Content-Type':'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json_lib.loads(resp.read().decode('utf-8'))

            response_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
            if '```' in response_text:
                for part in response_text.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): response_text = part; break

            questions = json_lib.loads(response_text)

            if output_mode == 'cbt':
                # BUG-004 FIX: prevent merging into existing test
                c.execute("SELECT COUNT(*) as cnt FROM questions WHERE class=? AND subject=? AND test_no=?",
                         (class_, subject, test_no))
                if c.fetchone()['cnt'] > 0:
                    conn.close()
                    return jsonify({'status':'error',
                                    'message':f'Test No "{test_no}" already exists for {subject} Class {class_}. Use a different Test No.'}), 400
                inserted = 0
                for q in questions:
                    if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,q.get('question_type','MCQ'),
                                  q.get('question',''),q.get('option_a',''),q.get('option_b',''),
                                  q.get('option_c',''),q.get('option_d',''),q.get('correct_answer','')))
                        inserted += 1
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                         (f"{class_}{section}_{subject}_{test_no}",'admin','admin',subject,test_no,'admin','admin',inserted))
                conn.commit()

            # Save to history (admin uses teacher_id='admin')
            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     ('admin',class_,section,subject,chapter,test_no,output_mode,
                      len(questions),mcq_count,assertion_count,very_short_count,
                      short_count,long_count,case_study_count,remark))
            conn.commit()
            history_id = c.lastrowid
            conn.close()

            if output_mode == 'print':
                return jsonify({'status':'success','mode':'print',
                                'questions':questions,'history_id':history_id,
                                'message':f'{len(questions)} questions generated'})
            return jsonify({'status':'success','mode':'cbt',
                            'message':f'{len(questions)} questions generated. Test No: {test_no}'})

        except urllib.error.HTTPError as e:
            conn.close()
            return jsonify({'status':'error','message':f'Gemini error: {e.read().decode()}'}), 500
        except Exception as e:
            conn.close()
            return jsonify({'status':'error','message':str(e)}), 500

    conn.close()
    return render_template('admin_create_test.html', all_pairs=all_pairs)

@app.route('/admin/test_history')
@admin_required
def admin_test_history():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_generation_history ORDER BY created_at DESC")
    history = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_test_history.html', history=history, is_admin=True)

# ── SHARED AI PROMPT BUILDER ──────────────────────────────────────────────────

def _build_ai_prompt(class_, section, subject, chapter, test_no, remark, type_instructions):
    return f"""You are an expert CBSE/ICSE question paper generator.
Class: {class_}{section} | Subject: {subject} | Chapter: {chapter} | Test No: {test_no}
Teacher instructions: {remark if remark else 'Standard difficulty, balanced coverage'}

Generate EXACTLY these question types:
{chr(10).join(type_instructions)}

OUTPUT: Return ONLY a valid JSON array. No markdown. No text before or after.
Every object must have ALL keys: question_type, content_type, question, option_a, option_b, option_c, option_d, correct_answer, smiles, image_prompt, marks

MATH LaTeX RULES:
- Inline: $...$ | Block: $$...$$
- Matrices: $\\begin{{pmatrix}} a & b \\\\ c & d \\end{{pmatrix}}$
- Fractions: \\frac{{num}}{{den}} | Vectors: $\\vec{{F}} = 3\\hat{{i}}$
- Escape backslashes in JSON: \\\\frac not \\frac

CHEMISTRY mhchem RULES:
- Equations: "$\\\\ce{{H2SO4 + 2NaOH -> Na2SO4 + 2H2O}}$"
- Structural diagrams: set "smiles" field (Benzene="c1ccccc1")

NON-MCQ: set option_a/b/c/d="" and correct_answer="N/A"
smiles="" and image_prompt="" when not needed.

Return ONLY the JSON array starting with [ and ending with ]."""


@app.route('/descriptive_paper', methods=['GET', 'POST'])
def descriptive_paper():
    """Descriptive AI paper generator — accessible to both admin and teacher."""
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return redirect(url_for('teacher_login'))

    # GET — show the form
    if request.method == 'GET':
        teacher_id = session.get('teacher_id')
        assignments = []
        all_pairs   = []
        if teacher_id:
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
            assignments = [dict(r) for r in c.fetchall()]
            conn.close()
        elif session.get('admin_logged_in'):
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
            all_pairs = [dict(r) for r in c.fetchall()]
            conn.close()

        school_name      = get_setting('school_name', 'RRB Group of Schools')
        school_address   = get_setting('school_address', '')
        academic_session = get_setting('academic_session', '')
        logo_path        = get_setting('logo_path', '')

        return render_template('descriptive_paper.html',
                               assignments=assignments,
                               all_pairs=all_pairs,
                               school_name=school_name,
                               school_address=school_address,
                               academic_session=academic_session,
                               logo_path=logo_path,
                               is_admin=bool(session.get('admin_logged_in')))

    # POST — generate via Gemini and return structured JSON for the preview
    import urllib.request, json as json_lib

    # ── Collect all form fields ──────────────────────────────────────────────
    class_         = request.form.get('class', '').strip()
    subject        = request.form.get('subject', '').strip()
    topics         = request.form.get('topics', '').strip()
    full_syllabus  = request.form.get('full_syllabus') == 'yes'
    exam_type      = request.form.get('exam_type', 'Unit Test').strip()
    duration       = request.form.get('duration', '3 Hours').strip()
    max_marks      = request.form.get('max_marks', '100').strip()
    academic_sess  = request.form.get('academic_session', '').strip() or get_setting('academic_session', '')
    school_nm      = request.form.get('school_name', '').strip() or get_setting('school_name', 'RRB Group of Schools')
    school_addr    = request.form.get('school_address', '').strip() or get_setting('school_address', '')
    remarks        = request.form.get('remarks', '').strip()

    # Question counts per type
    mcq_count        = int(request.form.get('mcq_count', 0) or 0)
    ar_count         = int(request.form.get('ar_count', 0) or 0)
    vs_count         = int(request.form.get('vs_count', 0) or 0)
    sh_count         = int(request.form.get('sh_count', 0) or 0)
    lg_count         = int(request.form.get('lg_count', 0) or 0)
    cs_count         = int(request.form.get('cs_count', 0) or 0)
    total_q = mcq_count + ar_count + vs_count + sh_count + lg_count + cs_count

    if total_q == 0:
        return jsonify({'status': 'error', 'message': 'Please enter at least 1 question'}), 400

    # ── Optional image upload ────────────────────────────────────────────────
    image_context = ''
    img_file = request.files.get('image_file')
    if img_file and img_file.filename:
        allowed_img = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        ext = img_file.filename.rsplit('.', 1)[-1].lower()
        if ext in allowed_img:
            img_bytes   = img_file.read()
            img_b64     = base64.b64encode(img_bytes).decode()
            img_mime    = f'image/{ext}' if ext != 'jpg' else 'image/jpeg'
            image_context = f'\n\nAn image has been provided (base64 omitted here). Use it conceptually to inspire questions involving diagrams, graphs, or visual-based problems relevant to {subject}.'

    # ── Build section descriptions ───────────────────────────────────────────
    sections = []
    if mcq_count: sections.append(f"Section A — MCQ: {mcq_count} questions, 1 mark each")
    if ar_count:  sections.append(f"Section B — Assertion-Reason: {ar_count} questions, 1 mark each")
    if vs_count:  sections.append(f"Section C — Very Short Answer: {vs_count} questions, 2 marks each")
    if sh_count:  sections.append(f"Section D — Short Answer: {sh_count} questions, 3 marks each")
    if lg_count:  sections.append(f"Section E — Long Answer: {lg_count} questions, 5 marks each")
    if cs_count:  sections.append(f"Section F — Case Study Based: {cs_count} questions, 4 marks each")

    topic_text = 'Complete Syllabus' if full_syllabus else (topics or 'All Topics')

    # ── Gemini prompt ─────────────────────────────────────────────────────────
    prompt = f"""You are an expert CBSE/ICSE question paper setter. Generate a complete descriptive question paper.

PAPER DETAILS:
- Class: {class_}
- Subject: {subject}
- Topics/Chapters: {topic_text}
- Exam Type: {exam_type}
- Duration: {duration}
- Maximum Marks: {max_marks}
- Academic Session: {academic_sess}

SECTIONS TO GENERATE (generate EXACTLY these):
{chr(10).join(sections)}

TEACHER INSTRUCTIONS / REMARKS:
{remarks if remarks else 'Standard difficulty. Mix of theory and application. Follow CBSE pattern.'}
{image_context}

OUTPUT FORMAT — return ONLY a valid JSON object like this (no markdown, no code fences):
{{
  "sections": [
    {{
      "section_label": "Section A",
      "section_title": "Multiple Choice Questions",
      "marks_per_question": 1,
      "instruction": "Choose the correct answer.",
      "questions": [
        {{
          "number": 1,
          "question": "Full question text here.",
          "options": ["(a) Option A", "(b) Option B", "(c) Option C", "(d) Option D"],
          "sub_questions": []
        }}
      ]
    }}
  ]
}}

RULES:
- NO LaTeX. Use plain English text only.
- For MCQ/AR: include "options" array with 4 choices formatted as (a), (b), (c), (d).
- For Short/Long/Case Study: "options" should be empty [], use "sub_questions" for multi-part questions.
- For Case Study: write a reading passage first in "question", then list sub-questions in "sub_questions" array (each with number and question text).
- Assertion-Reason: write Assertion and Reason clearly in the question text. Options must be the standard 4 AR options.
- Keep language age-appropriate for Class {class_}.
- Return ONLY the JSON object. Nothing else."""

    gemini_key = _get_gemini_key()
    if not gemini_key:
        return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY not configured'}), 500

    try:
        url     = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
        payload = json_lib.dumps({
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.7, "maxOutputTokens": 8192}
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload,
                                     headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = json_lib.loads(resp.read().decode('utf-8'))

        raw = result['candidates'][0]['content']['parts'][0]['text'].strip()

        # Strip markdown fences if present
        if '```' in raw:
            for part in raw.split('```'):
                part = part.strip()
                if part.startswith('json'): part = part[4:].strip()
                if part.startswith('{'): raw = part; break

        paper_data = json_lib.loads(raw)

        # Build logo base64 for the preview
        logo_path   = get_setting('logo_path', '')
        logo_base64 = ''
        if logo_path:
            abs_path = os.path.join(app.static_folder, logo_path)
            if os.path.exists(abs_path):
                with open(abs_path, 'rb') as f:
                    lb  = f.read()
                    ext = os.path.splitext(abs_path)[1].lower()
                    mt  = 'image/png' if ext == '.png' else 'image/jpeg'
                    logo_base64 = f"data:{mt};base64,{base64.b64encode(lb).decode()}"

        return jsonify({
            'status': 'success',
            'paper':  paper_data,
            'meta': {
                'school_name':      school_nm,
                'school_address':   school_addr,
                'academic_session': academic_sess,
                'exam_type':        exam_type,
                'class':            class_,
                'subject':          subject,
                'duration':         duration,
                'max_marks':        max_marks,
                'logo_base64':      logo_base64,
                'date':             datetime.datetime.now().strftime('%d / %m / %Y'),
            }
        })

    except urllib.error.HTTPError as e:
        return jsonify({'status': 'error', 'message': f'Gemini API: {e.read().decode()}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/csv_template')
def download_csv_template():
    """FEAT-001: Downloadable CSV template for question paper upload."""
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['question', 'option a', 'option b', 'option c', 'option d',
                     'correct_answer', 'question_type', 'marks', 'negative_mark'])
    writer.writerow([
        'What is the speed of light in vacuum?',
        '3 x 10^8 m/s', '3 x 10^6 m/s', '3 x 10^10 m/s', '3 x 10^4 m/s',
        'option_a', 'MCQ', '1', '0.33'
    ])
    writer.writerow([
        'Assertion: Water boils at 100°C. Reason: Boiling point depends on atmospheric pressure.',
        'Both A and R are true and R is the correct explanation of A',
        'Both A and R are true but R is NOT the correct explanation of A',
        'A is true but R is false',
        'A is false but R is true',
        'option_a', 'Assertion-Reason', '1', '0.33'
    ])
    writer.writerow([
        'Explain the process of photosynthesis.', '', '', '', '',
        'N/A', 'Short', '3', '0'
    ])
    writer.writerow([
        'Define Newton\'s first law of motion and give two examples.', '', '', '', '',
        'N/A', 'Very Short', '2', '0'
    ])
    writer.writerow([
        'Discuss the structure and functions of DNA.', '', '', '', '',
        'N/A', 'Long', '5', '0'
    ])
    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=question_upload_template.csv'}
    )

@app.route('/announcement/<int:bid>')
def announcement_permalink(bid):
    """FEAT-005: Permanent link to a single announcement."""
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT b.*, COALESCE(t.name,'Admin') as poster_name
                 FROM bulletins b LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
                 WHERE b.id=? AND b.is_active=1""", (bid,))
    bulletin = c.fetchone()
    conn.close()
    if not bulletin:
        return "Announcement not found or has been removed.", 404
    bulletin = dict(bulletin)
    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    return render_template('announcement_permalink.html',
                           bulletin=bulletin, school_name=school_name, logo_path=logo_path)

@app.route('/api/shared_tests')
def api_shared_tests():
    """FEAT-008: List all shared/available tests for a class+subject."""
    class_   = request.args.get('class','').strip()
    subject  = request.args.get('subject','').strip()
    if not class_ or not subject:
        return jsonify({'tests': []})
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT tp.id, tp.filename, tp.test_no, tp.section, tp.question_count,
                        tp.is_active, tp.created_at, tp.uploader_type,
                        COALESCE(t.name,'Admin') as creator_name
                 FROM test_papers tp
                 LEFT JOIN teachers t ON tp.uploaded_by=t.id AND tp.uploader_type='teacher'
                 WHERE tp.class=? AND tp.subject=?
                 ORDER BY tp.created_at DESC""", (class_, subject))
    tests = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'tests': tests})

@app.route('/api/assign_existing_test', methods=['POST'])
def assign_existing_test():
    """FEAT-008: Link an existing test to a class/section without duplicating questions."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status':'error'}), 401
    paper_id  = request.form.get('paper_id', type=int)
    section   = request.form.get('section','').strip()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_papers WHERE id=?", (paper_id,))
    paper = c.fetchone()
    if not paper:
        conn.close()
        return jsonify({'status':'error','message':'Test not found'}), 404
    # Update section assignment (or leave as shared)
    c.execute("UPDATE test_papers SET section=?, is_active=1 WHERE id=?", (section, paper_id))
    conn.commit()
    conn.close()
    return jsonify({'status':'success','message':f'Test "{paper["filename"]}" assigned to Section {section or "All"}'})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)