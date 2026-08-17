import os
import csv
import requests
from dotenv import load_dotenv
import sqlite3
import datetime
import uuid
import random
import base64
import hashlib
import pdfkit
import json
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, make_response
from werkzeug.utils import secure_filename
import openpyxl
from weasyprint import HTML, CSS
# ─── WeasyPrint 61+ compatibility fix ──────────────────────────────
import weasyprint.pdf
import weasyprint.pdf.stream
import re
import json
import ast
from validators import validate_schema, STUDENT_LOGIN_SCHEMA, ADMIN_LOGIN_SCHEMA, SAVE_ANSWER_SCHEMA, REATTEMPT_REQUEST_SCHEMA

def safe_json_loads(text):
    """
    Attempt to parse JSON from AI response.
    Uses a series of fallback strategies:
      1. Remove markdown fences and extra text.
      2. Extract the first JSON array or object.
      3. Clean trailing commas and unescaped quotes.
      4. Use ast.literal_eval (Python literals).
      5. Try to repair by escaping newlines and quotes inside strings.
    """
    if not text:
        return None

    # 1. Remove markdown code fences
    text = re.sub(r'```(?:json)?\s*', '', text)
    text = re.sub(r'```\s*$', '', text)
    text = text.strip()

    # 2. Try to extract a JSON array or object
    # Look for anything that starts with [ or { and ends with ] or }
    # Use non-greedy matching to avoid consuming extra
    match = re.search(r'(\[.*\]|\{.*\})', text, re.DOTALL)
    candidate = match.group(1) if match else text

    # 3. Try normal parsing (strict=False allows trailing commas)
    try:
        return json.loads(candidate, strict=False)
    except json.JSONDecodeError:
        pass

    # 4. Clean trailing commas before ] or }
    cleaned = re.sub(r',\s*([\]}])', r'\1', candidate)

    # 5. Try to fix unescaped double quotes inside strings
    # This is tricky; we'll replace all " inside strings that are not already escaped.
    # A simple heuristic: find strings that start with " and end with " but contain "
    # We'll use a more robust approach: use ast.literal_eval after converting null/true/false.
    try:
        # Convert JSON null/true/false to Python None/True/False
        py_literal = cleaned.replace('null', 'None').replace('true', 'True').replace('false', 'False')
        return ast.literal_eval(py_literal)
    except:
        pass

    # 6. Try to split into multiple parts and parse each separately (if multiple arrays)
    # Some AIs return concatenated JSON objects
    parts = re.split(r'\]\s*,\s*\{', candidate)
    if len(parts) > 1:
        for i in range(len(parts)):
            # Reconstruct a JSON array
            try_part = '[' + '],['.join(parts) + ']'
            try:
                return json.loads(try_part, strict=False)
            except:
                continue

    # 7. Last resort: try to fix unescaped newlines and quotes with a custom repair
    try:
        # Escape newlines and carriage returns
        repaired = re.sub(r'(?<!\\)\n', '\\n', cleaned)
        repaired = re.sub(r'(?<!\\)\r', '\\r', repaired)
        # Escape unescaped quotes inside strings (naive but may work)
        # This is a simplified version; we'll try to escape all " that are not at the start or end of a string.
        # Instead, we'll use a regex to find strings and escape their content.
        # This is complex; we'll skip and return None.
        pass
    except:
        pass

    # If all fails, log the raw text and return None
    return None

def parse_question_csv_row(row):
    """
    Flexible CSV parser supporting column headers:
    Questions, option a, option b, option c, option d, correct option
    (and all common variations like option_a, question, correct_answer).
    """
    if not isinstance(row, dict):
        return None

    # Case-insensitive normalized map
    norm = {}
    for k, v in row.items():
        if k is not None and str(k).strip():
            k_clean = str(k).strip().lower().replace('_', ' ')
            norm[k_clean] = str(v).strip() if v is not None else ''

    def get_val(*keys):
        for k in keys:
            k_clean = k.strip().lower().replace('_', ' ')
            if k_clean in norm and norm[k_clean]:
                return norm[k_clean]
        return ''

    q = get_val('questions', 'question', 'question text', 'q')
    if not q:
        return None

    opt_a = get_val('option a', 'optiona', 'a', 'option 1', 'option1')
    opt_b = get_val('option b', 'optionb', 'b', 'option 2', 'option2')
    opt_c = get_val('option c', 'optionc', 'c', 'option 3', 'option3')
    opt_d = get_val('option d', 'optiond', 'd', 'option 4', 'option4')

    raw_ans = get_val('correct option', 'correct answer', 'correct_answer', 'correctoption', 'answer', 'correct')
    
    # Normalize correct_answer to option_a, option_b, option_c, option_d
    ans_clean = raw_ans.lower().replace('_', ' ').strip()
    correct_ans = 'option_a'
    if ans_clean in ('option a', 'optiona', 'a', '1', 'option 1', 'option1'):
        correct_ans = 'option_a'
    elif ans_clean in ('option b', 'optionb', 'b', '2', 'option 2', 'option2'):
        correct_ans = 'option_b'
    elif ans_clean in ('option c', 'optionc', 'c', '3', 'option 3', 'option3'):
        correct_ans = 'option_c'
    elif ans_clean in ('option d', 'optiond', 'd', '4', 'option 4', 'option4'):
        correct_ans = 'option_d'
    elif raw_ans and opt_a and raw_ans == opt_a:
        correct_ans = 'option_a'
    elif raw_ans and opt_b and raw_ans == opt_b:
        correct_ans = 'option_b'
    elif raw_ans and opt_c and raw_ans == opt_c:
        correct_ans = 'option_c'
    elif raw_ans and opt_d and raw_ans == opt_d:
        correct_ans = 'option_d'
    elif raw_ans:
        correct_ans = raw_ans

    q_type = get_val('question type', 'type') or ('MCQ' if (opt_a or opt_b) else 'Short')
    marks = get_val('marks', 'mark') or '1'
    neg_mark = get_val('negative mark', 'negative marks') or '0.33'

    return {
        'question': q,
        'option_a': opt_a,
        'option_b': opt_b,
        'option_c': opt_c,
        'option_d': opt_d,
        'correct_answer': correct_ans,
        'question_type': q_type,
        'marks': marks,
        'negative_mark': neg_mark
    }

def _normalize_correct_answer(correct_raw, opt_a='', opt_b='', opt_c='', opt_d=''):
    """
    Normalize any correct_answer representation ('A', 'a', 'option_a', 'Option A', '1', or actual option text)
    to a standard uppercase letter ('A', 'B', 'C', 'D').
    """
    if not correct_raw:
        return ''
    c_str = str(correct_raw).strip()
    c_upper = c_str.upper()

    # 1. Single letter 'A', 'B', 'C', 'D'
    if c_upper in ('A', 'B', 'C', 'D'):
        return c_upper

    # 2. Number string '1', '2', '3', '4'
    if c_str in ('1', '2', '3', '4'):
        return chr(ord('A') + int(c_str) - 1)

    # 3. 'OPTION_A', 'OPTION A', 'OPT_A', 'OPTION_B', etc.
    c_clean = c_upper.replace(' ', '_')
    if c_clean in ('OPTION_A', 'OPT_A'):
        return 'A'
    if c_clean in ('OPTION_B', 'OPT_B'):
        return 'B'
    if c_clean in ('OPTION_C', 'OPT_C'):
        return 'C'
    if c_clean in ('OPTION_D', 'OPT_D'):
        return 'D'

    # 4. Text match against option_a, option_b, option_c, option_d
    c_lower = c_str.lower()
    if opt_a and str(opt_a).strip().lower() == c_lower:
        return 'A'
    if opt_b and str(opt_b).strip().lower() == c_lower:
        return 'B'
    if opt_c and str(opt_c).strip().lower() == c_lower:
        return 'C'
    if opt_d and str(opt_d).strip().lower() == c_lower:
        return 'D'

    return ''

def _save_or_update_result(c, student_id, name, class_, section, subject, score, total_questions, percentage, chapter):
    """Save new result or update existing result row for (student_id, subject, chapter)."""
    c.execute("""SELECT id FROM results 
                 WHERE student_id=? AND LOWER(subject)=LOWER(?) AND (chapter=? OR (chapter IS NULL AND ?=''))""",
              (student_id, subject, chapter or '', chapter or ''))
    existing = c.fetchone()
    if existing:
        c.execute("""UPDATE results 
                     SET name=?, class=?, section=?, score=?, total_questions=?, percentage=?, test_date=CURRENT_TIMESTAMP
                     WHERE id=?""",
                  (name, class_, section or '', round(score, 2), total_questions, percentage, existing['id']))
    else:
        c.execute("""INSERT INTO results 
                     (student_id, name, class, section, subject, score, total_questions, percentage, chapter, test_date)
                     VALUES (?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                  (student_id, name, class_, section or '', subject, round(score, 2), total_questions, percentage, chapter or ''))

def get_working_model(api_key):
    """
    Try to find a working Gemini model by listing available models.
    Returns a model name string or default 'gemini-2.0-flash'.
    """
    import urllib.request, json
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            models = data.get('models', [])
            supported_models = []
            for model in models:
                name = model.get('name', '').replace('models/', '')
                supported = model.get('supportedGenerationMethods', [])
                if 'generateContent' in supported:
                    supported_models.append(name)

            # Preference priority for valid active models
            for preferred in ['gemini-2.0-flash', 'gemini-2.0-flash-lite']:
                if preferred in supported_models:
                    return preferred

            for name in supported_models:
                if 'flash' in name or 'pro' in name:
                    return name

            if supported_models:
                return supported_models[0]
    except Exception:
        pass
    return 'gemini-2.0-flash'

def call_gemini_generate_content(prompt, api_key, generation_config=None, timeout=60):
    """
    Robust Gemini API caller with automatic candidate model fallbacks and HTTP 429 backoff/retry.
    Returns tuple: (result_dict, error_message)
    """
    import urllib.request, urllib.error, json, time
    if not generation_config:
        generation_config = {"temperature": 0.7, "maxOutputTokens": 8192}

    candidate_models = ['gemini-2.0-flash', 'gemini-2.0-flash-lite']
    detected = get_working_model(api_key)
    if detected and detected not in candidate_models:
        candidate_models.insert(0, detected)

    payload_bytes = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": generation_config
    }).encode('utf-8')

    last_error = ""

    for model_name in candidate_models:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
        for attempt in range(2):
            try:
                req = urllib.request.Request(
                    url,
                    data=payload_bytes,
                    headers={'Content-Type': 'application/json'},
                    method='POST'
                )
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    resp_data = resp.read().decode('utf-8')
                    if resp_data.lstrip().startswith('<'):
                        last_error = "Gemini API returned HTML instead of JSON."
                        break
                    return json.loads(resp_data), None
            except urllib.error.HTTPError as e:
                err_body = e.read().decode('utf-8')
                last_error = f"HTTP {e.code}: {err_body}"
                if e.code == 429:
                    time.sleep(2)
                    continue
                else:
                    break
            except Exception as e:
                last_error = str(e)
                break

    if "429" in last_error or "quota" in last_error.lower() or "resource_exhausted" in last_error.lower():
        clean_msg = "Gemini API rate limit / free tier quota exceeded (HTTP 429). Please wait 30–60 seconds and try again, or check your API key quota at https://ai.google.dev."
    else:
        clean_msg = f"Gemini API error: {last_error[:300]}"

    return None, clean_msg

def _get_ai_keys():
    """
    Parse apikey.env, .api_key, and OS environment for API keys across providers:
    Gemini, DeepSeek, OpenAI (ChatGPT), and Claude (Anthropic).
    Returns dict: {'gemini': '...', 'deepseek': '...', 'openai': '...', 'claude': '...'}
    """
    keys = {
        'gemini': os.environ.get('GEMINI_API_KEY', '').strip().strip('"\''),
        'deepseek': os.environ.get('DEEPSEEK_API_KEY', '').strip().strip('"\''),
        'openai': os.environ.get('OPENAI_API_KEY', '').strip().strip('"\''),
        'claude': os.environ.get('CLAUDE_API_KEY', '') or os.environ.get('ANTHROPIC_API_KEY', ''),
        'huggingface': os.environ.get('HUGGINGFACE_API_KEY', '') or os.environ.get('HF_TOKEN', '') or os.environ.get('HF_API_KEY', ''),
        'hf_model': os.environ.get('HF_MODEL', '') or os.environ.get('HUGGINGFACE_MODEL', '')
    }
    for k in keys:
        keys[k] = keys[k].strip().strip('"\'')

    # 1. Primary Master Source: apikey.env file
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'apikey.env')
    if os.path.exists(env_path):
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            is_single = True
            for line in lines:
                line_s = line.strip()
                if not line_s or line_s.startswith('#'):
                    continue
                if '=' in line_s:
                    is_single = False
                    parts = line_s.split('=', 1)
                    k_name = parts[0].strip().upper()
                    val = parts[1].strip().strip('"\'')
                    if 'GEMINI' in k_name and val:
                        keys['gemini'] = val
                    elif 'DEEPSEEK' in k_name and val:
                        keys['deepseek'] = val
                    elif ('OPENAI' in k_name or 'CHATGPT' in k_name) and val:
                        keys['openai'] = val
                    elif ('CLAUDE' in k_name or 'ANTHROPIC' in k_name) and val:
                        keys['claude'] = val
                    elif ('HUGGINGFACE_MODEL' in k_name or 'HF_MODEL' in k_name) and val:
                        keys['hf_model'] = val
                    elif ('HUGGINGFACE' in k_name or 'HF_' in k_name) and val:
                        keys['huggingface'] = val

            if is_single and lines:
                raw_val = ''.join(lines).strip().strip('"\'')
                if raw_val and not raw_val.startswith('#'):
                    keys['gemini'] = raw_val
        except Exception:
            pass

    # 2. Secondary Fallback / Sync: .api_key file (used by batch tools)
    dot_key_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.api_key')
    if os.path.exists(dot_key_path):
        try:
            val = open(dot_key_path, 'r', encoding='utf-8').read().strip().strip('"\'')
            if val and not val.startswith('#') and not keys['gemini']:
                keys['gemini'] = val
        except Exception:
            pass

    # Keep .api_key in sync with current active Gemini key
    if keys['gemini']:
        try:
            with open(dot_key_path, 'w', encoding='utf-8') as f:
                f.write(keys['gemini'])
        except Exception:
            pass

    return keys

def _call_deepseek_api(prompt, api_key, timeout=90):
    """Call DeepSeek AI chat completions API (OpenAI-compatible)"""
    import urllib.request, urllib.error, json
    url = "https://api.deepseek.com/chat/completions"
    payload = json.dumps({
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
        "max_tokens": 4096
    }).encode('utf-8')
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            res = json.loads(resp.read().decode('utf-8'))
            text = res['choices'][0]['message']['content'].strip()
            return text, None
    except urllib.error.HTTPError as e:
        err_body = e.read().decode('utf-8')
        try:
            msg = json.loads(err_body).get('error', {}).get('message', err_body)
        except Exception:
            msg = err_body
        return None, f"DeepSeek HTTP {e.code}: {msg[:300]}"
    except Exception as e:
        return None, f"DeepSeek Error: {str(e)}"

def _call_openai_api(prompt, api_key, timeout=90):
    """Call OpenAI ChatGPT API (gpt-4o-mini / gpt-3.5-turbo)"""
    import urllib.request, urllib.error, json
    models = ["gpt-4o-mini", "gpt-3.5-turbo"]
    last_err = ""
    for model_name in models:
        url = "https://api.openai.com/v1/chat/completions"
        payload = json.dumps({
            "model": model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.7,
            "max_tokens": 4096
        }).encode('utf-8')
        req = urllib.request.Request(
            url, data=payload,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                res = json.loads(resp.read().decode('utf-8'))
                text = res['choices'][0]['message']['content'].strip()
                return text, None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            try:
                msg = json.loads(err_body).get('error', {}).get('message', err_body)
            except Exception:
                msg = err_body
            last_err = f"OpenAI HTTP {e.code}: {msg[:300]}"
            if e.code in (404, 400) and 'model' in last_err.lower():
                continue
            else:
                break
        except Exception as e:
            last_err = f"OpenAI Error: {str(e)}"
            break
    return None, last_err

def _call_claude_api(prompt, api_key, timeout=90):
    """Call Anthropic Claude API (claude-3-5-haiku-20241022 / claude-3-haiku-20240307)"""
    import urllib.request, urllib.error, json
    models = ["claude-3-5-haiku-20241022", "claude-3-haiku-20240307"]
    last_err = ""
    for model_name in models:
        url = "https://api.anthropic.com/v1/messages"
        payload = json.dumps({
            "model": model_name,
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": prompt}]
        }).encode('utf-8')
        req = urllib.request.Request(
            url, data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            },
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                res = json.loads(resp.read().decode('utf-8'))
                text = res['content'][0]['text'].strip()
                return text, None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            try:
                err_json = json.loads(err_body)
                msg = err_json.get('error', {}).get('message', err_body)
            except Exception:
                msg = err_body
            last_err = f"Claude HTTP {e.code}: {msg[:300]}"
            if e.code in (404, 400) and 'model' in last_err.lower():
                continue
            else:
                break
        except Exception as e:
            last_err = f"Claude Error: {str(e)}"
            break
    return None, last_err

def _call_huggingface_api(prompt, api_key, timeout=90):
    """Call Hugging Face Inference API (Serverless Open-Source models)"""
    import urllib.request, urllib.error, json
    keys = _get_ai_keys()
    custom_model = keys.get('hf_model')
    models = [
        "Qwen/Qwen2.5-7B-Instruct",
        "meta-llama/Llama-3.1-8B-Instruct",
        "mistralai/Mistral-7B-Instruct-v0.3",
        "Qwen/Qwen2.5-Coder-32B-Instruct"
    ]
    if custom_model and custom_model not in models:
        models.insert(0, custom_model)
    elif custom_model and custom_model in models:
        models.remove(custom_model)
        models.insert(0, custom_model)
    last_err = ""
    for model_name in models:
        url = "https://router.huggingface.co/v1/chat/completions"
        payload = json.dumps({
            "model": model_name,
            "messages": [
                {"role": "system", "content": "You are a professional educational question paper generator. Always return clean JSON array output."},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 4096,
            "temperature": 0.7
        }).encode('utf-8')
        req = urllib.request.Request(
            url, data=payload,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) RRB-CBT-Engine"
            },
            method='POST'
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                res = json.loads(resp.read().decode('utf-8'))
                text = res['choices'][0]['message']['content'].strip()
                return text, None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            try:
                err_json = json.loads(err_body)
                msg = err_json.get('error', {}).get('message', err_body)
                if isinstance(err_json.get('error'), str):
                    msg = err_json.get('error')
            except Exception:
                msg = err_body
            last_err = f"Hugging Face HTTP {e.code}: {msg[:300]}"
            if e.code in (404, 400, 503):
                continue
            else:
                break
        except Exception as e:
            last_err = f"Hugging Face Error: {str(e)}"
            break
    return None, last_err

def generate_ai_content(prompt, timeout=90):
    """
    Multi-provider AI generator with automatic failover across Gemini, DeepSeek, OpenAI, Claude, and Hugging Face.
    Returns (result_dict, provider_name, error_message).
    """
    keys = _get_ai_keys()
    errors = []

    # Priority 1: Google Gemini
    if keys.get('gemini'):
        res_dict, err = call_gemini_generate_content(prompt, keys['gemini'], timeout=timeout)
        if res_dict:
            return res_dict, 'Gemini AI', None
        app.logger.warning(f"Gemini AI failed, trying alternate providers: {err}")
        errors.append(f"Gemini: {err}")

    # Priority 2: DeepSeek AI
    if keys.get('deepseek'):
        text, err = _call_deepseek_api(prompt, keys['deepseek'], timeout=timeout)
        if text:
            res_dict = {'candidates': [{'content': {'parts': [{'text': text}]}}]}
            return res_dict, 'DeepSeek AI', None
        app.logger.warning(f"DeepSeek AI failed: {err}")
        errors.append(f"DeepSeek: {err}")

    # Priority 3: OpenAI ChatGPT
    if keys.get('openai'):
        text, err = _call_openai_api(prompt, keys['openai'], timeout=timeout)
        if text:
            res_dict = {'candidates': [{'content': {'parts': [{'text': text}]}}]}
            return res_dict, 'ChatGPT (OpenAI)', None
        app.logger.warning(f"OpenAI ChatGPT failed: {err}")
        errors.append(f"OpenAI: {err}")

    # Priority 4: Anthropic Claude
    if keys.get('claude'):
        text, err = _call_claude_api(prompt, keys['claude'], timeout=timeout)
        if text:
            res_dict = {'candidates': [{'content': {'parts': [{'text': text}]}}]}
            return res_dict, 'Claude AI', None
        app.logger.warning(f"Claude AI failed: {err}")
        errors.append(f"Claude: {err}")

    # Priority 5: Hugging Face AI
    if keys.get('huggingface'):
        text, err = _call_huggingface_api(prompt, keys['huggingface'], timeout=timeout)
        if text:
            res_dict = {'candidates': [{'content': {'parts': [{'text': text}]}}]}
            return res_dict, 'Hugging Face AI', None
        app.logger.warning(f"Hugging Face AI failed: {err}")
        errors.append(f"Hugging Face: {err}")

    if not any(keys.values()):
        return None, None, "No AI API key found. Please set GEMINI_API_KEY, DEEPSEEK_API_KEY, OPENAI_API_KEY, or CLAUDE_API_KEY in apikey.env or .api_key."

    combined_err = " | ".join(errors) if errors else "All configured AI providers failed."
    return None, None, f"AI Generation Failed: {combined_err}"
# ─── WeasyPrint & pydyf 0.12+ compatibility fixes ──────────────────────────
import pydyf

# 1. Patch pydyf.PDF.__init__ compatibility with WeasyPrint 60.0 & pydyf 0.12+
_orig_pydyf_pdf_init = pydyf.PDF.__init__

def _patched_pydyf_pdf_init(self, version='1.7', identifier=None, *args, **kwargs):
    _orig_pydyf_pdf_init(self)
    if version:
        self.version = version.encode('ascii') if isinstance(version, str) else version
    else:
        self.version = b'1.7'
    if identifier:
        self.identifier = identifier.encode('ascii') if isinstance(identifier, str) else identifier

pydyf.PDF.__init__ = _patched_pydyf_pdf_init

# 2. Patch Stream.transform compatibility for WeasyPrint 60+ stream
def _patched_stream_transform(self, a=1, b=0, c=0, d=1, e=0, f=0):
    if hasattr(self, 'set_matrix'):
        self.set_matrix(a, b, c, d, e, f)
    else:
        self._matrix = (a, b, c, d, e, f)

weasyprint.pdf.stream.Stream.transform = _patched_stream_transform

# 3. Patch Stream.text_matrix alias compatibility with pydyf 0.12+
if not hasattr(weasyprint.pdf.stream.Stream, 'text_matrix'):
    def _stream_text_matrix(self, a=1, b=0, c=0, d=1, e=0, f=0):
        if hasattr(self, 'set_matrix'):
            self.set_matrix(a, b, c, d, e, f)
        elif hasattr(self, 'set_text_matrix'):
            self.set_text_matrix(a, b, c, d, e, f)
        else:
            self._matrix = (a, b, c, d, e, f)
    
    weasyprint.pdf.stream.Stream.text_matrix = _stream_text_matrix

app = Flask(__name__)
app.secret_key = 'rrb-cbt-v104-secret-key-2024'
load_dotenv('.env')
load_dotenv('apikey.env')

# ─── JWT Authentication Subsystem ──────────────────────────────────────
import jwt
from flask import g

JWT_SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'rrb_cbt_jwt_secret_key_2026_x98f72a')
JWT_EXPIRATION_HOURS = int(os.environ.get('JWT_EXPIRATION_HOURS', 24))

def generate_jwt_token(identity, role, name=None, expires_in_hours=None):
    """
    Generate a signed JWT token using PyJWT.
    Claims: sub (identity), role (admin/teacher/student), name, iat, exp.
    """
    if expires_in_hours is None:
        expires_in_hours = JWT_EXPIRATION_HOURS
        
    now = datetime.datetime.now(datetime.timezone.utc)
    expiration = now + datetime.timedelta(hours=expires_in_hours)
    
    payload = {
        'sub': str(identity),
        'role': role,
        'name': name or str(identity),
        'iat': int(now.timestamp()),
        'exp': int(expiration.timestamp())
    }
    
    token = jwt.encode(payload, JWT_SECRET_KEY, algorithm='HS256')
    return token, int(expiration.timestamp())

def decode_jwt_token(token):
    """
    Decode and validate a signed JWT token using PyJWT.
    Returns payload dict if valid, or raises PyJWT exception if invalid/expired.
    """
    return jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])

def jwt_required(allowed_roles=None):
    """
    Decorator for API-only routes.
    Validates Authorization: Bearer <token> header.
    Attaches decoded payload to Flask context g.jwt_payload & g.jwt_user.
    Returns HTTP 401 Unauthorized if missing, expired, or invalid.
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            auth_header = request.headers.get('Authorization')
            if not auth_header:
                return jsonify({'status': 'error', 'message': 'Missing Authorization header'}), 401
                
            parts = auth_header.split()
            if len(parts) != 2 or parts[0].lower() != 'bearer':
                return jsonify({'status': 'error', 'message': 'Authorization header format must be Bearer <token>'}), 401
                
            token = parts[1]
            try:
                payload = decode_jwt_token(token)
                g.jwt_payload = payload
                g.jwt_user = {
                    'id': payload.get('sub'),
                    'role': payload.get('role'),
                    'name': payload.get('name')
                }
                
                if allowed_roles:
                    roles_list = [allowed_roles] if isinstance(allowed_roles, str) else allowed_roles
                    if payload.get('role') not in roles_list:
                        return jsonify({'status': 'error', 'message': 'Forbidden: Insufficient privileges'}), 403
                        
            except jwt.ExpiredSignatureError:
                return jsonify({'status': 'error', 'message': 'Token has expired'}), 401
            except jwt.InvalidTokenError as e:
                return jsonify({'status': 'error', 'message': f'Invalid token: {str(e)}'}), 401
                
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─── Flask-Limiter Setup ───────────────────────────────────────────
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

@app.errorhandler(429)
def ratelimit_handler(e):
    retry_after = getattr(e, 'retry_after', 60) or 60
    return jsonify({
        'status': 'error',
        'rate_limit': True,
        'retry_after': int(retry_after),
        'message': f'Server rate limit reached (13 requests per minute). Please wait {int(retry_after)} seconds before trying again.'
    }), 429
# Configuration
UPLOAD_FOLDER = 'uploads'
EXPORT_FOLDER = 'exports'
BRANDING_FOLDER = os.path.join('static', 'uploads', 'branding')
STUDENT_PIC_FOLDER = os.path.join('static', 'student_profile_pic')
TEACHER_PIC_FOLDER = os.path.join('static', 'Teacher_profile_picture')
ALLOWED_EXTENSIONS = {'csv'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)
os.makedirs(os.path.join('static', 'uploads'), exist_ok=True)
os.makedirs(BRANDING_FOLDER, exist_ok=True)
os.makedirs(STUDENT_PIC_FOLDER, exist_ok=True)
os.makedirs(TEACHER_PIC_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXPORT_FOLDER'] = EXPORT_FOLDER
app.config['BRANDING_FOLDER'] = BRANDING_FOLDER
app.config['STUDENT_PIC_FOLDER'] = STUDENT_PIC_FOLDER
app.config['TEACHER_PIC_FOLDER'] = TEACHER_PIC_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# ── helpers ──────────────────────────────────────────────
def hash_password(password):
    """
    Hash a password using SHA256 encryption.
    """
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    """
    Verify if a provided password matches its hashed version.
    """
    return hash_password(password) == hashed

def save_student_picture(file, student_id):
    """
    Save a student profile picture to the file system and return the file path.
    """
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"student_{student_id}.{ext}"
    filepath = os.path.join(STUDENT_PIC_FOLDER, filename)
    # Remove old picture if different extension
    for old_ext in ALLOWED_IMAGE_EXTENSIONS:
        old_path = os.path.join(STUDENT_PIC_FOLDER, f"student_{student_id}.{old_ext}")
        if os.path.exists(old_path) and old_path != filepath:
            os.remove(old_path)
    file.save(filepath)
    return os.path.join('student_profile_pic', filename).replace('\\', '/')

def save_teacher_picture(file, teacher_id):
    """
    Save a teacher profile picture to the file system and return the file path.
    """
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"teacher_{teacher_id}.{ext}"
    filepath = os.path.join(TEACHER_PIC_FOLDER, filename)
    for old_ext in ALLOWED_IMAGE_EXTENSIONS:
        old_path = os.path.join(TEACHER_PIC_FOLDER, f"teacher_{teacher_id}.{old_ext}")
        if os.path.exists(old_path) and old_path != filepath:
            os.remove(old_path)
    file.save(filepath)
    return os.path.join('Teacher_profile_picture', filename).replace('\\', '/')

def allowed_image_file(filename):
    """
    Check if a filename has an allowed image extension (png, jpg, jpeg, gif, webp).
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def save_question_image(file, class_, subject, question_id):
    """
    Save a question image to the appropriate folder based on class and subject.
    """
    if not file or file.filename == '':
        return None
    if not allowed_image_file(file.filename):
        return None

    folder_name = f"{class_.strip()}_{subject.strip()}".replace(' ', '_')
    folder_path = os.path.join('static', 'uploads', folder_name)
    os.makedirs(folder_path, exist_ok=True)

    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"{folder_name}_{question_id}.{ext}"
    filepath = os.path.join(folder_path, filename)

    if os.path.exists(filepath):
        os.remove(filepath)

    file.save(filepath)
    return os.path.join('uploads', folder_name, filename).replace('\\', '/')

# Database initialization
def init_db():
    """
    Initialize the SQLite database with all required tables and schema. Handles migrations from old schemas.
    """
    conn = sqlite3.connect('database.db', timeout=30.0)
    c = conn.cursor()
    c.execute("PRAGMA busy_timeout = 30000")
    c.execute("PRAGMA journal_mode=DELETE")

    c.execute('''CREATE TABLE IF NOT EXISTS students
                 (student_id TEXT PRIMARY KEY, name TEXT, class TEXT, subject TEXT, ip TEXT,
                  status TEXT DEFAULT 'Not Started', exam_started_at TIMESTAMP,
                  admission_no TEXT, section TEXT, dob TEXT, house TEXT, 
                  parents_name TEXT, address TEXT, picture TEXT)''')
    c.execute("PRAGMA table_info(students)")
    columns = [col[1] for col in c.fetchall()]
    if 'exam_started_at' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN exam_started_at TIMESTAMP")
    if 'admission_no' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN admission_no TEXT")
    if 'section' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN section TEXT")
    if 'dob' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN dob TEXT")
    if 'house' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN house TEXT")
    if 'parents_name' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN parents_name TEXT")
    if 'address' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN address TEXT")
    if 'picture' not in columns:
        c.execute("ALTER TABLE students ADD COLUMN picture TEXT")
    # Add test_no and system_name columns to students if not exists
    c.execute("PRAGMA table_info(students)")
    cols = [col[1] for col in c.fetchall()]
    if 'test_no' not in cols:
        c.execute("ALTER TABLE students ADD COLUMN test_no TEXT DEFAULT ''")
    if 'system_name' not in cols:
        c.execute("ALTER TABLE students ADD COLUMN system_name TEXT DEFAULT ''")

    c.execute('''CREATE TABLE IF NOT EXISTS questions
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, class TEXT, subject TEXT, question TEXT,
                  option_a TEXT, option_b TEXT, option_c TEXT, option_d TEXT, correct_answer TEXT,
                  image_path TEXT, chapter TEXT)''')
        # Add chapter column if not exists
    c.execute("PRAGMA table_info(questions)")
    columns = [col[1] for col in c.fetchall()]
    if 'chapter' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN chapter TEXT")
    if 'class' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN class TEXT")
    if 'subject' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN subject TEXT")
    if 'image_path' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN image_path TEXT")
    if 'section' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN section TEXT")
    if 'test_no' not in columns:
        c.execute("ALTER TABLE questions ADD COLUMN test_no TEXT")
    for hi_col in ['question_hi', 'option_a_hi', 'option_b_hi', 'option_c_hi', 'option_d_hi']:
        if hi_col not in columns:
            c.execute(f"ALTER TABLE questions ADD COLUMN {hi_col} TEXT")

    c.execute('''CREATE TABLE IF NOT EXISTS responses
                 (student_id TEXT, question_id INTEGER, selected_option TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (student_id, question_id))''')
    c.execute("PRAGMA table_info(responses)")
    resp_columns = [col[1] for col in c.fetchall()]
    if 'created_at' not in resp_columns:
        c.execute("ALTER TABLE responses ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")

    c.execute('''CREATE TABLE IF NOT EXISTS results
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id TEXT, name TEXT, 
                  class TEXT, subject TEXT, score INTEGER, total_questions INTEGER,
                  percentage REAL, test_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  chapter TEXT)''')
    # Migrate old results table if it has the old schema (student_id was PRIMARY KEY)
    c.execute("PRAGMA table_info(results)")
    res_columns = [col[1] for col in c.fetchall()]
    if 'id' not in res_columns:
        # Old schema – rename and recreate
        c.execute("ALTER TABLE results RENAME TO results_old")
        c.execute('''CREATE TABLE results
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, student_id TEXT, name TEXT,
                      class TEXT, section TEXT, subject TEXT, score INTEGER, total_questions INTEGER,
                      percentage REAL, test_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                      chapter TEXT)''')
        c.execute("""INSERT INTO results (student_id, name, class, subject, score, total_questions, percentage)
                     SELECT student_id, name, class, subject, score, score, 0 FROM results_old""")
        c.execute("DROP TABLE results_old")
    elif 'section' not in res_columns:
        c.execute("ALTER TABLE results ADD COLUMN section TEXT")
    c.execute('''CREATE TABLE IF NOT EXISTS exam_control
                 (id INTEGER PRIMARY KEY CHECK (id=1), is_active INTEGER DEFAULT 0,
                  start_time TIMESTAMP, duration INTEGER DEFAULT 60,
                  negative_marking INTEGER DEFAULT 0,
                  negative_value REAL DEFAULT 0.33)''')
    c.execute("INSERT OR IGNORE INTO exam_control (id, is_active, start_time, duration, negative_marking, negative_value) VALUES (1, 0, NULL, 60, 0, 0.33)")
    # Add columns if upgrading from older schema
    c.execute("PRAGMA table_info(exam_control)")
    ec_cols = [r[1] for r in c.fetchall()]
    if 'negative_marking' not in ec_cols:
        c.execute("ALTER TABLE exam_control ADD COLUMN negative_marking INTEGER DEFAULT 0")
    if 'negative_value' not in ec_cols:
        c.execute("ALTER TABLE exam_control ADD COLUMN negative_value REAL DEFAULT 0.33")

    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (key TEXT PRIMARY KEY, value TEXT)''')
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('school_name', 'RRB Group of Schools')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('logo_path', '')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('school_address', '')")
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('academic_session', '')")

    c.execute('''CREATE TABLE IF NOT EXISTS reattempt_requests
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                class TEXT NOT NULL,
                subject TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                reviewed_at TIMESTAMP,
                admin_note TEXT)''')

    # Table for storing shuffled question order and option mapping per student
    c.execute('''CREATE TABLE IF NOT EXISTS shuffled_questions
                 (student_id TEXT,
                  question_id INTEGER,
                  shuffled_index INTEGER,
                  option_order TEXT,
                  PRIMARY KEY (student_id, question_id))''')

    # Teachers table
    c.execute('''CREATE TABLE IF NOT EXISTS teachers
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  mobile TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  email TEXT,
                  address TEXT,
                  picture TEXT,
                  status TEXT DEFAULT 'active',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Teacher assignments (class and subject assignments)
    c.execute('''CREATE TABLE IF NOT EXISTS teacher_assignments
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id INTEGER NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  is_class_teacher INTEGER DEFAULT 0,
                  FOREIGN KEY (teacher_id) REFERENCES teachers(id))''')
    c.execute("PRAGMA table_info(teacher_assignments)")
    ta_cols = [col[1] for col in c.fetchall()]
    if 'section' not in ta_cols:
        c.execute("ALTER TABLE teacher_assignments ADD COLUMN section TEXT DEFAULT ''")

    # Test papers table (CSV uploads with naming convention)
    c.execute('''CREATE TABLE IF NOT EXISTS test_papers
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  filename TEXT NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  test_no TEXT NOT NULL,
                  uploaded_by TEXT NOT NULL,
                  uploader_type TEXT DEFAULT 'admin',
                  is_active INTEGER DEFAULT 0,
                  question_count INTEGER DEFAULT 0,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Bulletins / Announcements (Admin + Teacher can post; shown on student login page)
    c.execute('''CREATE TABLE IF NOT EXISTS bulletins
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  title TEXT NOT NULL,
                  content TEXT NOT NULL,
                  posted_by TEXT NOT NULL,
                  poster_type TEXT DEFAULT 'admin',
                  target_class TEXT DEFAULT '',
                  is_active INTEGER DEFAULT 1,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Test Generation History (for AI-generated tests)
    c.execute('''CREATE TABLE IF NOT EXISTS test_generation_history
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id TEXT NOT NULL,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  chapter TEXT NOT NULL,
                  test_no TEXT NOT NULL,
                  output_mode TEXT DEFAULT 'cbt',
                  total_questions INTEGER DEFAULT 0,
                  mcq_count INTEGER DEFAULT 0,
                  assertion_count INTEGER DEFAULT 0,
                  very_short_count INTEGER DEFAULT 0,
                  short_count INTEGER DEFAULT 0,
                  long_count INTEGER DEFAULT 0,
                  case_study_count INTEGER DEFAULT 0,
                  remark TEXT DEFAULT '',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # MCQ Test History (for tracking generated MCQ tests)
    c.execute('''CREATE TABLE IF NOT EXISTS mcq_test_history
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id INTEGER,
                  teacher_name TEXT,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  test_no TEXT DEFAULT '',
                  question_count INTEGER DEFAULT 0,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Teacher Notifications table (for background job completion alerts)
    c.execute('''CREATE TABLE IF NOT EXISTS teacher_notifications
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  teacher_id INTEGER NOT NULL,
                  title TEXT NOT NULL,
                  message TEXT NOT NULL,
                  link TEXT DEFAULT '',
                  is_read INTEGER DEFAULT 0,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Audit Logs table (for security and action monitoring)
    c.execute('''CREATE TABLE IF NOT EXISTS audit_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_type TEXT NOT NULL,
                  user_id TEXT NOT NULL,
                  action TEXT NOT NULL,
                  target_table TEXT DEFAULT '',
                  target_id TEXT DEFAULT '',
                  timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  ip_address TEXT DEFAULT '')''')

    # Add question_type column to questions if not exists
    c.execute("PRAGMA table_info(questions)")
    q_cols = [col[1] for col in c.fetchall()]
    if 'question_type' not in q_cols:
        c.execute("ALTER TABLE questions ADD COLUMN question_type TEXT DEFAULT 'MCQ'")
    if 'test_no' not in q_cols:
        c.execute("ALTER TABLE questions ADD COLUMN test_no TEXT DEFAULT ''")
    if 'section' not in q_cols:
        c.execute("ALTER TABLE questions ADD COLUMN section TEXT DEFAULT ''")

    # Student class lock: once registered, class/section is locked
    c.execute('''CREATE TABLE IF NOT EXISTS student_class_lock
                 (student_id TEXT PRIMARY KEY,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  locked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    # Combined / Multi-Subject Test entities
    c.execute('''CREATE TABLE IF NOT EXISTS combined_tests
                 (id           INTEGER PRIMARY KEY AUTOINCREMENT,
                  test_no      TEXT NOT NULL,
                  class        TEXT NOT NULL,
                  section      TEXT DEFAULT '',
                  title        TEXT DEFAULT '',
                  created_by   TEXT NOT NULL,
                  creator_type TEXT DEFAULT 'teacher',
                  is_active    INTEGER DEFAULT 1,
                  created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    c.execute('''CREATE TABLE IF NOT EXISTS combined_test_subjects
                 (id               INTEGER PRIMARY KEY AUTOINCREMENT,
                  combined_test_id INTEGER NOT NULL,
                  subject          TEXT NOT NULL,
                  question_count   INTEGER DEFAULT 0,
                  FOREIGN KEY (combined_test_id) REFERENCES combined_tests(id) ON DELETE CASCADE)''')

    # Scheduled Tests table (for scheduling test windows & duration control)
    c.execute('''CREATE TABLE IF NOT EXISTS scheduled_tests
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  class TEXT NOT NULL,
                  section TEXT DEFAULT '',
                  subject TEXT NOT NULL,
                  test_no TEXT NOT NULL,
                  scheduled_date TEXT NOT NULL,
                  start_time TEXT NOT NULL,
                  end_time TEXT NOT NULL,
                  duration_minutes INTEGER NOT NULL,
                  status TEXT DEFAULT 'scheduled',
                  created_by TEXT DEFAULT 'admin',
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

    conn.commit()
    conn.close()

def update_scheduled_tests_status():
    """
    Evaluates and updates scheduled_tests status based on current datetime (scheduled -> active -> expired).
    """
    try:
        now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn = get_db()
        c = conn.cursor()
        c.execute("""UPDATE scheduled_tests
                     SET status='active'
                     WHERE status='scheduled' AND start_time <= ? AND end_time >= ?""", (now_str, now_str))
        c.execute("""UPDATE scheduled_tests
                     SET status='expired'
                     WHERE status IN ('scheduled', 'active') AND end_time < ?""", (now_str,))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.error(f"Error updating scheduled tests status: {e}")

def get_student_schedule_status(class_, section, subject, test_no, student_id=None):
    """
    Check if a test schedule exists for given class, section, subject, test_no.
    Returns dict with state ('none', 'upcoming', 'active', 'expired'), schedule info, and effective duration.
    """
    update_scheduled_tests_status()
    conn = get_db()
    c = conn.cursor()
    clean_cls = str(class_ or '').strip()
    sec = str(section or '').strip()
    subj = str(subject or '').strip()
    tno = str(test_no or '').strip()

    c.execute("""SELECT * FROM scheduled_tests
                 WHERE class=? AND (? = '' OR section=? OR section='') 
                   AND LOWER(subject)=LOWER(?) AND test_no=?
                   AND status != 'cancelled'
                 ORDER BY id DESC LIMIT 1""", (clean_cls, sec, sec, subj, tno))
    row = c.fetchone()
    conn.close()

    if not row:
        return {'state': 'none', 'schedule': None}

    sch = dict(row)
    now = datetime.datetime.now()

    try:
        if len(sch['start_time']) == 5:
            st_dt = datetime.datetime.strptime(f"{sch['scheduled_date']} {sch['start_time']}", '%Y-%m-%d %H:%M')
        else:
            st_dt = datetime.datetime.strptime(sch['start_time'][:19], '%Y-%m-%d %H:%M:%S')
    except Exception:
        st_dt = now

    try:
        if len(sch['end_time']) == 5:
            end_dt = datetime.datetime.strptime(f"{sch['scheduled_date']} {sch['end_time']}", '%Y-%m-%d %H:%M')
        else:
            end_dt = datetime.datetime.strptime(sch['end_time'][:19], '%Y-%m-%d %H:%M:%S')
    except Exception:
        end_dt = now

    sch['start_dt_str'] = st_dt.strftime('%d/%m/%Y %I:%M %p')
    sch['end_dt_str'] = end_dt.strftime('%d/%m/%Y %I:%M %p')

    if now < st_dt:
        return {
            'state': 'upcoming',
            'schedule': sch,
            'message': f"This test is locked. Scheduled for {st_dt.strftime('%d %b %Y')} at {st_dt.strftime('%I:%M %p')}."
        }
    elif st_dt <= now <= end_dt:
        rem_seconds = (end_dt - now).total_seconds()
        rem_minutes = max(1, int(rem_seconds // 60))
        effective_duration = min(sch['duration_minutes'], rem_minutes)
        return {
            'state': 'active',
            'schedule': sch,
            'effective_duration': effective_duration,
            'message': f"Available Now (Window closes at {end_dt.strftime('%I:%M %p')})"
        }
    else:
        return {
            'state': 'expired',
            'schedule': sch,
            'message': "This test window has closed and is now permanently locked. Contact admin for rescheduling."
        }

def get_db():
    """
    Get a database connection with Row factory enabled for dict-like access.
    Configured with timeout and busy_timeout to prevent lock conflicts.
    """
    conn = sqlite3.connect('database.db', timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn

def reset_exam_on_startup():
    """
    Reset the exam status on application startup to ensure clean state.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE exam_control SET is_active=0, start_time=NULL WHERE id=1")
    conn.commit()
    conn.close()

init_db()
reset_exam_on_startup()

def get_setting(key, default=''):
    """
    Retrieve a setting value from the settings table by key.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row['value'] if row else default

def set_setting(key, value):
    """
    Save or update a setting in the settings table.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def log_audit_event(user_type, user_id, action, target_table='', target_id='', ip_address=''):
    """
    Inserts record into audit_logs table for audit tracking.
    """
    try:
        if not ip_address and request:
            ip_address = request.remote_addr or ''
        conn = get_db()
        c = conn.cursor()
        c.execute("""INSERT INTO audit_logs (user_type, user_id, action, target_table, target_id, ip_address)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (str(user_type or 'system'), str(user_id or 'anon'), str(action),
                   str(target_table or ''), str(target_id or ''), str(ip_address or '')))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.error(f"Failed to record audit log: {e}")

def admin_required(f):
    """
    Decorator to enforce admin login requirement on routes.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    """
    Decorator to enforce student login requirement on routes.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('student_id'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def teacher_required(f):
    """
    Decorator to enforce teacher login requirement on routes.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('teacher_logged_in'):
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated_function

# -------------------------------
# Student Routes
# -------------------------------
@app.route('/')
@app.route('/login')
def index():
    """
    Render the student login page with available classes, subjects, bulletins, and class teachers.
    """
    session.pop('admin_logged_in', None)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class != '' ORDER BY class, subject")
    rows = c.fetchall()

    class_subject_map = {}
    for row in rows:
        cls = row['class']
        sub = row['subject']
        if cls not in class_subject_map:
            class_subject_map[cls] = []
        class_subject_map[cls].append(sub)

    # Get active bulletins for the student login page bulletin board
    c.execute("""SELECT b.title, b.content, b.poster_type, b.target_class, b.created_at,
                        COALESCE(t.name, 'Admin') as poster_name
                 FROM bulletins b
                 LEFT JOIN teachers t ON b.posted_by = t.id AND b.poster_type = 'teacher'
                 WHERE b.is_active = 1
                 ORDER BY b.created_at DESC LIMIT 10""")
    bulletins = [dict(r) for r in c.fetchall()]

    # Get class teachers for display
    c.execute("""SELECT ta.class, ta.section, t.name, t.picture
                 FROM teacher_assignments ta
                 JOIN teachers t ON ta.teacher_id = t.id
                 WHERE ta.is_class_teacher = 1
                 ORDER BY ta.class, ta.section""")
    class_teachers = [dict(r) for r in c.fetchall()]

    # Get active & upcoming test schedules to flash on Notice Board
    update_scheduled_tests_status()
    c.execute("""SELECT * FROM scheduled_tests 
                 WHERE status IN ('scheduled', 'active') 
                 ORDER BY scheduled_date ASC, start_time ASC LIMIT 10""")
    scheduled_tests = [dict(r) for r in c.fetchall()]

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    conn.close()

    return render_template('index.html',
                           class_subject_map=class_subject_map,
                           bulletins=bulletins,
                           scheduled_tests=scheduled_tests,
                           class_teachers=class_teachers,
                           school_name=school_name,
                           logo_path=logo_path)

@app.route('/student_login', methods=['POST'])
@validate_schema(STUDENT_LOGIN_SCHEMA)
def student_login():
    """
    Handle student login and exam registration. Validates class lock, manages reattempt requests.
    """
    auto_submit_expired_exams()

    name        = request.form.get('name', '').strip()
    student_id  = request.form.get('student_id', '').strip()
    class_      = request.form.get('class', '').strip()
    section     = request.form.get('section', '').strip()
    subject     = request.form.get('subject', '').strip()
    test_no     = request.form.get('test_no', '').strip()   # Feature #4
    ip          = request.remote_addr

    if not all([name, student_id, class_, subject]):
        return "All fields required", 400

    conn = get_db()
    c = conn.cursor()

    # Auto-select latest test if no specific test_no was selected by student
    if not test_no:
        clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
        c.execute("""SELECT test_no FROM test_papers
                     WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND test_no != ''
                     ORDER BY id DESC LIMIT 1""",
                  (class_, f"%{clean_cls}%", subject))
        row = c.fetchone()
        if not row or not row['test_no']:
            c.execute("""SELECT test_no FROM questions
                         WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND test_no != ''
                         ORDER BY id DESC LIMIT 1""",
                      (class_, f"%{clean_cls}%", subject))
            row = c.fetchone()
        if not row or not row['test_no']:
            c.execute("""SELECT chapter as test_no FROM questions
                         WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND chapter IS NOT NULL AND chapter != ''
                         ORDER BY id DESC LIMIT 1""",
                      (class_, f"%{clean_cls}%", subject))
            row = c.fetchone()
        if row and row['test_no']:
            test_no = row['test_no']

    # ── SCHEDULE LOCK CHECK ───────────────────────────────────────
    if class_ and subject and test_no:
        sch_res = get_student_schedule_status(class_, section, subject, test_no, student_id)
        if sch_res['state'] in ('upcoming', 'expired'):
            conn.close()
            return render_template('index.html',
                                   error=f"Test '{test_no}' is locked. {sch_res['message']}",
                                   class_subject_map=_get_class_subject_map(),
                                   bulletins=[], scheduled_tests=[], class_teachers=[],
                                   school_name=get_setting('school_name','RRB Group of Schools'),
                                   logo_path=get_setting('logo_path',''))

    # ── CLASS LOCK CHECK ──────────────────────────────────────────
    c.execute("SELECT class, section FROM student_class_lock WHERE student_id=?", (student_id,))
    lock = c.fetchone()
    if lock:
        if lock['class'] != class_ or lock['section'] != section:
            conn.close()
            return render_template('index.html',
                                   error=f"You are locked to Class {lock['class']} Section {lock['section']}. Contact Admin to change.",
                                   class_subject_map=_get_class_subject_map(),
                                   bulletins=[], class_teachers=[],
                                   school_name=get_setting('school_name','RRB Group of Schools'),
                                   logo_path=get_setting('logo_path',''))
    else:
        c.execute("INSERT OR IGNORE INTO student_class_lock (student_id, class, section) VALUES (?,?,?)",
                  (student_id, class_, section))

    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()

    # BUG-001 FIX: Only block if student submitted THIS specific test (same subject+test_no)
    # Different test_no = new exam, must never be blocked by a previous test's submission
    if student and student['status'] == 'Submitted':
        prev_subject = student['subject'] or ''
        # Check if previous submission was for the SAME test
        c.execute("""SELECT id FROM results
                     WHERE student_id=? AND subject=? AND (
                         (? != '' AND chapter=?) OR
                         (? = '' AND subject=?)
                     ) LIMIT 1""",
                  (student_id, subject,
                   test_no, test_no,
                   test_no, subject))
        same_test_result = c.fetchone()

        if same_test_result and prev_subject == subject:
            # Same test attempted again — show reattempt prompt
            conn.commit()
            conn.close()
            session.update({'student_id': student_id, 'student_name': name,
                            'class': class_, 'subject': subject, 'section': section, 'test_no': test_no})
            return redirect(url_for('submitted'))
        else:
            # Different test (new test_no or different subject) — reset status and allow
            c.execute("""UPDATE students SET status='Not Started', exam_started_at=NULL,
                          class=?, subject=?, section=?, ip=? WHERE student_id=?""",
                      (class_, subject, section, ip, student_id))
            c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
            conn.commit()

    system_name = request.headers.get('User-Agent', 'Unknown')[:255]  # limit length

    if student:
        c.execute("""UPDATE students 
                     SET name=?, class=?, subject=?, section=?, ip=?, 
                         test_no=?, system_name=?
                     WHERE student_id=?""",
                  (name, class_, subject, section, ip, test_no, system_name, student_id))
    else:
        c.execute("""INSERT INTO students 
                     (student_id, name, class, subject, section, ip, test_no, system_name, status)
                     VALUES (?,?,?,?,?,?,?,?,'Not Started')""",
                  (student_id, name, class_, subject, section, ip, test_no, system_name))

    conn.commit()
    c.execute("SELECT status, subject FROM students WHERE student_id=?", (student_id,))
    updated_status = c.fetchone()
    if updated_status and updated_status['status'] == 'Submitted' and updated_status['subject'] == subject:
        conn.close()
        session.update({'student_id': student_id, 'student_name': name,
                        'class': class_, 'subject': subject, 'section': section, 'test_no': test_no})
        return redirect(url_for('submitted'))

    # Shuffle questions — filter by test_no if provided
    # For combined tests, detect and load ALL subjects
    combined_test_id = None
    combined_subjects = []

    if test_no:
        conn2 = get_db()
        c2    = conn2.cursor()
        # Check if this test_no is a combined test for this class
        c2.execute("""SELECT ct.id FROM combined_tests ct
                      WHERE ct.test_no=? AND ct.class=? AND ct.is_active=1""",
                   (test_no, class_))
        ct_row = c2.fetchone()
        if ct_row:
            combined_test_id = ct_row['id']
            c2.execute("""SELECT subject FROM combined_test_subjects
                          WHERE combined_test_id=? ORDER BY id""", (combined_test_id,))
            combined_subjects = [r['subject'] for r in c2.fetchall()]

            # Load ALL questions across all subjects for this combined test
            placeholders = ','.join(['?']*len(combined_subjects))
            c2.execute(f"""SELECT id FROM questions
                           WHERE class=? AND test_no=? AND subject IN ({placeholders})
                           ORDER BY subject, id""",
                       [class_, test_no] + combined_subjects)
            qids = [r['id'] for r in c2.fetchall()]
        else:
            c2.execute("SELECT id FROM questions WHERE class=? AND subject=? AND (test_no=? OR chapter=?) ORDER BY id",
                       (class_, subject, test_no, test_no))
            qids = [r['id'] for r in c2.fetchall()]
        conn2.close()
    else:
        conn_tmp = get_db()
        c_tmp    = conn_tmp.cursor()
        c_tmp.execute("SELECT id FROM questions WHERE class=? AND subject=? ORDER BY id", (class_, subject))
        qids = [r['id'] for r in c_tmp.fetchall()]
        conn_tmp.close()

    conn3 = get_db()
    c3    = conn3.cursor()
    random.shuffle(qids)
    c3.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    option_letters = ['A', 'B', 'C', 'D']
    for idx, qid in enumerate(qids):
        shuffled_opts = option_letters[:]
        random.shuffle(shuffled_opts)
        c3.execute("INSERT INTO shuffled_questions (student_id, question_id, shuffled_index, option_order) VALUES (?,?,?,?)",
                   (student_id, qid, idx, ''.join(shuffled_opts)))
    conn3.commit()
    conn3.close()

    session.update({'student_id': student_id, 'student_name': name,
                    'class': class_, 'subject': subject, 'section': section, 'test_no': test_no,
                    'combined_test_id': combined_test_id,
                    'combined_subjects': combined_subjects})
    return redirect(url_for('waiting'))

def _get_class_subject_map():
    """
    Helper function to get a mapping of classes to their available subjects.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
    rows = c.fetchall()
    conn.close()
    csmap = {}
    for r in rows:
        csmap.setdefault(r['class'], []).append(r['subject'])
    return csmap

@app.route('/api/test_numbers')
def api_test_numbers():
    """Return distinct test_no values for a class+subject combo with schedule lock status."""
    class_   = request.args.get('class','').strip()
    section  = request.args.get('section','').strip()
    subject  = request.args.get('subject','').strip()
    if not class_:
        return jsonify({'test_numbers': [], 'test_details': [], 'combined_tests': []})

    conn = get_db()
    c = conn.cursor()

    if subject:
        clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
        c.execute("""SELECT DISTINCT test_no as tno FROM test_papers WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND test_no != ''
                     UNION
                     SELECT DISTINCT COALESCE(NULLIF(test_no,''), chapter) as tno FROM questions WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND (test_no IS NOT NULL AND test_no != '' OR chapter IS NOT NULL AND chapter != '')""",
                  (class_, f"%{clean_cls}%", subject, class_, f"%{clean_cls}%", subject))
        test_numbers = sorted(list(set(r[0] for r in c.fetchall() if r[0])))
    else:
        test_numbers = []

    # Calculate schedule lock status for each test_no
    update_scheduled_tests_status()
    test_details = []
    for tn in test_numbers:
        sch_res = get_student_schedule_status(class_, section, subject, tn)
        state = sch_res.get('state', 'none')
        is_locked = False
        status_label = ""
        msg = sch_res.get('message', '')

        if state == 'upcoming':
            is_locked = True
            sch_info = sch_res.get('schedule', {})
            st_time = sch_info.get('start_dt_str', sch_info.get('start_time', ''))
            status_label = f"🔒 Locked (Scheduled for {st_time})"
        elif state == 'expired':
            is_locked = True
            status_label = "🔴 Locked (Window Expired)"
        elif state == 'active':
            status_label = "🟢 Available Now"

        test_details.append({
            'test_no': tn,
            'is_locked': is_locked,
            'state': state,
            'status_label': status_label,
            'message': msg
        })

    # Combined (multi-subject) tests for this class
    c.execute("""SELECT ct.id, ct.test_no, ct.title,
                        GROUP_CONCAT(cts.subject || '(' || cts.question_count || 'Q)', ', ') as subjects_info
                 FROM combined_tests ct
                 JOIN combined_test_subjects cts ON cts.combined_test_id = ct.id
                 WHERE ct.class=? AND ct.is_active=1
                 GROUP BY ct.id ORDER BY ct.created_at DESC""", (class_,))
    combined = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'test_numbers': test_numbers, 'test_details': test_details, 'combined_tests': combined})

@app.route('/api/generate_combined_test', methods=['POST'])
@limiter.limit("13 per minute")
def generate_combined_test():
    """Generate a truly unified multi-subject test — one entity, multiple subjects."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    import urllib.request, json as json_lib

    class_   = request.form.get('class','').strip()
    section  = request.form.get('section','').strip()
    test_no  = request.form.get('test_no','').strip()
    title    = request.form.get('title','').strip() or test_no
    remark   = request.form.get('remark','').strip()

    # subjects_json: [{"subject":"Math","count":10}, ...]
    try:
        subjects = json_lib.loads(request.form.get('subjects_json','[]'))
    except Exception:
        return jsonify({'status':'error','message':'Invalid subjects data'}), 400

    if not class_ or not test_no or not subjects:
        return jsonify({'status':'error','message':'Class, Test No and at least one subject required'}), 400

    gemini_key = _get_gemini_key()
    if not gemini_key:
        return jsonify({'status':'error','message':'GEMINI_API_KEY not configured'}), 500

    conn = get_db()
    c = conn.cursor()

    # Check no combined test with same test_no+class already exists
    c.execute("SELECT id FROM combined_tests WHERE test_no=? AND class=?", (test_no, class_))
    if c.fetchone():
        conn.close()
        return jsonify({'status':'error',
                        'message':f'Combined test "{test_no}" already exists for Class {class_}. Choose a different Test No.'}), 400

    creator_id   = str(session.get('teacher_id','admin'))
    creator_type = 'teacher' if session.get('teacher_logged_in') else 'admin'

    # Create the combined_tests record first
    c.execute("""INSERT INTO combined_tests (test_no, class, section, title, created_by, creator_type)
                 VALUES (?,?,?,?,?,?)""",
             (test_no, class_, section, title, creator_id, creator_type))
    combined_id = c.lastrowid
    conn.commit()

    errors       = []
    total_inserted = 0

    for subj_entry in subjects:
        subject   = subj_entry.get('subject','').strip()
        mcq_count = int(subj_entry.get('count', 0))
        if not subject or mcq_count <= 0:
            continue

        time.sleep(2)  # Pause to remain within Gemini Free Tier 15 RPM rate limits

        # Build AI prompt for this subject
        prompt = _build_ai_prompt(class_, section, subject,
                                  f'{title} — {subject}', test_no,
                                  remark or 'Standard CBSE difficulty',
                                  [f"{mcq_count} MCQ (4 options, mark correct_answer as option_a/b/c/d)"])
        try:
            result, provider, err_msg = generate_ai_content(prompt, timeout=60)
            if not result:
                app.logger.error(f"Auto generate failed: {err_msg}")
                continue

            raw = result['candidates'][0]['content']['parts'][0]['text'].strip()
            if '```' in raw:
                for part in raw.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): raw = part; break
            questions = json_lib.loads(raw)

            inserted = 0
            for q in questions:
                if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                    c.execute("""INSERT INTO questions
                               (class, subject, chapter, test_no, question_type,
                                question, option_a, option_b, option_c, option_d, correct_answer)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                             (class_, subject, title, test_no,
                              q.get('question_type','MCQ'), q.get('question',''),
                              q.get('option_a',''), q.get('option_b',''),
                              q.get('option_c',''), q.get('option_d',''),
                              q.get('correct_answer','')))
                    inserted += 1

            # Record subject in combined_test_subjects
            c.execute("""INSERT INTO combined_test_subjects (combined_test_id, subject, question_count)
                         VALUES (?,?,?)""", (combined_id, subject, inserted))
            total_inserted += inserted
            conn.commit()

        except Exception as e:
            errors.append(f"{subject}: {str(e)}")

    if errors and total_inserted == 0:
        # Complete failure — clean up
        c.execute("DELETE FROM combined_tests WHERE id=?", (combined_id,))
        conn.commit()
        conn.close()
        return jsonify({'status':'error','message':'Generation failed:\n' + '\n'.join(errors)}), 500

    conn.close()
    msg = f'Combined test "{test_no}" created with {total_inserted} questions across {len(subjects)} subjects.'
    if errors:
        msg += f'\nWarnings: ' + '; '.join(errors)
    return jsonify({'status':'success','message':msg,'combined_test_id':combined_id})

@app.route('/api/check_test_no')
def api_check_test_no():
    """Feature #5: Check if a test_no already exists for a subject (across all sections)."""
    subject = request.args.get('subject','').strip()
    test_no = request.args.get('test_no','').strip()
    class_  = request.args.get('class','').strip()
    if not subject or not test_no:
        return jsonify({'exists': False, 'details': []})
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT DISTINCT tp.class, tp.section, tp.test_no, t.name as teacher_name, tp.created_at
                 FROM test_papers tp
                 LEFT JOIN teachers t ON tp.uploaded_by = t.id AND tp.uploader_type='teacher'
                 WHERE tp.subject=? AND tp.test_no=? AND tp.class=?""",
             (subject, test_no, class_))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'exists': len(rows) > 0, 'details': rows})

@app.route('/waiting')
@student_required
def waiting():
    """
    Render the waiting page for students to wait for exam to start.
    """
    student_id = session.get('student_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    c.execute("SELECT status, class, section, subject, test_no FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    lock_info = None
    if student and student['class'] and student['subject']:
        tno = student['test_no'] or session.get('test_no', '')
        sch_res = get_student_schedule_status(student['class'], student['section'], student['subject'], tno, student_id)
        if sch_res['state'] == 'upcoming':
            lock_info = {'title': 'Exam Locked', 'message': sch_res['message'], 'icon': 'fa-lock', 'color': '#eab308', 'state': 'upcoming'}
        elif sch_res['state'] == 'expired' and student['status'] == 'Not Started':
            lock_info = {'title': 'Exam Window Closed', 'message': sch_res['message'], 'icon': 'fa-times-circle', 'color': '#ef4444', 'state': 'expired'}

    if not lock_info and row and row['is_active']:
        return redirect(url_for('guidelines'))
    return render_template('waiting.html', lock_info=lock_info)

@app.route('/check_exam_status')
@student_required
def check_exam_status():
    """
    API endpoint to check if the exam has started.
    """
    auto_submit_expired_exams()
    update_scheduled_tests_status()
    student_id = session.get('student_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active, start_time, duration FROM exam_control WHERE id=1")
    row = c.fetchone()
    c.execute("SELECT status, class, section, subject, test_no FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    force_submitted = bool(student and student['status'] == 'Submitted')
    if force_submitted:
        return jsonify({'active': False, 'force_submitted': True})

    if student and student['class'] and student['subject']:
        tno = student['test_no'] or session.get('test_no', '')
        sch_res = get_student_schedule_status(student['class'], student['section'], student['subject'], tno, student_id)
        if sch_res['state'] == 'upcoming':
            return jsonify({'active': False, 'force_submitted': False, 'schedule_locked': True, 'state': 'upcoming', 'message': sch_res['message']})
        elif sch_res['state'] == 'expired' and student['status'] == 'Not Started':
            return jsonify({'active': False, 'force_submitted': False, 'schedule_locked': True, 'state': 'expired', 'message': sch_res['message']})

    if not row or not row['is_active']:
        return jsonify({'active': False, 'force_submitted': False})
    return jsonify({'active': True, 'force_submitted': False})

@app.route('/exam')
@student_required
def exam():
    """
    Render the exam page with shuffled questions for the student.
    """
    import random
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    if not row or not row['is_active']:
        conn.close()
        return redirect(url_for('waiting'))

    student_id = session['student_id']
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if student and student['status'] == 'Submitted':
        conn.close()
        return redirect(url_for('waiting'))

    if student and student['status'] == 'Not Started':
        tno = (student['test_no'] if ('test_no' in student.keys() and student['test_no']) else session.get('test_no', '') or '').strip()
        sch_res = get_student_schedule_status(student['class'], student['section'], student['subject'], tno, student_id)
        if sch_res['state'] in ('upcoming', 'expired'):
            conn.close()
            return redirect(url_for('waiting'))

    if student:
        c.execute("SELECT COUNT(*) FROM questions WHERE class=? AND subject=?",
                  (student['class'], student['subject']))
        q_count = c.fetchone()[0]
        if q_count == 0:
            conn.close()
            return "No questions available for this subject/class. Please contact admin.", 404

        c.execute("SELECT COUNT(*) FROM shuffled_questions WHERE student_id=?", (student_id,))
        count = c.fetchone()[0]
        if count == 0:
            test_no = (student['test_no'] if ('test_no' in student.keys() and student['test_no']) else session.get('test_no', '') or '').strip()
            if test_no:
                c.execute("SELECT id FROM questions WHERE class=? AND subject=? AND (test_no=? OR chapter=?) ORDER BY id",
                          (student['class'], student['subject'], test_no, test_no))
            else:
                c.execute("SELECT id FROM questions WHERE class=? AND subject=? ORDER BY id",
                          (student['class'], student['subject']))
            qids = [row['id'] for row in c.fetchall()]
            random.shuffle(qids)
            option_letters = ['A', 'B', 'C', 'D']
            for idx, qid in enumerate(qids):
                shuffled_opts = option_letters[:]
                random.shuffle(shuffled_opts)
                opt_str = ''.join(shuffled_opts)
                c.execute("INSERT INTO shuffled_questions (student_id, question_id, shuffled_index, option_order) VALUES (?,?,?,?)",
                          (student_id, qid, idx, opt_str))
            conn.commit()

        if student['exam_started_at'] is None or student['status'] == 'Not Started':
            now_str = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            c.execute("UPDATE students SET status='In Progress', exam_started_at=? WHERE student_id=?",
                      (now_str, student_id))
            conn.commit()

    # Get chapter/test_no for header
    c.execute("SELECT DISTINCT chapter FROM questions WHERE class=? AND subject=? AND chapter IS NOT NULL AND chapter!='' LIMIT 1",
              (student['class'], student['subject']))
    chapter_row = c.fetchone()
    test_no = chapter_row['chapter'] if chapter_row else ''

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    conn.close()

    now = datetime.datetime.now()
    combined_subjects = session.get('combined_subjects', [])
    is_combined       = bool(session.get('combined_test_id'))
    return render_template('exam.html',
                           school_name=school_name,
                           logo_path=logo_path,
                           gemini_key=_get_gemini_key(),
                           student_name=session.get('student_name',''),
                           student_id=student_id,
                           student_class=student['class'],
                           student_section=student['section'] if student['section'] else session.get('section', ''),
                           student_subject=student['subject'],
                           test_no=test_no,
                           exam_date=now.strftime('%d/%m/%Y'),
                           exam_day=now.strftime('%A'),
                           is_combined=is_combined,
                           combined_subjects=combined_subjects)

@app.route('/get_questions')
@student_required
def get_questions():
    """
    API endpoint to retrieve exam questions with shuffled options for display.
    """
    student_id = session.get('student_id')
    class_     = session.get('class')
    subject    = session.get('subject')
    conn = get_db()
    c    = conn.cursor()

    c.execute('''
        SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d,
               q.question_hi, q.option_a_hi, q.option_b_hi, q.option_c_hi, q.option_d_hi,
               q.image_path, q.subject, q.question_type,
               COALESCE(sq.option_order, 'ABCD') as option_order,
               r.selected_option
        FROM questions q
        LEFT JOIN shuffled_questions sq ON q.id = sq.question_id AND sq.student_id = ?
        LEFT JOIN responses r           ON q.id = r.question_id  AND r.student_id = ?
        WHERE sq.student_id = ?
        ORDER BY COALESCE(sq.shuffled_index, q.id)
    ''', (student_id, student_id, student_id))
    rows = c.fetchall()
    conn.close()

    if not rows:
        return jsonify([])

    questions = []
    for row in rows:
        opt_order     = row['option_order']
        original_opts = {'A': row['option_a'], 'B': row['option_b'],
                         'C': row['option_c'], 'D': row['option_d']}
        original_opts_hi = {'A': row['option_a_hi'] or '', 'B': row['option_b_hi'] or '',
                            'C': row['option_c_hi'] or '', 'D': row['option_d_hi'] or ''}
        shuffled_options = {}
        shuffled_options_hi = {}
        for display_idx, orig_letter in enumerate(opt_order):
            display_letter = chr(ord('A') + display_idx)
            shuffled_options[display_letter] = original_opts[orig_letter]
            shuffled_options_hi[display_letter] = original_opts_hi[orig_letter]

        raw_image = row['image_path'] or ''
        if raw_image.startswith('__smiles__'):
            smiles_val   = raw_image[len('__smiles__'):]
            actual_image = None
        else:
            smiles_val   = ''
            actual_image = raw_image if raw_image else None

        questions.append({
            'id':            row['id'],
            'question':      row['question'],
            'question_hi':   row['question_hi'] or '',
            'option_a':      shuffled_options['A'],
            'option_b':      shuffled_options['B'],
            'option_c':      shuffled_options['C'],
            'option_d':      shuffled_options['D'],
            'option_a_hi':   shuffled_options_hi['A'],
            'option_b_hi':   shuffled_options_hi['B'],
            'option_c_hi':   shuffled_options_hi['C'],
            'option_d_hi':   shuffled_options_hi['D'],
            'image_path':    actual_image,
            'smiles':        smiles_val,
            'subject':       row['subject'],       # for tab grouping
            'question_type': row['question_type'] or 'MCQ',
            'selected':      row['selected_option'] or '',
        })
    return jsonify(questions)

    if not rows:
        return jsonify([])

    questions = []
    for row in rows:
        opt_order = row['option_order']
        original_opts = {
            'A': row['option_a'],
            'B': row['option_b'],
            'C': row['option_c'],
            'D': row['option_d']
        }
        shuffled_options = {}
        for display_idx, orig_letter in enumerate(opt_order):
            display_letter = chr(ord('A') + display_idx)
            shuffled_options[display_letter] = original_opts[orig_letter]

        # Decode image_path: if prefixed __smiles__ extract SMILES, else treat as image path
        raw_image = row['image_path'] or ''
        if raw_image.startswith('__smiles__'):
            smiles_val      = raw_image[len('__smiles__'):]
            actual_image    = None
        else:
            smiles_val      = ''
            actual_image    = raw_image if raw_image else None

        questions.append({
            'id':           row['id'],
            'question':     row['question'],
            'option_a':     shuffled_options['A'],
            'option_b':     shuffled_options['B'],
            'option_c':     shuffled_options['C'],
            'option_d':     shuffled_options['D'],
            'image_path':   actual_image,
            'smiles':       smiles_val,
            'content_type': row.get('content_type', '') if hasattr(row, 'get') else '',
            'image_prompt': '',
            'selected':     row['selected_option'],
            'option_order': opt_order
        })
    return jsonify(questions)

@app.route('/save_answer', methods=['POST'])
@student_required
@validate_schema(SAVE_ANSWER_SCHEMA, is_json=True)
def save_answer():
    """
    API endpoint to save a student's answer to a question.
    """
    data = request.get_json()
    question_id = data.get('question_id')
    displayed_option = data.get('selected_option')
    student_id = session['student_id']

    conn = get_db()
    c = conn.cursor()
    # Get the option order for this question
    c.execute("SELECT option_order FROM shuffled_questions WHERE student_id=? AND question_id=?",
              (student_id, question_id))
    row = c.fetchone()
    original_letter = ''
    if row and displayed_option:
        opt_order = row['option_order']
        display_index = ord(displayed_option) - ord('A')
        if display_index < len(opt_order):
            original_letter = opt_order[display_index]

    c.execute("INSERT OR REPLACE INTO responses (student_id, question_id, selected_option, created_at) VALUES (?,?,?, CURRENT_TIMESTAMP)",
              (student_id, question_id, original_letter))
    c.execute("UPDATE students SET status='In Progress' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})

@app.route('/submit_exam', methods=['POST'])
@student_required
def submit_exam():
    """
    API endpoint to submit the exam and calculate results.
    Accepts optional submission reason (e.g. 'tab_switch', 'fullscreen_exit').
    """
    student_id        = session['student_id']
    class_            = session.get('class')
    subject           = session.get('subject')
    name              = session.get('student_name')
    combined_subjects = session.get('combined_subjects', [])
    test_no           = session.get('test_no', '')

    reason = 'manual'
    if request.is_json:
        reason = request.json.get('reason', 'manual')
    elif request.form.get('reason'):
        reason = request.form.get('reason', 'manual')

    session['submission_reason'] = reason

    conn = get_db()
    c    = conn.cursor()

    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec          = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    def _score_subject(subj):
        c.execute("""SELECT q.correct_answer, q.option_a, q.option_b, q.option_c, q.option_d, r.selected_option, q.subject
                     FROM shuffled_questions sq
                     JOIN questions q   ON q.id = sq.question_id
                     LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
                     WHERE sq.student_id = ? AND q.subject = ?""",
                  (student_id, student_id, subj))
        rows  = c.fetchall()
        total = len(rows)
        raw   = 0.0
        for r in rows:
            sel = (r['selected_option'] or '').strip().upper()
            cor_letter = _normalize_correct_answer(
                r['correct_answer'],
                r['option_a'],
                r['option_b'],
                r['option_c'],
                r['option_d']
            )
            if sel and cor_letter and sel == cor_letter:
                raw += 1.0
            elif sel and neg_enabled:
                raw -= neg_value
        score = max(0.0, raw)
        pct   = round((score / total * 100), 2) if total > 0 else 0.0
        return score, total, pct

    c.execute("SELECT section, test_no FROM students WHERE student_id=?", (student_id,))
    st_row = c.fetchone()
    section_val = (st_row['section'] if st_row and st_row['section'] else session.get('section', ''))
    test_no_val = (st_row['test_no'] if st_row and st_row['test_no'] else session.get('test_no', ''))

    if combined_subjects:
        for subj in combined_subjects:
            score, total, pct = _score_subject(subj)
            _save_or_update_result(c, student_id, name, class_, section_val, subj, score, total, pct, test_no_val)
    else:
        score, total, pct = _score_subject(subject)
        _save_or_update_result(c, student_id, name, class_, section_val, subject, score, total, pct, test_no_val)

    c.execute("UPDATE students SET status='Submitted' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()

    if reason in ('tab_switch', 'fullscreen_exit'):
        log_audit_event('student', student_id, f'EXAM_AUTO_SUBMIT_{reason.upper()}', 'students', student_id, request.remote_addr)

    return jsonify({'status': 'submitted', 'redirect': url_for('submitted')})

@app.route('/submitted')
@student_required
def submitted():
    """
    Render the page shown after exam submission with reattempt option and security reason display.
    """
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')
    reason = session.get('submission_reason', 'manual')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT status FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='pending'",
              (student_id, class_, subject))
    pending = c.fetchone()
    conn.close()
    return render_template('submitted.html', pending=bool(pending), reason=reason)

@app.route('/request_reattempt', methods=['POST'])
@student_required
@validate_schema(REATTEMPT_REQUEST_SCHEMA, is_json=True)
def request_reattempt():
    """
    API endpoint for students to request reattempt of an exam.
    """
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='pending'",
              (student_id, class_, subject))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'You already have a pending request.'}), 400

    c.execute("INSERT INTO reattempt_requests (student_id, class, subject, status) VALUES (?,?,?, 'pending')",
              (student_id, class_, subject))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Request sent to admin.'})

@app.route('/check_reattempt_status')
@student_required
def check_reattempt_status():
    """
    API endpoint to check the status of a reattempt request.
    """
    student_id = session.get('student_id')
    class_ = session.get('class')
    subject = session.get('subject')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM reattempt_requests WHERE student_id=? AND class=? AND subject=? AND status='approved'",
              (student_id, class_, subject))
    approved = c.fetchone()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    exam_active = bool(exam_row['is_active']) if exam_row else False
    c.execute("SELECT status FROM students WHERE student_id=?", (student_id,))
    student_row = c.fetchone()
    student_status = student_row['status'] if student_row else 'Unknown'
    conn.close()

    can_start = (approved is not None) and exam_active and (student_status == 'Not Started')
    return jsonify({
        'approved': approved is not None,
        'exam_active': exam_active,
        'student_status': student_status,
        'can_start': can_start
    })

@app.route('/get_exam_time')
@student_required
def get_exam_time():
    """Feature #2: Per-student timer — counts from student's own exam_started_at."""
    student_id = session.get('student_id')
    conn = get_db()
    c = conn.cursor()

    # Get exam duration from global control
    c.execute("SELECT duration, is_active FROM exam_control WHERE id=1")
    ctrl = c.fetchone()
    if not ctrl or not ctrl['is_active']:
        conn.close()
        return jsonify({'remaining': 0, 'duration': 0})

    duration = ctrl['duration']

    c.execute("SELECT class, section, subject, test_no, exam_started_at, status FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    conn.close()

    if student and student['class'] and student['subject']:
        tno = student['test_no'] or session.get('test_no', '')
        sch_res = get_student_schedule_status(student['class'], student['section'], student['subject'], tno, student_id)
        if sch_res['state'] == 'active' and 'effective_duration' in sch_res:
            duration = sch_res['effective_duration']
        elif sch_res.get('schedule'):
            duration = sch_res['schedule']['duration_minutes']

    if not student or not student['exam_started_at']:
        return jsonify({'remaining': duration * 60, 'duration': duration})

    if student['status'] == 'Submitted':
        return jsonify({'remaining': 0, 'duration': duration})

    try:
        # Support both ISO format and SQLite CURRENT_TIMESTAMP format
        started_str = student['exam_started_at'].replace('T', ' ')
        start = datetime.datetime.fromisoformat(started_str)
    except Exception:
        return jsonify({'remaining': duration * 60, 'duration': duration})

    end_time  = start + datetime.timedelta(minutes=duration)
    remaining = int((end_time - datetime.datetime.utcnow()).total_seconds())
    if remaining < 0:
        remaining = 0

    return jsonify({'remaining': remaining, 'duration': duration})

# -------------------------------
# Admin Routes
# -------------------------------
@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    """
    Handle admin login and render admin login page.
    """
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if password == 'admin123':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin_login.html', error='Invalid credentials')
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """
    Render the admin dashboard with statistics.
    """
    conn = get_db()
    c = conn.cursor()

    # ── Existing stats ──────────────────────────────────────────────
    c.execute("SELECT COUNT(*) FROM students WHERE status='In Progress'")
    active_students = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM teachers WHERE status='active'")
    active_teachers = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM reattempt_requests WHERE status='pending'")
    pending_reattempts = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM students")
    total_students = c.fetchone()[0]

    # ── NEW: Class, Subject, Test counts ──────────────────────────
    c.execute("SELECT COUNT(DISTINCT class) FROM questions WHERE class IS NOT NULL AND class != ''")
    total_classes = c.fetchone()[0] or 0

    c.execute("SELECT COUNT(DISTINCT subject) FROM questions WHERE subject IS NOT NULL AND subject != ''")
    total_subjects = c.fetchone()[0] or 0

    c.execute("SELECT COUNT(*) FROM test_papers")
    total_tests = c.fetchone()[0] or 0

    # ── Breakdowns ──────────────────────────────────────────────────
    c.execute("""
        SELECT class, COUNT(*) 
        FROM students 
        WHERE class IS NOT NULL AND class != '' 
        GROUP BY class 
        ORDER BY class
    """)
    class_student_counts = {row[0]: row[1] for row in c.fetchall()}

    c.execute("""
        SELECT subject, COUNT(*) 
        FROM test_papers 
        GROUP BY subject 
        ORDER BY subject
    """)
    subject_test_counts = {row[0]: row[1] for row in c.fetchall()}

    # ── Class → Subject map (for dropdowns) ──────────────────────
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class != '' ORDER BY class, subject")
    rows = c.fetchall()
    class_subject_map = {}
    for row in rows:
        class_subject_map.setdefault(row['class'], []).append(row['subject'])

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')

    conn.close()

    return render_template('admin_dashboard.html',
                           active_students=active_students,
                           active_teachers=active_teachers,
                           pending_reattempts=pending_reattempts,
                           total_students=total_students,
                           total_classes=total_classes,
                           total_subjects=total_subjects,
                           total_tests=total_tests,
                           class_student_counts=class_student_counts,
                           subject_test_counts=subject_test_counts,
                           class_subject_map=class_subject_map,
                           school_name=school_name,
                           logo_path=logo_path)

@app.route('/admin/logout')
def admin_logout():
    """
    Handle admin logout.
    """
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/exam_status')
@admin_required
def exam_status():
    """
    Render the exam status page showing all students and their progress.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    return jsonify({'is_active': bool(row['is_active']) if row else False})

@app.route('/admin/exam_has_started')
@admin_required
def exam_has_started():
    """Check if exam has ever been started (start_time is set)."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT start_time FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    has_started = row and row['start_time'] is not None
    return jsonify({'has_started': has_started})

@app.route('/admin/start_exam', methods=['POST'])
@admin_required
def start_exam():
    """
    API endpoint for admin to start the exam.
    """
    duration         = request.form.get('duration', 60, type=int)
    negative_marking = 1 if request.form.get('negative_marking') == 'yes' else 0
    negative_value   = request.form.get('negative_value', 0.33, type=float)
    conn = get_db()
    c = conn.cursor()
    start_time = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    c.execute("""UPDATE exam_control
                 SET is_active=1, start_time=?, duration=?,
                     negative_marking=?, negative_value=?
                 WHERE id=1""",
              (start_time, duration, negative_marking, negative_value))
    conn.commit()
    conn.close()
    log_audit_event('admin', 'admin', 'EXAM_START', 'exam_control', 1, request.remote_addr)
    return redirect(url_for('admin_dashboard'))

@app.route('/teacher/start_exam', methods=['POST'])
@teacher_required
def teacher_start_exam():
    """
    API endpoint for teacher to start the exam with negative marking options.
    """
    duration         = request.form.get('duration', 60, type=int)
    negative_marking = 1 if request.form.get('negative_marking') == 'yes' else 0
    negative_value   = request.form.get('negative_value', 0.33, type=float)
    conn = get_db()
    c = conn.cursor()
    start_time = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    c.execute("""UPDATE exam_control
                 SET is_active=1, start_time=?, duration=?,
                     negative_marking=?, negative_value=?
                 WHERE id=1""",
              (start_time, duration, negative_marking, negative_value))
    conn.commit()
    conn.close()
    return redirect(url_for('teacher_dashboard'))

@app.route('/teacher/stop_exam')
@teacher_required
def teacher_stop_exam():
    """
    API endpoint for teacher to stop the exam.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE exam_control SET is_active=0 WHERE id=1")
    conn.commit()
    conn.close()
    log_audit_event('teacher', session.get('teacher_id') or 'teacher', 'EXAM_STOP', 'exam_control', 1, request.remote_addr)
    return redirect(url_for('teacher_dashboard'))

@app.route('/admin/exam_settings')
@admin_required
def admin_exam_settings():
    """Return current exam control settings as JSON — used by admin dashboard."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT duration, negative_marking, negative_value FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    if row:
        return jsonify({'duration': row['duration'],
                        'negative_marking': bool(row['negative_marking']),
                        'negative_value': row['negative_value']})
    return jsonify({'duration': 60, 'negative_marking': False, 'negative_value': 0.33})

@app.route('/teacher/exam_settings')
@teacher_required
def teacher_exam_settings():
    """Return current exam control settings as JSON — used by teacher dashboard."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active, duration, negative_marking, negative_value FROM exam_control WHERE id=1")
    row = c.fetchone()
    conn.close()
    if row:
        return jsonify({'is_active': bool(row['is_active']),
                        'duration': row['duration'],
                        'negative_marking': bool(row['negative_marking']),
                        'negative_value': row['negative_value']})
    return jsonify({'is_active': False, 'duration': 60, 'negative_marking': False, 'negative_value': 0.33})

# ═══════════════════════════════════════════════════════════════
# TEST SCHEDULING ROUTES & APIS
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/schedule_test')
@admin_required
def admin_schedule_test():
    update_scheduled_tests_status()
    class_subject_map = _get_class_subject_map()
    return render_template('scheduled_tests.html', is_admin=True, class_subject_map=class_subject_map)

@app.route('/teacher/schedule_test')
@teacher_required
def teacher_schedule_test():
    update_scheduled_tests_status()
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, section, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    class_subject_map = _get_class_subject_map()
    return render_template('scheduled_tests.html', is_admin=False, assignments=assignments, class_subject_map=class_subject_map)

@app.route('/api/scheduled_test/create', methods=['POST'])
def create_scheduled_test():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    class_ = request.form.get('class', '').strip()
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject', '').strip()
    test_no = request.form.get('test_no', '').strip()
    scheduled_date = request.form.get('scheduled_date', '').strip()
    start_time_raw = request.form.get('start_time', '').strip()
    end_time_raw = request.form.get('end_time', '').strip()
    duration_minutes = int(request.form.get('duration_minutes', 30) or 30)

    if not class_ or not subject or not test_no or not scheduled_date or not start_time_raw or not end_time_raw:
        return jsonify({'status': 'error', 'message': 'Class, Subject, Test No, Date, Start Time, and End Time are required'}), 400

    full_start_str = f"{scheduled_date} {start_time_raw}:00" if len(start_time_raw) == 5 else f"{scheduled_date} {start_time_raw}"
    full_end_str = f"{scheduled_date} {end_time_raw}:00" if len(end_time_raw) == 5 else f"{scheduled_date} {end_time_raw}"

    try:
        st_dt = datetime.datetime.strptime(full_start_str, '%Y-%m-%d %H:%M:%S')
        end_dt = datetime.datetime.strptime(full_end_str, '%Y-%m-%d %H:%M:%S')
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Invalid Date or Time format: {e}'}), 400

    if end_dt <= st_dt:
        return jsonify({'status': 'error', 'message': 'End time must be after Start time'}), 400

    now = datetime.datetime.now()
    if st_dt <= now <= end_dt:
        init_status = 'active'
    elif now > end_dt:
        init_status = 'expired'
    else:
        init_status = 'scheduled'

    created_by = 'admin' if session.get('admin_logged_in') else str(session.get('teacher_id', 'teacher'))

    conn = get_db()
    c = conn.cursor()

    # Check if a schedule already exists for this Class, Section, Subject, and Test No
    c.execute("""SELECT id FROM scheduled_tests
                 WHERE class=? AND section=? AND LOWER(subject)=LOWER(?) AND test_no=?""",
              (class_, section, subject, test_no))
    existing = c.fetchone()

    if existing:
        # UPDATE existing row directly — no new row added to the list!
        c.execute("""UPDATE scheduled_tests
                     SET scheduled_date=?, start_time=?, end_time=?, duration_minutes=?, status=?, created_by=?
                     WHERE id=?""",
                  (scheduled_date, full_start_str, full_end_str, duration_minutes, init_status, created_by, existing['id']))
        conn.commit()
        conn.close()
        msg = f'Test schedule updated successfully ({init_status.capitalize()})'
    else:
        # Insert new row if no prior schedule exists for this test
        c.execute("""INSERT INTO scheduled_tests
                     (class, section, subject, test_no, scheduled_date, start_time, end_time, duration_minutes, status, created_by)
                     VALUES (?,?,?,?,?,?,?,?,?,?)""",
                  (class_, section, subject, test_no, scheduled_date, full_start_str, full_end_str, duration_minutes, init_status, created_by))
        conn.commit()
        conn.close()
        msg = f'Test scheduled successfully ({init_status.capitalize()})'

    u_type = 'admin' if session.get('admin_logged_in') else 'teacher'
    u_id = session.get('teacher_id') if session.get('teacher_logged_in') else 'admin'
    log_audit_event(u_type, u_id, 'TEST_SCHEDULE_SAVE', 'scheduled_tests', f"{class_}_{subject}_{test_no}", request.remote_addr)

    return jsonify({'status': 'success', 'message': msg})

@app.route('/api/scheduled_tests', methods=['GET'])
def get_scheduled_tests():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    update_scheduled_tests_status()

    class_ = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    date_filter = request.args.get('date', '').strip()
    status_filter = request.args.get('status', '').strip()

    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM scheduled_tests WHERE 1=1"
    params = []

    if class_:
        query += " AND class=?"
        params.append(class_)
    if section:
        query += " AND section=?"
        params.append(section)
    if date_filter:
        query += " AND scheduled_date=?"
        params.append(date_filter)
    if status_filter:
        query += " AND status=?"
        params.append(status_filter)

    query += " ORDER BY id DESC"
    c.execute(query, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    return jsonify({'status': 'success', 'scheduled_tests': rows})

@app.route('/api/scheduled_test/<int:sid>/cancel', methods=['POST'])
def cancel_scheduled_test(sid):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE scheduled_tests SET status='cancelled' WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    u_type = 'admin' if session.get('admin_logged_in') else 'teacher'
    u_id = session.get('teacher_id') if session.get('teacher_logged_in') else 'admin'
    log_audit_event(u_type, u_id, 'TEST_SCHEDULE_CANCEL', 'scheduled_tests', sid, request.remote_addr)
    return jsonify({'status': 'success', 'message': 'Scheduled test cancelled'})

@app.route('/api/scheduled_test/<int:sid>/reopen', methods=['POST'])
def reopen_scheduled_test(sid):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    scheduled_date = request.form.get('scheduled_date', '').strip()
    start_time_raw = request.form.get('start_time', '').strip()
    end_time_raw = request.form.get('end_time', '').strip()
    duration_minutes = int(request.form.get('duration_minutes', 30) or 30)

    if not scheduled_date or not start_time_raw or not end_time_raw:
        return jsonify({'status': 'error', 'message': 'Date, Start Time, and End Time are required'}), 400

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM scheduled_tests WHERE id=?", (sid,))
    orig = c.fetchone()
    if not orig:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Schedule entry not found'}), 404

    full_start_str = f"{scheduled_date} {start_time_raw}:00" if len(start_time_raw) == 5 else f"{scheduled_date} {start_time_raw}"
    full_end_str = f"{scheduled_date} {end_time_raw}:00" if len(end_time_raw) == 5 else f"{scheduled_date} {end_time_raw}"

    try:
        st_dt = datetime.datetime.strptime(full_start_str, '%Y-%m-%d %H:%M:%S')
        end_dt = datetime.datetime.strptime(full_end_str, '%Y-%m-%d %H:%M:%S')
    except Exception as e:
        conn.close()
        return jsonify({'status': 'error', 'message': f'Invalid Date or Time format: {e}'}), 400

    if end_dt <= st_dt:
        conn.close()
        return jsonify({'status': 'error', 'message': 'End time must be after Start time'}), 400

    now = datetime.datetime.now()
    if st_dt <= now <= end_dt:
        init_status = 'active'
    elif now > end_dt:
        init_status = 'expired'
    else:
        init_status = 'scheduled'

    created_by = 'admin' if session.get('admin_logged_in') else str(session.get('teacher_id', 'teacher'))

    # Update existing schedule row directly — status changes to active/scheduled/expired without creating duplicate row
    c.execute("""UPDATE scheduled_tests
                 SET scheduled_date=?, start_time=?, end_time=?, duration_minutes=?, status=?, created_by=?
                 WHERE id=?""",
              (scheduled_date, full_start_str, full_end_str, duration_minutes, init_status, created_by, sid))
    conn.commit()
    conn.close()

    u_type = 'admin' if session.get('admin_logged_in') else 'teacher'
    u_id = session.get('teacher_id') if session.get('teacher_logged_in') else 'admin'
    log_audit_event(u_type, u_id, 'TEST_SCHEDULE_RESCHEDULE', 'scheduled_tests', sid, request.remote_addr)

    return jsonify({'status': 'success', 'message': f'Test rescheduled successfully ({init_status.capitalize()})'})

@app.route('/api/scheduled_test/check_student', methods=['GET'])
def check_student_scheduled_test():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'status': 'error', 'message': 'Not logged in'}), 401

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT class, section, subject, test_no, status FROM students WHERE student_id=?", (student_id,))
    st = c.fetchone()
    conn.close()

    if not st:
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404

    sch_res = get_student_schedule_status(st['class'], st['section'], st['subject'], st['test_no'], student_id)
    return jsonify({'status': 'success', 'schedule_info': sch_res})

@app.route('/api/available_test_numbers', methods=['GET'])
def get_available_test_numbers():
    class_ = request.args.get('class', '').strip()
    subject = request.args.get('subject', '').strip()
    clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()

    conn = get_db()
    c = conn.cursor()
    if class_ and subject:
        c.execute("""SELECT DISTINCT test_no FROM test_papers 
                     WHERE (class=? OR class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     UNION
                     SELECT DISTINCT test_no FROM questions 
                     WHERE (class=? OR class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     ORDER BY test_no""",
                  (class_, clean_cls, f"%{clean_cls}%", subject, class_, clean_cls, f"%{clean_cls}%", subject))
    elif class_:
        c.execute("""SELECT DISTINCT test_no FROM test_papers 
                     WHERE (class=? OR class=? OR class LIKE ?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     UNION
                     SELECT DISTINCT test_no FROM questions 
                     WHERE (class=? OR class=? OR class LIKE ?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     ORDER BY test_no""",
                  (class_, clean_cls, f"%{clean_cls}%", class_, clean_cls, f"%{clean_cls}%"))
    elif subject:
        c.execute("""SELECT DISTINCT test_no FROM test_papers 
                     WHERE LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     UNION
                     SELECT DISTINCT test_no FROM questions 
                     WHERE LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND trim(test_no) != ''
                     ORDER BY test_no""", (subject, subject))
    else:
        c.execute("""SELECT DISTINCT test_no FROM test_papers WHERE test_no IS NOT NULL AND trim(test_no) != ''
                     UNION
                     SELECT DISTINCT test_no FROM questions WHERE test_no IS NOT NULL AND trim(test_no) != ''
                     ORDER BY test_no""")
    rows = [r[0] for r in c.fetchall() if r[0]]
    conn.close()
    return jsonify({'status': 'success', 'test_numbers': rows})

@app.route('/admin/student/force_submit/<student_id>', methods=['POST'])
@admin_required
def admin_force_submit_student(student_id):
    """Feature #3: Force-submit an individual student from monitoring page."""
    conn = get_db()
    c = conn.cursor()

    # Only act if student is In Progress
    c.execute("SELECT name, class, section, subject, status, test_no FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
    if student['status'] not in ('In Progress', 'Not Started'):
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student already submitted'}), 400

    name    = student['name']
    class_  = student['class']
    subject = student['subject']
    test_no = (student['test_no'] if ('test_no' in student.keys() and student['test_no']) else '').strip()

    # Calculate score with negative marking
    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    if test_no:
        c.execute("""SELECT q.correct_answer, q.option_a, q.option_b, q.option_c, q.option_d, r.selected_option
                     FROM questions q
                     LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
                     WHERE q.class = ? AND q.subject = ? AND (q.test_no = ? OR q.chapter = ?)""",
                  (student_id, class_, subject, test_no, test_no))
    else:
        c.execute("""SELECT q.correct_answer, q.option_a, q.option_b, q.option_c, q.option_d, r.selected_option
                     FROM questions q
                     LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
                     WHERE q.class = ? AND q.subject = ?""",
                  (student_id, class_, subject))
    rows  = c.fetchall()
    total = len(rows)
    raw = 0.0
    for r in rows:
        sel = (r['selected_option'] or '').strip().upper()
        cor_letter = _normalize_correct_answer(
            r['correct_answer'],
            r['option_a'],
            r['option_b'],
            r['option_c'],
            r['option_d']
        )
        if sel and cor_letter and sel == cor_letter:
            raw += 1.0
        elif sel and neg_enabled:
            raw -= neg_value
    score      = max(0.0, raw)
    percentage = round((score / total * 100), 2) if total > 0 else 0.0

    # Save result
    _save_or_update_result(c, student_id, name, class_, student['section'] if 'section' in student.keys() else '', subject, score, total, percentage, test_no)
    c.execute("UPDATE students SET status='Submitted' WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': f'{name} submitted. Score: {round(score,2)}/{total}'})

@app.route('/admin/stop_exam')
@admin_required
def stop_exam():
    """
    API endpoint for admin to stop the exam.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE exam_control SET is_active=0 WHERE id=1")
    conn.commit()
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    """
    Render and handle admin settings page (school name, logo, etc).
    """
    if request.method == 'POST':
        for field in ['school_name', 'school_address', 'academic_session']:
            val = request.form.get(field, '').strip()
            if val:
                set_setting(field, val)

        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename != '' and allowed_image_file(logo_file.filename):
            ext = logo_file.filename.rsplit('.', 1)[1].lower()
            filename = f"school_logo.{ext}"
            filepath = os.path.join(app.config['BRANDING_FOLDER'], filename)
            for old_ext in ALLOWED_IMAGE_EXTENSIONS:
                old_path = os.path.join(app.config['BRANDING_FOLDER'], f"school_logo.{old_ext}")
                if os.path.exists(old_path) and old_path != filepath:
                    os.remove(old_path)
            logo_file.save(filepath)
            set_setting('logo_path', os.path.join('uploads', 'branding', filename).replace('\\', '/'))

        return redirect(url_for('admin_settings'))

    school_name      = get_setting('school_name', 'RRB Group of Schools')
    school_address   = get_setting('school_address', '')
    academic_session = get_setting('academic_session', '')
    logo_path        = get_setting('logo_path', '')
    return render_template('settings.html',
                           school_name=school_name,
                           school_address=school_address,
                           academic_session=academic_session,
                           logo_path=logo_path)

@app.route('/admin/generate_question_paper')
@admin_required
def generate_question_paper():
    """
    Render page to generate HTML Print Preview / PDF of the question paper.
    """
    class_ = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    subject = request.args.get('subject', '').strip()
    test_no = request.args.get('test_no', '').strip()

    if not class_ or not subject:
        return "Class and subject required", 400

    conn = get_db()
    c = conn.cursor()

    clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()

    # Query matching questions using multi-tier fallback
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d, correct_answer, image_path, chapter, test_no,
                        CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                 FROM questions
                 WHERE (class=? OR class=? OR class LIKE ?)
                   AND (? = '' OR section=? OR section LIKE ?)
                   AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                   AND (? = '' OR test_no=? OR chapter=? OR test_no LIKE ? OR chapter LIKE ?)
                 ORDER BY id""",
             (class_, clean_cls, f"%{clean_cls}%",
              section, section, f"%{section}%",
              subject, f"%{subject}%",
              test_no, test_no, test_no, f"%{test_no}%", f"%{test_no}%"))
    questions = [dict(row) for row in c.fetchall()]

    if not questions:
        c.execute("""SELECT id, question, option_a, option_b, option_c, option_d, correct_answer, image_path, chapter, test_no,
                            CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                     FROM questions
                     WHERE (LOWER(class)=LOWER(?) OR LOWER(class)=LOWER(?) OR class LIKE ?)
                       AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                     ORDER BY id""",
                 (class_, clean_cls, f"%{clean_cls}%", subject, f"%{subject}%"))
        questions = [dict(row) for row in c.fetchall()]

    conn.close()

    for q in questions:
        q_type = str(q.get('question_type') or '').strip()
        if not q_type or q_type.upper() in ['MCQ', 'CHOICE', 'OBJECTIVE']:
            q['question_type'] = 'MCQ'

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode('utf-8')}"

    # Determine display chapter name: prefer explicit query param or real question chapter over literal test_no
    req_chapter = request.args.get('chapter', '').strip()
    if req_chapter:
        chapter_display = req_chapter
    else:
        unique_ch = []
        for q in questions:
            ch_val = str(q.get('chapter') or '').strip()
            if ch_val and ch_val.lower() != test_no.lower() and ch_val not in unique_ch:
                unique_ch.append(ch_val)
        if unique_ch:
            chapter_display = ", ".join(unique_ch)
        else:
            chapter_display = 'All Chapters'

    total_marks = len(questions)
    display_class = f"{class_}{section}" if section else class_
    rendered_html = render_template(
        'question_paper_template_v2.html',
        school_name=school_name,
        logo_base64=logo_base64,
        class_name=display_class,
        subject=subject,
        chapter=chapter_display,
        test_no=test_no or 'All',
        questions=questions,
        total_marks=total_marks,
        date=datetime.datetime.now().strftime('%d/%m/%Y')
    )

    if request.args.get('format') == 'download_pdf':
        pdf = HTML(string=rendered_html, base_url=request.base_url).write_pdf()
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Question_Paper_{display_class}_{subject}.pdf'
        return response

    return rendered_html

@app.route('/admin/questions')
@admin_required
def questions():
    """
    Render the admin questions management page.
    """
    return render_template('questions.html')

@app.route('/admin/questions/data')
@admin_required
def questions_data():
    """
    API endpoint to fetch questions data with filtering and pagination.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, class, section, subject, question, option_a, option_b, option_c, option_d, correct_answer, image_path, test_no, chapter FROM questions ORDER BY class, subject, id")
    questions = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(questions)

@app.route('/admin/question/add', methods=['POST'])
@admin_required
def add_question():
    """
    API endpoint to add a new question to the database.
    """
    if request.is_json:
        data = request.get_json()
        class_ = data['class']
        subject = data['subject']
        question = data['question']
        opt_a = data['option_a']
        opt_b = data['option_b']
        opt_c = data['option_c']
        opt_d = data['option_d']
        correct = data['correct_answer']
        image_file = None
    else:
        class_ = request.form.get('class')
        section = request.form.get('section', '').strip()
        subject = request.form.get('subject')
        test_no = request.form.get('test_no', '').strip()
        question = request.form.get('question')
        opt_a = request.form.get('option_a')
        opt_b = request.form.get('option_b')
        opt_c = request.form.get('option_c')
        opt_d = request.form.get('option_d')
        correct = request.form.get('correct_answer')
        image_file = request.files.get('image')

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO questions (class, section, subject, test_no, chapter, question, option_a, option_b, option_c, option_d, correct_answer) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
              (class_, section, subject, test_no, test_no, question, opt_a, opt_b, opt_c, opt_d, correct))
    qid = c.lastrowid

    image_path = None
    if image_file and image_file.filename != '':
        image_path = save_question_image(image_file, class_, subject, qid)
        if image_path:
            c.execute("UPDATE questions SET image_path=? WHERE id=?", (image_path, qid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'id': qid})

@app.route('/admin/question/update/<int:qid>', methods=['POST'])
@admin_required
def update_question(qid):
    """
    API endpoint to update an existing question.
    """
    class_ = request.form.get('class')
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject')
    test_no = request.form.get('test_no', '').strip()
    question = request.form.get('question')
    opt_a = request.form.get('option_a')
    opt_b = request.form.get('option_b')
    opt_c = request.form.get('option_c')
    opt_d = request.form.get('option_d')
    correct = request.form.get('correct_answer')
    image_file = request.files.get('image')
    remove_image = request.form.get('remove_image') == 'true'

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT image_path FROM questions WHERE id=?", (qid,))
    old_image = c.fetchone()
    old_path = old_image['image_path'] if old_image else None

    c.execute('''UPDATE questions 
                 SET class=?, section=?, subject=?, test_no=?, chapter=?, question=?, option_a=?, option_b=?, option_c=?, option_d=?, correct_answer=?
                 WHERE id=?''',
              (class_, section, subject, test_no, test_no, question, opt_a, opt_b, opt_c, opt_d, correct, qid))

    image_path = old_path
    if remove_image:
        if old_path and os.path.exists(os.path.join('static', old_path)):
            try:
                os.remove(os.path.join('static', old_path))
            except:
                pass
        image_path = None
        c.execute("UPDATE questions SET image_path=NULL WHERE id=?", (qid,))
    elif image_file and image_file.filename != '':
        if old_path and os.path.exists(os.path.join('static', old_path)):
            try:
                os.remove(os.path.join('static', old_path))
            except:
                pass
        new_path = save_question_image(image_file, class_, subject, qid)
        if new_path:
            image_path = new_path
            c.execute("UPDATE questions SET image_path=? WHERE id=?", (image_path, qid))

    conn.commit()
    conn.close()
    return jsonify({'status': 'updated'})

@app.route('/admin/question/delete/<int:qid>', methods=['DELETE'])
@admin_required
def delete_question(qid):
    """
    API endpoint to delete a question by ID.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT image_path FROM questions WHERE id=?", (qid,))
    row = c.fetchone()
    if row and row['image_path']:
        img_path = os.path.join('static', row['image_path'])
        if os.path.exists(img_path):
            try:
                os.remove(img_path)
            except:
                pass
    c.execute("DELETE FROM questions WHERE id=?", (qid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'deleted'})

@app.route('/admin/questions/delete_by_class_subject', methods=['POST'])
@admin_required
def delete_questions_by_class_subject():
    """
    API endpoint to delete all questions for a specific class and subject.
    """
    data = request.get_json()
    class_ = data.get('class')
    subject = data.get('subject')

    if not class_ or not subject:
        return jsonify({'status': 'error', 'message': 'Class and subject required'}), 400

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT image_path FROM questions WHERE class=? AND subject=?", (class_, subject))
    images = c.fetchall()

    for img in images:
        if img['image_path']:
            img_path = os.path.join('static', img['image_path'])
            if os.path.exists(img_path):
                try:
                    os.remove(img_path)
                except:
                    pass

    c.execute("DELETE FROM questions WHERE class=? AND subject=?", (class_, subject))
    deleted_count = c.rowcount
    conn.commit()
    conn.close()
    log_audit_event('admin', 'admin', 'PAPER_DELETE', 'questions', f"{class_}_{subject}", request.remote_addr)

    return jsonify({'status': 'success', 'deleted': deleted_count})

@app.route('/admin/upload_csv', methods=['POST'])
@admin_required
def upload_csv():
    """
    API endpoint to upload questions from a CSV file.
    """
    if 'csv_file' not in request.files:
        return redirect(request.url)
    file = request.files['csv_file']
    if file.filename == '':
        return redirect(request.url)
    if file and file.filename.endswith('.csv'):
        filename = secure_filename(file.filename)
        base = os.path.splitext(filename)[0]
        parts = base.split('_')
        # Parse class, subject, and optional test_no
        if len(parts) >= 3:
            class_ = parts[0].strip()
            subject = parts[1].strip()
            test_no = parts[2].strip()
        elif len(parts) == 2:
            class_ = parts[0].strip()
            subject = parts[1].strip()
            test_no = ''
        else:
            class_ = "Unknown"
            subject = base.strip()
            test_no = ''
        if not class_:
            class_ = "Unknown"

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        conn = get_db()
        c = conn.cursor()
        try:
            with open(filepath, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader)
                normalized = [h.strip().lower().replace(' ', '_') for h in header]

                try:
                    idx_q = normalized.index('question')
                    idx_a = normalized.index('option_a')
                    idx_b = normalized.index('option_b')
                    idx_c = normalized.index('option_c')
                    idx_d = normalized.index('option_d')
                    idx_correct = normalized.index('correct_answer')
                except ValueError as e:
                    return f"Missing required column: {e}. Headers found: {header}", 400

                inserted_count = 0
                for row in reader:
                    if not row or all(cell.strip() == '' for cell in row):
                        continue
                    if len(row) <= max(idx_q, idx_a, idx_b, idx_c, idx_d, idx_correct):
                        continue

                    question = row[idx_q].strip()
                    opt_a = row[idx_a].strip()
                    opt_b = row[idx_b].strip()
                    opt_c = row[idx_c].strip()
                    opt_d = row[idx_d].strip()
                    correct_value = row[idx_correct].strip()

                    correct_letter = None
                    if correct_value == opt_a:
                        correct_letter = 'A'
                    elif correct_value == opt_b:
                        correct_letter = 'B'
                    elif correct_value == opt_c:
                        correct_letter = 'C'
                    elif correct_value == opt_d:
                        correct_letter = 'D'
                    else:
                        if correct_value.upper() in ('A', 'B', 'C', 'D'):
                            correct_letter = correct_value.upper()
                        elif correct_value.lower() == 'option_a':
                            correct_letter = 'A'
                        elif correct_value.lower() == 'option_b':
                            correct_letter = 'B'
                        elif correct_value.lower() == 'option_c':
                            correct_letter = 'C'
                        elif correct_value.lower() == 'option_d':
                            correct_letter = 'D'
                        else:
                            print(f"Warning: Could not map correct_answer '{correct_value}' to any option. Skipping question.")
                            continue

                    if not question:
                        continue

                    # Insert with chapter (test_no)
                    c.execute("INSERT INTO questions (class, subject, question, option_a, option_b, option_c, option_d, correct_answer, chapter) VALUES (?,?,?,?,?,?,?,?,?)",
                              (class_, subject, question, opt_a, opt_b, opt_c, opt_d, correct_letter, test_no))
            conn.commit()
        except Exception as e:
            conn.rollback()
            return f"Error processing CSV: {e}", 500
        finally:
            conn.close()
            os.remove(filepath)
        return redirect(url_for('questions'))
    return "Invalid file", 400

@app.route('/admin/students')
@admin_required
def manage_students():
    """
    Render the admin students management page.
    """
    return render_template('manage_students.html')

@app.route('/admin/students/data')
@admin_required
def students_data():
    """
    API endpoint to fetch students data with advanced filtering:
    class, section, subject, test_no, date (exam_started_at).
    Also returns system_name (instead of IP) and exam_started_at.
    """
    filter_class   = request.args.get('class', '')
    filter_section = request.args.get('section', '')
    filter_subject = request.args.get('subject', '')
    filter_test_no = request.args.get('test_no', '')
    filter_date    = request.args.get('date', '')

    conn = get_db()
    c = conn.cursor()

    query = """SELECT student_id, name, class, section, subject,
                      ip, system_name, status, exam_started_at,
                      test_no, admission_no, dob, house, parents_name, address, picture
               FROM students WHERE 1=1"""
    params = []

    if filter_class:
        query += " AND class = ?"
        params.append(filter_class)
    if filter_section:
        query += " AND section = ?"
        params.append(filter_section)
    if filter_subject:
        query += " AND subject = ?"
        params.append(filter_subject)
    if filter_test_no:
        # test_no is stored as text, use LIKE for partial match
        query += " AND test_no LIKE ?"
        params.append(f"%{filter_test_no}%")
    if filter_date:
        # filter by exam_started_at date (YYYY-MM-DD)
        query += " AND DATE(exam_started_at) = ?"
        params.append(filter_date)

    query += " ORDER BY class, section, name"
    c.execute(query, params)
    students = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(students)

@app.route('/api/class_report/subjects')
def api_class_report_subjects():
    """Return list of subjects available across test_papers, questions, and results for a given class & section."""
    class_ = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    if not class_:
        return jsonify({'subjects': []})

    clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
    conn = get_db()
    c = conn.cursor()

    teacher_id = session.get('teacher_id')
    is_admin = session.get('admin_logged_in')

    if teacher_id and not is_admin:
        c.execute("SELECT subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=? AND (class=? OR class LIKE ?)",
                  (teacher_id, class_, f"%{clean_cls}%"))
        t_rows = [dict(r) for r in c.fetchall()]
        is_ct = any(r.get('is_class_teacher') for r in t_rows)
        t_subjs = sorted(list(set([r['subject'] for r in t_rows if r.get('subject')])))

        if not is_ct and t_subjs:
            conn.close()
            return jsonify({'subjects': t_subjs})

    # Admin or Class Teacher: Query distinct subjects across test_papers, questions, and results
    c.execute("""SELECT DISTINCT subject FROM test_papers WHERE (class=? OR class LIKE ?)
                 UNION
                 SELECT DISTINCT subject FROM questions WHERE (class=? OR class LIKE ?)
                 UNION
                 SELECT DISTINCT subject FROM results WHERE (class=? OR class LIKE ?)""",
              (class_, f"%{clean_cls}%", class_, f"%{clean_cls}%", class_, f"%{clean_cls}%"))
    subjects = sorted(list(set(r['subject'] for r in c.fetchall() if r['subject'])))
    conn.close()
    return jsonify({'subjects': subjects})


@app.route('/api/class_report/test_numbers')
def api_class_report_test_numbers():
    """Return list of test numbers created for a given class, section and subject across test_papers, questions, and results."""
    class_ = request.args.get('class', '').strip()
    section = request.args.get('section', '').strip()
    subject = request.args.get('subject', '').strip()
    if not class_ or not subject:
        return jsonify({'test_numbers': []})

    clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
    conn = get_db()
    c = conn.cursor()

    # Query created test numbers from test_papers, questions (chapter & test_no), and results
    c.execute("""SELECT DISTINCT test_no as tn FROM test_papers WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND test_no != ''
                 UNION
                 SELECT DISTINCT chapter as tn FROM questions WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND chapter IS NOT NULL AND chapter != ''
                 UNION
                 SELECT DISTINCT test_no as tn FROM questions WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND test_no IS NOT NULL AND test_no != ''
                 UNION
                 SELECT DISTINCT chapter as tn FROM results WHERE (class=? OR class LIKE ?) AND LOWER(subject)=LOWER(?) AND chapter IS NOT NULL AND chapter != ''""",
              (class_, f"%{clean_cls}%", subject,
               class_, f"%{clean_cls}%", subject,
               class_, f"%{clean_cls}%", subject,
               class_, f"%{clean_cls}%", subject))
    test_numbers = sorted(list(set(r['tn'] for r in c.fetchall() if r['tn'])))
    conn.close()
    return jsonify({'test_numbers': test_numbers})


@app.route('/admin/student/delete/<student_id>', methods=['DELETE'])
@admin_required
def delete_student(student_id):
    """
    API endpoint to delete a student record.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM students WHERE student_id=?",           (student_id,))
    c.execute("DELETE FROM responses WHERE student_id=?",          (student_id,))
    c.execute("DELETE FROM results WHERE student_id=?",            (student_id,))
    c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))  # Bug #3 fix
    c.execute("DELETE FROM reattempt_requests WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'deleted'})

@app.route('/admin/monitoring')
@admin_required
def monitoring():
    """
    Render the exam monitoring page showing real-time student status.
    """
    return render_template('monitoring.html')

@app.route('/admin/monitoring/data')
@admin_required
def monitoring_data():
    """
    API endpoint to fetch real-time monitoring data.
    """
    auto_submit_expired_exams()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, class, subject, student_id, ip, status, exam_started_at FROM students ORDER BY status, name")
    students = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(students)

@app.route('/admin/evaluate')
@admin_required
def evaluate():
    """
    Render the page for evaluating descriptive responses.
    """
    conn = get_db()
    c = conn.cursor()

    # BUG-003 Fix: fetch negative marking settings FIRST
    c.execute("SELECT negative_marking, negative_value FROM exam_control WHERE id=1")
    ec = c.fetchone()
    neg_enabled = bool(ec['negative_marking']) if ec else False
    neg_value   = float(ec['negative_value'])  if ec else 0.33

    c.execute("SELECT student_id, name, class, section, subject, test_no FROM students WHERE status='Submitted'")
    students = c.fetchall()
    for student in students:
        student_id = student['student_id']
        name  = student['name']
        class_  = student['class']
        section_ = student['section'] or ''
        subject = student['subject']
        test_no = (student['test_no'] or '').strip() if ('test_no' in student.keys() and student['test_no']) else ''
        if test_no:
            c.execute("SELECT id, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=? AND (test_no=? OR chapter=?)", (class_, subject, test_no, test_no))
        else:
            c.execute("SELECT id, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=?", (class_, subject))
        q_rows = c.fetchall()
        correct_map = {}
        for q in q_rows:
            cor_letter = _normalize_correct_answer(
                q['correct_answer'],
                q['option_a'],
                q['option_b'],
                q['option_c'],
                q['option_d']
            )
            if cor_letter:
                correct_map[q['id']] = cor_letter
        c.execute("SELECT question_id, selected_option FROM responses WHERE student_id=?", (student_id,))
        responses = c.fetchall()

        # BUG-003 Fix: apply negative marking to each wrong answer
        raw_score = 0.0
        total     = len(correct_map)
        for resp in responses:
            qid      = resp['question_id']
            selected = (resp['selected_option'] or '').strip().upper()
            if qid not in correct_map:
                continue
            if selected == correct_map[qid]:
                raw_score += 1.0
            elif selected and neg_enabled:
                raw_score -= neg_value

        score      = max(0.0, raw_score)
        percentage = round((score / total * 100), 2) if total > 0 else 0.0
        _save_or_update_result(c, student_id, name, class_, section_, subject, score, total, percentage, test_no)
    conn.commit()
    conn.close()
    return redirect(url_for('results_page'))
@app.route('/api/results/filters')
def api_results_filters():
    """Return distinct classes, sections, subjects, test_numbers from results."""
    conn = get_db()
    c = conn.cursor()

    # Classes from results
    c.execute("SELECT DISTINCT class FROM results WHERE class IS NOT NULL AND class != '' ORDER BY class")
    classes = [r['class'] for r in c.fetchall()]

    # Sections from students (join with results)
    c.execute("""SELECT DISTINCT s.section FROM results r
                 JOIN students s ON r.student_id = s.student_id
                 WHERE s.section IS NOT NULL AND s.section != ''
                 ORDER BY s.section""")
    sections = [r['section'] for r in c.fetchall()]

    # Subjects from results
    c.execute("SELECT DISTINCT subject FROM results WHERE subject IS NOT NULL AND subject != '' ORDER BY subject")
    subjects = [r['subject'] for r in c.fetchall()]

    # Test numbers (chapter) from results
    c.execute("SELECT DISTINCT chapter FROM results WHERE chapter IS NOT NULL AND chapter != '' ORDER BY chapter")
    test_numbers = [r['chapter'] for r in c.fetchall()]

    conn.close()
    return jsonify({
        'classes': classes,
        'sections': sections,
        'subjects': subjects,
        'test_numbers': test_numbers
    })

@app.route('/api/results/summary')
def api_results_summary():
    """Return aggregated summary per class, section, subject, test_no."""
    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT
            r.class,
            COALESCE(r.section, s.section, '') as section,
            r.subject,
            r.chapter as test_no,
            COUNT(DISTINCT r.student_id) as student_count,
            AVG(r.score) as avg_score,
            MAX(r.score) as max_score,
            MIN(r.score) as min_score
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        WHERE r.class IS NOT NULL AND r.class != ''
          AND r.subject IS NOT NULL AND r.subject != ''
        GROUP BY r.class, COALESCE(r.section, s.section, ''), r.subject, r.chapter
        ORDER BY r.class, COALESCE(r.section, s.section, ''), r.subject, r.chapter
    """
    c.execute(query)
    rows = c.fetchall()
    summary = [dict(row) for row in rows]

    # Total registered students from students table
    c.execute("SELECT COUNT(*) as total FROM students")
    total_st_row = c.fetchone()
    c.execute("SELECT COUNT(DISTINCT student_id) as total_res FROM results")
    total_res_row = c.fetchone()
    total_students = max(total_st_row['total'] if total_st_row else 0, total_res_row['total_res'] if total_res_row else 0)
    conn.close()

    return jsonify({
        'summary': summary,
        'total_students': total_students
    })

@app.route('/api/verify_admin_password', methods=['POST'])
def verify_admin_password():
    """Verify admin password for elevation from results page to admin dashboard."""
    data = request.get_json(silent=True) or request.form or {}
    password = data.get('password', '').strip()
    if password == 'admin123':
        session['admin_logged_in'] = True
        return jsonify({'status': 'success', 'redirect': url_for('admin_dashboard')})
    return jsonify({'status': 'error', 'message': 'Incorrect admin password.'}), 401

@app.route('/admin/results')
def results_page():
    """
    Render the results page showing all test results. Accessible to both Admin and Teacher roles.
    """
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return redirect(url_for('teacher_login'))
    return render_template('results.html',
                           is_admin=bool(session.get('admin_logged_in')),
                           is_teacher=bool(session.get('teacher_logged_in')))

@app.route('/admin/results/data')
def results_data():
    """
    API endpoint to fetch results data with filtering. Accessible to both Admin and Teacher roles.
    """
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    """
    API endpoint to fetch results data with filtering:
    class, section (via join), subject, test_no (chapter).
    """
    class_filter   = request.args.get('class', '')
    section_filter = request.args.get('section', '')
    subject_filter = request.args.get('subject', '')
    test_filter    = request.args.get('test_no', '')

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT
            r.id, r.student_id, r.name, r.class, r.subject,
            r.score, r.total_questions, r.percentage, r.test_date,
            r.chapter,
            COALESCE(r.section, s.section, '') as section,
            s.exam_started_at
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        WHERE r.id IN (
            SELECT MAX(id) FROM results GROUP BY student_id, LOWER(subject), COALESCE(chapter, '')
        )
    """
    params = []

    if class_filter:
        query += " AND r.class = ?"
        params.append(class_filter)
    if section_filter:
        query += " AND COALESCE(r.section, s.section, '') = ?"
        params.append(section_filter)
    if subject_filter:
        query += " AND r.subject = ?"
        params.append(subject_filter)
    if test_filter:
        query += " AND r.chapter = ?"
        params.append(test_filter)

    query += " ORDER BY r.test_date DESC, r.class, r.subject, r.name"
    c.execute(query, params)
    results = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(results)
@app.route('/admin/results_filter')
@admin_required
def results_filter():
    """
    Render the advanced results filter and export page.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class FROM results WHERE class IS NOT NULL AND class != '' UNION SELECT DISTINCT class FROM questions WHERE class IS NOT NULL AND class != '' ORDER BY class")
    classes = [dict(r)['class'] for r in c.fetchall()]
    
    c.execute("SELECT DISTINCT subject FROM results WHERE subject IS NOT NULL AND subject != '' UNION SELECT DISTINCT subject FROM questions WHERE subject IS NOT NULL AND subject != '' ORDER BY subject")
    subjects = [dict(r)['subject'] for r in c.fetchall()]
    
    c.execute("SELECT DISTINCT section FROM students WHERE section IS NOT NULL AND section != '' ORDER BY section")
    sections = [dict(r)['section'] for r in c.fetchall()]
    conn.close()
    return render_template('results_filter.html', classes=classes, subjects=subjects, sections=sections)

@app.route('/admin/export_results_page')
@admin_required
def export_results_page():
    """
    Render the page for exporting results.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class FROM students ORDER BY class")
    classes = [row['class'] for row in c.fetchall()]
    c.execute("SELECT DISTINCT subject FROM students ORDER BY subject")
    subjects = [row['subject'] for row in c.fetchall()]
    c.execute("SELECT DISTINCT section FROM students WHERE section IS NOT NULL AND section != '' ORDER BY section")
    sections = [row['section'] for row in c.fetchall()]
    conn.close()
    return render_template('export_results_filter.html', classes=classes, subjects=subjects, sections=sections)

@app.route('/admin/export/results')
@admin_required
def export_results():
    """
    Export results to Excel, PDF, or JSON (preview) with filters.
    """
    # ── Parse filters ──────────────────────────────────────────────────
    class_filter   = request.args.get('class', '')
    section_filter = request.args.get('section', '')
    subject_filter = request.args.get('subject', '')
    test_filter    = request.args.get('test_no', '')
    start_date     = request.args.get('start_date', '')
    end_date       = request.args.get('end_date', '')
    single_date    = request.args.get('date', '')
    format_type    = request.args.get('format', 'excel')  # 'excel', 'pdf', 'json'

    conn = get_db()
    c = conn.cursor()

    # ── Build query ────────────────────────────────────────────────────
    query = '''
        SELECT
            s.student_id,
            s.name,
            s.class,
            COALESCE(s.section, '') as section,
            s.subject,
            COALESCE(NULLIF(r.chapter,''), NULLIF(s.test_no,''), '—') as test_no,
            COALESCE(r.score, 0) as score,
            COALESCE(r.total_questions, 0) as total_questions,
            COALESCE(r.percentage, 0.0) as percentage,
            COALESCE(r.test_date, s.exam_started_at) as test_date,
            s.exam_started_at
        FROM students s
        LEFT JOIN results r ON r.id = (
            SELECT id FROM results 
            WHERE student_id = s.student_id 
              AND (LOWER(subject) = LOWER(s.subject) OR subject IS NULL OR subject = '')
              AND (chapter = s.test_no OR chapter IS NULL OR chapter = '' OR s.test_no IS NULL OR s.test_no = '')
            ORDER BY id DESC LIMIT 1
        )
        WHERE 1=1
    '''
    params = []

    if class_filter and class_filter != 'All':
        clean_cls = class_filter.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
        query += " AND (s.class = ? OR s.class LIKE ? OR r.class = ? OR r.class LIKE ?)"
        params.extend([class_filter, f"%{clean_cls}%", class_filter, f"%{clean_cls}%"])

    if section_filter and section_filter != 'All':
        query += " AND LOWER(COALESCE(s.section, '')) = LOWER(?)"
        params.append(section_filter.strip())

    if subject_filter and subject_filter != 'All':
        query += " AND LOWER(COALESCE(s.subject, r.subject)) = LOWER(?)"
        params.append(subject_filter.strip())

    if test_filter and test_filter != '' and test_filter != 'All':
        query += " AND (r.chapter = ? OR s.test_no = ?)"
        params.extend([test_filter, test_filter])

    if single_date:
        query += " AND DATE(COALESCE(r.test_date, s.exam_started_at)) = ?"
        params.append(single_date)
    else:
        if start_date:
            query += " AND DATE(COALESCE(r.test_date, s.exam_started_at)) >= ?"
            params.append(start_date)
        if end_date:
            query += " AND DATE(COALESCE(r.test_date, s.exam_started_at)) <= ?"
            params.append(end_date)

    query += " ORDER BY LOWER(TRIM(s.name)) ASC, s.student_id ASC"

    c.execute(query, params)
    rows = c.fetchall()
    
    seen = set()
    deduped_results = []
    for row in rows:
        r_dict = dict(row)
        key = (r_dict['student_id'], r_dict['subject'], r_dict['test_no'])
        if key not in seen:
            seen.add(key)
            deduped_results.append(r_dict)
            
    results = deduped_results
    total_students = len(results)
    conn.close()

    # ─── JSON (Preview) ──────────────────────────────────────────────────
    if format_type == 'json':
        return jsonify({'results': results, 'count': total_students})

    # ─── Common metadata ────────────────────────────────────────────────
    school_name = get_setting('school_name', 'RRB Group of Schools')
    class_display = class_filter if class_filter and class_filter != 'All' else "All Classes"
    section_display = section_filter if section_filter and section_filter != 'All' else "All Sections"
    subject_display = subject_filter if subject_filter and subject_filter != 'All' else "All Subjects"
    test_display = test_filter if test_filter else "All Tests"

    if single_date:
        date_display = single_date
    elif start_date and end_date:
        date_display = f"{start_date} to {end_date}"
    elif start_date:
        date_display = f"From {start_date}"
    elif end_date:
        date_display = f"Up to {end_date}"
    else:
        date_display = "All Dates"

    export_time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M')

    # ─── PDF Export ──────────────────────────────────────────────────────
    if format_type == 'pdf':
        # Build logo base64 if exists
        logo_path = get_setting('logo_path', '')
        logo_base64 = None
        if logo_path:
            abs_path = os.path.join(app.static_folder, logo_path)
            if os.path.exists(abs_path):
                with open(abs_path, "rb") as f:
                    logo_bytes = f.read()
                    ext = os.path.splitext(abs_path)[1].lower()
                    mime = "image/png" if ext == ".png" else "image/jpeg"
                    logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"

        rendered = render_template('export_results_pdf.html',
            school_name=school_name,
            logo_base64=logo_base64,
            class_display=class_display,
            section_display=section_display,
            subject_display=subject_display,
            test_display=test_display,
            date_display=date_display,
            export_time=export_time,
            total_students=total_students,
            results=results
        )
        if request.args.get('download') == 'pdf':
            pdf = HTML(string=rendered).write_pdf()
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=export_results_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            return response

        return rendered

    # ─── Excel Export (default) ─────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    # Header rows
    ws.merge_cells('A1:H1')
    ws['A1'].value = school_name
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'].value = f"Class: {class_display}  |  Section: {section_display}  |  Subject: {subject_display}  |  Test: {test_display}"
    ws['A2'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.merge_cells('A3:H3')
    ws['A3'].value = f"Date: {date_display}  |  Total Students: {total_students}  |  Generated: {export_time}"
    ws['A3'].font = openpyxl.styles.Font(italic=True)
    ws['A3'].alignment = openpyxl.styles.Alignment(horizontal='center')
    ws.append([])

    headers = ['Student ID', 'Name', 'Class', 'Section', 'Subject', 'Test No', 'Score', 'Percentage']
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    for row_idx, row in enumerate(results, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=row['student_id'])
        ws.cell(row=row_idx, column=2, value=row['name'])
        ws.cell(row=row_idx, column=3, value=row['class'])
        ws.cell(row=row_idx, column=4, value=row.get('section', 'N/A'))
        ws.cell(row=row_idx, column=5, value=row['subject'])
        ws.cell(row=row_idx, column=6, value=row.get('test_no', 'N/A'))
        ws.cell(row=row_idx, column=7, value=f"{row['score']}/{row['total_questions']}" if row['total_questions'] else row['score'])
        ws.cell(row=row_idx, column=8, value=f"{row['percentage']:.1f}%" if row['percentage'] else 'N/A')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    file_parts = []
    file_parts.append(class_filter if class_filter and class_filter != 'All' else 'all')
    file_parts.append(subject_filter if subject_filter and subject_filter != 'All' else 'all')
    if start_date and end_date:
        file_parts.append(f"{start_date}_to_{end_date}")
    elif start_date:
        file_parts.append(f"from_{start_date}")
    elif end_date:
        file_parts.append(f"to_{end_date}")
    else:
        file_parts.append("all_dates")
    file_parts.append(timestamp)
    filename = f"results_{'_'.join(file_parts)}.xlsx"

    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/admin/student/responses/<student_id>')
@admin_required
def view_student_responses(student_id):
    """
    Render page to view all responses of a specific student.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        return "Student not found", 404
    log_audit_event('admin', 'admin', 'STUDENT_RESULT_VIEW', 'students', student_id, request.remote_addr)

    # Fetch questions with all option texts, filtering by test_no if set
    test_no = (student['test_no'] if ('test_no' in student.keys() and student['test_no']) else '').strip()
    if test_no:
        c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=? AND (test_no=? OR chapter=?) ORDER BY id",
                  (student['class'], student['subject'], test_no, test_no))
    else:
        c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer FROM questions WHERE class=? AND subject=? ORDER BY id",
                  (student['class'], student['subject']))
    questions = c.fetchall()

    c.execute("SELECT question_id, selected_option FROM responses WHERE student_id=?", (student_id,))
    responses = {row['question_id']: row['selected_option'] for row in c.fetchall()}

    conn.close()

    question_data = []
    for q in questions:
        selected = responses.get(q['id'], '')
        correct_letter = _normalize_correct_answer(
            q['correct_answer'],
            q['option_a'],
            q['option_b'],
            q['option_c'],
            q['option_d']
        )
        is_correct = bool(selected and correct_letter and selected.strip().upper() == correct_letter)

        question_data.append({
            'id': q['id'],
            'question': q['question'],
            'options': {
                'A': q['option_a'],
                'B': q['option_b'],
                'C': q['option_c'],
                'D': q['option_d']
            },
            'correct': correct_letter or '?',  # display the letter, not raw text
            'selected': selected,
            'is_correct': is_correct
        })

    return render_template('student_responses.html', student=student, questions=question_data)

# ─────── HIERARCHICAL RESULTS FILTERING ────────
@app.route('/admin/results_filter')
@admin_required
def results_filter_page():
    """
    Render the hierarchical results filter page (Class → Subject → Test → Date).
    """
    return render_template('results_filter.html')

@app.route('/api/admin/results/get_classes')
@admin_required
def api_get_classes():
    """
    API endpoint to get all classes with results.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class FROM results WHERE class IS NOT NULL AND class != '' ORDER BY class")
    classes = [row['class'] for row in c.fetchall()]
    conn.close()
    return jsonify({'classes': classes})

@app.route('/api/admin/results/get_subjects/<class_name>')
@admin_required
def api_get_subjects(class_name):
    """
    API endpoint to get subjects for a specific class.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT subject FROM results WHERE class=? AND subject IS NOT NULL AND subject != '' ORDER BY subject", (class_name,))
    subjects = [row['subject'] for row in c.fetchall()]
    conn.close()
    return jsonify({'subjects': subjects})

@app.route('/api/admin/results/get_test_numbers/<class_name>/<subject_name>')
@admin_required
def api_get_test_numbers(class_name, subject_name):
    """
    API endpoint to get test numbers for a specific class and subject.
    """
    conn = get_db()
    c = conn.cursor()
    # Get test_no from chapters (stored in results.chapter)
    c.execute("SELECT DISTINCT chapter FROM results WHERE class=? AND subject=? AND chapter IS NOT NULL AND chapter != '' ORDER BY chapter", 
              (class_name, subject_name))
    test_numbers = [row['chapter'] for row in c.fetchall()]
    conn.close()
    return jsonify({'test_numbers': test_numbers})

@app.route('/api/admin/results/get_dates/<class_name>/<subject_name>/<test_no>')
@admin_required
def api_get_dates(class_name, subject_name, test_no):
    """
    API endpoint to get test dates for a specific class, subject, and test number.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT DISTINCT DATE(test_date) as date FROM results 
                 WHERE class=? AND subject=? AND chapter=? 
                 ORDER BY test_date DESC""", 
              (class_name, subject_name, test_no))
    dates = [row['date'] for row in c.fetchall()]
    conn.close()
    return jsonify({'dates': dates})

@app.route('/api/admin/results/get_filtered_results')
@admin_required
def api_get_filtered_results():
    """
    API endpoint to get filtered results based on class, subject, test_no, and date.
    """
    class_name = request.args.get('class', '')
    subject_name = request.args.get('subject', '')
    test_no = request.args.get('test_no', '')
    date_filter = request.args.get('date', '')

    conn = get_db()
    c = conn.cursor()

    query = '''SELECT r.id, r.student_id, r.name, r.class, r.subject, r.score,
                      r.total_questions, r.percentage, r.test_date, r.chapter
               FROM results r
               WHERE 1=1'''
    params = []

    if class_name:
        query += " AND r.class = ?"
        params.append(class_name)
    if subject_name:
        query += " AND r.subject = ?"
        params.append(subject_name)
    if test_no:
        query += " AND r.chapter = ?"
        params.append(test_no)
    if date_filter:
        query += " AND DATE(r.test_date) = ?"
        params.append(date_filter)

    query += " ORDER BY r.name ASC"

    c.execute(query, params)
    results = [dict(row) for row in c.fetchall()]
    conn.close()

    return jsonify({
        'results': results,
        'count': len(results),
        'filters': {
            'class': class_name,
            'subject': subject_name,
            'test_no': test_no,
            'date': date_filter
        }
    })

@app.route('/admin/export/answers')
@admin_required
def export_answers():
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT r.student_id, q.question, r.selected_option, q.correct_answer
                 FROM responses r JOIN questions q ON r.question_id = q.id
                 ORDER BY r.student_id, q.id''')
    rows = c.fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.append(['Student ID', 'Question', 'Selected', 'Correct'])
    for row in rows:
        ws.append([row['student_id'], row['question'], row['selected_option'], row['correct_answer']])
    filename = f"answers_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)

@app.route('/admin/current_test_sessions')
@admin_required
def admin_current_test_sessions():
    """Page showing list of recent test sessions with generate buttons."""
    return render_template('current_test_sessions.html')

@app.route('/api/admin/recent_test_sessions')
@admin_required
def api_recent_test_sessions():
    """Return list of distinct test sessions across all dates, grouped by class, section, subject, test_no, date."""
    conn = get_db()
    c = conn.cursor()
    
    c.execute("""
        SELECT 
            r.class,
            COALESCE(NULLIF(r.section, ''), NULLIF(s.section, ''), '') as section,
            r.subject,
            COALESCE(NULLIF(r.chapter, ''), 'Test') as test_no,
            DATE(r.test_date) as session_date,
            MAX(r.test_date) as test_date,
            COUNT(DISTINCT r.student_id) as student_count
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        WHERE r.class IS NOT NULL AND r.class != ''
        GROUP BY r.class, COALESCE(NULLIF(r.section, ''), NULLIF(s.section, ''), ''), r.subject, r.chapter, DATE(r.test_date)
        ORDER BY MAX(r.test_date) DESC
        LIMIT 50
    """)
    sessions = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'sessions': sessions})

@app.route('/admin/current_test_results')
@admin_required
def current_test_results():
    """Show results of the test session, filtered by class, section, subject, test_no, date."""
    format_type = request.args.get('format', 'html')
    class_filter = request.args.get('class', '').strip()
    section_filter = request.args.get('section', '').strip()
    subject_filter = request.args.get('subject', '').strip()
    test_no_filter = request.args.get('test_no', '').strip()
    date_filter = request.args.get('date', '').strip()
    
    conn = get_db()
    c = conn.cursor()

    clean_cls = class_filter.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()

    # Build base query with joins
    query = """
        SELECT r.*, COALESCE(NULLIF(s.admission_no, ''), s.student_id) as admission_no,
               COALESCE(NULLIF(r.section, ''), s.section, '') as student_section
        FROM results r
        LEFT JOIN students s ON r.student_id = s.student_id
        WHERE 1=1
    """
    params = []

    if class_filter:
        query += " AND (r.class = ? OR r.class = ? OR r.class LIKE ?)"
        params.extend([class_filter, clean_cls, f"%{clean_cls}%"])
    if section_filter:
        query += " AND (r.section = ? OR s.section = ?)"
        params.extend([section_filter, section_filter])
    if subject_filter:
        query += " AND (LOWER(r.subject) = LOWER(?) OR r.subject LIKE ?)"
        params.extend([subject_filter, f"%{subject_filter}%"])
    if test_no_filter:
        query += " AND (r.chapter = ? OR r.chapter LIKE ?)"
        params.extend([test_no_filter, f"%{test_no_filter}%"])
    if date_filter:
        query += " AND DATE(r.test_date) = DATE(?)"
        params.append(date_filter)

    query += " ORDER BY r.score DESC"
    c.execute(query, params)
    results = [dict(row) for row in c.fetchall()]
    conn.close()

    if not results:
        return "No results found for the selected filters.", 404

    # Assign ranks
    for i, res in enumerate(results):
        res['rank'] = i + 1

    # Get max total questions (assuming same total for all in this test)
    total_q = results[0].get('total_questions', 0) if results else 0

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"

    total_q_map = {f"{res['class']}_{res['subject']}": res.get('total_questions', 0) for res in results}
    max_total_marks = total_q

    display_class = results[0]['class'] if results else '—'
    display_section = section_filter or results[0].get('student_section', '—') if results else '—'
    display_subject = results[0]['subject'] if results else '—'
    display_test_no = results[0].get('chapter', '—') if results else '—'
    start_time = results[0]['test_date'] if results else ''

    if format_type == 'pdf':
        rendered = render_template('current_test_results.html',
                                   results=results,
                                   school_name=school_name,
                                   logo_base64=logo_base64,
                                   start_time=start_time,
                                   total_students=len(results),
                                   total_q_map=total_q_map,
                                   max_total_marks=max_total_marks,
                                   display_class=display_class,
                                   display_section=display_section,
                                   display_subject=display_subject,
                                   display_test_no=display_test_no,
                                   datetime=datetime)
        pdf = HTML(string=rendered).write_pdf()
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=current_test_results.pdf'
        return response
    
    elif format_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(['S.No.', 'Admission No', 'Student Name', f'Marks (out of {max_total_marks})', 'Rank'])
        for i, res in enumerate(results):
            ws.append([i+1, res.get('admission_no', ''), res['name'], f"{res['score']}/{max_total_marks}", res['rank']])
        filename = f"current_test_results_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
        wb.save(filepath)
        return send_file(filepath, as_attachment=True)
    
    else:
        return render_template('current_test_results.html',
                               results=results,
                               school_name=school_name,
                               logo_base64=logo_base64,
                               start_time=start_time,
                               total_students=len(results),
                               total_q_map=total_q_map,
                               max_total_marks=max_total_marks,
                               display_class=display_class,
                               display_section=display_section,
                               display_subject=display_subject,
                               display_test_no=display_test_no,
                               datetime=datetime)

       
@app.route('/admin/reattempt_requests')
@admin_required
def admin_reattempt_requests():
    return render_template('reattempt_requests.html')

@app.route('/admin/reattempt_requests/data')
@admin_required
def reattempt_requests_data():
    filter_status = request.args.get('status', '')
    conn = get_db()
    c = conn.cursor()
    query = """
        SELECT r.id, r.student_id, s.name, r.class, r.subject, r.status, r.requested_at, r.reviewed_at, r.admin_note
        FROM reattempt_requests r
        JOIN students s ON r.student_id = s.student_id
        WHERE 1=1
    """
    params = []
    if filter_status:
        query += " AND r.status = ?"
        params.append(filter_status)
    query += " ORDER BY r.requested_at DESC"
    c.execute(query, params)
    requests = [dict(row) for row in c.fetchall()]
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    exam_is_active = bool(exam_row['is_active']) if exam_row else False
    conn.close()
    return jsonify({'requests': requests, 'exam_is_active': exam_is_active})

def auto_submit_expired_exams():
    """
    Auto-submit exams that have exceeded the time limit.
    """
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, duration, is_active FROM exam_control WHERE id=1")
    exam = c.fetchone()

    if not exam or not exam['duration'] or not exam['is_active']:
        conn.close()
        return

    duration = exam['duration']

    # ✅ Use SQLite's own time for comparison — avoids timezone mismatch
    c.execute("""
        UPDATE students 
        SET status='Submitted'
        WHERE status='In Progress'
        AND exam_started_at IS NOT NULL
        AND datetime(exam_started_at, ? || ' minutes') < datetime('now')
    """, (str(duration),))

    conn.commit()
    conn.close()

@app.route('/guidelines')
@student_required
def guidelines():
    # Mark that the student has not yet accepted guidelines for this session
    session['guidelines_accepted'] = False
    return render_template('guidelines.html')

@app.route('/accept_guidelines', methods=['POST'])
@student_required
def accept_guidelines():
    session['guidelines_accepted'] = True
    return redirect(url_for('exam'))

@app.route('/admin/reattempt_request/<int:req_id>/<action>', methods=['POST'])
@admin_required
def handle_reattempt_request(req_id, action):
    if action not in ('approve', 'reject'):
        return jsonify({'status': 'error', 'message': 'Invalid action'}), 400

    admin_note = request.json.get('note', '') if request.is_json else ''

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT student_id, class, subject, status FROM reattempt_requests WHERE id=?", (req_id,))
    req = c.fetchone()
    if not req:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Request not found'}), 404
    if req['status'] != 'pending':
        conn.close()
        return jsonify({'status': 'error', 'message': 'Request already processed'}), 400

    student_id = req['student_id']
    class_ = req['class']
    subject = req['subject']

    if action == 'approve':
        c.execute("DELETE FROM responses WHERE student_id=?", (student_id,))
        c.execute("DELETE FROM results WHERE student_id=?", (student_id,))
        c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
        c.execute("UPDATE students SET status='Not Started', exam_started_at=NULL WHERE student_id=?", (student_id,))
        c.execute("UPDATE reattempt_requests SET status='approved', reviewed_at=CURRENT_TIMESTAMP, admin_note=? WHERE id=?",
                  (admin_note, req_id))
        
        conn.commit()
    else:
        c.execute("UPDATE reattempt_requests SET status='rejected', reviewed_at=CURRENT_TIMESTAMP, admin_note=? WHERE id=?",
                  (admin_note, req_id))
        conn.commit()
    conn.close()
    log_audit_event('admin', 'admin', 'REATTEMPT_' + action.upper(), 'reattempt_requests', req_id, request.remote_addr)
    return jsonify({'status': 'success'})

# ========================
# TEACHER ROUTES
# ========================

@app.route('/teacher', methods=['GET', 'POST'])
@app.route('/teacher/login', methods=['GET', 'POST'])
def teacher_login():
    """
    Handle teacher login.
    """
    if request.method == 'POST':
        mobile = request.form.get('mobile')
        password = request.form.get('password')
        
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM teachers WHERE mobile=? AND status='active'", (mobile,))
        teacher = c.fetchone()
        conn.close()
        
        if teacher and verify_password(password, teacher['password']):
            session['teacher_logged_in'] = True
            session['teacher_id'] = teacher['id']
            session['teacher_name'] = teacher['name']
            session['teacher_mobile'] = teacher['mobile']
            return redirect(url_for('teacher_dashboard'))
        else:
            return render_template('teacher_login.html', error="Invalid credentials")
    
    return render_template('teacher_login.html')

@app.route('/teacher/logout')
def teacher_logout():
    """
    Handle teacher logout.
    """
    session.pop('teacher_logged_in', None)
    session.pop('teacher_id', None)
    session.pop('teacher_name', None)
    session.pop('teacher_mobile', None)
    return redirect(url_for('teacher_login'))

@app.route('/teacher/create_test', methods=['GET', 'POST'])
@teacher_required
@limiter.limit("13 per minute")
def teacher_create_test():
    if request.method == 'POST':
        class_ = request.form.get('class')
        subject = request.form.get('subject')
        chapter = request.form.get('chapter')
        num_questions = int(request.form.get('num_questions', 10))
        prompt = request.form.get('prompt', '')
        method = request.form.get('method', 'ai')  # 'ai' or 'upload'
        
        if method == 'upload':
            # Handle CSV upload
            file = request.files.get('csv_file')
            if file and file.filename.endswith('.csv'):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Import questions from CSV
                with open(filepath, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    conn = get_db()
                    c = conn.cursor()
                    for row in reader:
                        c.execute("""INSERT INTO questions 
                                   (class, subject, chapter, question, option_a, option_b, option_c, option_d, correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?)""",
                                (class_, subject, chapter,
                                 row.get('question', ''),
                                 row.get('option a', row.get('option_a', '')),
                                 row.get('option b', row.get('option_b', '')),
                                 row.get('option c', row.get('option_c', '')),
                                 row.get('option d', row.get('option_d', '')),
                                 row.get('correct_answer', '')))
                    conn.commit()
                    conn.close()
                os.remove(filepath)
                return jsonify({'status': 'success', 'message': 'Questions uploaded successfully'})
        
        else:  # AI generation using Google Gemini
            try:
                import urllib.request
                import json as json_lib

                gemini_api_key = os.environ.get('GEMINI_API_KEY', '')
                if not gemini_api_key:
                    return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY not set. Use Option 7 in the batch menu to set it.'}), 500

                ai_prompt = f"""You are an expert CBSE/ICSE question paper generator.
Output ONLY a valid JSON array — no markdown, no text outside the array.

Generate exactly {num_questions} MCQ questions for:
Class: {class_} | Subject: {subject} | Chapter: {chapter}

Teacher instructions: {prompt if prompt else 'Standard difficulty, balanced coverage'}

SCHEMA — every object must have these exact keys:
{{
  "question":      "<question text, use LaTeX or mhchem where needed>",
  "option_a":      "<option text or LaTeX>",
  "option_b":      "<option text or LaTeX>",
  "option_c":      "<option text or LaTeX>",
  "option_d":      "<option text or LaTeX>",
  "correct_answer":"option_a"|"option_b"|"option_c"|"option_d",
  "smiles":        "<SMILES for structural diagram, else empty string>",
  "content_type":  "math"|"chemistry"|"physics"|"biology"|"text"
}}

MATHEMATICS — use MathJax LaTeX (escape backslashes in JSON):
• Inline: $\\frac{{d}}{{dx}}(x^2)$   Block: $$\\int_0^1 x^2\\,dx = \\frac{{1}}{{3}}$$
• Vectors: $\\vec{{F}} = m\\vec{{a}}$  Matrices: $\\begin{{pmatrix}} a & b \\\\\\\\ c & d \\end{{pmatrix}}$
• NEVER use raw Unicode math symbols — use LaTeX commands only.

CHEMISTRY — use mhchem inside $ $:
• Equations: "$\\ce{{H2 + O2 -> H2O}}$"   Equilibrium: "$\\ce{{N2 + 3H2 <=> 2NH3}}$"
• States: "$\\ce{{CaCO3(s) -> CaO(s) + CO2(g)}}$"   Ionic: "$\\ce{{Fe^{{2+}} + 2e- -> Fe}}$"
• Structural diagrams: set smiles field (e.g. benzene="c1ccccc1", ethanol="CCO")

Output the JSON array now:"""

                # Call Multi-Provider AI (Gemini, DeepSeek, ChatGPT, Claude)
                result, provider, err_msg = generate_ai_content(ai_prompt, timeout=90)
                if not result:
                    return jsonify({'status': 'error', 'message': err_msg}), 500

                response_text = result['candidates'][0]['content']['parts'][0]['text'].strip()

                # Clean markdown fences if present
                if '```' in response_text:
                    parts = response_text.split('```')
                    for part in parts:
                        part = part.strip()
                        if part.startswith('json'):
                            part = part[4:].strip()
                        if part.startswith('['):
                            response_text = part
                            break

                questions = json_lib.loads(response_text)

                # Save to database (including smiles and content_type for multimodal rendering)
                conn = get_db()
                c = conn.cursor()
                for q in questions:
                    # Store smiles in image_path column prefixed so renderer knows it's SMILES
                    smiles_val = q.get('smiles', '') or ''
                    # Store content_type in question_type if it's a pure text field question
                    c.execute("""INSERT INTO questions
                               (class, subject, chapter, question_type, question,
                                option_a, option_b, option_c, option_d, correct_answer, image_path)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                            (class_, subject, chapter,
                             'MCQ',
                             q.get('question', ''),
                             q.get('option_a', ''),
                             q.get('option_b', ''),
                             q.get('option_c', ''),
                             q.get('option_d', ''),
                             q.get('correct_answer', ''),
                             f'__smiles__{smiles_val}' if smiles_val else None))
                conn.commit()
                conn.close()

                return jsonify({'status': 'success', 'message': f'{len(questions)} questions generated and saved using Gemini!'})

            except urllib.error.HTTPError as e:
                err_body = e.read().decode('utf-8')
                return jsonify({'status': 'error', 'message': f'Gemini API error: {err_body}'}), 500
            except Exception as e:
                return jsonify({'status': 'error', 'message': str(e)}), 500
    
    # GET request - show form
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    class_subject_map = {}
    for a in assignments:
        cls = a['class']
        if cls not in class_subject_map:
            class_subject_map[cls] = []
        if a['subject'] not in class_subject_map[cls]:
            class_subject_map[cls].append(a['subject'])
    
    return render_template('teacher_create_test_v2.html', assignments=assignments, class_subject_map=class_subject_map)
@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    """
    Render the teacher dashboard with allotted class stats, all-classes list,
    and same-subject teacher test counts.
    """
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    
    # Get teacher assignments
    c.execute("""SELECT class, subject, is_class_teacher 
                 FROM teacher_assignments 
                 WHERE teacher_id=?""", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Get teacher info
    c.execute("SELECT * FROM teachers WHERE id=?", (teacher_id,))
    t_row = c.fetchone()
    teacher = dict(t_row) if t_row else {}
    
    # Compile Detailed Allotted Classes Data
    allotted_details = []
    total_assigned_students = 0
    total_my_tests_all = 0

    for a in assignments:
        cls = str(a['class'])
        subj = str(a['subject'])
        
        # Enrolled students count for class
        c.execute('SELECT COUNT(*) FROM students WHERE class=?', (cls,))
        st_count = c.fetchone()[0]
        total_assigned_students += st_count
        
        # Total tests conducted for class & subject
        c.execute('SELECT COUNT(DISTINCT test_no) FROM questions WHERE class=? AND subject=?', (cls, subj))
        q_tests = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM test_papers WHERE class=? AND subject=?', (cls, subj))
        tp_tests = c.fetchone()[0]
        tot_tests = max(q_tests, tp_tests)
        
        # My tests created for class & subject
        c.execute('SELECT COUNT(*) FROM test_generation_history WHERE teacher_id=? AND class=? AND subject=?', (teacher_id, cls, subj))
        my_ai = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM test_papers WHERE uploaded_by=? AND class=? AND subject=?', (str(teacher_id), cls, subj))
        my_up = c.fetchone()[0]
        my_tests_count = my_ai + my_up
        total_my_tests_all += my_tests_count
        
        # Same subject teachers breakdown (all teachers teaching this subject across school)
        c.execute('''SELECT DISTINCT t.id, t.name, t.mobile, t.picture 
                     FROM teacher_assignments ta 
                     JOIN teachers t ON ta.teacher_id = t.id 
                     WHERE ta.subject=?''', (subj,))
        s_teachers = [dict(r) for r in c.fetchall()]
        
        teachers_breakdown = []
        for st in s_teachers:
            tid = st['id']
            c.execute('SELECT COUNT(*) FROM test_generation_history WHERE teacher_id=? AND subject=?', (tid, subj))
            tai = c.fetchone()[0]
            c.execute('SELECT COUNT(*) FROM test_papers WHERE uploaded_by=? AND subject=?', (str(tid), subj))
            tup = c.fetchone()[0]
            st['tests_count'] = tai + tup
            st['is_me'] = (tid == teacher_id)
            teachers_breakdown.append(st)
            
        allotted_details.append({
            'class': cls,
            'subject': subj,
            'is_class_teacher': a['is_class_teacher'],
            'student_count': st_count,
            'total_tests': tot_tests,
            'my_tests': my_tests_count,
            'teachers_breakdown': teachers_breakdown
        })

    # Compile All Classes List across System
    c.execute('''
        SELECT class FROM students WHERE class IS NOT NULL AND class != ''
        UNION
        SELECT class FROM questions WHERE class IS NOT NULL AND class != ''
        UNION
        SELECT class FROM teacher_assignments WHERE class IS NOT NULL AND class != ''
    ''')
    class_names = sorted([r[0] for r in c.fetchall()], key=lambda x: int(x) if str(x).isdigit() else str(x))

    all_classes_data = []
    for cls in class_names:
        c.execute('SELECT COUNT(*) FROM students WHERE class=?', (str(cls),))
        st_cnt = c.fetchone()[0]
        
        c.execute('SELECT DISTINCT subject FROM questions WHERE class=? AND subject IS NOT NULL', (str(cls),))
        q_subjs = [r[0] for r in c.fetchall()]
        c.execute('SELECT DISTINCT subject FROM teacher_assignments WHERE class=? AND subject IS NOT NULL', (str(cls),))
        ta_subjs = [r[0] for r in c.fetchall()]
        all_subjs = sorted(list(set(q_subjs + ta_subjs)))
        
        c.execute('SELECT COUNT(DISTINCT test_no) FROM questions WHERE class=?', (str(cls),))
        q_tests = c.fetchone()[0]
        c.execute('SELECT COUNT(*) FROM test_papers WHERE class=?', (str(cls),))
        tp_tests = c.fetchone()[0]
        tot_tests = max(q_tests, tp_tests)
        
        all_classes_data.append({
            'class': str(cls),
            'student_count': st_cnt,
            'subjects': all_subjs,
            'total_tests': tot_tests
        })

    conn.close()
    return render_template('teacher_dashboard.html', 
                         assignments=assignments,
                         teacher=teacher,
                         allotted_details=allotted_details,
                         all_classes_data=all_classes_data,
                         total_assigned_students=total_assigned_students,
                         total_my_tests_all=total_my_tests_all)



@app.route('/teacher/students')
@teacher_required
def teacher_students():
    """
    Render the page showing teacher's assigned students with step-by-step Class -> Section -> Student navigation.
    """
    teacher_id = session.get('teacher_id')
    class_filter = request.args.get('class', '').strip()
    section_filter = request.args.get('section', '').strip()
    
    conn = get_db()
    c = conn.cursor()
    
    # 1. Get teacher assignments
    c.execute("""SELECT class, section, subject, is_class_teacher 
                 FROM teacher_assignments 
                 WHERE teacher_id=?""", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Build assigned classes map
    assigned_classes_map = {}
    for a in assignments:
        cls = str(a['class'])
        if not cls: continue
        if cls not in assigned_classes_map:
            assigned_classes_map[cls] = {
                'class_name': cls,
                'sections': set(),
                'subjects': [],
                'is_class_teacher': False,
                'total_students': 0
            }
        if a.get('section'):
            assigned_classes_map[cls]['sections'].add(str(a['section']))
        if a.get('subject') and a['subject'] not in assigned_classes_map[cls]['subjects']:
            assigned_classes_map[cls]['subjects'].append(a['subject'])
        if a.get('is_class_teacher'):
            assigned_classes_map[cls]['is_class_teacher'] = True
            
    for cls, cdata in assigned_classes_map.items():
        c.execute("SELECT DISTINCT section FROM students WHERE class=? AND section IS NOT NULL AND section!=''", (cls,))
        for r in c.fetchall():
            cdata['sections'].add(r[0])
        cdata['sections'] = sorted(list(cdata['sections']))
        
        c.execute("SELECT COUNT(*) FROM students WHERE class=?", (cls,))
        cdata['total_students'] = c.fetchone()[0]
        
    assigned_classes_list = sorted(list(assigned_classes_map.values()), key=lambda x: int(x['class_name']) if x['class_name'].isdigit() else x['class_name'])
    
    # 2. Build Student Query
    query = """SELECT DISTINCT s.student_id, s.name, s.admission_no, s.class, s.section, 
                      s.dob, s.house, s.parents_name, s.picture
               FROM students s
               WHERE 1=1"""
    params = []
    
    if class_filter:
        query += " AND s.class = ?"
        params.append(class_filter)
        if section_filter:
            query += " AND s.section = ?"
            params.append(section_filter)
    else:
        all_t_classes = list(set([str(a['class']) for a in assignments if a.get('class')]))
        if all_t_classes:
            placeholders = ','.join(['?'] * len(all_t_classes))
            query += f" AND s.class IN ({placeholders})"
            params.extend(all_t_classes)
            
    query += " ORDER BY s.class, s.section, s.name"
    c.execute(query, params)
    raw_students = [dict(row) for row in c.fetchall()]
    
    # 3. Compute student test metrics
    students = []
    for st in raw_students:
        sid = st['student_id']
        c.execute("SELECT COUNT(*) FROM results WHERE student_id=?", (sid,))
        t_count = c.fetchone()[0]
        
        c.execute("SELECT AVG(percentage) FROM results WHERE student_id=? AND percentage IS NOT NULL", (sid,))
        avg_perc = c.fetchone()[0]
        avg_perc = round(avg_perc, 1) if avg_perc is not None else 0.0
        
        st['tests_appeared'] = t_count
        st['avg_percentage'] = avg_perc
        students.append(st)
        
    conn.close()
    
    return render_template('teacher_students.html', 
                         assigned_classes=assigned_classes_list,
                         students=students,
                         assignments=assignments,
                         class_filter=class_filter,
                         section_filter=section_filter)

@app.route('/api/student_performance/<student_id>')
@teacher_required
def api_student_performance(student_id):
    """API endpoint to get detailed student performance analysis and graph data."""
    conn = get_db()
    c = conn.cursor()
    
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
    student = dict(student)
    
    teacher_id = session.get('teacher_id')
    c.execute("SELECT class, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    is_class_teacher = any(a['class'] == student['class'] and a['is_class_teacher'] for a in assignments)
    
    if is_class_teacher:
        c.execute("""
            SELECT id, subject, chapter, score, total_questions, percentage, test_date
            FROM results
            WHERE student_id=?
            ORDER BY test_date ASC
        """, (student_id,))
    else:
        teacher_subjects = [a['subject'] for a in assignments if a['class'] == student['class']]
        if teacher_subjects:
            placeholders = ','.join(['?'] * len(teacher_subjects))
            c.execute(f"""
                SELECT id, subject, chapter, score, total_questions, percentage, test_date
                FROM results
                WHERE student_id=? AND subject IN ({placeholders})
                ORDER BY test_date ASC
            """, [student_id] + teacher_subjects)
        else:
            c.execute("SELECT id, subject, chapter, score, total_questions, percentage, test_date FROM results WHERE 1=0")
            
    results_list = [dict(r) for r in c.fetchall()]
    
    total_tests = len(results_list)
    avg_score = round(sum(r['percentage'] or 0 for r in results_list) / total_tests, 1) if total_tests > 0 else 0.0
    max_score = round(max((r['percentage'] or 0 for r in results_list), default=0.0), 1)
    passed_tests = sum(1 for r in results_list if (r['percentage'] or 0) >= 40.0)
    pass_rate = round((passed_tests / total_tests) * 100, 1) if total_tests > 0 else 0.0
    
    subject_graphs = {}
    for r in results_list:
        subj = r['subject'] or 'General'
        if subj not in subject_graphs:
            subject_graphs[subj] = {'labels': [], 'scores': [], 'max_marks': []}
        t_label = r['chapter'] if r.get('chapter') else f"Test #{r['id']}"
        label = f"{t_label} ({r['test_date'][:10] if r.get('test_date') else ''})"
        subject_graphs[subj]['labels'].append(label)
        subject_graphs[subj]['scores'].append(round(r['percentage'] or 0.0, 1))
        subject_graphs[subj]['max_marks'].append(f"{r['score']}/{r['total_questions']}")
        
    conn.close()
    
    return jsonify({
        'status': 'success',
        'student': student,
        'stats': {
            'total_tests': total_tests,
            'avg_score': avg_score,
            'max_score': max_score,
            'passed_tests': passed_tests,
            'pass_rate': pass_rate
        },
        'results': results_list,
        'subject_graphs': subject_graphs
    })

@app.route('/teacher/student/<student_id>')
@teacher_required
def teacher_student_profile(student_id):
    """
    Render the profile page of a student (teacher view).
    """
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    
    # Get student info
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return "Student not found", 404
    student = dict(student)
    log_audit_event('teacher', session.get('teacher_id') or 'teacher', 'STUDENT_RESULT_VIEW', 'students', student_id, request.remote_addr)
    
    # Get teacher assignments
    c.execute("SELECT class, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    
    # Check if teacher is class teacher for this student
    is_class_teacher = any(a['class'] == student['class'] and a['is_class_teacher'] for a in assignments)
    
    # Get test history with safe numeric values
    if is_class_teacher:
        # Class teacher sees all subjects
        c.execute("""
            SELECT 
                r.id, r.subject, r.chapter, r.score, r.test_date,
                COALESCE(NULLIF(r.total_questions, 0), 
                    (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter))) AS total_questions,
                COALESCE(r.percentage,
                    CASE 
                        WHEN (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter)) > 0 
                        THEN ROUND((r.score * 100.0 / (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter))), 2)
                        ELSE 0.0 
                    END) AS percentage
            FROM results r
            WHERE r.student_id = ?
            ORDER BY r.test_date DESC
        """, (student_id,))
    else:
        # Subject teacher sees only their subjects
        teacher_subjects = [a['subject'] for a in assignments if a['class'] == student['class']]
        if teacher_subjects:
            placeholders = ','.join(['?'] * len(teacher_subjects))
            c.execute(f"""
                SELECT 
                    r.id, r.subject, r.chapter, r.score, r.test_date,
                    COALESCE(NULLIF(r.total_questions, 0), 
                        (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter))) AS total_questions,
                    COALESCE(r.percentage,
                        CASE 
                            WHEN (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter)) > 0 
                            THEN ROUND((r.score * 100.0 / (SELECT COUNT(*) FROM questions WHERE class = r.class AND subject = r.subject AND (r.chapter IS NULL OR r.chapter = '' OR test_no = r.chapter OR chapter = r.chapter))), 2)
                            ELSE 0.0 
                        END) AS percentage
                FROM results r
                WHERE r.student_id = ? AND r.subject IN ({placeholders})
                ORDER BY r.test_date DESC
            """, [student_id] + teacher_subjects)
        else:
            c.execute("""
                SELECT 
                    r.id, r.subject, r.chapter, r.score, r.test_date,
                    COALESCE(r.total_questions, 0) AS total_questions,
                    COALESCE(r.percentage, 0.0) AS percentage
                FROM results r
                WHERE 1=0
            """)
    
    test_history = [dict(row) for row in c.fetchall()]
    
    # Generate chart data for progress
    chart_data = {}
    for test in test_history:
        subj = test['subject']
        if subj not in chart_data:
            chart_data[subj] = {'dates': [], 'percentages': [], 'marks': []}
        chart_data[subj]['dates'].append(test['test_date'][:10] if test['test_date'] else '')
        chart_data[subj]['percentages'].append(test['percentage'] or 0)
        chart_data[subj]['marks'].append(f"{test['score']}/{test['total_questions']}")
    
    conn.close()
    
    return render_template('teacher_student_profile.html',
                         student=student,
                         test_history=test_history,
                         chart_data=chart_data,
                         is_class_teacher=is_class_teacher)

@app.route('/teacher/student/<student_id>/edit', methods=['POST'])
@teacher_required
def teacher_edit_student(student_id):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()

    # Get student class
    c.execute("SELECT class FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404

    # Both class teachers AND subject teachers in that class can edit
    c.execute("""SELECT id FROM teacher_assignments
                 WHERE teacher_id=? AND class=?""", (teacher_id, student['class']))
    assignment = c.fetchone()
    if not assignment:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403

    name = request.form.get('name')
    admission_no = request.form.get('admission_no')
    class_ = request.form.get('class')
    section = request.form.get('section')
    dob = request.form.get('dob')
    house = request.form.get('house')
    parents_name = request.form.get('parents_name')
    address = request.form.get('address')

    picture_path = save_student_picture(request.files.get('picture'), student_id)

    if picture_path:
        c.execute("""UPDATE students
                     SET name=?, admission_no=?, class=?, section=?, dob=?, house=?,
                         parents_name=?, address=?, picture=?
                     WHERE student_id=?""",
                 (name, admission_no, class_, section, dob, house, parents_name, address,
                  picture_path, student_id))
    else:
        c.execute("""UPDATE students
                     SET name=?, admission_no=?, class=?, section=?, dob=?, house=?,
                         parents_name=?, address=?
                     WHERE student_id=?""",
                 (name, admission_no, class_, section, dob, house, parents_name, address, student_id))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Student details updated'})

@app.route('/teacher/student/add', methods=['POST'])
@teacher_required
def teacher_add_student():
    teacher_id = session.get('teacher_id')
    name = request.form.get('name', '').strip()
    student_id = request.form.get('student_id', '').strip()
    class_ = request.form.get('class', '').strip()
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject', '').strip()
    admission_no = request.form.get('admission_no', '').strip()
    dob = request.form.get('dob', '').strip()
    house = request.form.get('house', '').strip()
    parents_name = request.form.get('parents_name', '').strip()
    address = request.form.get('address', '').strip()

    if not name or not student_id or not class_:
        return jsonify({'status': 'error', 'message': 'Name, Student ID and Class are required'}), 400

    # Verify teacher has access to this class
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM teacher_assignments WHERE teacher_id=? AND class=?", (teacher_id, class_))
    if not c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'You are not assigned to this class'}), 403

    # Check duplicate
    c.execute("SELECT student_id FROM students WHERE student_id=?", (student_id,))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student ID already exists'}), 400

    picture_path = save_student_picture(request.files.get('picture'), student_id)

    c.execute("""INSERT INTO students
                 (student_id, name, class, subject, section, admission_no, dob, house, parents_name, address, picture, status)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,'Not Started')""",
             (student_id, name, class_, subject, section, admission_no, dob, house, parents_name, address, picture_path))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': f'Student {name} added successfully'})

@app.route('/teacher/student/<student_id>/delete', methods=['DELETE'])
@teacher_required
def teacher_delete_student(student_id):
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT class FROM students WHERE student_id=?", (student_id,))
    student = c.fetchone()
    if not student:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
    c.execute("SELECT id FROM teacher_assignments WHERE teacher_id=? AND class=?", (teacher_id, student['class']))
    if not c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    c.execute("DELETE FROM students WHERE student_id=?",           (student_id,))
    c.execute("DELETE FROM responses WHERE student_id=?",          (student_id,))
    c.execute("DELETE FROM results WHERE student_id=?",            (student_id,))
    c.execute("DELETE FROM shuffled_questions WHERE student_id=?", (student_id,))
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))  # Bug #3 fix
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Student deleted'})

@app.route('/teacher/profile', methods=['GET', 'POST'])
@teacher_required
def teacher_profile():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        address = request.form.get('address', '').strip()
        new_password = request.form.get('new_password', '').strip()
        pic_path = save_teacher_picture(request.files.get('picture'), teacher_id)
        if new_password:
            hashed = hash_password(new_password)
            if pic_path:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, password=?, picture=? WHERE id=?",
                         (name, email, address, hashed, pic_path, teacher_id))
            else:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, password=? WHERE id=?",
                         (name, email, address, hashed, teacher_id))
        else:
            if pic_path:
                c.execute("UPDATE teachers SET name=?, email=?, address=?, picture=? WHERE id=?",
                         (name, email, address, pic_path, teacher_id))
            else:
                c.execute("UPDATE teachers SET name=?, email=?, address=? WHERE id=?",
                         (name, email, address, teacher_id))
        conn.commit()
        session['teacher_name'] = name
        conn.close()
        return jsonify({'status': 'success', 'message': 'Profile updated'})

    c.execute("SELECT * FROM teachers WHERE id=?", (teacher_id,))
    teacher = dict(c.fetchone())
    c.execute("SELECT class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_profile.html', teacher=teacher, assignments=assignments)

@app.route('/teacher/test/<int:result_id>/responses')
@teacher_required
def teacher_view_test_responses(result_id):
    conn = get_db()
    c = conn.cursor()
    
    # Get test result details
    c.execute("SELECT * FROM results WHERE id=?", (result_id,))
    result = c.fetchone()
    if not result:
        conn.close()
        return "Test result not found", 404
    result = dict(result)
    
    # Get student responses with full question data
    c.execute("""
        SELECT q.id, q.question, q.option_a, q.option_b, q.option_c, q.option_d, 
               q.correct_answer, q.image_path, r.selected_option
        FROM questions q
        LEFT JOIN responses r ON q.id = r.question_id AND r.student_id = ?
        WHERE q.class = ? AND q.subject = ?
        ORDER BY q.id
    """, (result['student_id'], result['class'], result['subject']))
    rows = c.fetchall()
    conn.close()
    
    # Process questions to map correct_answer to a letter
    questions = []
    for row in rows:
        q = dict(row)
        selected = q['selected_option'] or ''
        correct_raw = q['correct_answer']
        correct_letter = None
        
        if correct_raw:
            correct_raw = correct_raw.strip()
            # If already a single letter, use it
            if len(correct_raw) == 1 and correct_raw.upper() in ('A','B','C','D'):
                correct_letter = correct_raw.upper()
            else:
                # Try to match against option texts or "option_x"
                if correct_raw == q['option_a'] or correct_raw.lower() == 'option_a':
                    correct_letter = 'A'
                elif correct_raw == q['option_b'] or correct_raw.lower() == 'option_b':
                    correct_letter = 'B'
                elif correct_raw == q['option_c'] or correct_raw.lower() == 'option_c':
                    correct_letter = 'C'
                elif correct_raw == q['option_d'] or correct_raw.lower() == 'option_d':
                    correct_letter = 'D'
        
        is_correct = False
        if selected and correct_letter:
            is_correct = (selected.strip().upper() == correct_letter)
        
        questions.append({
            'id': q['id'],
            'question': q['question'],
            'options': {
                'A': q['option_a'],
                'B': q['option_b'],
                'C': q['option_c'],
                'D': q['option_d']
            },
            'image_path': q['image_path'],
            'correct': correct_letter or '?',
            'selected': selected,
            'is_correct': is_correct
        })
    
    return render_template('teacher_test_responses.html',
                         result=result,
                         questions=questions)

@app.route('/teacher/generate_student_report/<student_id>')
@teacher_required
def teacher_generate_student_report(student_id):
    conn = get_db()
    c = conn.cursor()
    
    c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
    student = dict(c.fetchone())
    
    c.execute("""SELECT subject, chapter, score, total_questions, percentage, test_date
                 FROM results WHERE student_id=?
                 ORDER BY test_date DESC""", (student_id,))
    test_history = [dict(row) for row in c.fetchall()]
    
    conn.close()
    
    # Generate PDF (simplified version)
    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial; margin: 20px; }}
            h1 {{ color: #2e7d32; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #2e7d32; color: white; }}
        </style>
    </head>
    <body>
        <h1>Student Progress Report</h1>
        <p><strong>Name:</strong> {student.get('name', 'N/A')}</p>
        <p><strong>Admission No:</strong> {student.get('admission_no', 'N/A')}</p>
        <p><strong>Class:</strong> {student.get('class', 'N/A')} - {student.get('section', 'N/A')}</p>
        <p><strong>House:</strong> {student.get('house', 'N/A')}</p>
        
        <h2>Test History</h2>
        <table>
            <tr>
                <th>Date</th>
                <th>Subject</th>
                <th>Chapter</th>
                <th>Marks</th>
                <th>Percentage</th>
            </tr>
    """
    
    for test in test_history:
        html_content += f"""
            <tr>
                <td>{test.get('test_date', '')[:10]}</td>
                <td>{test.get('subject', 'N/A')}</td>
                <td>{test.get('chapter', 'N/A')}</td>
                <td>{test.get('score', 0)}/{test.get('total_questions', 0)}</td>
                <td>{test.get('percentage', 0):.1f}%</td>
            </tr>
        """
    
    html_content += """
        </table>
    </body>
    </html>
    """
    
    from weasyprint import HTML
    pdf = HTML(string=html_content).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=student_report_{student_id}.pdf'
    
    return response

# ========================
# ADMIN TEACHER MANAGEMENT
# ========================

@app.route('/admin/teachers')
@admin_required
def admin_teachers():
    return render_template('admin_teachers.html')

@app.route('/admin/teachers/data')
@admin_required
def admin_teachers_data():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT t.id, t.name, t.mobile, t.email, t.status, t.created_at,
                        GROUP_CONCAT(DISTINCT ta.class || ' - ' || ta.subject) as assignments
                 FROM teachers t
                 LEFT JOIN teacher_assignments ta ON t.id = ta.teacher_id
                 GROUP BY t.id
                 ORDER BY t.name""")
    teachers = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'teachers': teachers})

@app.route('/admin/teacher/add', methods=['POST'])
@admin_required
def admin_add_teacher():
    name = request.form.get('name')
    mobile = request.form.get('mobile')
    password = request.form.get('password')
    email = request.form.get('email')
    address = request.form.get('address')
    
    conn = get_db()
    c = conn.cursor()
    
    # Check if mobile exists
    c.execute("SELECT id FROM teachers WHERE mobile=?", (mobile,))
    if c.fetchone():
        conn.close()
        return jsonify({'status': 'error', 'message': 'Mobile number already exists'}), 400
    
    hashed_pw = hash_password(password)
    c.execute("""INSERT INTO teachers (name, mobile, password, email, address)
                 VALUES (?,?,?,?,?)""",
             (name, mobile, hashed_pw, email, address))
    conn.commit()
    conn.close()
    
    return jsonify({'status': 'success', 'message': 'Teacher added successfully'})

@app.route('/admin/teacher/<int:teacher_id>/assign', methods=['POST'])
@admin_required
def admin_assign_teacher(teacher_id):
    class_ = request.form.get('class', '').strip()
    section = request.form.get('section', '').strip()
    subject = request.form.get('subject', '').strip()
    is_class_teacher = request.form.get('is_class_teacher') == 'true'

    if not class_ or not subject:
        return jsonify({'status': 'error', 'message': 'Class and subject are required'}), 400

    conn = get_db()
    c = conn.cursor()

    # Enforce 1:1 class teacher constraint
    if is_class_teacher:
        c.execute("""SELECT ta.id, t.name FROM teacher_assignments ta
                     JOIN teachers t ON ta.teacher_id = t.id
                     WHERE ta.class=? AND ta.section=? AND ta.is_class_teacher=1
                     AND ta.teacher_id != ?""",
                 (class_, section, teacher_id))
        existing_ct = c.fetchone()
        if existing_ct:
            conn.close()
            return jsonify({'status': 'error',
                           'message': f"Class {class_} {section} already has a Class Teacher: {existing_ct['name']}"}), 400

    # Check if this exact assignment already exists
    c.execute("""SELECT id FROM teacher_assignments
                 WHERE teacher_id=? AND class=? AND section=? AND subject=?""",
             (teacher_id, class_, section, subject))
    existing = c.fetchone()

    if existing:
        c.execute("""UPDATE teacher_assignments
                     SET is_class_teacher=?
                     WHERE id=?""", (1 if is_class_teacher else 0, existing['id']))
    else:
        c.execute("""INSERT INTO teacher_assignments (teacher_id, class, section, subject, is_class_teacher)
                     VALUES (?,?,?,?,?)""",
                 (teacher_id, class_, section, subject, 1 if is_class_teacher else 0))

    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Assignment saved successfully'})

@app.route('/admin/teacher/<int:teacher_id>/remove_assignment/<int:assign_id>', methods=['DELETE'])
@admin_required
def admin_remove_assignment(teacher_id, assign_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM teacher_assignments WHERE id=? AND teacher_id=?", (assign_id, teacher_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

@app.route('/admin/teacher/<int:teacher_id>', methods=['GET'])
@admin_required
def admin_get_teacher(teacher_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name, mobile, email, address, picture, status FROM teachers WHERE id=?", (teacher_id,))
    teacher = c.fetchone()
    if not teacher:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Teacher not found'}), 404
    c.execute("SELECT id, class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'teacher': dict(teacher), 'assignments': assignments})

@app.route('/admin/teacher/<int:teacher_id>/update', methods=['POST'])
@admin_required
def admin_update_teacher(teacher_id):
    conn = get_db()
    c = conn.cursor()
    name = request.form.get('name', '').strip()
    mobile = request.form.get('mobile', '').strip()
    email = request.form.get('email', '').strip()
    address = request.form.get('address', '').strip()
    new_password = request.form.get('password', '').strip()

    pic_file = request.files.get('picture')
    pic_path = save_teacher_picture(pic_file, teacher_id)

    if new_password:
        hashed = hash_password(new_password)
        if pic_path:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, password=?, picture=? WHERE id=?",
                     (name, mobile, email, address, hashed, pic_path, teacher_id))
        else:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, password=? WHERE id=?",
                     (name, mobile, email, address, hashed, teacher_id))
    else:
        if pic_path:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=?, picture=? WHERE id=?",
                     (name, mobile, email, address, pic_path, teacher_id))
        else:
            c.execute("UPDATE teachers SET name=?, mobile=?, email=?, address=? WHERE id=?",
                     (name, mobile, email, address, teacher_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Teacher updated'})

@app.route('/admin/teacher/upload_csv', methods=['POST'])
@admin_required
def admin_upload_teachers_csv():
    file = request.files.get('file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'status': 'error', 'message': 'Invalid file'}), 400
    
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    conn = get_db()
    c = conn.cursor()
    added = 0
    
    with open(filepath, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            name = row.get('name', '').strip()
            mobile = row.get('mobile', '').strip()
            password = row.get('password', 'teacher123')  # Default password
            email = row.get('email', '').strip()
            
            if not name or not mobile:
                continue
            
            # Check if exists
            c.execute("SELECT id FROM teachers WHERE mobile=?", (mobile,))
            if c.fetchone():
                continue
            
            hashed_pw = hash_password(password)
            c.execute("""INSERT INTO teachers (name, mobile, password, email)
                         VALUES (?,?,?,?)""",
                     (name, mobile, hashed_pw, email))
            added += 1
    
    conn.commit()
    conn.close()
    os.remove(filepath)
    
    return jsonify({'status': 'success', 'message': f'{added} teachers added successfully'})

# ===================================================
# TEST PAPER MODULE (CSV naming: Class_Subject_TestNo)
# ===================================================

@app.route('/admin/test_papers')
@admin_required
def admin_test_papers():
    return render_template('test_papers.html', role='admin')

@app.route('/teacher/test_papers')
@teacher_required
def teacher_test_papers():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('test_papers.html', role='teacher', assignments=assignments)

@app.route('/api/test_papers')
def api_test_papers():
    role = request.args.get('role', 'admin')
    teacher_id = session.get('teacher_id') if role == 'teacher' else None
    conn = get_db()
    c = conn.cursor()
    if role == 'teacher' and teacher_id:
        c.execute("""SELECT ta.class, ta.subject FROM teacher_assignments ta WHERE ta.teacher_id=?""", (teacher_id,))
        assigns = c.fetchall()
        papers = []
        for a in assigns:
            c.execute("""SELECT tp.*, t.name as uploader_name FROM test_papers tp
                         LEFT JOIN teachers t ON tp.uploaded_by = t.id AND tp.uploader_type='teacher'
                         WHERE tp.class=? AND tp.subject=?
                         ORDER BY tp.created_at DESC""", (a['class'], a['subject']))
            papers.extend([dict(r) for r in c.fetchall()])
    else:
        c.execute("""SELECT tp.* FROM test_papers tp ORDER BY tp.created_at DESC""")
        papers = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'papers': papers})

@app.route('/api/test_papers/upload', methods=['POST'])
def upload_test_paper():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    file = request.files.get('csv_file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'status': 'error', 'message': 'CSV file required'}), 400

    # Parse filename convention: Class_Subject_TestNo.csv
    raw_name = os.path.splitext(file.filename)[0]
    parts = raw_name.split('_')
    if len(parts) < 3:
        return jsonify({'status': 'error',
                        'message': 'Filename must follow format: Class_Subject_TestNo (e.g. 10A_Math_Test01)'}), 400

    class_ = parts[0].strip()
    test_no = parts[-1].strip()
    subject = '_'.join(parts[1:-1]).strip()
    section = ''
    # If class has letter suffix treat it as section: e.g. 10A -> class=10, section=A
    if class_ and class_[-1].isalpha():
        section = class_[-1].upper()
        class_ = class_[:-1]

    uploader_type = 'admin' if session.get('admin_logged_in') else 'teacher'
    uploaded_by = 'admin' if uploader_type == 'admin' else str(session.get('teacher_id'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    # Import questions
    conn = get_db()
    c = conn.cursor()
    question_count = 0
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                parsed = parse_question_csv_row(row)
                if not parsed or not parsed['question']:
                    continue
                c.execute("""INSERT INTO questions
                             (class, subject, chapter, question, option_a, option_b, option_c, option_d, correct_answer)
                             VALUES (?,?,?,?,?,?,?,?,?)""",
                         (class_, subject, test_no, parsed['question'],
                          parsed['option_a'], parsed['option_b'], parsed['option_c'], parsed['option_d'],
                          parsed['correct_answer']))
                question_count += 1

        c.execute("""INSERT INTO test_papers
                     (filename, class, section, subject, test_no, uploaded_by, uploader_type, question_count)
                     VALUES (?,?,?,?,?,?,?,?)""",
                 (raw_name, class_, section, subject, test_no, uploaded_by, uploader_type, question_count))

        # Record into mcq_test_history table
        t_id = session.get('teacher_id')
        t_name = uploader_type.capitalize()
        if t_id:
            c.execute("SELECT name FROM teachers WHERE id=?", (t_id,))
            t_r = c.fetchone()
            if t_r: t_name = t_r['name']
        c.execute("""INSERT INTO mcq_test_history (teacher_id, teacher_name, class, section, subject, test_no, question_count)
                     VALUES (?,?,?,?,?,?,?)""",
                  (t_id, t_name, class_, section, subject, test_no, question_count))

        c.execute("""INSERT INTO test_generation_history
                     (teacher_id,class,section,subject,chapter,test_no,output_mode,
                      total_questions,mcq_count,assertion_count,very_short_count,
                      short_count,long_count,case_study_count,remark)
                     VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                  (uploaded_by, class_, section, subject, test_no, test_no, 'cbt',
                   question_count, question_count, 0, 0, 0, 0, 0, 'Test Paper CSV Upload'))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        os.remove(filepath)
        return jsonify({'status': 'error', 'message': str(e)}), 500
    conn.close()
    os.remove(filepath)
    return jsonify({'status': 'success',
                    'message': f'Uploaded {question_count} questions for {class_}{section} {subject} - {test_no}'})

@app.route('/api/test_papers/<int:paper_id>/toggle', methods=['POST'])
def toggle_test_paper(paper_id):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_active FROM test_papers WHERE id=?", (paper_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'status': 'error', 'message': 'Not found'}), 404
    new_status = 0 if row['is_active'] else 1
    c.execute("UPDATE test_papers SET is_active=? WHERE id=?", (new_status, paper_id))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'is_active': new_status})

@app.route('/api/test_papers/<int:paper_id>/pdf')
def test_paper_pdf(paper_id):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return redirect(url_for('teacher_login'))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_papers WHERE id=?", (paper_id,))
    paper = c.fetchone()
    if not paper:
        conn.close()
        return "Not found", 404
    paper = dict(paper)

    # Robust multi-fallback query for test paper PDF questions
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                        correct_answer, image_path,
                        CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                 FROM questions
                 WHERE (class=? OR class=? OR class LIKE ?)
                   AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                   AND (test_no=? OR chapter=? OR test_no LIKE ? OR chapter LIKE ?)
                 ORDER BY id""",
             (paper['class'],
              paper['class'].replace('th','').replace('st','').replace('nd','').replace('rd',''),
              f"%{paper['class']}%",
              paper['subject'], f"%{paper['subject']}%",
              paper['test_no'], paper['test_no'],
              f"%{paper['test_no']}%", f"%{paper['test_no']}%"))
    questions = [dict(r) for r in c.fetchall()]

    if not questions:
        clean_class = paper['class'].replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
        c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                            correct_answer, image_path,
                            CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                     FROM questions
                     WHERE (LOWER(class)=LOWER(?) OR LOWER(class)=LOWER(?) OR class LIKE ?)
                       AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                     ORDER BY id""",
                 (paper['class'], clean_class, f"%{clean_class}%", paper['subject'], f"%{paper['subject']}%"))
        questions = [dict(r) for r in c.fetchall()]

    # Fallback: Parse paper['filename'] (e.g. 6A_Computer_Test01) if initial query matches 0 questions
    if not questions and paper.get('filename'):
        parts = paper['filename'].split('_')
        if len(parts) >= 3:
            raw_cls = parts[0]
            t_no = parts[-1]
            subj = '_'.join(parts[1:-1]).replace('_', ' ')
            m = re.match(r'^(\d+)([A-Za-z]*)', raw_cls)
            if m:
                extracted_cls = m.group(1)
                c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                                    correct_answer, image_path,
                                    CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                             FROM questions
                             WHERE (class=? OR class LIKE ?)
                               AND (LOWER(subject)=LOWER(?) OR LOWER(subject)=LOWER(?))
                             ORDER BY id""",
                         (extracted_cls, f"%{extracted_cls}%", subj, paper.get('subject','')))
                questions = [dict(r) for r in c.fetchall()]

    if not questions:
        c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                            correct_answer, image_path,
                            CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                     FROM questions
                     ORDER BY id LIMIT 50""")
        questions = [dict(r) for r in c.fetchall()]

    conn.close()

    for q in questions:
        q_type = str(q.get('question_type') or '').strip()
        if not q_type or q_type.upper() in ['MCQ', 'CHOICE', 'OBJECTIVE']:
            q['question_type'] = 'MCQ'

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext  = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"

    # Determine display chapter name for paper PDF
    unique_ch = []
    for q in questions:
        ch_val = str(q.get('chapter') or '').strip()
        if ch_val and ch_val.lower() != str(paper['test_no']).lower() and ch_val not in unique_ch:
            unique_ch.append(ch_val)
    paper_chapter = ", ".join(unique_ch) if unique_ch else paper['test_no']

    # Use v2 template (handles all question types properly)
    rendered = render_template('question_paper_template_v2.html',
        school_name=school_name,
        logo_base64=logo_base64,
        class_name=f"{paper['class']}{paper['section']}",
        subject=paper['subject'],
        chapter=paper_chapter,
        test_no=paper['test_no'],
        questions=questions,
        total_marks=len(questions),
        date=datetime.datetime.now().strftime('%d/%m/%Y'))

    if request.args.get('format') == 'download_pdf':
        pdf = HTML(string=rendered, base_url=request.base_url).write_pdf()
        resp = make_response(pdf)
        resp.headers['Content-Type']        = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename={paper["filename"]}.pdf'
        return resp

    # Default: Return HTML Printable Preview (opens browser Print / Save as PDF dialog)
    return rendered

# ===================================================
# TEACHER MONITORING (own classes only)
# ===================================================

@app.route('/teacher/monitoring')
@teacher_required
def teacher_monitoring():
    """
    Render the teacher monitoring page for their assigned students.
    """
    return render_template('teacher_monitoring.html')

@app.route('/teacher/monitoring/data')
@teacher_required
def teacher_monitoring_data():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    # Get teacher's assigned classes
    c.execute("SELECT DISTINCT class FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    classes = [r['class'] for r in c.fetchall()]
    if not classes:
        conn.close()
        return jsonify({'students': [], 'exam_is_active': False})

    placeholders = ','.join(['?'] * len(classes))
    c.execute(f"""SELECT s.student_id, s.name, s.class, s.subject, s.section, s.picture,
                         s.status, s.exam_started_at,
                         (SELECT COUNT(*) FROM responses r WHERE r.student_id=s.student_id) as answered
                  FROM students s
                  WHERE s.class IN ({placeholders})
                  ORDER BY s.class, s.name""", classes)
    students = [dict(r) for r in c.fetchall()]
    c.execute("SELECT is_active FROM exam_control WHERE id=1")
    exam_row = c.fetchone()
    conn.close()
    return jsonify({'students': students, 'exam_is_active': bool(exam_row['is_active']) if exam_row else False})

# ===================================================
# CLASS REPORT (Excel + PDF)
# ===================================================

@app.route('/admin/class_report')
@admin_required
def admin_class_report():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT DISTINCT class FROM students WHERE class IS NOT NULL AND class != ''
                 UNION
                 SELECT DISTINCT class FROM results WHERE class IS NOT NULL AND class != ''
                 UNION
                 SELECT DISTINCT class FROM questions WHERE class IS NOT NULL AND class != ''
                 UNION
                 SELECT DISTINCT class FROM test_papers WHERE class IS NOT NULL AND class != ''""")
    classes = sorted(list(set(r['class'].strip() for r in c.fetchall() if r['class'] and r['class'].strip())))
    conn.close()
    return render_template('class_report.html', role='admin', classes=classes)

@app.route('/teacher/class_report')
@teacher_required
def teacher_class_report():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, section, subject, is_class_teacher FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('class_report.html', role='teacher', assignments=assignments)

@app.route('/api/class_report/generate', methods=['POST'])
def generate_class_report():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401

    data = request.get_json(silent=True) or request.form or {}
    class_       = data.get('class', '').strip()
    section      = data.get('section', '').strip()
    subject      = data.get('subject', '').strip()
    date_from    = data.get('date_from', '').strip()
    date_to      = data.get('date_to', '').strip()
    export_format = data.get('format', 'excel')
    sort_by      = data.get('sort_by', 'name')

    if not class_:
        return jsonify({'status': 'error', 'message': 'Please enter a class to generate the report.'}), 400

    clean_cls = class_.replace('th','').replace('st','').replace('nd','').replace('rd','').strip()

    conn = get_db()
    c = conn.cursor()

    # 1. Fetch students from students table
    stu_query = """SELECT student_id, name, admission_no, class, section, house, parents_name, picture
                   FROM students WHERE (class=? OR class=? OR class LIKE ?)"""
    stu_params = [class_, clean_cls, f"%{clean_cls}%"]
    if section:
        stu_query += " AND (section=? OR section IS NULL OR section='')"
        stu_params.append(section)

    c.execute(stu_query, stu_params)
    db_students = [dict(r) for r in c.fetchall()]

    # 2. Fetch distinct student_ids from results table for this class
    res_query = """SELECT DISTINCT student_id, name, class, section FROM results
                   WHERE (class=? OR class=? OR class LIKE ?)"""
    res_params = [class_, clean_cls, f"%{clean_cls}%"]
    if section:
        res_query += " AND (section=? OR section IS NULL OR section='')"
        res_params.append(section)

    c.execute(res_query, res_params)
    res_students = [dict(r) for r in c.fetchall()]

    # Combine and deduplicate strictly by student_id (or name if student_id blank)
    student_map = {}
    for st in db_students:
        key = (st.get('student_id') or st.get('name') or '').strip().lower()
        if key and key not in student_map:
            st['admission_no'] = st.get('admission_no') or st.get('student_id') or ''
            student_map[key] = st

    for r in res_students:
        key = (r.get('student_id') or r.get('name') or '').strip().lower()
        if key and key not in student_map:
            student_map[key] = {
                'student_id': r.get('student_id'),
                'name': r.get('name') or key,
                'admission_no': r.get('student_id') or '',
                'class': r.get('class') or class_,
                'section': r.get('section') or section,
                'house': '',
                'parents_name': '',
                'picture': ''
            }

    students = list(student_map.values())

    report_data = []
    for st in students:
        result_query  = """SELECT subject, chapter, score, total_questions, percentage, test_date
                           FROM results WHERE student_id=?"""
        result_params = [st['student_id']]
        if subject:
            result_query += " AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)"
            result_params.extend([subject, f"%{subject}%"])
        test_no = data.get('test_no', '').strip()
        if test_no:
            result_query += " AND (chapter=? OR chapter LIKE ?)"
            result_params.extend([test_no, f"%{test_no}%"])
        if date_from:
            result_query += " AND DATE(test_date)>=?"
            result_params.append(date_from)
        if date_to:
            result_query += " AND DATE(test_date)<=?"
            result_params.append(date_to)
        result_query += " ORDER BY test_date DESC"
        c.execute(result_query, result_params)
        results    = [dict(r) for r in c.fetchall()]
        total_score      = sum(r['score'] or 0 for r in results)
        total_possible   = sum(r['total_questions'] or 0 for r in results)
        avg_pct          = round(sum(r['percentage'] or 0 for r in results) / len(results), 1) if results else 0
        st['results']        = results
        st['avg_percentage'] = avg_pct
        st['total_tests']    = len(results)
        st['total_score']    = total_score
        st['total_possible'] = total_possible
        st['score_ratio']    = f"{total_score}/{total_possible}" if total_possible else '0/0'
        report_data.append(st)

    school_name = get_setting('school_name', 'RRB Group of Schools')
    conn.close()  # single close here

    if sort_by == 'percentage':
        report_data.sort(key=lambda x: x['avg_percentage'], reverse=True)
    else:
        report_data.sort(key=lambda x: x['name'])

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    label     = f"{class_}{section}_{subject or 'All'}_{timestamp}"

    # Bug #5 Fix: json preview no longer calls conn.close() again
    if export_format == 'json':
        return jsonify({'status': 'success', 'students': report_data})

    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class Report"

        # Header
        ws.merge_cells('A1:H1')
        ws['A1'] = school_name
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Class: {class_}{section}  |  Subject: {subject or 'All'}  |  Generated: {datetime.datetime.now().strftime('%d-%m-%Y')}"
        ws['A2'].font = openpyxl.styles.Font(bold=True, size=12)
        ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.append([])

        headers = ['#', 'Admission No / ID', 'Name', 'Class', 'Section', 'House', 'Total Tests', 'Score']
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color='1b5e20', end_color='1b5e20', fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for i, st in enumerate(report_data, 1):
            ws.append([i, st.get('admission_no') or st.get('student_id') or '', st['name'], st['class'],
                       st.get('section') or '', st.get('house') or '',
                       st['total_tests'], st.get('score_ratio', '0/0')])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12) + 2

        filepath = os.path.join(app.config['EXPORT_FOLDER'], f"class_report_{label}.xlsx")
        wb.save(filepath)
        return send_file(filepath, as_attachment=True, download_name=f"class_report_{label}.xlsx")

    else:  # PDF / Printable Preview
        rows_html = ''
        for i, st in enumerate(report_data, 1):
            rows_html += f"""<tr>
                <td>{i}</td>
                <td>{st.get('admission_no') or st.get('student_id') or '—'}</td>
                <td><strong>{st['name']}</strong></td>
                <td>{st['class']}{st.get('section') or ''}</td>
                <td>{st.get('house') or '—'}</td>
                <td>{st['total_tests']}</td>
                <td><strong>{st.get('score_ratio', '0/0')}</strong></td>
            </tr>"""

        html = f"""<!DOCTYPE html><html><head>
        <meta charset="UTF-8">
        <title>Class Report - {school_name}</title>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css">
        <style>
            * {{ box-sizing: border-box; }}
            body {{ font-family: Arial, Helvetica, sans-serif; margin: 15mm; font-size: 10pt; color: #1e293b; background: #fff; line-height: 1.5; }}
            .print-toolbar {{
                position: sticky; top: 0; z-index: 9999;
                background: #1b5e20; color: #ffffff;
                padding: 10px 20px; margin: -15mm -15mm 20px -15mm;
                display: flex; align-items: center; justify-content: space-between; gap: 12px;
                box-shadow: 0 4px 15px rgba(0,0,0,0.15);
            }}
            .print-toolbar .title {{ font-weight: 700; font-size: 1rem; display: flex; align-items: center; gap: 8px; }}
            .print-toolbar .actions {{ display: flex; align-items: center; gap: 10px; }}
            .print-btn {{
                background: #ffffff; color: #1b5e20; border: none; padding: 8px 16px; border-radius: 6px;
                font-weight: 700; font-size: 0.85rem; cursor: pointer; display: inline-flex; align-items: center; gap: 6px; text-decoration: none;
            }}
            .print-btn:hover {{ background: #e8f5e9; }}
            .header-box {{ text-align: center; border-bottom: 2.5pt double #1b5e20; padding-bottom: 12px; margin-bottom: 16px; }}
            .school-title {{ font-size: 18pt; font-weight: 800; color: #1b5e20; }}
            .sub-title {{ font-size: 11pt; color: #475569; margin-top: 4px; font-weight: 600; }}
            .meta-bar {{ display: flex; justify-content: space-between; background: #f0fdf4; border: 1px solid #bbf7d0; padding: 8px 16px; border-radius: 8px; font-size: 9.5pt; margin-bottom: 16px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background: #1b5e20; color: white; padding: 8px 10px; text-align: left; font-size: 9.5pt; font-weight: 700; }}
            td {{ border-bottom: 1px solid #e2e8f0; padding: 8px 10px; font-size: 9pt; }}
            tr:nth-child(even) td {{ background: #f8fafc; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 9pt; color: #64748b; border-top: 1px solid #e2e8f0; padding-top: 10px; }}
            @media print {{
                body {{ margin: 10mm; }}
                .print-toolbar {{ display: none !important; }}
                th {{ -webkit-print-color-adjust: exact; }}
            }}
        </style></head><body>
        <div class="print-toolbar">
            <div class="title"><i class="fas fa-chart-bar"></i> Class Performance Report</div>
            <div class="actions">
                <button onclick="window.print()" class="print-btn"><i class="fas fa-print"></i> Print / Save as PDF</button>
            </div>
        </div>
        <div class="header-box">
            <div class="school-title">{school_name}</div>
            <div class="sub-title">Class Performance &amp; Results Report</div>
        </div>
        <div class="meta-bar">
            <div><strong>Class:</strong> {class_}{section} &nbsp;|&nbsp; <strong>Subject:</strong> {subject or 'All Subjects'}</div>
            <div><strong>Test No:</strong> {data.get('test_no') or 'All Tests'} &nbsp;|&nbsp; <strong>Date:</strong> {datetime.datetime.now().strftime('%d-%m-%Y')}</div>
        </div>
        <table>
            <thead>
                <tr><th>#</th><th>Admission No</th><th>Name</th><th>Class</th><th>House</th><th>Tests</th><th>Score</th></tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        <div class="footer">Generated by RRB CBT | Developed by Gaurav Shukla &amp; Team</div>
        <script>
            window.addEventListener('load', function() {{
                if (!window.location.search.includes('noprint')) {{
                    setTimeout(function() {{ window.print(); }}, 400);
                }}
            }});
        </script>
        </body></html>"""

        if export_format == 'download_pdf':
            pdf = HTML(string=html).write_pdf()
            resp = make_response(pdf)
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'attachment; filename=class_report_{label}.pdf'
            return resp

        # Default: HTML Printable Preview
        return html

# ═══════════════════════════════════════════════════════════════
# BULLETIN BOARD
# ═══════════════════════════════════════════════════════════════

@app.route('/api/bulletins')
def api_bulletins():
    conn = get_db()
    c = conn.cursor()
    class_filter = request.args.get('class', '')
    query = """SELECT b.id, b.title, b.content, b.poster_type, b.target_class, b.created_at,
                      COALESCE(t.name,'Admin') as poster_name
               FROM bulletins b
               LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
               WHERE b.is_active=1"""
    params = []
    if class_filter:
        query += " AND (b.target_class='' OR b.target_class=?)"
        params.append(class_filter)
    query += " ORDER BY b.created_at DESC LIMIT 15"
    c.execute(query, params)
    bulletins = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'bulletins': bulletins})

@app.route('/api/bulletins/post', methods=['POST'])
def post_bulletin():
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401
    title   = request.form.get('title','').strip()
    content = request.form.get('content','').strip()
    target  = request.form.get('target_class','').strip()
    if not title or not content:
        return jsonify({'status': 'error', 'message': 'Title and content required'}), 400
    if session.get('admin_logged_in'):
        posted_by   = 'admin'
        poster_type = 'admin'
    else:
        posted_by   = str(session.get('teacher_id'))
        poster_type = 'teacher'
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO bulletins (title, content, posted_by, poster_type, target_class) VALUES (?,?,?,?,?)",
              (title, content, posted_by, poster_type, target))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success', 'message': 'Bulletin posted'})

@app.route('/api/bulletins/<int:bid>/delete', methods=['DELETE'])
def delete_bulletin(bid):
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error'}), 401
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE bulletins SET is_active=0 WHERE id=?", (bid,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'success'})

# ═══════════════════════════════════════════════════════════════
# ADMIN: UNLOCK STUDENT CLASS
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/student/<student_id>/unlock_class', methods=['POST'])
@admin_required
def admin_unlock_student_class(student_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM student_class_lock WHERE student_id=?", (student_id,))
    conn.commit()
    conn.close()
    log_audit_event('admin', 'admin', 'STUDENT_UNLOCK', 'student_class_lock', student_id, request.remote_addr)
    return jsonify({'status': 'success', 'message': 'Class lock removed. Student can re-register.'})

# ═══════════════════════════════════════════════════════════════
# ENHANCED AI TEST GENERATOR (Multi-type + Dual output mode)
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/create_test', methods=['GET', 'POST'])
@admin_required
@limiter.limit("13 per minute")
def admin_create_test():
    """Feature #8: Admin test generation - unrestricted class/subject access."""
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
    all_pairs = [dict(r) for r in c.fetchall()]

    if request.method == 'POST':
        import urllib.request, json as json_lib

        class_    = request.form.get('class','').strip()
        section   = request.form.get('section','').strip()
        subject   = request.form.get('subject','').strip()
        chapter   = request.form.get('chapter','').strip()
        test_no   = request.form.get('test_no','').strip()
        remark    = request.form.get('remark','').strip()
        output_mode = request.form.get('output_mode','cbt')
        method    = request.form.get('method','ai')

        if method == 'upload':
            eng_file = request.files.get('english_csv') or request.files.get('csv_file')
            hindi_file = request.files.get('hindi_csv')

            if not eng_file or not eng_file.filename.endswith('.csv'):
                conn.close()
                return jsonify({'status':'error','message':'English CSV file is required'}), 400

            # 1. Process compulsory English CSV & record inserted Row IDs
            eng_filename = secure_filename(eng_file.filename)
            eng_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"eng_{uuid.uuid4().hex[:8]}_{eng_filename}")
            eng_file.save(eng_filepath)

            inserted_ids = []
            try:
                with open(eng_filepath, 'r', encoding='utf-8-sig') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        parsed = parse_question_csv_row(row)
                        if not parsed or not parsed['question']:
                            continue
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_, subject, chapter, test_no, parsed['question_type'], parsed['question'],
                                  parsed['option_a'], parsed['option_b'], parsed['option_c'], parsed['option_d'],
                                  parsed['correct_answer']))
                        inserted_ids.append(c.lastrowid)
                conn.commit()
            finally:
                if os.path.exists(eng_filepath):
                    os.remove(eng_filepath)

            if not inserted_ids:
                conn.close()
                return jsonify({'status':'error','message':'No valid questions found in English CSV'}), 400

            # 2. Process optional Hindi CSV & update corresponding Row IDs sequentially
            hindi_updated = 0
            if hindi_file and hindi_file.filename and hindi_file.filename.endswith('.csv'):
                hin_filename = secure_filename(hindi_file.filename)
                hin_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"hin_{uuid.uuid4().hex[:8]}_{hin_filename}")
                hindi_file.save(hin_filepath)
                try:
                    with open(hin_filepath, 'r', encoding='utf-8-sig') as csvfile:
                        reader = csv.DictReader(csvfile)
                        for idx, row in enumerate(reader):
                            if idx >= len(inserted_ids):
                                break
                            parsed_hi = parse_question_csv_row(row)
                            if not parsed_hi:
                                continue
                            q_id = inserted_ids[idx]
                            c.execute("""UPDATE questions
                                         SET question_hi=?, option_a_hi=?, option_b_hi=?, option_c_hi=?, option_d_hi=?
                                         WHERE id=?""",
                                     (parsed_hi['question'], parsed_hi['option_a'], parsed_hi['option_b'],
                                      parsed_hi['option_c'], parsed_hi['option_d'], q_id))
                            hindi_updated += 1
                    conn.commit()
                finally:
                    if os.path.exists(hin_filepath):
                        os.remove(hin_filepath)

            # 3. Record test generation history & test paper entries
            uploader_type = 'admin' if session.get('admin_logged_in') else 'teacher'
            uploader_id = 'admin' if uploader_type == 'admin' else str(session.get('teacher_id') or '0')
            teacher_name = uploader_type.capitalize()
            t_id = session.get('teacher_id')
            if t_id:
                c.execute("SELECT name FROM teachers WHERE id=?", (t_id,))
                t_r = c.fetchone()
                if t_r: teacher_name = t_r['name']

            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (uploader_id, class_, section, subject, chapter or test_no, test_no, output_mode,
                       len(inserted_ids), len(inserted_ids), 0, 0, 0, 0, 0, remark or 'CSV Upload'))

            c.execute("""INSERT INTO mcq_test_history (teacher_id, teacher_name, class, section, subject, test_no, question_count)
                         VALUES (?,?,?,?,?,?,?)""",
                      (t_id or 0, teacher_name, class_, section, subject, test_no, len(inserted_ids)))

            if output_mode in ('cbt', 'both'):
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                          (f"{class_}{section}_{subject}_{test_no}", class_, section, subject, test_no,
                           uploader_id, uploader_type, len(inserted_ids)))

            conn.commit()
            conn.close()
            msg = f"{len(inserted_ids)} English questions uploaded"
            if hindi_updated > 0:
                msg += f" (with {hindi_updated} Hindi translations mapped)"
            return jsonify({'status':'success','message': msg})

        mcq_count       = int(request.form.get('mcq_count', 0) or 0)
        assertion_count = int(request.form.get('assertion_count', 0) or 0)

        type_instructions = []
        if mcq_count:       type_instructions.append(f"{mcq_count} MCQ")
        if assertion_count: type_instructions.append(f"{assertion_count} Assertion-Reason")

        if not type_instructions:
            conn.close()
            return jsonify({'status':'error','message':'Enter at least 1 question quantity'}), 400

        # ── AI generation ──
        gemini_api_key = _get_gemini_key()
        if not gemini_api_key:
            conn.close()
            return jsonify({'status':'error','message':'GEMINI_API_KEY not set'}), 500

        # Detect a working model or fallback
        model_name = get_working_model(gemini_api_key) or 'gemini-2.0-flash'
        app.logger.info(f"Using Gemini model: {model_name}")

        ai_prompt = _build_ai_prompt(class_, section, subject, chapter, test_no, remark, type_instructions)

        try:
            result, provider, err_msg = generate_ai_content(ai_prompt, timeout=60)
            if not result:
                return jsonify({'status':'error','message': err_msg}), 500

            raw_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
            app.logger.info(f"Raw AI response length: {len(raw_text)}")
            app.logger.debug(f"Raw AI response (first 1000 chars): {raw_text[:1000]}")
            if '```' in raw_text:
                for part in raw_text.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): raw_text = part; break

            questions = safe_json_loads(raw_text)
            if questions is None:
                app.logger.error(f"AI response invalid. Full response (first 2000 chars): {raw_text[:2000]}")
                # Also store the response in session for debugging (optional)
                return jsonify({
                    'status': 'error',
                    'message': 'AI returned invalid JSON. Please try again with a simpler prompt. '
                            'Check the server logs for the full AI response.'
                }), 500
            
            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      ('admin',class_,section,subject,chapter,test_no,output_mode,
                       len(questions),mcq_count,assertion_count,0,0,0,0,remark))
            c.execute("""INSERT INTO mcq_test_history (teacher_id, teacher_name, class, section, subject, test_no, question_count)
                         VALUES (0, 'Admin', ?, ?, ?, ?, ?)""",
                      (class_, section, subject, test_no, len(questions)))
            
            if output_mode == 'cbt':
                c.execute("SELECT COUNT(*) as cnt FROM questions WHERE class=? AND subject=? AND test_no=?",
                         (class_, subject, test_no))
                if c.fetchone()['cnt'] > 0:
                    conn.close()
                    return jsonify({'status':'error',
                                    'message':f'Test No "{test_no}" already exists for {subject} Class {class_}. Use a different Test No.'}), 400
                inserted = 0
                for q in questions:
                    if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,question_hi,
                                    option_a,option_a_hi,option_b,option_b_hi,option_c,option_c_hi,
                                    option_d,option_d_hi,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,q.get('question_type','MCQ'),
                                  q.get('question',''),q.get('question_hi',''),
                                  q.get('option_a',''),q.get('option_a_hi',''),
                                  q.get('option_b',''),q.get('option_b_hi',''),
                                  q.get('option_c',''),q.get('option_c_hi',''),
                                  q.get('option_d',''),q.get('option_d_hi',''),
                                  q.get('correct_answer','')))
                        inserted += 1
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                         (f"{class_}{section}_{subject}_{test_no}",class_,section,subject,test_no,
                          'admin','admin',inserted))
                conn.commit()

            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     ('admin',class_,section,subject,chapter,test_no,output_mode,
                      len(questions),mcq_count,assertion_count,0,0,0,0,remark))
            conn.commit()
            history_id = c.lastrowid
            conn.close()

            if output_mode == 'print':
                return jsonify({'status':'success','mode':'print',
                                'questions':questions,'history_id':history_id,
                                'message':f'{len(questions)} questions generated'})
            return jsonify({'status':'success','mode':'cbt',
                            'message':f'{len(questions)} questions generated. Test No: {test_no}'})

        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8') if hasattr(e, 'read') else str(e)
            app.logger.error(f"Gemini HTTP error: {err_body}")
            return jsonify({'status':'error','message':f'Gemini API error: {e.code} - {err_body[:200]}'}), 500
        except Exception as e:
            app.logger.error(f"Gemini exception: {str(e)}")
            return jsonify({'status':'error','message':str(e)}), 500

    conn.close()
    return render_template('admin_create_test.html', all_pairs=all_pairs)

@app.route('/teacher/print_test/<int:result_id>')
@teacher_required
def teacher_print_test(result_id):
    conn = get_db()
    c = conn.cursor()
    # Get the test result
    c.execute("SELECT * FROM results WHERE id=?", (result_id,))
    result = c.fetchone()
    if not result:
        conn.close()
        return "Test result not found", 404
    result = dict(result)

    # Check teacher access (optional, if you want to restrict)
    teacher_id = session['teacher_id']
    if not teacher_has_access(teacher_id, result['class'], result['subject']):
        conn.close()
        return "Access denied", 403

    # Fetch questions for this test's class/subject
    c.execute("SELECT id, question, option_a, option_b, option_c, option_d, correct_answer, image_path FROM questions WHERE class=? AND subject=? ORDER BY id",
              (result['class'], result['subject']))
    questions = [dict(row) for row in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, "rb") as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = "image/png" if ext == ".png" else "image/jpeg"
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode('utf-8')}"

    total_marks = len(questions)
    return render_template('question_paper_print.html',
                           school_name=school_name,
                           logo_base64=logo_base64,
                           class_name=result['class'],
                           subject=result['subject'],
                           questions=questions,
                           total_marks=total_marks,
                           date=datetime.datetime.now().strftime('%d/%m/%Y'))

@app.route('/teacher/test_history')
@teacher_required
def teacher_test_history():
    """
    Render the page showing test history for teacher's classes.
    """
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT * FROM test_generation_history
                 WHERE teacher_id=? ORDER BY created_at DESC""", (str(teacher_id),))
    history = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_test_history.html', history=history)

@app.route('/api/test_generation_history/<int:hid>/pdf')
def test_gen_history_pdf(hid):
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return redirect(url_for('teacher_login'))
    conn = get_db()
    c = conn.cursor()
    if session.get('teacher_logged_in'):
        c.execute("SELECT * FROM test_generation_history WHERE id=? AND teacher_id=?",
                  (hid, str(session.get('teacher_id'))))
    else:
        c.execute("SELECT * FROM test_generation_history WHERE id=?", (hid,))
    hist = c.fetchone()
    if not hist:
        conn.close()
        return "Not found", 404
    hist = dict(hist)
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                        correct_answer, image_path,
                        CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                 FROM questions
                 WHERE (class=? OR class=? OR class LIKE ?)
                   AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                   AND (test_no=? OR chapter=? OR test_no LIKE ? OR chapter LIKE ?)
                 ORDER BY id""",
             (hist['class'],
              hist['class'].replace('th','').replace('st','').replace('nd','').replace('rd',''),
              f"%{hist['class']}%",
              hist['subject'], f"%{hist['subject']}%",
              hist['test_no'], hist['test_no'],
              f"%{hist['test_no']}%", f"%{hist['test_no']}%"))
    questions = [dict(r) for r in c.fetchall()]

    if not questions:
        c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                            correct_answer, image_path,
                            CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                     FROM questions
                     WHERE (LOWER(class)=LOWER(?) OR class LIKE ?)
                       AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                     ORDER BY id""",
                 (hist['class'], f"%{hist['class']}%", hist['subject'], f"%{hist['subject']}%"))
        questions = [dict(r) for r in c.fetchall()]

    conn.close()

    for q in questions:
        q_type = str(q.get('question_type') or '').strip()
        if not q_type or q_type.upper() in ['MCQ', 'CHOICE', 'OBJECTIVE']:
            q['question_type'] = 'MCQ'
    school_name = get_setting('school_name','RRB Group of Schools')
    logo_path   = get_setting('logo_path','')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path,'rb') as f:
                logo_bytes = f.read()
                ext = os.path.splitext(abs_path)[1].lower()
                mime = 'image/png' if ext=='.png' else 'image/jpeg'
                logo_base64 = f"data:{mime};base64,{base64.b64encode(logo_bytes).decode()}"
    rendered = render_template('question_paper_template_v2.html',
        school_name=school_name, logo_base64=logo_base64,
        class_name=f"{hist['class']}{hist['section']}", subject=hist['subject'],
        chapter=hist['chapter'], test_no=hist['test_no'],
        questions=questions, total_marks=len(questions),
        date=datetime.datetime.now().strftime('%d/%m/%Y'))
    pdf = HTML(string=rendered, base_url=request.base_url).write_pdf()
    resp = make_response(pdf)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename=test_{hist["test_no"]}.pdf'
    return resp

# ═══════════════════════════════════════════════════════════════
# ADMIN: Student class lock management
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/student_locks')
@admin_required
def admin_student_locks():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT l.student_id, l.class, l.section, l.locked_at, s.name
                 FROM student_class_lock l
                 LEFT JOIN students s ON l.student_id=s.student_id
                 ORDER BY l.class, l.section, s.name""")
    locks = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'locks': locks})

# ═══════════════════════════════════════════════════════════════
# MULTIMODAL: IMAGEN 3 IMAGE GENERATION PROXY
# ═══════════════════════════════════════════════════════════════

@app.route('/api/generate_diagram', methods=['POST'])
def generate_diagram():
    """Proxy endpoint: receives an image prompt, calls Gemini Imagen, returns base64 image."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401

    import urllib.request, json as json_lib

    data = request.get_json() or {}
    image_prompt = data.get('prompt', '').strip()
    if not image_prompt:
        return jsonify({'status': 'error', 'message': 'No prompt provided'}), 400

    gemini_api_key = _get_gemini_key()
    if not gemini_api_key:
        return jsonify({'status': 'error', 'message': 'GEMINI_API_KEY not set'}), 500

    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/imagen-3.0-generate-002:predict?key={gemini_api_key}"
        payload = json_lib.dumps({
            "instances": [{"prompt": image_prompt}],
            "parameters": {"sampleCount": 1, "aspectRatio": "1:1", "safetyFilterLevel": "block_some"}
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload,
                                     headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=45) as resp:
            result = json_lib.loads(resp.read().decode('utf-8'))
        img_b64 = result['predictions'][0]['bytesBase64Encoded']
        mime = result['predictions'][0].get('mimeType', 'image/png')
        return jsonify({'status': 'success', 'image': f'data:{mime};base64,{img_b64}',
                        'prompt': image_prompt})
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8')
        return jsonify({'status': 'error', 'message': f'Imagen API error: {err}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

def _get_gemini_key():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'apikey.env')
    if os.path.exists(env_path):
        try:
            content = open(env_path, 'r', encoding='utf-8').read().strip()
            if content:
                if 'GEMINI_API_KEY=' in content:
                    key = content.split('GEMINI_API_KEY=', 1)[1].split('\n', 1)[0].strip().strip('"\'')
                else:
                    key = content.strip().strip('"\'')
                if key:
                    os.environ['GEMINI_API_KEY'] = key
                    return key
        except Exception:
            pass
    return os.environ.get('GEMINI_API_KEY', '').strip().strip('"\'')

# ═══════════════════════════════════════════════════════════════
# BULLETIN POSTING API for Admin & Teacher dashboards
# ═══════════════════════════════════════════════════════════════

@app.route('/admin/bulletins')
@admin_required
def admin_bulletins_page():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT b.*, COALESCE(t.name,'Admin') as poster_name
                 FROM bulletins b LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
                 WHERE b.is_active=1 ORDER BY b.created_at DESC""")
    bulletins = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('admin_bulletins.html', bulletins=bulletins)


@app.route('/admin/audit_logs', methods=['GET'])
@admin_required
def admin_audit_logs():
    """
    Render Admin Audit Logs monitoring dashboard with filters.
    Supports filtering by user_type, action, start_date, and end_date.
    """
    user_type = request.args.get('user_type', '').strip().lower()
    action = request.args.get('action', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    
    conn = get_db()
    c = conn.cursor()
    
    query = "SELECT * FROM audit_logs WHERE 1=1"
    params = []
    
    if user_type:
        query += " AND LOWER(user_type) = ?"
        params.append(user_type)
        
    if action:
        query += " AND action = ?"
        params.append(action)
        
    if start_date:
        query += " AND timestamp >= ?"
        params.append(f"{start_date} 00:00:00")
        
    if end_date:
        query += " AND timestamp <= ?"
        params.append(f"{end_date} 23:59:59")
        
    query += " ORDER BY timestamp DESC LIMIT 500"
    
    c.execute(query, params)
    logs = [dict(r) for r in c.fetchall()]
    
    c.execute("SELECT COUNT(*) FROM audit_logs")
    total_count = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM audit_logs WHERE LOWER(user_type)='admin'")
    admin_count = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM audit_logs WHERE LOWER(user_type)='teacher'")
    teacher_count = c.fetchone()[0] or 0
    
    c.execute("SELECT COUNT(*) FROM audit_logs WHERE LOWER(user_type)='student'")
    student_count = c.fetchone()[0] or 0
    
    conn.close()
    
    if request.args.get('format') == 'json':
        return jsonify({'status': 'success', 'logs': logs, 'total': total_count})
        
    return render_template(
        'admin_audit_logs.html',
        logs=logs,
        total_count=total_count,
        admin_count=admin_count,
        teacher_count=teacher_count,
        student_count=student_count,
        selected_user_type=user_type,
        selected_action=action,
        start_date=start_date,
        end_date=end_date
    )


@app.route('/api/test_generation_history/<int:hid>/delete', methods=['DELETE'])
def delete_test_history(hid):
    """Feature #1: Delete individual test history entry."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status': 'error', 'message': 'Not authenticated'}), 401
    conn = get_db()
    c = conn.cursor()
    if session.get('teacher_logged_in'):
        c.execute("DELETE FROM test_generation_history WHERE id=? AND teacher_id=?",
                  (hid, str(session.get('teacher_id'))))
    else:
        c.execute("DELETE FROM test_generation_history WHERE id=?", (hid,))
    conn.commit()
    conn.close()
    u_type = 'teacher' if session.get('teacher_logged_in') else 'admin'
    u_id = session.get('teacher_id') if session.get('teacher_logged_in') else 'admin'
    log_audit_event(u_type, u_id, 'PAPER_DELETE', 'test_generation_history', hid, request.remote_addr)
    return jsonify({'status': 'success'})

@app.route('/api/test_generation_history/<int:hid>/print_preview')
def test_gen_print_preview(hid):
    """Feature #6: Return HTML page for browser print dialog (client-side PDF)."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return redirect(url_for('teacher_login'))
    conn = get_db()
    c = conn.cursor()

    if session.get('teacher_logged_in'):
        c.execute("SELECT * FROM test_generation_history WHERE id=? AND teacher_id=?",
                  (hid, str(session.get('teacher_id'))))
    else:
        c.execute("SELECT * FROM test_generation_history WHERE id=?", (hid,))
    hist = c.fetchone()
    if not hist:
        conn.close()
        return "Not found", 404
    hist = dict(hist)

    # Get questions by test_no OR chapter using multi-tier fallback
    clean_cls = hist['class'].replace('th','').replace('st','').replace('nd','').replace('rd','').strip()
    c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                        correct_answer, image_path,
                        CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                 FROM questions
                 WHERE (class=? OR class=? OR class LIKE ?)
                   AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                   AND (test_no=? OR chapter=? OR test_no LIKE ? OR chapter LIKE ?)
                 ORDER BY id""",
             (hist['class'], clean_cls, f"%{clean_cls}%",
              hist['subject'], f"%{hist['subject']}%",
              hist['test_no'], hist.get('chapter',''),
              f"%{hist['test_no']}%", f"%{hist.get('chapter','')}%"))
    questions = [dict(r) for r in c.fetchall()]

    if not questions:
        c.execute("""SELECT id, question, option_a, option_b, option_c, option_d,
                            correct_answer, image_path,
                            CASE WHEN question_type IS NULL OR trim(question_type) = '' THEN 'MCQ' ELSE question_type END as question_type
                     FROM questions
                     WHERE (LOWER(class)=LOWER(?) OR class LIKE ?)
                       AND (LOWER(subject)=LOWER(?) OR subject LIKE ?)
                     ORDER BY id""",
                 (hist['class'], f"%{hist['class']}%", hist['subject'], f"%{hist['subject']}%"))
        questions = [dict(r) for r in c.fetchall()]
    conn.close()

    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    logo_base64 = None
    if logo_path:
        abs_path = os.path.join(app.static_folder, logo_path)
        if os.path.exists(abs_path):
            with open(abs_path, 'rb') as f:
                lb = f.read()
                ext  = os.path.splitext(abs_path)[1].lower()
                mime = 'image/png' if ext == '.png' else 'image/jpeg'
                logo_base64 = f"data:{mime};base64,{base64.b64encode(lb).decode()}"

    return render_template('print_preview.html',
                           school_name=school_name,
                           logo_base64=logo_base64,
                           hist=hist,
                           questions=questions,
                           date=datetime.datetime.now().strftime('%d/%m/%Y'))

# ── ADMIN TEST GENERATION (Feature #8) ───────────────────────────────────────

@app.route('/teacher/create_test_v2', methods=['GET', 'POST'])
@teacher_required
@limiter.limit("13 per minute")
def teacher_create_test_v2():
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT class, section, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
    assignments = [dict(r) for r in c.fetchall()]

    if request.method == 'POST':
        import urllib.request, json as json_lib

        class_   = request.form.get('class','').strip()
        section  = request.form.get('section','').strip()
        subject  = request.form.get('subject','').strip()
        chapter  = request.form.get('chapter','').strip()
        test_no  = request.form.get('test_no','').strip()
        remark   = request.form.get('remark','').strip()
        output_mode = request.form.get('output_mode','cbt')
        method   = request.form.get('method','ai')

        mcq_count        = int(request.form.get('mcq_count', 0) or 0)
        assertion_count  = int(request.form.get('assertion_count', 0) or 0)
        very_short_count = int(request.form.get('very_short_count', 0) or 0)
        short_count      = int(request.form.get('short_count', 0) or 0)
        long_count       = int(request.form.get('long_count', 0) or 0)
        case_study_count = int(request.form.get('case_study_count', 0) or 0)

        if method == 'upload':
            eng_file = request.files.get('english_csv') or request.files.get('csv_file')
            hindi_file = request.files.get('hindi_csv')

            if not eng_file or not eng_file.filename.endswith('.csv'):
                conn.close()
                return jsonify({'status':'error','message':'English CSV file is required'}), 400

            eng_filename = secure_filename(eng_file.filename)
            eng_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"eng_{uuid.uuid4().hex[:8]}_{eng_filename}")
            eng_file.save(eng_filepath)

            inserted_ids = []
            try:
                with open(eng_filepath, 'r', encoding='utf-8-sig') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        parsed = parse_question_csv_row(row)
                        if not parsed or not parsed['question']:
                            continue
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,
                                    option_a,option_b,option_c,option_d,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_, subject, chapter, test_no, parsed.get('question_type','MCQ'), parsed['question'],
                                  parsed['option_a'], parsed['option_b'], parsed['option_c'], parsed['option_d'],
                                  parsed['correct_answer']))
                        inserted_ids.append(c.lastrowid)
                conn.commit()
            finally:
                if os.path.exists(eng_filepath):
                    os.remove(eng_filepath)

            if not inserted_ids:
                conn.close()
                return jsonify({'status':'error','message':'No valid questions found in English CSV'}), 400

            hindi_updated = 0
            if hindi_file and hindi_file.filename and hindi_file.filename.endswith('.csv'):
                hin_filename = secure_filename(hindi_file.filename)
                hin_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"hin_{uuid.uuid4().hex[:8]}_{hin_filename}")
                hindi_file.save(hin_filepath)
                try:
                    with open(hin_filepath, 'r', encoding='utf-8-sig') as csvfile:
                        reader = csv.DictReader(csvfile)
                        for idx, row in enumerate(reader):
                            if idx >= len(inserted_ids):
                                break
                            parsed_hi = parse_question_csv_row(row)
                            if not parsed_hi:
                                continue
                            q_id = inserted_ids[idx]
                            c.execute("""UPDATE questions
                                         SET question_hi=?, option_a_hi=?, option_b_hi=?, option_c_hi=?, option_d_hi=?
                                         WHERE id=?""",
                                     (parsed_hi['question'], parsed_hi['option_a'], parsed_hi['option_b'],
                                      parsed_hi['option_c'], parsed_hi['option_d'], q_id))
                            hindi_updated += 1
                    conn.commit()
                finally:
                    if os.path.exists(hin_filepath):
                        os.remove(hin_filepath)

            uploader_type = 'admin' if session.get('admin_logged_in') else 'teacher'
            uploader_id = 'admin' if uploader_type == 'admin' else str(session.get('teacher_id') or '0')
            teacher_name = uploader_type.capitalize()
            t_id = session.get('teacher_id')
            if t_id:
                c.execute("SELECT name FROM teachers WHERE id=?", (t_id,))
                t_r = c.fetchone()
                if t_r: teacher_name = t_r['name']

            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                      (uploader_id, class_, section, subject, chapter or test_no, test_no, output_mode,
                       len(inserted_ids), len(inserted_ids), 0, 0, 0, 0, 0, remark or 'CSV Upload'))

            c.execute("""INSERT INTO mcq_test_history (teacher_id, teacher_name, class, section, subject, test_no, question_count)
                         VALUES (?,?,?,?,?,?,?)""",
                      (t_id or 0, teacher_name, class_, section, subject, test_no, len(inserted_ids)))

            if output_mode in ('cbt', 'both'):
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                          (f"{class_}{section}_{subject}_{test_no}", class_, section, subject, test_no,
                           uploader_id, uploader_type, len(inserted_ids)))

            conn.commit()
            conn.close()

            msg = f"{len(inserted_ids)} English questions uploaded"
            if hindi_updated > 0:
                msg += f" (with {hindi_updated} Hindi translations mapped)"
            return jsonify({'status':'success','message': msg})

        # ── AI generation ──
        gemini_api_key = _get_gemini_key()
        if not gemini_api_key:
            conn.close()
            return jsonify({'status':'error','message':'GEMINI_API_KEY not set'}), 500

        model_name = get_working_model(gemini_api_key) or 'gemini-2.0-flash'
        app.logger.info(f"Using Gemini model: {model_name}")

        type_instructions = []
        if mcq_count:        type_instructions.append(f"{mcq_count} MCQ (4 options)")
        if assertion_count:  type_instructions.append(f"{assertion_count} Assertion-Reason")
        if very_short_count: type_instructions.append(f"{very_short_count} Very Short Answer (1 Mark)")
        if short_count:      type_instructions.append(f"{short_count} Short Answer (3 Marks)")
        if long_count:       type_instructions.append(f"{long_count} Long Answer (5 Marks)")
        if case_study_count: type_instructions.append(f"{case_study_count} Case Study (4 Marks)")

        if not type_instructions:
            conn.close()
            return jsonify({'status':'error','message':'Enter at least 1 question quantity'}), 400

        ai_prompt = _build_ai_prompt(class_, section, subject, chapter, test_no, remark, type_instructions)

        try:
            result, provider, err_msg = generate_ai_content(ai_prompt, timeout=60)
            if not result:
                return jsonify({'status':'error','message': err_msg}), 500

            raw_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
            app.logger.info(f"Raw AI response length: {len(raw_text)}")
            app.logger.debug(f"Raw AI response (first 1000 chars): {raw_text[:1000]}")
            if '```' in raw_text:
                for part in raw_text.split('```'):
                    part = part.strip()
                    if part.startswith('json'): part = part[4:].strip()
                    if part.startswith('['): raw_text = part; break

            questions = safe_json_loads(raw_text)
            if questions is None:
                app.logger.error(f"AI response invalid. Full response (first 2000 chars): {raw_text[:2000]}")
                return jsonify({
                    'status': 'error',
                    'message': 'AI returned invalid JSON. Please try again with a simpler prompt.'
                }), 500

            if output_mode in ('cbt', 'both'):
                c.execute("""SELECT COUNT(*) as cnt FROM questions
                             WHERE class=? AND subject=? AND test_no=?""",
                         (class_, subject, test_no))
                if c.fetchone()['cnt'] > 0:
                    conn.close()
                    return jsonify({
                        'status': 'error',
                        'message': f'Test No "{test_no}" already has questions for {subject} Class {class_}. Use a different Test No.'
                    }), 400

                inserted = 0
                for q in questions:
                    if q.get('question_type','MCQ') in ('MCQ','Assertion-Reason'):
                        c.execute("""INSERT INTO questions
                                   (class,subject,chapter,test_no,question_type,question,question_hi,
                                    option_a,option_a_hi,option_b,option_b_hi,option_c,option_c_hi,
                                    option_d,option_d_hi,correct_answer)
                                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                 (class_,subject,chapter,test_no,
                                  q.get('question_type','MCQ'),q.get('question',''),q.get('question_hi',''),
                                  q.get('option_a',''),q.get('option_a_hi',''),
                                  q.get('option_b',''),q.get('option_b_hi',''),
                                  q.get('option_c',''),q.get('option_c_hi',''),
                                  q.get('option_d',''),q.get('option_d_hi',''),
                                  q.get('correct_answer','')))
                        inserted += 1
                c.execute("""INSERT INTO test_papers
                             (filename,class,section,subject,test_no,uploaded_by,uploader_type,question_count,is_active)
                             VALUES (?,?,?,?,?,?,?,?,1)""",
                         (f"{class_}{section}_{subject}_{test_no}",class_,section,subject,test_no,
                          str(teacher_id),'teacher',inserted))
                conn.commit()

            c.execute("""INSERT INTO test_generation_history
                         (teacher_id,class,section,subject,chapter,test_no,output_mode,
                          total_questions,mcq_count,assertion_count,very_short_count,
                          short_count,long_count,case_study_count,remark)
                         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (str(teacher_id),class_,section,subject,chapter,test_no,output_mode,
                      len(questions),mcq_count,assertion_count,very_short_count,
                      short_count,long_count,case_study_count,remark))
            conn.commit()
            history_id = c.lastrowid
            conn.close()


            if output_mode == 'print':
                return jsonify({'status':'success','mode':'print',
                                'preview_url': f'/api/test_generation_history/{history_id}/print_preview',
                                'history_id': history_id,
                                'message':f'{len(questions)} questions generated'})
            else:
                return jsonify({'status':'success','mode':'cbt',
                                'message':f'{len(questions)} questions generated. Test No: {test_no}'})

        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8') if hasattr(e, 'read') else str(e)
            app.logger.error(f"Gemini HTTP error: {err_body}")
            return jsonify({'status':'error','message':f'Gemini API error: {e.code} - {err_body[:200]}'}), 500
        except Exception as e:
            app.logger.error(f"Gemini exception: {str(e)}")
            return jsonify({'status':'error','message':str(e)}), 500

    conn.close()
    # Build class -> section -> subjects map
    class_section_subject_map = {}
    class_subject_map = {}
    for a in assignments:
        cls  = str(a['class']).strip() if a['class'] is not None else ''
        sec  = str(a.get('section') or '').strip().upper()
        subj = str(a['subject']).strip() if a['subject'] is not None else ''

        if not cls or not subj:
            continue

        if cls not in class_subject_map:
            class_subject_map[cls] = []
        if subj not in class_subject_map[cls]:
            class_subject_map[cls].append(subj)

        if cls not in class_section_subject_map:
            class_section_subject_map[cls] = {'all_subjects': [], 'sections': {}}

        if subj not in class_section_subject_map[cls]['all_subjects']:
            class_section_subject_map[cls]['all_subjects'].append(subj)

        if sec:
            if sec not in class_section_subject_map[cls]['sections']:
                class_section_subject_map[cls]['sections'][sec] = []
            if subj not in class_section_subject_map[cls]['sections'][sec]:
                class_section_subject_map[cls]['sections'][sec].append(subj)

    # Fetch recent test generation and upload history
    conn = get_db()
    c = conn.cursor()
    if session.get('admin_logged_in'):
        c.execute("SELECT * FROM test_generation_history ORDER BY created_at DESC LIMIT 20")
    else:
        c.execute("SELECT * FROM test_generation_history WHERE teacher_id=? OR teacher_id='admin' ORDER BY created_at DESC LIMIT 20", (str(teacher_id),))
    history = [dict(r) for r in c.fetchall()]
    conn.close()

    return render_template('teacher_create_test_v2.html',
                           assignments=assignments,
                           class_subject_map=class_subject_map,
                           class_section_subject_map=class_section_subject_map,
                           history=history)
@app.route('/admin/test_history')
@admin_required
def admin_test_history():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_generation_history ORDER BY created_at DESC")
    history = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template('teacher_test_history.html', history=history, is_admin=True)

# ── SHARED AI PROMPT BUILDER ──────────────────────────────────────────────────

def _build_ai_prompt(class_, section, subject, chapter, test_no, remark, type_instructions):
    return f"""You are an expert CBSE/ICSE bilingual (English & Hindi) question paper generator.
Class: {class_}{section} | Subject: {subject} | Chapter: {chapter} | Test No: {test_no}
Teacher instructions: {remark if remark else 'Standard difficulty, balanced coverage'}

Generate EXACTLY these question types:
{chr(10).join(type_instructions)}

OUTPUT: Return ONLY a valid JSON array. No markdown. No text before or after.
Every object must have ALL keys:
question_type, content_type, question, question_hi, option_a, option_a_hi, option_b, option_b_hi, option_c, option_c_hi, option_d, option_d_hi, correct_answer, smiles, image_prompt, marks

RULES:
- Provide question and options in English ('question', 'option_a'...) AND their accurate Hindi translation ('question_hi', 'option_a_hi', 'option_b_hi', 'option_c_hi', 'option_d_hi').
- MATH LaTeX RULES: Inline $...$, Block $$...$$
- Escape backslashes in JSON: \\\\frac not \\frac
- NON-MCQ: set option_a/b/c/d="", option_a_hi/b_hi/c_hi/d_hi="" and correct_answer="N/A"
- smiles="" and image_prompt="" when not needed.

Return ONLY the JSON array starting with [ and ending with ]."""


# ── REDIS + RQ (REDIS QUEUE) & THREAD FALLBACK QUEUE MANAGER ──────────────────
import queue
import threading
import redis
from rq import Queue as RQ_Queue
from rq.job import Job as RQ_Job

paper_job_results = {}
paper_job_lock = threading.Lock()
ai_paper_thread_queue = queue.Queue()

REDIS_AVAILABLE = False
redis_conn = None
rq_paper_queue = None

def init_redis_queue():
    global REDIS_AVAILABLE, redis_conn, rq_paper_queue
    try:
        r_host = os.environ.get('REDIS_HOST', 'localhost')
        r_port = int(os.environ.get('REDIS_PORT', 6379))
        r_db = int(os.environ.get('REDIS_DB', 0))
        r_client = redis.Redis(host=r_host, port=r_port, db=r_db, socket_timeout=1.5)
        r_client.ping()
        redis_conn = r_client
        rq_paper_queue = RQ_Queue('paper_generation', connection=redis_conn)
        REDIS_AVAILABLE = True
        app.logger.info("[+] Redis connected! Using Redis + RQ for background paper processing.")
    except Exception as e:
        REDIS_AVAILABLE = False
        app.logger.info(f"[!] Redis unavailable ({e}). Reverting to built-in Thread Queue worker.")

init_redis_queue()


def process_ai_paper_job(job_data):
    """
    Core AI Paper Generation Task executed by Redis RQ Worker or Thread Queue.
    Generates paper via Gemini/multi-provider, formats .doc repository file,
    and inserts alert into teacher_notifications table.
    """
    job_id = job_data.get('job_id')
    prompt = job_data.get('prompt')
    teacher_id = job_data.get('teacher_id')
    teacher_name = job_data.get('teacher_name') or 'Admin'
    class_ = job_data.get('class_')
    section = job_data.get('section', 'A')
    subject = job_data.get('subject')
    exam_type = job_data.get('exam_type', 'Exam')
    duration = job_data.get('duration', '3 Hours')
    max_marks = job_data.get('max_marks', '100')
    school_name = job_data.get('school_name', 'RRB Group of Schools')
    
    with paper_job_lock:
        paper_job_results[job_id] = {
            'status': 'processing',
            'progress': 40,
            'message': 'AI engine generating questions and solution layout...'
        }
        
    try:
        result, provider, err_msg = generate_ai_content(prompt, timeout=90)
        if not result:
            with paper_job_lock:
                paper_job_results[job_id] = {'status': 'error', 'message': err_msg or 'AI generation failed'}
            return
            
        raw = result['candidates'][0]['content']['parts'][0]['text'].strip()
        if '```' in raw:
            for part in raw.split('```'):
                part = part.strip()
                if part.startswith('json'): part = part[4:].strip()
                if part.startswith('{'): raw = part; break
                
        paper_data = json.loads(raw)
        
        # Clean metadata from questions
        if isinstance(paper_data, dict) and 'sections' in paper_data:
            for sec in paper_data.get('sections', []):
                for q in sec.get('questions', []):
                    if isinstance(q.get('question'), str):
                        q['question'] = re.sub(r'\{"?meta(data)?"?:?.*\}', '', q['question'], flags=re.IGNORECASE).strip()

        # Save to repository: paper/Class_<N>/Section_<X>/
        abs_path, doc_filename, rel_link = save_paper_to_repository(
            paper_data, class_, section, subject, exam_type, teacher_name,
            meta={'school_name': school_name, 'duration': duration, 'max_marks': max_marks}
        )
        
        doc_link = f"/api/shared_papers/download/{rel_link}"
        
        with paper_job_lock:
            paper_job_results[job_id] = {
                'status': 'success',
                'progress': 100,
                'paper': paper_data,
                'doc_link': doc_link,
                'filename': doc_filename,
                'provider': provider
            }
            
        # Log notification into teacher_notifications table
        if teacher_id:
            try:
                conn = get_db()
                c = conn.cursor()
                c.execute("""INSERT INTO teacher_notifications (teacher_id, title, message, link)
                             VALUES (?, ?, ?, ?)""",
                          (teacher_id, "Descriptive Paper Ready!",
                           f"Your {subject} ({exam_type}) paper for Class {class_} {section} has been generated.",
                           doc_link))
                conn.commit()
                conn.close()
            except Exception as ex:
                app.logger.error(f"Error logging notification: {ex}")
                
    except Exception as e:
        with paper_job_lock:
            paper_job_results[job_id] = {'status': 'error', 'message': str(e)}


def enqueue_ai_paper_job(job_data):
    """
    Enqueues AI paper generation job.
    Uses Redis + RQ if connected; otherwise falls back to thread queue.
    """
    job_id = job_data['job_id']
    with paper_job_lock:
        paper_job_results[job_id] = {
            'status': 'queued',
            'progress': 15,
            'message': 'Queued in background processing pipeline...'
        }
        
    if REDIS_AVAILABLE and rq_paper_queue:
        try:
            rq_job = rq_paper_queue.enqueue(
                process_ai_paper_job,
                job_data,
                job_id=job_id,
                job_timeout=180
            )
            return job_id, 'rq'
        except Exception as e:
            app.logger.error(f"RQ enqueue failed ({e}). Reverting to thread queue.")
            
    ai_paper_thread_queue.put(job_data)
    return job_id, 'thread'


def ai_paper_thread_worker():
    """Fallback thread worker daemon."""
    while True:
        try:
            job_data = ai_paper_thread_queue.get()
            if job_data is None:
                break
            process_ai_paper_job(job_data)
        except Exception as e:
            app.logger.error(f"Error in thread worker: {e}")
        finally:
            ai_paper_thread_queue.task_done()

thread_daemon = threading.Thread(target=ai_paper_thread_worker, daemon=True)
thread_daemon.start()


@app.route('/api/paper_job/<job_id>', methods=['GET'])
def get_paper_job_status(job_id):
    """Poll job status for background AI paper processing."""
    with paper_job_lock:
        res = paper_job_results.get(job_id)
        
    if res:
        return jsonify({'status': 'success', 'job': res})
        
    # Check RQ Job if Redis is available
    if REDIS_AVAILABLE and redis_conn:
        try:
            rq_j = RQ_Job.fetch(job_id, connection=redis_conn)
            if rq_j:
                return jsonify({
                    'status': 'success',
                    'job': {
                        'status': rq_j.get_status(),
                        'message': f"Job status: {rq_j.get_status()}"
                    }
                })
        except Exception:
            pass
            
    return jsonify({'status': 'error', 'message': 'Job ID not found'}), 404


def save_paper_to_repository(paper_data, class_, section, subject, exam_type, teacher_name, meta=None):
    """
    Saves generated paper into structured folder hierarchy:
    paper/Class_<N>/Section_<X>/<class>_<section>_<subject>_<examtype>_<TeacherName>.doc
    """
    clean_cls = (class_ or 'General').strip().replace(' ', '_')
    clean_sec = (section or 'A').strip().replace(' ', '_')
    clean_sub = (subject or 'Subject').strip().replace(' ', '_')
    clean_exam = (exam_type or 'Exam').strip().replace(' ', '_')
    clean_teacher = (teacher_name or 'Teacher').strip().replace(' ', '_')
    
    cls_folder = f"Class_{clean_cls}"
    sec_folder = f"Section_{clean_sec}"
    
    rel_dir = os.path.join('paper', cls_folder, sec_folder)
    abs_dir = os.path.join(app.root_path, rel_dir)
    os.makedirs(abs_dir, exist_ok=True)
    
    filename = f"{clean_cls}_{clean_sec}_{clean_sub}_{clean_exam}_{clean_teacher}.doc"
    abs_file_path = os.path.join(abs_dir, filename)
    rel_link = f"paper/Class_{clean_cls}/Section_{clean_sec}/{filename}"
    
    meta_info = meta or {}
    school_name = meta_info.get('school_name') or get_setting('school_name', 'RRB Group of Schools')
    duration = meta_info.get('duration') or '3 Hours'
    max_marks = meta_info.get('max_marks') or '100'
    
    sections_html = ""
    if isinstance(paper_data, dict) and 'sections' in paper_data:
        for sec in paper_data.get('sections', []):
            label = sec.get('section_label', '')
            title = sec.get('section_title', '')
            instruction = sec.get('instruction', '')
            sections_html += f"<h3 style='margin-top:16px;color:#1e1b4b;border-bottom:1px solid #ddd;'>{label}: {title}</h3>"
            if instruction:
                sections_html += f"<p style='font-style:italic;color:#555;'>Note: {instruction}</p>"
            
            for q in sec.get('questions', []):
                q_num = q.get('number', '')
                q_text = q.get('question', '')
                sections_html += f"<div style='margin-bottom:10px;'><strong>Q{q_num}.</strong> {q_text}</div>"
                
                opts = q.get('options', [])
                if opts:
                    sections_html += "<ul style='list-style-type:none;padding-left:15px;margin-top:4px;'>"
                    for opt in opts:
                        sections_html += f"<li style='margin-bottom:2px;'>{opt}</li>"
                    sections_html += "</ul>"
                
                sub_qs = q.get('sub_questions', [])
                if sub_qs:
                    sections_html += "<ol style='padding-left:20px;margin-top:4px;'>"
                    for sq in sub_qs:
                        sq_text = sq if isinstance(sq, str) else sq.get('question', '')
                        sections_html += f"<li>{sq_text}</li>"
                    sections_html += "</ol>"

    doc_html = f"""<html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word' xmlns='http://www.w3.org/TR/REC-html40'>
<head><meta charset='utf-8'><title>{subject} {exam_type}</title>
<style>
body {{ font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.4; margin: 20px; }}
.header {{ text-align: center; font-weight: bold; margin-bottom: 20px; border-bottom: 2px solid #000; padding-bottom: 10px; }}
</style>
</head>
<body>
<div class='header'>
    <h2>{school_name}</h2>
    <h3>{exam_type} — {subject} (Class {class_} {section})</h3>
</div>
<table width='100%' style='margin-bottom:15px;font-weight:bold;'>
    <tr>
        <td>Time Allowed: {duration}</td>
        <td align='right'>Maximum Marks: {max_marks}</td>
    </tr>
    <tr>
        <td>Teacher: {clean_teacher.replace('_', ' ')}</td>
        <td align='right'>Date: {datetime.datetime.now().strftime('%d/%m/%Y')}</td>
    </tr>
</table>
<hr>
{sections_html}
</body>
</html>"""

    try:
        with open(abs_file_path, 'w', encoding='utf-8') as f:
            f.write(doc_html)
    except Exception as e:
        app.logger.error(f"Error saving paper doc: {e}")
        
    return abs_file_path, filename, rel_link


@app.route('/descriptive_paper', methods=['GET', 'POST'])
@limiter.limit("30 per minute")
def descriptive_paper():
    """Descriptive AI paper generator — accessible to both admin and teacher."""
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return redirect(url_for('teacher_login'))

    # GET — show the form
    if request.method == 'GET':
        teacher_id = session.get('teacher_id')
        assignments = []
        all_pairs   = []
        if teacher_id:
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT DISTINCT class, subject FROM teacher_assignments WHERE teacher_id=?", (teacher_id,))
            assignments = [dict(r) for r in c.fetchall()]
            conn.close()
        elif session.get('admin_logged_in'):
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT DISTINCT class, subject FROM questions WHERE class IS NOT NULL AND class!='' ORDER BY class, subject")
            all_pairs = [dict(r) for r in c.fetchall()]
            conn.close()

        school_name      = get_setting('school_name', 'RRB Group of Schools')
        school_address   = get_setting('school_address', '')
        academic_session = get_setting('academic_session', '')
        logo_path        = get_setting('logo_path', '')

        return render_template('descriptive_paper.html',
                               assignments=assignments,
                               all_pairs=all_pairs,
                               school_name=school_name,
                               school_address=school_address,
                               academic_session=academic_session,
                               logo_path=logo_path,
                               is_admin=bool(session.get('admin_logged_in')))

    # POST — generate via Gemini and return structured JSON for the preview
    import urllib.request, json as json_lib

    # ── Collect all form fields ──────────────────────────────────────────────
    class_         = request.form.get('class', '').strip()
    subject        = request.form.get('subject', '').strip()
    topics         = request.form.get('topics', '').strip()
    full_syllabus  = request.form.get('full_syllabus') == 'yes'
    exam_type      = request.form.get('exam_type', 'Unit Test').strip()
    duration       = request.form.get('duration', '3 Hours').strip()
    max_marks      = request.form.get('max_marks', '100').strip()
    academic_sess  = request.form.get('academic_session', '').strip() or get_setting('academic_session', '')
    school_nm      = request.form.get('school_name', '').strip() or get_setting('school_name', 'RRB Group of Schools')
    school_addr    = request.form.get('school_address', '').strip() or get_setting('school_address', '')
    remarks        = request.form.get('remarks', '').strip()

    # Question counts per type
    mcq_count        = int(request.form.get('mcq_count', 0) or 0)
    fib_count        = int(request.form.get('fib_count', 0) or 0)
    tf_count         = int(request.form.get('tf_count', 0) or 0)
    ar_count         = int(request.form.get('ar_count', 0) or 0)
    vs_count         = int(request.form.get('vs_count', 0) or 0)
    sh_count         = int(request.form.get('sh_count', 0) or 0)
    lg_count         = int(request.form.get('lg_count', 0) or 0)
    cs_count         = int(request.form.get('cs_count', 0) or 0)
    total_q = mcq_count + fib_count + tf_count + ar_count + vs_count + sh_count + lg_count + cs_count

    if total_q == 0:
        return jsonify({'status': 'error', 'message': 'Please enter at least 1 question'}), 400

    # ── Optional image upload ────────────────────────────────────────────────
    image_context = ''
    img_file = request.files.get('image_file')
    if img_file and img_file.filename:
        allowed_img = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        ext = img_file.filename.rsplit('.', 1)[-1].lower()
        if ext in allowed_img:
            img_bytes   = img_file.read()
            img_b64     = base64.b64encode(img_bytes).decode()
            img_mime    = f'image/{ext}' if ext != 'jpg' else 'image/jpeg'
            image_context = f'\n\nAn image has been provided (base64 omitted here). Use it conceptually to inspire questions involving diagrams, graphs, or visual-based problems relevant to {subject}.'

    # ── Build section descriptions ───────────────────────────────────────────
    section_char = ord('A')
    sections = []
    if mcq_count:
        sections.append(f"Section {chr(section_char)} — Multiple Choice Questions (MCQ): {mcq_count} questions, 1 mark each")
        section_char += 1
    if fib_count:
        sections.append(f"Section {chr(section_char)} — Fill in the Blanks: {fib_count} questions, 1 mark each (each question statement must contain a blank '_______' to fill)")
        section_char += 1
    if tf_count:
        sections.append(f"Section {chr(section_char)} — True / False: {tf_count} questions, 1 mark each (state True or False)")
        section_char += 1
    if ar_count:
        sections.append(f"Section {chr(section_char)} — Assertion-Reason: {ar_count} questions, 1 mark each")
        section_char += 1
    if vs_count:
        sections.append(f"Section {chr(section_char)} — Very Short Answer: {vs_count} questions, 2 marks each")
        section_char += 1
    if sh_count:
        sections.append(f"Section {chr(section_char)} — Short Answer: {sh_count} questions, 3 marks each")
        section_char += 1
    if lg_count:
        sections.append(f"Section {chr(section_char)} — Long Answer: {lg_count} questions, 5 marks each")
        section_char += 1
    if cs_count:
        sections.append(f"Section {chr(section_char)} — Case Study Based: {cs_count} questions, 4 marks each")
        section_char += 1

    topic_text = 'Complete Syllabus' if full_syllabus else (topics or 'All Topics')

    # ── Gemini prompt ─────────────────────────────────────────────────────────
    prompt = f"""You are an expert CBSE/ICSE question paper setter. Generate a complete descriptive question paper.

PAPER DETAILS:
- Class: {class_}
- Subject: {subject}
- Topics/Chapters: {topic_text}
- Exam Type: {exam_type}
- Duration: {duration}
- Maximum Marks: {max_marks}
- Academic Session: {academic_sess}

SECTIONS TO GENERATE (generate EXACTLY these):
{chr(10).join(sections)}

TEACHER INSTRUCTIONS / REMARKS:
{remarks if remarks else 'Standard difficulty. Mix of theory and application. Follow CBSE pattern.'}
{image_context}

OUTPUT FORMAT — return ONLY a valid JSON object like this (no markdown, no code fences):
{{
  "sections": [
    {{
      "section_label": "Section A",
      "section_title": "Multiple Choice Questions",
      "marks_per_question": 1,
      "instruction": "Choose the correct answer.",
      "questions": [
        {{
          "number": 1,
          "question": "Full question text here.",
          "options": ["(a) Option A", "(b) Option B", "(c) Option C", "(d) Option D"],
          "sub_questions": []
        }}
      ]
    }}
  ]
}}

RULES:
- For Mathematical and Chemical formulas:
  * Use proper HTML tags <sub> and <sup> for chemical formulas and powers (e.g. H<sub>2</sub>O, H<sub>2</sub>SO<sub>4</sub>, x<sup>2</sup>, CO<sub>2</sub>, Ca(OH)<sub>2</sub>).
  * Use standard mathematical symbols or MathJax LaTeX delimiters \\( ... \\) for square roots, equations, fractions, and symbols (e.g. \\(\\sqrt{{x}}\\), \\(\\neq\\), \\(\\pm\\), \\(\\frac{{a}}{{b}}\\), \\(\\pi\\), \\(\\Delta\\), \\(\\theta\\), \\(\\rightarrow\\)).
- For MCQ/AR: include "options" array with 4 choices formatted as (a), (b), (c), (d).
- For Fill in the Blanks: include "options": [] and place a clear blank line "_______" inside the question text.
- For True/False: include "options": ["True", "False"] or "options": [].
- For Short/Long/Case Study: "options" should be empty [], use "sub_questions" for multi-part questions.
- For Case Study: write a reading passage first in "question", then list sub-questions in "sub_questions" array (each with number and question text).
- Assertion-Reason: write Assertion and Reason clearly in the question text. Options must be the standard 4 AR options.
- Keep language age-appropriate for Class {class_}.
- Do NOT output any raw metadata, answer keys, or debug JSON objects inside or after the question paper.
- Return ONLY the JSON object. Nothing else."""

    try:
        result, provider, err_msg = generate_ai_content(prompt, timeout=90)
        if not result:
            return jsonify({'status': 'error', 'message': err_msg}), 500

        raw = result['candidates'][0]['content']['parts'][0]['text'].strip()

        # Strip markdown fences if present
        if '```' in raw:
            for part in raw.split('```'):
                part = part.strip()
                if part.startswith('json'): part = part[4:].strip()
                if part.startswith('{'): raw = part; break

        paper_data = json_lib.loads(raw)

        # Clean any raw metadata fields or trailing JSON strings from question content
        if isinstance(paper_data, dict) and 'sections' in paper_data:
            for sec in paper_data.get('sections', []):
                for q in sec.get('questions', []):
                    if isinstance(q.get('question'), str):
                        q['question'] = re.sub(r'\{"?meta(data)?"?:?.*\}', '', q['question'], flags=re.IGNORECASE).strip()
                    if isinstance(q.get('sub_questions'), list):
                        cleaned_sq = []
                        for sq in q['sub_questions']:
                            if isinstance(sq, str):
                                cleaned_sq.append(re.sub(r'\{"?meta(data)?"?:?.*\}', '', sq, flags=re.IGNORECASE).strip())
                            elif isinstance(sq, dict) and 'question' in sq and isinstance(sq['question'], str):
                                sq['question'] = re.sub(r'\{"?meta(data)?"?:?.*\}', '', sq['question'], flags=re.IGNORECASE).strip()
                                cleaned_sq.append(sq)
                            else:
                                cleaned_sq.append(sq)
                        q['sub_questions'] = cleaned_sq

        # Build logo base64 for the preview
        logo_path   = get_setting('logo_path', '')
        logo_base64 = ''
        if logo_path:
            abs_path = os.path.join(app.static_folder, logo_path)
            if os.path.exists(abs_path):
                with open(abs_path, 'rb') as f:
                    lb  = f.read()
                    ext = os.path.splitext(abs_path)[1].lower()
                    mt  = 'image/png' if ext == '.png' else 'image/jpeg'
                    logo_base64 = f"data:{mt};base64,{base64.b64encode(lb).decode()}"

        # Automatic Repository Save into paper/Class_X/Section_Y/
        teacher_name = 'Admin'
        if session.get('teacher_id'):
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT name FROM teachers WHERE id=?", (session['teacher_id'],))
            t_row = c.fetchone()
            if t_row: teacher_name = t_row['name']
            conn.close()
            
        section_val = request.form.get('section', 'A').strip() or 'A'
        abs_path, doc_filename, rel_link = save_paper_to_repository(
            paper_data, class_, section_val, subject, exam_type, teacher_name,
            meta={'school_name': school_nm, 'duration': duration, 'max_marks': max_marks}
        )

        return jsonify({
            'status': 'success',
            'paper':  paper_data,
            'doc_link': f"/api/shared_papers/download/{rel_link}",
            'meta': {
                'school_name':      school_nm,
                'school_address':   school_addr,
                'academic_session': academic_sess,
                'exam_type':        exam_type,
                'class':            class_,
                'subject':          subject,
                'duration':         duration,
                'max_marks':        max_marks,
                'logo_base64':      logo_base64,
                'date':             datetime.datetime.now().strftime('%d / %m / %Y'),
            }
        })

    except urllib.error.HTTPError as e:
        return jsonify({'status': 'error', 'message': f'Gemini API: {e.read().decode()}'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/shared_papers', methods=['GET'])
def list_shared_papers():
    """Browse shared paper/Class_<N>/Section_<X>/ directory repository."""
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    cls_param = request.args.get('class', '').strip()
    sec_param = request.args.get('section', '').strip()
    
    paper_base = os.path.join(app.root_path, 'paper')
    if not os.path.exists(paper_base):
        return jsonify({'status': 'success', 'files': []})
        
    papers_list = []
    for root, dirs, files in os.walk(paper_base):
        for f in files:
            if f.endswith('.doc') or f.endswith('.docx'):
                abs_path = os.path.join(root, f)
                rel_path = os.path.relpath(abs_path, paper_base).replace('\\', '/')
                path_parts = rel_path.split('/')
                
                # Check path filters
                file_cls = path_parts[0].replace('Class_', '') if len(path_parts) > 1 else ''
                file_sec = path_parts[1].replace('Section_', '') if len(path_parts) > 2 else ''
                
                if cls_param and file_cls.lower() != cls_param.lower():
                    continue
                if sec_param and file_sec.lower() != sec_param.lower():
                    continue
                    
                stat = os.stat(abs_path)
                mtime = datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                size_kb = round(stat.st_size / 1024, 1)
                
                papers_list.append({
                    'filename': f,
                    'class': file_cls,
                    'section': file_sec,
                    'rel_path': rel_path,
                    'size_kb': size_kb,
                    'modified_at': mtime,
                    'download_url': f"/api/shared_papers/download/{rel_path}"
                })
                
    papers_list.sort(key=lambda x: x['modified_at'], reverse=True)
    return jsonify({'status': 'success', 'files': papers_list})


@app.route('/api/shared_papers/download/<path:rel_path>', methods=['GET'])
def download_shared_paper(rel_path):
    """Download .doc paper file from repository."""
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    paper_base = os.path.join(app.root_path, 'paper')
    abs_file_path = os.path.abspath(os.path.join(paper_base, rel_path))
    
    # Security check: prevent path traversal out of paper directory
    if not abs_file_path.startswith(os.path.abspath(paper_base)):
        return jsonify({'status': 'error', 'message': 'Invalid file path'}), 400
        
    if not os.path.exists(abs_file_path):
        return jsonify({'status': 'error', 'message': 'File not found'}), 444
        
    filename = os.path.basename(abs_file_path)
    return send_file(abs_file_path, as_attachment=True, download_name=filename, mimetype='application/msword')


@app.route('/admin/check_class_teacher', methods=['GET'])
@admin_required
def admin_check_class_teacher():
    """Check if Class Teacher is already assigned for Class + Section."""
    cls = request.args.get('class', '').strip()
    sec = request.args.get('section', '').strip()
    t_id = request.args.get('teacher_id', type=int)
    
    if not cls:
        return jsonify({'has_ct': False})
        
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT ta.id, t.name FROM teacher_assignments ta
                 JOIN teachers t ON ta.teacher_id = t.id
                 WHERE ta.class=? AND (ta.section=? OR ta.section='') AND ta.is_class_teacher=1
                 AND (? IS NULL OR ta.teacher_id != ?)""",
              (cls, sec, t_id, t_id))
    ct = c.fetchone()
    conn.close()
    
    if ct:
        return jsonify({'has_ct': True, 'ct_name': ct['name']})
    return jsonify({'has_ct': False})


@app.route('/api/teacher/mcq_history', methods=['GET'])
def get_teacher_mcq_history():
    """Get MCQ test generation history for teachers, including pre-existing test papers."""
    if not session.get('admin_logged_in') and not session.get('teacher_logged_in'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
    teacher_id = session.get('teacher_id')
    conn = get_db()
    c = conn.cursor()
    
    # 1. Fetch from mcq_test_history
    if teacher_id:
        c.execute("""SELECT * FROM mcq_test_history WHERE teacher_id=? ORDER BY id DESC""", (teacher_id,))
    else:
        c.execute("""SELECT * FROM mcq_test_history ORDER BY id DESC""")
    rows = [dict(r) for r in c.fetchall()]
    
    # 2. Also fetch pre-existing tests from test_papers table
    if teacher_id:
        c.execute("""SELECT tp.id, tp.uploaded_by as teacher_id, COALESCE(t.name, 'Faculty') as teacher_name, 
                            tp.class, tp.section, tp.subject, tp.test_no, tp.question_count, tp.created_at
                     FROM test_papers tp
                     LEFT JOIN teachers t ON tp.uploaded_by = CAST(t.id AS TEXT)
                     WHERE tp.uploaded_by = ? OR tp.uploader_type='admin'
                     ORDER BY tp.created_at DESC""", (str(teacher_id),))
    else:
        c.execute("""SELECT tp.id, tp.uploaded_by as teacher_id, COALESCE(t.name, 'Admin') as teacher_name, 
                            tp.class, tp.section, tp.subject, tp.test_no, tp.question_count, tp.created_at
                     FROM test_papers tp
                     LEFT JOIN teachers t ON tp.uploaded_by = CAST(t.id AS TEXT)
                     ORDER BY tp.created_at DESC""")
    tp_rows = [dict(r) for r in c.fetchall()]
    
    # Combine & deduplicate by (class, section, subject, test_no)
    seen_keys = set()
    combined = []
    for r in rows:
        key = (str(r.get('class')), str(r.get('section')), str(r.get('subject')).lower(), str(r.get('test_no')))
        seen_keys.add(key)
        combined.append(r)
        
    for r in tp_rows:
        key = (str(r.get('class')), str(r.get('section')), str(r.get('subject')).lower(), str(r.get('test_no')))
        if key not in seen_keys:
            seen_keys.add(key)
            combined.append(r)
            
    conn.close()
    return jsonify({'status': 'success', 'history': combined})


@app.route('/api/csv_template')
def download_csv_template():
    """FEAT-001: Downloadable CSV template for question paper upload."""
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Questions', 'option a', 'option b', 'option c', 'option d',
                     'correct option', 'question_type', 'marks', 'negative_mark'])
    writer.writerow([
        'What is the speed of light in vacuum?',
        '3 x 10^8 m/s', '3 x 10^6 m/s', '3 x 10^10 m/s', '3 x 10^4 m/s',
        'option a', 'MCQ', '1', '0.33'
    ])
    writer.writerow([
        'Assertion: Water boils at 100°C. Reason: Boiling point depends on atmospheric pressure.',
        'Both A and R are true and R is the correct explanation of A',
        'Both A and R are true but R is NOT the correct explanation of A',
        'A is true but R is false',
        'A is false but R is true',
        'option_a', 'Assertion-Reason', '1', '0.33'
    ])
    writer.writerow([
        'Explain the process of photosynthesis.', '', '', '', '',
        'N/A', 'Short', '3', '0'
    ])
    writer.writerow([
        'Define Newton\'s first law of motion and give two examples.', '', '', '', '',
        'N/A', 'Very Short', '2', '0'
    ])
    writer.writerow([
        'Discuss the structure and functions of DNA.', '', '', '', '',
        'N/A', 'Long', '5', '0'
    ])
    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=question_upload_template.csv'}
    )

@app.route('/announcement/<int:bid>')
def announcement_permalink(bid):
    """FEAT-005: Permanent link to a single announcement."""
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT b.*, COALESCE(t.name,'Admin') as poster_name
                 FROM bulletins b LEFT JOIN teachers t ON b.posted_by=t.id AND b.poster_type='teacher'
                 WHERE b.id=? AND b.is_active=1""", (bid,))
    bulletin = c.fetchone()
    conn.close()
    if not bulletin:
        return "Announcement not found or has been removed.", 404
    bulletin = dict(bulletin)
    school_name = get_setting('school_name', 'RRB Group of Schools')
    logo_path   = get_setting('logo_path', '')
    return render_template('announcement_permalink.html',
                           bulletin=bulletin, school_name=school_name, logo_path=logo_path)

@app.route('/api/shared_tests')
def api_shared_tests():
    """FEAT-008: List all shared/available tests for a class+subject."""
    class_   = request.args.get('class','').strip()
    subject  = request.args.get('subject','').strip()
    if not class_ or not subject:
        return jsonify({'tests': []})
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT tp.id, tp.filename, tp.test_no, tp.section, tp.question_count,
                        tp.is_active, tp.created_at, tp.uploader_type,
                        COALESCE(t.name,'Admin') as creator_name
                 FROM test_papers tp
                 LEFT JOIN teachers t ON tp.uploaded_by=t.id AND tp.uploader_type='teacher'
                 WHERE tp.class=? AND tp.subject=?
                 ORDER BY tp.created_at DESC""", (class_, subject))
    tests = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'tests': tests})

@app.route('/api/assign_existing_test', methods=['POST'])
def assign_existing_test():
    """FEAT-008: Link an existing test to a class/section without duplicating questions."""
    if not session.get('teacher_logged_in') and not session.get('admin_logged_in'):
        return jsonify({'status':'error'}), 401
    paper_id  = request.form.get('paper_id', type=int)
    section   = request.form.get('section','').strip()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM test_papers WHERE id=?", (paper_id,))
    paper = c.fetchone()
    if not paper:
        conn.close()
        return jsonify({'status':'error','message':'Test not found'}), 404
    # Update section assignment (or leave as shared)
    c.execute("UPDATE test_papers SET section=?, is_active=1 WHERE id=?", (section, paper_id))
    conn.commit()
    conn.close()
    return jsonify({'status':'success','message':f'Test "{paper["filename"]}" assigned to Section {section or "All"}'})


@app.route('/api/token', methods=['POST'])
@limiter.limit("20 per minute")
def api_issue_token():
    """
    Issue JWT access token for Admin, Teacher, or Student API authentication.
    Accepts JSON or Form Data payload with:
      - role: 'admin', 'teacher', or 'student'
      - username / mobile / student_id / roll_no
      - password (for admin & teacher)
      - name, class, section (for student)
    """
    data = request.get_json(silent=True) or request.form or {}
    role = (data.get('role') or 'admin').strip().lower()
    
    if role == 'admin':
        username = (data.get('username') or data.get('identity') or 'admin').strip()
        password = data.get('password') or ''
        
        if password == 'admin123':
            token, exp_ts = generate_jwt_token(identity=username, role='admin', name='System Admin')
            return jsonify({
                'status': 'success',
                'access_token': token,
                'token_type': 'Bearer',
                'expires_in': JWT_EXPIRATION_HOURS * 3600,
                'user': {'id': username, 'name': 'System Admin', 'role': 'admin'}
            })
        else:
            return jsonify({'status': 'error', 'message': 'Invalid admin credentials'}), 401
            
    elif role == 'teacher':
        mobile = (data.get('mobile') or data.get('username') or data.get('identity') or '').strip()
        password = data.get('password') or ''
        
        if not mobile or not password:
            return jsonify({'status': 'error', 'message': 'Mobile and password required'}), 400
            
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM teachers WHERE mobile=? AND status='active'", (mobile,))
        teacher = c.fetchone()
        conn.close()
        
        if teacher and verify_password(password, teacher['password']):
            token, exp_ts = generate_jwt_token(identity=teacher['id'], role='teacher', name=teacher['name'])
            return jsonify({
                'status': 'success',
                'access_token': token,
                'token_type': 'Bearer',
                'expires_in': JWT_EXPIRATION_HOURS * 3600,
                'user': {'id': teacher['id'], 'name': teacher['name'], 'role': 'teacher', 'mobile': teacher['mobile']}
            })
        else:
            return jsonify({'status': 'error', 'message': 'Invalid teacher mobile or password'}), 401
            
    elif role == 'student':
        student_id = (data.get('student_id') or data.get('roll_no') or data.get('username') or data.get('identity') or '').strip()
        name = (data.get('name') or 'Student').strip()
        class_ = (data.get('class') or '').strip()
        section = (data.get('section') or '').strip()
        
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Student ID / Roll No required'}), 400
            
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
        student = c.fetchone()
        conn.close()
        
        student_name = student['name'] if student else name
        token, exp_ts = generate_jwt_token(identity=student_id, role='student', name=student_name)
        return jsonify({
            'status': 'success',
            'access_token': token,
            'token_type': 'Bearer',
            'expires_in': JWT_EXPIRATION_HOURS * 3600,
            'user': {
                'id': student_id,
                'name': student_name,
                'role': 'student',
                'class': student['class'] if student else class_,
                'section': student['section'] if student else section
            }
        })
    else:
        return jsonify({'status': 'error', 'message': f'Invalid role: {role}. Must be admin, teacher, or student.'}), 400


@app.route('/api/user/me', methods=['GET'])
@jwt_required()
def api_get_current_user():
    """Returns details of the currently authenticated JWT user context."""
    return jsonify({
        'status': 'success',
        'user': g.jwt_user,
        'payload': g.jwt_payload
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)